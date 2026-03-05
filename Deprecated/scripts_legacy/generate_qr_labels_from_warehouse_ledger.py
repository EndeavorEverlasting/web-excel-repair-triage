#!/usr/bin/env python3
"""
Offline QR generator for the Warehouse Ledger workbook.

Reads:
- PalletIDs from:      10_Warehouse_Ledger (header: PalletID)
- LocationCodes from:  11_Warehouse_Locations (header: LocationCode; fallback: column A)

Outputs:
- <out>/pallets/<PalletID>.png
- <out>/locations/<LocationCode>.png
- <out>/manifest.csv

Dependencies (pick one QR backend):
  pip install segno pillow openpyxl
or
  pip install qrcode[pil] openpyxl
"""

import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl

_QR_BACKEND = None
try:
    import segno  # type: ignore
    _QR_BACKEND = "segno"
except Exception:
    try:
        import qrcode  # type: ignore
        _QR_BACKEND = "qrcode"
    except Exception:
        _QR_BACKEND = None


def sanitize_filename(s: str) -> str:
    s = str(s).strip()
    s = re.sub(r'[\\/:*?"<>|]+', "_", s)
    s = re.sub(r"\s+", " ", s)
    return s


def find_header_col(ws, candidates):
    headers = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        headers[str(cell.value).strip().lower()] = cell.column
    for h in candidates:
        if h.lower() in headers:
            return headers[h.lower()]
    return None


def read_col(ws, col_idx, max_rows=20000):
    vals = []
    for r in range(2, min(ws.max_row + 1, max_rows + 1)):
        v = ws.cell(row=r, column=col_idx).value
        if v is None:
            continue
        s = str(v).strip()
        if not s or s.upper() == "0":
            continue
        vals.append(s)
    return vals


def gen_png(payload: str, out_png: Path, scale: int, border: int):
    out_png.parent.mkdir(parents=True, exist_ok=True)

    if _QR_BACKEND == "segno":
        qr = segno.make(payload, error="m")
        qr.save(str(out_png), kind="png", scale=scale, border=border)
        return

    if _QR_BACKEND == "qrcode":
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=scale,
            border=border,
        )
        qr.add_data(payload)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        img.save(str(out_png))
        return

    raise RuntimeError(
        "No QR backend installed. Install one:\n"
        "  pip install segno pillow openpyxl\n"
        "or\n"
        "  pip install qrcode[pil] openpyxl\n"
    )


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Warehouse Ledger .xlsx path")
    ap.add_argument("--out", default="./qr_out", help="Output directory")
    ap.add_argument("--scale", type=int, default=6, help="QR scale (default 6)")
    ap.add_argument("--border", type=int, default=2, help="QR border (default 2)")
    ap.add_argument("--limit", type=int, default=0, help="Limit count per type (0 = no limit)")
    args = ap.parse_args()

    xlsx = Path(args.xlsx).expanduser().resolve()
    out = Path(args.out).expanduser().resolve()
    if not xlsx.exists():
        print(f"ERROR: workbook not found: {xlsx}")
        return 2
    if _QR_BACKEND is None:
        print("ERROR: no QR backend installed.")
        print("Install one:\n  pip install segno pillow openpyxl\n  pip install qrcode[pil] openpyxl")
        return 2

    wb = openpyxl.load_workbook(xlsx, data_only=True)

    if "10_Warehouse_Ledger" not in wb.sheetnames:
        print("ERROR: missing sheet 10_Warehouse_Ledger")
        return 2
    ws_p = wb["10_Warehouse_Ledger"]
    col_pid = find_header_col(ws_p, ["PalletID", "Pallet ID", "Pallet Id"])
    if not col_pid:
        print("ERROR: PalletID header not found in 10_Warehouse_Ledger")
        return 2
    pallets = read_col(ws_p, col_pid)
    if args.limit and len(pallets) > args.limit:
        pallets = pallets[: args.limit]

    if "11_Warehouse_Locations" not in wb.sheetnames:
        print("ERROR: missing sheet 11_Warehouse_Locations")
        return 2
    ws_l = wb["11_Warehouse_Locations"]
    col_loc = find_header_col(ws_l, ["LocationCode", "Location Code", "Location Code (standard)"])
    if not col_loc:
        col_loc = 1  # fallback
    locs = read_col(ws_l, col_loc)
    if args.limit and len(locs) > args.limit:
        locs = locs[: args.limit]

    out.mkdir(parents=True, exist_ok=True)
    pallets_dir = out / "pallets"
    locs_dir = out / "locations"
    manifest = out / "manifest.csv"

    rows = []
    for pid in pallets:
        fn = sanitize_filename(pid) + ".png"
        p = pallets_dir / fn
        gen_png(pid, p, args.scale, args.border)
        rows.append(["pallet", pid, str(p)])

    for loc in locs:
        fn = sanitize_filename(loc) + ".png"
        p = locs_dir / fn
        gen_png(loc, p, args.scale, args.border)
        rows.append(["location", loc, str(p)])

    with manifest.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["type", "payload", "file_path"])
        w.writerows(rows)

    print(f"Done. Wrote {len(rows)} QR images. Manifest: {manifest}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
