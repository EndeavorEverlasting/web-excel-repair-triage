"""Rebuild billing-facing summaries from an admin context workbook.

This is the reverse direction from the 2026-05-20 reconciliation workflow:

    hours-tracker-safe admin context -> billing-facing summary

The script reads the tracker import tab, groups numeric carried hours, preserves
Friday batch posture, and writes a compact billing summary workbook. Blank-hour
OOO/context-only rows are tracked as cleared context and are not included in the
numeric billing total.
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from triage.admin_billing_context_rules import (
    APPROVED_EXCEPTION_SUMMARY,
    FRAMING_LINE,
    friday_batch_for,
    validate_context_row,
)

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=16)


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("_", " ")


def _header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[1]:
        key = _normalize_header(cell.value)
        if key:
            headers[key] = cell.column
    return headers


def _first_existing(headers: dict[str, int], names: list[str]) -> int | None:
    for name in names:
        col = headers.get(_normalize_header(name))
        if col:
            return col
    return None


def read_tracker_rows(source_path: Path, sheet_name: str = "02 Tracker Import") -> list[dict[str, Any]]:
    wb = load_workbook(source_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Missing required sheet: {sheet_name}")

    ws = wb[sheet_name]
    headers = _header_map(ws)

    date_col = _first_existing(headers, ["Date", "Work Date"])
    tech_col = _first_existing(headers, ["Technician", "Tech", "Name"])
    category_col = _first_existing(headers, ["Billing Category", "Category"])
    hours_col = _first_existing(headers, ["Hours", "Carried Hours", "Billing Hours"])
    status_col = _first_existing(headers, ["Hours Status", "Status"])
    action_col = _first_existing(headers, ["Admin Action", "Action"])
    note_col = _first_existing(headers, ["Safe Billing Note", "Reviewed Comment", "Note"])

    required = {
        "date": date_col,
        "technician": tech_col,
        "billing_category": category_col,
        "hours": hours_col,
    }
    missing = [name for name, col in required.items() if col is None]
    if missing:
        raise ValueError(f"Tracker import is missing required columns: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    for idx in range(2, ws.max_row + 1):
        work_date = ws.cell(idx, date_col).value
        technician = ws.cell(idx, tech_col).value
        category = ws.cell(idx, category_col).value
        hours = ws.cell(idx, hours_col).value
        if not any([work_date, technician, category, hours]):
            continue

        row = {
            "row_number": idx,
            "date": work_date,
            "technician": technician,
            "billing_category": category,
            "hours": hours,
            "status": ws.cell(idx, status_col).value if status_col else "",
            "admin_action": ws.cell(idx, action_col).value if action_col else "",
            "safe_note": ws.cell(idx, note_col).value if note_col else "",
        }
        rows.append(row)

    return rows


def write_summary(rows: list[dict[str, Any]], output_path: Path) -> None:
    totals_by_category: dict[str, float] = defaultdict(float)
    totals_by_tech: dict[str, float] = defaultdict(float)
    totals_by_friday: dict[str, float] = defaultdict(float)
    validation_issues = []
    cleared_blank_rows = 0

    for row in rows:
        issues = validate_context_row(
            row_number=row.get("row_number"),
            work_date=row.get("date"),
            hours=row.get("hours"),
            status=row.get("status"),
            admin_action=row.get("admin_action"),
            safe_note=row.get("safe_note"),
        )
        validation_issues.extend(issues)

        try:
            hours = float(row.get("hours") or 0)
        except (TypeError, ValueError):
            hours = 0.0

        if hours <= 0:
            cleared_blank_rows += 1
            continue

        category = str(row.get("billing_category") or "Uncategorized")
        technician = str(row.get("technician") or "Unassigned")
        batch = friday_batch_for(row.get("date")).isoformat()
        totals_by_category[category] += hours
        totals_by_tech[technician] += hours
        totals_by_friday[batch] += hours

    wb = Workbook()
    ws = wb.active
    ws.title = "Billing Summary"
    ws["A1"] = "Billing Summary Rebuilt from Admin Context"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = FRAMING_LINE
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws["A3"] = APPROVED_EXCEPTION_SUMMARY

    total = sum(totals_by_category.values())
    ws["A5"] = "Reviewed carried hours"
    ws["B5"] = total
    ws["B5"].number_format = "0.00"
    ws["A6"] = "Cleared blank-hour context rows"
    ws["B6"] = cleared_blank_rows
    ws["A7"] = "Validation issues"
    ws["B7"] = len(validation_issues)

    def write_table(start_row: int, title: str, items: dict[str, float]) -> int:
        ws.cell(start_row, 1, title).font = Font(bold=True, size=13)
        ws.cell(start_row + 1, 1, "Name")
        ws.cell(start_row + 1, 2, "Hours")
        for cell in ws[start_row + 1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        row_idx = start_row + 2
        for name, hours in sorted(items.items()):
            ws.cell(row_idx, 1, name)
            ws.cell(row_idx, 2, hours).number_format = "0.00"
            row_idx += 1
        return row_idx + 2

    next_row = write_table(10, "Hours by Billing Category", totals_by_category)
    next_row = write_table(next_row, "Hours by Technician", totals_by_tech)
    next_row = write_table(next_row, "Hours by Friday Batch", totals_by_friday)

    ws.cell(next_row, 1, "Validation Issues").font = Font(bold=True, size=13)
    ws.cell(next_row + 1, 1, "Severity")
    ws.cell(next_row + 1, 2, "Row")
    ws.cell(next_row + 1, 3, "Issue")
    ws.cell(next_row + 1, 4, "Message")
    for cell in ws[next_row + 1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row_idx = next_row + 2
    for issue in validation_issues:
        ws.cell(row_idx, 1, issue.severity)
        ws.cell(row_idx, 2, issue.row_number)
        ws.cell(row_idx, 3, issue.issue_code)
        ws.cell(row_idx, 4, issue.message)
        row_idx += 1

    for col_letter, width in {"A": 34, "B": 16, "C": 28, "D": 60}.items():
        ws.column_dimensions[col_letter].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def rebuild_billing_summary(source_path: Path, output_path: Path) -> None:
    rows = read_tracker_rows(source_path)
    write_summary(rows, output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rebuild_billing_summary(args.source, args.output)


if __name__ == "__main__":
    main()
