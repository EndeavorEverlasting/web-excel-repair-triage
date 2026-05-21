"""Build an admin-facing context workbook from a reviewed billing summary.

This is the forward direction from the 2026-05-20 reconciliation workflow:

    updated billing summary -> hours-tracker-safe admin context

The script keeps the 04 QC Pipeline tab internal by default. The exported admin
submission should contain only tabs 01-03.

This script is intentionally conservative. It copies clean source tabs when they
already exist and applies posture language plus tab filtering. Richer row mapping
can be layered in once the source workbook schemas stabilize.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from triage.admin_billing_context_rules import (
    ADMIN_SUBMISSION_TABS,
    APPROVED_EXCEPTION_SUMMARY,
    FRAMING_LINE,
    INTERNAL_ONLY_TABS,
)

TITLE_FILL = PatternFill("solid", fgColor="1F2937")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=16)
POSTURE_FONT = Font(italic=True, size=11)


def apply_submission_posture(ws) -> None:
    """Apply the approved 2026-05-20 framing to the top of a worksheet."""

    ws["A1"] = ws["A1"].value or "April 2026 Admin Billing Context Control"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = FRAMING_LINE
    ws["A2"].font = POSTURE_FONT
    ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)


def remove_internal_tabs(wb) -> None:
    """Remove all tabs outside the admin-facing submission set."""

    for sheet_name in list(wb.sheetnames):
        if sheet_name in INTERNAL_ONLY_TABS or sheet_name not in ADMIN_SUBMISSION_TABS:
            wb.remove(wb[sheet_name])


def patch_summary_language(wb) -> None:
    """Set approved posture and summary language without defensive wording."""

    if "01 Admin Summary" in wb.sheetnames:
        ws = wb["01 Admin Summary"]
        apply_submission_posture(ws)

        # Keep this in a visible control note area when possible. The exact cell
        # is intentionally simple so it survives fragile workbook layouts.
        ws["A3"] = APPROVED_EXCEPTION_SUMMARY
        ws["A3"].font = Font(italic=True, size=10)
        ws["A3"].alignment = Alignment(wrap_text=True)


def build_admin_context(source_path: Path, output_path: Path, include_qc: bool = False) -> None:
    """Build an admin context workbook from an existing reviewed workbook."""

    wb = load_workbook(source_path)
    patch_summary_language(wb)

    if not include_qc:
        remove_internal_tabs(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path, help="Reviewed billing/admin workbook to export")
    parser.add_argument("output", type=Path, help="Output .xlsx path")
    parser.add_argument(
        "--include-qc",
        action="store_true",
        help="Keep internal QC tabs. Default exports the admin-facing 3-tab artifact.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    build_admin_context(args.source, args.output, include_qc=args.include_qc)


if __name__ == "__main__":
    main()
