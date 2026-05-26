"""Inspect an admin billing workbook before transformation or submission.

This is the preflight gate for the 2026-05-20 admin billing context pipeline.
It does not modify the workbook. It answers the question that matters before a
Friday submission:

    Is this workbook structurally safe enough to transform or submit?

The inspector reports:

- sheet names;
- required admin tabs present/missing;
- internal tabs present;
- required tracker columns present/missing;
- total carried numeric hours;
- blank-hour rows;
- unresolved REVIEW rows;
- suspicious language hits;
- OOO rows that still imply performed work;
- Friday batch totals derived from row dates.
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from triage.admin_billing_context_rules import (
    ADMIN_SUBMISSION_TABS,
    INTERNAL_ONLY_TABS,
    contains_suspicious_language,
    contains_work_performed_language,
    friday_batch_for,
    row_looks_ooo,
    validate_context_row,
)

TRACKER_SHEET = "02 Tracker Import"
REQUIRED_TRACKER_COLUMNS = {
    "date": ["Date", "Work Date"],
    "technician": ["Technician", "Tech", "Name"],
    "billing_category": ["Billing Category", "Category"],
    "hours": ["Hours", "Carried Hours", "Billing Hours"],
}
OPTIONAL_TRACKER_COLUMNS = {
    "status": ["Hours Status", "Status"],
    "admin_action": ["Admin Action", "Action"],
    "safe_note": ["Safe Billing Note", "Reviewed Comment", "Note"],
}


@dataclass(frozen=True)
class WorkbookIssue:
    severity: str
    issue_code: str
    message: str
    sheet: str | None = None
    row: int | None = None
    column: str | None = None


@dataclass
class WorkbookInspection:
    path: str
    sheet_names: list[str]
    required_admin_tabs_present: list[str]
    required_admin_tabs_missing: list[str]
    internal_tabs_present: list[str]
    tracker_sheet_found: bool
    tracker_columns_found: dict[str, str] = field(default_factory=dict)
    tracker_columns_missing: list[str] = field(default_factory=list)
    total_carried_hours: float = 0.0
    numeric_hour_rows: int = 0
    blank_hour_rows: int = 0
    unresolved_review_rows: int = 0
    suspicious_language_hits: int = 0
    ooo_work_language_hits: int = 0
    friday_batch_totals: dict[str, float] = field(default_factory=dict)
    issues: list[WorkbookIssue] = field(default_factory=list)

    @property
    def ok_to_transform(self) -> bool:
        return not any(issue.severity == "error" for issue in self.issues)

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["ok_to_transform"] = self.ok_to_transform
        return data


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("_", " ")


def _header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[1]:
        key = _normalize_header(cell.value)
        if key:
            headers[key] = cell.column
    return headers


def _first_existing(headers: dict[str, int], names: list[str]) -> tuple[str, int] | tuple[None, None]:
    for name in names:
        col = headers.get(_normalize_header(name))
        if col:
            return name, col
    return None, None


def _cell_value(ws, row: int, col: int | None) -> Any:
    if col is None:
        return ""
    return ws.cell(row, col).value


def inspect_workbook(path: Path) -> WorkbookInspection:
    wb = load_workbook(path, data_only=True)
    sheet_names = list(wb.sheetnames)

    inspection = WorkbookInspection(
        path=str(path),
        sheet_names=sheet_names,
        required_admin_tabs_present=[s for s in ADMIN_SUBMISSION_TABS if s in sheet_names],
        required_admin_tabs_missing=[s for s in ADMIN_SUBMISSION_TABS if s not in sheet_names],
        internal_tabs_present=[s for s in sheet_names if s in INTERNAL_ONLY_TABS],
        tracker_sheet_found=TRACKER_SHEET in sheet_names,
    )

    for missing_tab in inspection.required_admin_tabs_missing:
        inspection.issues.append(
            WorkbookIssue(
                "error",
                "missing_admin_tab",
                f"Missing required admin tab: {missing_tab}",
                sheet=missing_tab,
            )
        )

    if inspection.internal_tabs_present:
        inspection.issues.append(
            WorkbookIssue(
                "warning",
                "internal_tabs_present",
                "Internal tabs are present. Remove them for admin-facing submission exports.",
            )
        )

    if not inspection.tracker_sheet_found:
        inspection.issues.append(
            WorkbookIssue("error", "missing_tracker_sheet", f"Missing required sheet: {TRACKER_SHEET}")
        )
        return inspection

    ws = wb[TRACKER_SHEET]
    headers = _header_map(ws)
    resolved_required: dict[str, int] = {}
    resolved_optional: dict[str, int | None] = {}

    for logical_name, candidates in REQUIRED_TRACKER_COLUMNS.items():
        found_name, col = _first_existing(headers, candidates)
        if col is None:
            inspection.tracker_columns_missing.append(logical_name)
            inspection.issues.append(
                WorkbookIssue(
                    "error",
                    "missing_required_column",
                    f"Missing required tracker column for {logical_name}. Expected one of: {', '.join(candidates)}",
                    sheet=TRACKER_SHEET,
                    column=logical_name,
                )
            )
        else:
            inspection.tracker_columns_found[logical_name] = str(found_name)
            resolved_required[logical_name] = col

    for logical_name, candidates in OPTIONAL_TRACKER_COLUMNS.items():
        found_name, col = _first_existing(headers, candidates)
        if col is not None:
            inspection.tracker_columns_found[logical_name] = str(found_name)
        resolved_optional[logical_name] = col

    if inspection.tracker_columns_missing:
        return inspection

    friday_totals: dict[str, float] = defaultdict(float)

    for row_idx in range(2, ws.max_row + 1):
        row_values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if not any(row_values):
            continue

        work_date = _cell_value(ws, row_idx, resolved_required["date"])
        hours_value = _cell_value(ws, row_idx, resolved_required["hours"])
        status = _cell_value(ws, row_idx, resolved_optional.get("status"))
        admin_action = _cell_value(ws, row_idx, resolved_optional.get("admin_action"))
        safe_note = _cell_value(ws, row_idx, resolved_optional.get("safe_note"))

        row_issues = validate_context_row(
            row_number=row_idx,
            work_date=work_date,
            hours=hours_value,
            status=status,
            admin_action=admin_action,
            safe_note=safe_note,
        )
        for issue in row_issues:
            inspection.issues.append(
                WorkbookIssue(
                    issue.severity,
                    issue.issue_code,
                    issue.message,
                    sheet=TRACKER_SHEET,
                    row=issue.row_number,
                )
            )

        if str(status or "").strip().upper().startswith("REVIEW"):
            inspection.unresolved_review_rows += 1

        if any(contains_suspicious_language(v) for v in [status, admin_action, safe_note]):
            inspection.suspicious_language_hits += 1

        if row_looks_ooo([status, admin_action, safe_note]) and contains_work_performed_language(safe_note):
            inspection.ooo_work_language_hits += 1

        try:
            hours = float(hours_value or 0)
        except (TypeError, ValueError):
            hours = 0.0

        if hours > 0:
            inspection.numeric_hour_rows += 1
            inspection.total_carried_hours += hours
            try:
                batch = friday_batch_for(work_date).isoformat()
                friday_totals[batch] += hours
            except ValueError as exc:
                inspection.issues.append(
                    WorkbookIssue(
                        "error",
                        "invalid_work_date",
                        str(exc),
                        sheet=TRACKER_SHEET,
                        row=row_idx,
                    )
                )
        else:
            inspection.blank_hour_rows += 1

    inspection.total_carried_hours = round(inspection.total_carried_hours, 2)
    inspection.friday_batch_totals = {k: round(v, 2) for k, v in sorted(friday_totals.items())}
    return inspection


def print_human_report(inspection: WorkbookInspection) -> None:
    status = "PASS" if inspection.ok_to_transform else "FAIL"
    print(f"Admin billing workbook inspection: {status}")
    print(f"Path: {inspection.path}")
    print(f"Sheets: {', '.join(inspection.sheet_names)}")
    print(f"Required tabs present: {', '.join(inspection.required_admin_tabs_present) or '(none)'}")
    print(f"Required tabs missing: {', '.join(inspection.required_admin_tabs_missing) or '(none)'}")
    print(f"Internal tabs present: {', '.join(inspection.internal_tabs_present) or '(none)'}")
    print(f"Tracker sheet found: {inspection.tracker_sheet_found}")
    print(f"Total carried hours: {inspection.total_carried_hours:.2f}")
    print(f"Numeric-hour rows: {inspection.numeric_hour_rows}")
    print(f"Blank-hour rows: {inspection.blank_hour_rows}")
    print(f"Unresolved REVIEW rows: {inspection.unresolved_review_rows}")
    print(f"Suspicious language hits: {inspection.suspicious_language_hits}")
    print(f"OOO rows implying work: {inspection.ooo_work_language_hits}")
    print("Friday batch totals:")
    for batch, hours in inspection.friday_batch_totals.items():
        print(f"  {batch}: {hours:.2f}")
    if inspection.issues:
        print("Issues:")
        for issue in inspection.issues:
            location = f" sheet={issue.sheet or '-'} row={issue.row or '-'} column={issue.column or '-'}"
            print(f"  [{issue.severity}] {issue.issue_code}:{location} - {issue.message}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--json", action="store_true", help="Print machine-readable JSON instead of the human report")
    parser.add_argument("--output-json", type=Path, help="Optional path to write the JSON inspection report")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    inspection = inspect_workbook(args.workbook)
    payload = inspection.to_dict()

    if args.output_json:
        args.output_json.parent.mkdir(parents=True, exist_ok=True)
        args.output_json.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    if args.json:
        print(json.dumps(payload, indent=2))
    else:
        print_human_report(inspection)

    if not inspection.ok_to_transform:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
