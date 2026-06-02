"""Synthetic roster for Admin Billing Summary tests.

Covers multi-project resolution and the Assignments Overrides sub-table:

April (resolution precedence demonstrated):
  Mensa Dee   - default Neuron; Worked Projects override Apr 02 -> Projects Team;
                Assignments Override Apr 03 -> Neuron Deployments (beats worked)
  Rao Tully   - default Delivery/Transport; Worked Projects Apr 01 -> Neuron
  Solo Vant   - default Neuron all days (one long shift on Apr 02)
May:
  Mensa Dee   - default Neuron; iPhone Support via Worked Projects May 01
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook


def write_roster(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # ── Live - April 2026 ──
    apr = wb.create_sheet("Live - April 2026")
    apr.append(["April 2026 - Attendance"])
    apr.append(["Staff Name", "Project",
                "Apr 01 - Clock In", "Apr 01 - Clock Out",
                "Apr 02 - Clock In", "Apr 02 - Clock Out",
                "Apr 03 - Clock In", "Apr 03 - Clock Out"])
    apr.append(["Mensa Dee", "Neuron Deployments",
                "9:00 AM", "6:00 PM", "9:00 AM", "6:00 PM", "9:00 AM", "6:00 PM"])
    apr.append(["Rao Tully", "Delivery / Transport / Disposal",
                "8:00 AM", "5:00 PM", "", "", "", ""])
    apr.append(["Solo Vant", "Neuron Deployments",
                "", "", "6:00 AM", "11:00 PM", "", ""])

    # ── Worked Projects - April 2026 ──
    wapr = wb.create_sheet("Worked Projects - April 2026")
    wapr.append(["April 2026 - Worked Projects"])
    wapr.append(["Staff Name", "Default Project",
                 datetime(2026, 4, 1), datetime(2026, 4, 2), datetime(2026, 4, 3)])
    # Mensa worked Projects Team on Apr 02 (worked beats default);
    # Apr 03 worked says Projects Team too, but an Override will beat it.
    wapr.append(["Mensa Dee", "Neuron Deployments", "", "Projects Team", "Projects Team"])
    wapr.append(["Rao Tully", "Delivery / Transport / Disposal", "Neuron Deployments", "", ""])
    wapr.append(["Solo Vant", "Neuron Deployments", "", "", ""])

    # ── Assignments - April 2026 (main table + Overrides sub-table) ──
    aapr = wb.create_sheet("Assignments - April 2026")
    aapr.append(["April 2026 - Project Assignments"])
    aapr.append(["Staff Name", "Default Project",
                 datetime(2026, 4, 1), datetime(2026, 4, 2), datetime(2026, 4, 3)])
    aapr.append(["Mensa Dee", "Neuron Deployments", "", "", ""])
    aapr.append(["Rao Tully", "Delivery / Transport / Disposal", "", "", ""])
    aapr.append(["Solo Vant", "Neuron Deployments", "", "", ""])
    aapr.append(["Overrides (only if different from Default Project)"])
    aapr.append(["Override Staff Name", "Override Date", "Override Project", "Notes"])
    aapr.append(["Mensa Dee", datetime(2026, 4, 3), "Neuron Deployments",
                 "Richard review: Neurons confirmed"])

    # ── Live - May 2026 ──
    may = wb.create_sheet("Live - May 2026")
    may.append(["May 2026 - Attendance"])
    may.append(["Staff Name", "Project",
                "May 01 - Clock In", "May 01 - Clock Out",
                "May 02 - Clock In", "May 02 - Clock Out"])
    may.append(["Mensa Dee", "Neuron Deployments", "9:00 AM", "5:00 PM", "9:00 AM", "6:00 PM"])
    may.append(["Solo Vant", "Neuron Deployments", "", "", "8:00 AM", "4:00 PM"])

    wmay = wb.create_sheet("Worked Projects - May 2026")
    wmay.append(["May 2026 - Worked Projects"])
    wmay.append(["Staff Name", "Default Project", datetime(2026, 5, 1)])
    wmay.append(["Mensa Dee", "Neuron Deployments", "iPhone Support"])

    amay = wb.create_sheet("Assignments - May 2026")
    amay.append(["May 2026 - Project Assignments"])
    amay.append(["Staff Name", "Default Project", datetime(2026, 5, 1)])
    amay.append(["Mensa Dee", "Neuron Deployments", ""])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_prior_april(path: Path) -> None:
    """A minimal prior April 'preferred-format' copy for delta tests.

    Only needs a 'Project Summary' tab with a 'Net Hours' column.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Summary"
    ws.append(["April 2026 Project Summary (prior copy)"])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append(["Project", "Tech Count", "Worked Days", "Gross Span Hours",
               "Lunch Deducted", "Net Hours"])
    # Prior submission claimed different numbers -> drives a delta.
    ws.append(["Neuron Deployments", 2, 4, 30.0, 3.0, 27.0])
    ws.append(["Projects Team", 1, 1, 9.0, 1.0, 8.0])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build(base: Path) -> dict:
    base.mkdir(parents=True, exist_ok=True)
    roster = base / "abs_roster.xlsx"
    prior = base / "abs_prior_april.xlsx"
    write_roster(roster)
    write_prior_april(prior)
    return {"roster": roster, "prior": prior}
