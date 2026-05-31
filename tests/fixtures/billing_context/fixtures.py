"""Synthetic workbook builders for billing context CLI E2E tests."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook


def write_track_hours(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "April May 2026"
    ws.append(["Tech", "Date", "Hours", "Assignment", "In", "Out"])
    ws.append(["Example Tech", date(2026, 4, 25), 8, "Neuron Installation", "09:00", "17:00"])
    ws.append(["Example Tech", date(2026, 5, 14), 8, "Neuron Installation", "09:00", "17:00"])
    wb.save(path)


def write_april_context(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tracker Import"
    ws.append(["Tech", "Date", "Task Notes"])
    ws.append(
        [
            "Example Tech",
            date(2026, 4, 25),
            "Deployed floor support for go-live.",
        ]
    )
    ws.append(
        [
            "Example Tech",
            date(2026, 5, 14),
            "Configured devices and staged inventory.",
        ]
    )
    wb.save(path)


def write_roster_log(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"
    ws.append(["Tech", "Date", "Hours"])
    ws.append(["Example Tech", date(2026, 4, 25), 8])
    ws.append(["Example Tech", date(2026, 5, 14), 8])
    wb.save(path)


def write_admin_copy(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Admin Hours"
    ws.append(["Tech", "Date", "Hours"])
    ws.append(["Example Tech", date(2026, 4, 25), 8])
    ws.append(["Example Tech", date(2026, 5, 14), 8])
    wb.save(path)


def write_dashboard(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Active_Admin_Targets"
    headers = [
        "Review Status",
        "Tech",
        "Date",
        "Roster Latest Hours",
        "Current Admin Value",
        "Expected Total",
    ]
    ws.append(headers)
    ws.append(["RESOLVED GREEN", "Example Tech", "2026-04-25", 8, 8, 8])
    ws.append(["RESOLVED GREEN", "Example Tech", "2026-05-14", 8, 8, 8])
    wb.create_sheet("CF_Dictionary")
    wb["CF_Dictionary"].append(["Rule ID"])
    wb.save(path)


def write_all_fixtures(base: Path) -> dict[str, Path]:
    base.mkdir(parents=True, exist_ok=True)
    paths = {
        "track_hours": base / "track_hours.xlsx",
        "april_context": base / "april_context.xlsx",
        "roster_log": base / "roster_log.xlsx",
        "admin_copy": base / "admin_copy.xlsx",
        "dashboard": base / "dashboard.xlsx",
    }
    write_track_hours(paths["track_hours"])
    write_april_context(paths["april_context"])
    write_roster_log(paths["roster_log"])
    write_admin_copy(paths["admin_copy"])
    write_dashboard(paths["dashboard"])
    return paths
