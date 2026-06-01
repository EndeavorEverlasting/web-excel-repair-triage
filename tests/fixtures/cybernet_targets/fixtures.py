"""Synthetic workbook builders for Cybernet target sprint tests."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


NEURON_HEADER = [
    "",
    "Wave",
    "No",
    "Site",
    "Location",
    "Neuron Network Port#",
    "Cybernet Network Port#",
    "Power Outlet",
    "Anesthesia Machine",
    "Serial Cable",
    "Anesthesia Machine SN",
    "Patient Monitor",
    "MIB CARD",
    "Serial Cable2",
    "Neuron Name",
    "Neuron Serial #",
    "PC Name",
    "Device Type",
]


def _pad_row(cells: list) -> list:
    row = [""] * len(NEURON_HEADER)
    for i, v in enumerate(cells):
        if i + 1 < len(row):
            row[i + 1] = v
    return row


def write_mini_all_wave(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Neuron Cybernet"
    for _ in range(5):
        ws.append([""] * len(NEURON_HEADER))
    ws.append(NEURON_HEADER)
    # Wave 1 — should be excluded
    ws.append(_pad_row(["1", "1", "3HQ", "EPIC LAB", "", "", "", "", "", "", "", "", "", "", "", "", ""]))
    # Wave 2 — excluded
    ws.append(_pad_row(["2", "1", "CCMC", "OR1", "", "", "", "", "", "", "", "", "", "", "WNH001", "Cybernet"]))
    # Wave 3 HH x2
    ws.append(_pad_row(["3", "1", "HH", "OR 8", "", "", "", "", "", "", "", "", "", "", "", "Cybernet"]))
    ws.append(_pad_row(["3", "2", "HH", "IR/CT", "", "", "", "", "", "", "", "", "", "", "", "Cybernet"]))
    # Wave 3 JTM x2
    ws.append(_pad_row(["3", "1", "JTM", "ENDO 1", "", "", "", "", "", "", "", "", "", "", "", "Cybernet"]))
    ws.append(_pad_row(["3", "2", "JTM", "ENDO 2", "", "", "", "", "", "", "", "", "", "", "", "Cybernet"]))
    # Wave 3 AMB x3
    ws.append(_pad_row(["3", "1", "Wave 3 AMB", "Endoscopy Suite", "", "", "", "", "", "", "", "", "", "", "", "Cybernet"]))
    ws.append(_pad_row(["3", "2", "Wave 3 AMB", "Pain Mgmt Brentwood", "", "", "", "", "", "", "", "", "", "", "", "Cybernet"]))
    ws.append(_pad_row(["3", "3", "Wave 3 AMB", "OB/GYN Lake Success", "", "", "", "", "", "", "", "", "", "", "", "Cybernet"]))
    # Wave 3 SSH x2
    ws.append(_pad_row(["3", "1", "SSH", "CATH LAB", "", "", "", "", "", "", "", "", "", "", "SSH-001", "Cybernet"]))
    ws.append(_pad_row(["3", "2", "SSH", "MRI", "", "", "", "", "", "", "", "", "", "", "SSH-002", "Cybernet"]))
    # Out of scope
    ws.append(_pad_row(["3", "99", "Out of Scop HH", "Spare", "", "", "", "", "", "", "", "", "", "", "", "Cybernet"]))

    ws2 = wb.create_sheet("ANE Ambulatory Locations")
    ws2.append(["ANE Ambulatory Locations Hardware Assessment"])
    ws2.append(["Wave", "Practice Name", "Address", "AMB DEP Name", "Device Type", "Serial Cable", "Cybernet ", "GCX Arm"])
    ws2.append(["2", "Practice A", "Addr", "SYOSS OPIMG [100171006]", "", "", "1", "1"])
    ws2.append(["2", "Practice B", "Addr", "ORLIN COHEN [100372001]", "", "", "1", "1"])
    ws2.append(["3", "Practice C", "Addr", "OBGYN [100435001]", "", "", "", ""])
    wb.save(path)


def write_mini_sprint_dashboard(path: Path) -> None:
    headers = [
        "Site",
        "Location",
        "Hostname (WBS)",
        "Workstream",
        "Status",
        "Imaged",
        "Labeled",
        "Boxed",
        "Ready for Delivery",
        "Completed Date",
    ]

    def site_sheet(title: str, rows: list[list]) -> None:
        ws = wb.create_sheet(title)
        ws.append(headers)
        for r in rows:
            ws.append([title] + r)

    wb = Workbook()
    wb.remove(wb.active)
    site_sheet("HH", [
        ["OR 8", "", "HH", "Open", "Yes", "Yes", "", "", ""],
        ["IR/CT", "", "HH", "Open", "", "", "", "", ""],
    ])
    site_sheet("JTM", [
        ["ENDO 1", "", "JTM", "Open", "", "", "", "", ""],
        ["ENDO 2", "", "JTM", "Open", "", "", "", "", ""],
    ])
    site_sheet("AMB", [
        ["Endoscopy Suite", "", "AMB", "Open", "", "", "", "", ""],
        ["Pain Mgmt Brentwood", "", "AMB", "Open", "", "", "", "", ""],
    ])
    site_sheet("SSUH", [
        ["Imaging Pipeline", "WBS-001", "Imaging", "Open", "", "", "", "", ""],
        ["Imaging Pipeline", "WBS-002", "Imaging", "Open", "", "", "", "", ""],
    ])
    wb.create_sheet("Dashboard")
    wb.create_sheet("00_CF_Dictionary").append(["Rule ID", "Description"])
    wb.save(path)


def write_mini_deployment_tracker(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Deployments"
    ws.append([
        "Device Type",
        "Deployed",
        "Current Building",
        "Room",
        "Cybernet Hostname",
        "Neuron Hostname",
    ])
    ws.append(["Cybernet-Neuron", "Yes", "SSH", "CATH LAB", "SSH-001", "SSH-MACH-A"])
    ws.append(["Cybernet-Neuron", "Yes", "HH", "OR 8", "WMH300OPR365", "HH-MACH-A"])
    wb.save(path)


def build_all_fixtures(base: Path) -> dict[str, Path]:
    base.mkdir(parents=True, exist_ok=True)
    paths = {
        "all_wave": base / "mini_all_wave.xlsx",
        "sprint_dashboard": base / "mini_sprint_dashboard.xlsx",
        "deployment_tracker": base / "mini_deployment_tracker.xlsx",
    }
    write_mini_all_wave(paths["all_wave"])
    write_mini_sprint_dashboard(paths["sprint_dashboard"])
    write_mini_deployment_tracker(paths["deployment_tracker"])
    return paths
