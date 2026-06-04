"""Synthetic .xlsx builders for the may_roster_webexcel tests.

All workbooks are tiny and generated into a tmp_path at test time so no real
client workbooks are ever committed. The Live sheet mirrors the real roster
layout: row-2 date headers like "May 17 - Clock In" with Sunday May 17 in
columns AI/AJ and Monday May 18 in columns AK/AL.
"""
from __future__ import annotations

import io
import zipfile
from pathlib import Path

# Real-roster column positions for May 2026.
# AG=33, AH=34, AI=35, AJ=36, AK=37, AL=38
_HEADERS = {
    33: "May 16 - Clock In",
    34: "May 16 - Clock Out",
    35: "May 17 - Clock In",   # Sunday
    36: "May 17 - Clock Out",
    37: "May 18 - Clock In",   # Monday
    38: "May 18 - Clock Out",
}


def _base_live_sheet():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Live - May 2026"
    ws.cell(1, 1, "Active Roster Log")
    ws.cell(2, 1, "Staff Name")
    ws.cell(2, 2, "Project")
    for col, label in _HEADERS.items():
        ws.cell(2, col, label)
    return wb, ws


def build_good_cf_workbook(path: str) -> str:
    """Sunday rules reference only Sunday columns -> no bleed."""
    from openpyxl.formatting.rule import FormulaRule

    wb, ws = _base_live_sheet()
    ws.cell(3, 1, "Jane Tech")
    ws.cell(3, 2, "Neuron Deployments")
    # Sunday (AI/AJ) malformed-single-punch rule referencing only Sunday cols.
    ws.conditional_formatting.add(
        "AI3:AI50",
        FormulaRule(formula=['AND($AI3="",$AJ3<>"")'], stopIfTrue=False, fill=None),
    )
    ws.conditional_formatting.add(
        "AK3:AK50",
        FormulaRule(formula=['AND($AK3="",$AL3<>"")'], stopIfTrue=False, fill=None),
    )
    wb.save(path)
    return path


def build_bad_cf_workbook(path: str) -> str:
    """Three Sunday-bleed defects: cross-ref, always-true blanket, merged range."""
    from openpyxl.formatting.rule import FormulaRule

    wb, ws = _base_live_sheet()
    ws.cell(3, 1, "Jane Tech")
    ws.cell(3, 2, "Neuron Deployments")
    # Defect 1: Sunday rule references Monday column AK.
    ws.conditional_formatting.add(
        "AI3:AI50",
        FormulaRule(formula=['AND($AI3="",$AK3<>"")'], stopIfTrue=False, fill=None),
    )
    # Defect 2: always-true blanket over Sunday columns.
    ws.conditional_formatting.add(
        "AI3:AJ50",
        FormulaRule(formula=["1"], stopIfTrue=False, fill=None),
    )
    # Defect 3: single merged range spanning Sunday (AI/AJ) and Monday (AK/AL).
    ws.conditional_formatting.add(
        "AI3:AL50",
        FormulaRule(formula=['$AI3<>""'], stopIfTrue=False, fill=None),
    )
    wb.save(path)
    return path


def build_punch_workbook(path: str) -> str:
    """Live sheet with overnight, malformed, blank, and unassigned punches."""
    wb, ws = _base_live_sheet()
    # Jane: overnight on Sunday May 17 (8:30 AM -> 1:00 AM), Neuron project.
    ws.cell(3, 1, "Jane Tech")
    ws.cell(3, 2, "Neuron Deployments")
    ws.cell(3, 35, "8:30 AM")
    ws.cell(3, 36, "1:00 AM")
    # Bob: unassigned (paid hours, project "0") same-day shift on May 16.
    ws.cell(4, 1, "Bob Tech")
    ws.cell(4, 2, "0")
    ws.cell(4, 33, "9:00 AM")
    ws.cell(4, 34, "5:00 PM")
    # Cara: single missing punch (malformed) on May 18.
    ws.cell(5, 1, "Cara Tech")
    ws.cell(5, 2, "Neuron Deployments")
    ws.cell(5, 37, "9:00 AM")
    wb.save(path)
    return path


def build_clean_simple_workbook(path: str) -> str:
    """A minimal values-only workbook expected to pass package preflight.

    openpyxl emits ``inlineStr`` cells; the repo's web-safe exporters run
    ``fix_inlinestr`` to eliminate them, so we mirror that here.
    """
    import openpyxl

    from triage.xlsx_utils import fix_inlinestr

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Hours"])
    ws.append(["Jane", 8])
    ws.append(["Bob", 7.5])
    wb.save(path)
    fix_inlinestr(path)
    return path


def build_formula_workbook(path: str) -> str:
    """A workbook containing a formula (fails share-safe purity)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["A", "B"])
    ws["A2"] = 2
    ws["B2"] = "=A2+1"
    wb.save(path)
    return path


def inject_text_into_sheet(src_path: str, dst_path: str, token: str) -> str:
    """Copy a workbook, inserting *token* as an XML comment in sheet1.xml.

    Keeps XML well-formed so only the targeted gate (namespace/function token)
    trips, not xml_wellformed.
    """
    with zipfile.ZipFile(src_path, "r") as zin:
        names = zin.namelist()
        data = {n: zin.read(n) for n in names}
    target = "xl/worksheets/sheet1.xml"
    text = data[target].decode("utf-8", errors="ignore")
    text = text.replace("</worksheet>", f"<!-- {token} --></worksheet>")
    data[target] = text.encode("utf-8")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, data[n])
    Path(dst_path).write_bytes(buf.getvalue())
    return dst_path
