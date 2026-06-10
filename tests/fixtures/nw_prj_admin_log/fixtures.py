"""Synthetic fixtures for NW PRJ Admin Log Project Team generator tests."""
from __future__ import annotations

import io
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _apr_row(name: str, project: str, punches: dict[int, tuple[str, str]]) -> list:
    days = (1, 2, 3, 4, 5, 6, 15, 30)
    row = [name, project]
    for day in days:
        cin, cout = punches.get(day, ("", ""))
        row.extend([cin, cout])
    return row


def write_roster_fixture(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    apr = wb.create_sheet("Live - April 2026")
    apr.append(["April 2026 - Attendance"])
    apr.append([
        "Staff Name", "Project",
        "Apr 01 - Clock In", "Apr 01 - Clock Out",
        "Apr 02 - Clock In", "Apr 02 - Clock Out",
        "Apr 03 - Clock In", "Apr 03 - Clock Out",
        "Apr 04 - Clock In", "Apr 04 - Clock Out",
        "Apr 05 - Clock In", "Apr 05 - Clock Out",
        "Apr 06 - Clock In", "Apr 06 - Clock Out",
        "Apr 15 - Clock In", "Apr 15 - Clock Out",
        "Apr 30 - Clock In", "Apr 30 - Clock Out",
    ])
    apr.append(_apr_row("Alpha Tech", "Projects Team", {
        1: ("9:00 AM", "6:00 PM"),
        6: ("8:00 AM", "4:00 PM"),
    }))
    apr.append(_apr_row("Bravo Tech", "Projects Team", {
        3: ("8:00 AM", "4:00 PM"),
    }))

    wapr = wb.create_sheet("Worked Projects - April 2026")
    wapr.append(["April 2026 - Worked Projects"])
    wapr.append(["Staff Name", "Default Project", datetime(2026, 4, 1)])
    wapr.append(["Alpha Tech", "Projects Team", ""])

    aapr = wb.create_sheet("Assignments - April 2026")
    aapr.append(["Staff Name", datetime(2026, 4, 1)])
    aapr.append(["Alpha Tech", "Projects Team"])

    may = wb.create_sheet("Live - May 2026")
    may.append(["May 2026 - Attendance"])
    may.append([
        "Staff Name", "Project",
        "May 31 - Clock In", "May 31 - Clock Out",
    ])
    may.append(["Alpha Tech", "Projects Team", "8:00 AM", "4:00 PM"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_visual_donor_fixture(path: Path, *, with_logo: bool = True) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Project Team")
    wb.create_sheet("Hidden Tab")
    header_fill = PatternFill("solid", fgColor="1F365C")
    for r in range(1, 4):
        ws.cell(r, 2, value="")
        ws.cell(r, 2).fill = header_fill
    ws.cell(10, 2, value="Techs ")
    ws.cell(10, 2).font = Font(bold=True, color="FFFFFF")
    ws.cell(10, 2).fill = header_fill
    # Donor starts Apr 6 — exporter must insert Apr 1-5.
    ws.cell(10, 3, value=datetime(2026, 4, 6))
    ws.cell(10, 7, value=datetime(2026, 4, 7))
    ws.cell(10, 11, value=datetime(2026, 5, 31))
    ws.cell(11, 3, value="In")
    ws.cell(11, 4, value="Out")
    ws.cell(11, 5, value="Total")
    ws.cell(12, 2, value="Alpha Tech")
    ws.cell(12, 3, value="STALE IN")
    ws.cell(12, 4, value="STALE OUT")
    ws.cell(12, 5, value=99)
    ws.cell(13, 2, value="Bravo Tech")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10

    if with_logo:
        try:
            from openpyxl.drawing.image import Image
            # 1x1 red PNG
            png = (
                b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01"
                b"\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx"
                b"\x9cc\xf8\xcf\xc0\x00\x00\x00\x03\x00\x01\x00\x05\xfe\xd4\x00\x00\x00"
                b"\x00IEND\xaeB`\x82"
            )
            ws.add_image(Image(io.BytesIO(png)), "B1")
        except Exception:
            pass

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_accepted_reference_fixture(path: Path) -> None:
    write_visual_donor_fixture(path, with_logo=True)
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["Project Team"]
    ws.column_dimensions["A"].hidden = True
    ws.freeze_panes = "C1"
    for name in wb.sheetnames:
        wb[name].sheet_state = "visible" if name == "Project Team" else "hidden"
    wb.save(path)
    wb.close()


def build_fixtures(base: Path) -> dict:
    base.mkdir(parents=True, exist_ok=True)
    roster = base / "roster.xlsx"
    donor = base / "visual_donor.xlsx"
    reference = base / "accepted_reference.xlsx"
    write_roster_fixture(roster)
    write_visual_donor_fixture(donor, with_logo=True)
    write_accepted_reference_fixture(reference)
    return {"roster": roster, "donor": donor, "reference": reference}
