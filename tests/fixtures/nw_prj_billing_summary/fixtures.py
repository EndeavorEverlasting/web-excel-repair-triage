"""Synthetic roster builder for NW PRJ billing summary tests (no private data).

Deterministic April + May 2026 scope. Calendar anchors:
  Apr 01 2026 = Wednesday  -> Friday batch Apr 03
  Apr 02 2026 = Thursday   -> Friday batch Apr 03
  Apr 04 2026 = Saturday   -> rolls to next Friday Apr 10
  May 01 2026 = Friday      -> Friday batch May 01
  May 02 2026 = Saturday    -> rolls to next Friday May 08

Kept (Project Team) rows and hours:
  Alpha  Apr01  9:00 AM-6:00 PM   gross 9  net 8   (live default Neuron)
  Beta   Apr02  8:00 AM/note-12PM gross 4  net 4   (Worked-Project override)
  Carol  Apr04  9:00 AM-1:00 PM   gross 4  net 4   (weekend)
  Rich   Apr01  1:00 PM-4:00 PM   gross 3  net 3   (pinned: short day NOT flagged)
  Alpha  May01  9:00 AM-6:00 PM   gross 9  net 8
  Alpha  May02  8:00 AM-12:00 PM  gross 4  net 4   (weekend)
  Combined: gross 33, lunch 2, net 31

Routed to review (excluded from admin totals):
  Dan    Apr02  9:00 AM-(blank)   partial punch
  Yostinn Minaya / Steven Marques (Inventory)  excluded non-member names
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook


def write_mini_roster(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # ── Live - April 2026 ───────────────────────────────────────────
    apr = wb.create_sheet("Live - April 2026")
    apr.append(["April 2026 - Attendance"])
    apr.append(["Staff Name", "Project",
                "Apr 01 - Clock In", "Apr 01 - Clock Out",
                "Apr 02 - Clock In", "Apr 02 - Clock Out",
                "Apr 04 - Clock In", "Apr 04 - Clock Out"])
    apr.append(["Alpha Tech", "Neuron Deployments", "9:00 AM", "6:00 PM", "", "", "", ""])
    apr.append(["Beta Tech", "Delivery / Transport", "", "", "8:00:00 AM/ Bonita", "12:00 PM", "", ""])
    apr.append(["Carol Tech", "Neuron Deployments", "", "", "", "", "9:00 AM", "1:00 PM"])
    apr.append(["Dan Tech", "Neuron Deployments", "", "", "9:00 AM", "", "", ""])
    apr.append(["Rich Perez", "Admin Project Team", "1:00 PM", "4:00 PM", "", "", "", ""])
    apr.append(["Yostinn Minaya", "Neuron Deployments", "9:00 AM", "5:00 PM", "", "", "", ""])
    apr.append(["Steven Marques (Inventory)", "Inventory", "9:00 AM", "5:00 PM", "", "", "", ""])

    # ── Worked Projects - April 2026 (override Beta's Apr 02) ────────
    wapr = wb.create_sheet("Worked Projects - April 2026")
    wapr.append(["April 2026 - Worked Projects"])
    wapr.append(["Staff Name", "Default Project",
                 datetime(2026, 4, 1), datetime(2026, 4, 2), datetime(2026, 4, 4)])
    wapr.append(["Beta Tech", "Delivery / Transport", "", "Admin Project Team", ""])

    # ── Live - May 2026 ─────────────────────────────────────────────
    may = wb.create_sheet("Live - May 2026")
    may.append(["May 2026 - Attendance"])
    may.append(["Staff Name", "Project",
                "May 01 - Clock In", "May 01 - Clock Out",
                "May 02 - Clock In", "May 02 - Clock Out"])
    may.append(["Alpha Tech", "Neuron Deployments", "9:00 AM", "6:00 PM", "8:00 AM", "12:00 PM"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_fixtures(base: Path) -> dict:
    base.mkdir(parents=True, exist_ok=True)
    roster = base / "mini_billing_roster.xlsx"
    write_mini_roster(roster)
    return {"roster": roster}


SAMPLE_INVOICES = [
    {"vendor": "Acme Cabling", "total": 1200.0, "cost_category": "Cabling",
     "po_number": "PO-1", "service_date": "2026-04-15"},
    {"vendor": "Acme Cabling", "total": 800.0, "cost_category": "Cabling",
     "po_number": "PO-2", "service_date": "2026-05-10"},
    {"vendor": "Globex Logistics", "total": 500.0, "cost_category": "Logistics",
     "po_number": "PO-3", "service_date": "2026-04-20"},
]
