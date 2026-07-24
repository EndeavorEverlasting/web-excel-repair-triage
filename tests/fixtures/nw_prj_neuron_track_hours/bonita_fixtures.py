"""Synthetic roster for the Bonita Neuron Track Hours generator tests.

One workbook with Live / Worked Projects / Assignments tabs for April and May,
engineered to exercise every inclusion and exclusion branch of the resolver:

April (counted shifts):
  Alpha Tech   Apr 01  9h   default Neuron (full-month start)
  Alpha Tech   Apr 30  8h   default Neuron (full-month end; proves span > Apr 1-4)
  Bravo Tech   Apr 03  8h   note-bearing punch (note must not leak to cells)
  Delta Tech   Apr 02  8h   default Delivery, Worked-Projects override -> Neuron
  Foxtrot Tech Apr 03  8h   Neuron, Delivery/Transport activity sub-label
  Hotel Tech   Apr 15  17h  long shift (counted + flagged)
  India Tech   Apr 02  8h   default Delivery, Assignments override -> Neuron
  Mixed Tech   Apr 06  8h   inventory+configuration note -> split (2 rows)
  => 9 rows, 74.0h

April (excluded, recorded in review) — continued:
  Kilo Tech    Apr 05  client coordination outside allowlist

April (excluded, recorded in review):
  Charlie Tech Apr 01  / Bonita off-project coverage punch
  Echo Tech    Apr 02  default Neuron, Worked-Projects override -> Delivery
  Golf Tech    Apr 04  PTO non-work marker
  Yostinn Minaya Apr 01  excluded name

May (counted shifts):
  Alpha Tech   May 01  8h
  Bravo Tech   May 02  8h
  => 2 rows, 16.0h
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

_APR_DAYS = (1, 2, 3, 4, 5, 6, 15, 30)


def _apr_row(name: str, project: str, punches: dict[int, tuple[str, str]]) -> list:
    """Build one Live-April row aligned to the eight punch columns in the header."""
    row = [name, project]
    for day in _APR_DAYS:
        cin, cout = punches.get(day, ("", ""))
        row.extend([cin, cout])
    return row


def write_bonita_roster(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # ── Live - April 2026 ────────────────────────────────────────────
    apr = wb.create_sheet("Live - April 2026")
    apr.append(["April 2026 - Attendance"])
    apr.append([
        "Staff Name", "Project",
        "Apr 01 - Clock In", "Apr 01 - Clock Out",
        "Apr 02 - Clock In", "Apr 02 - Clock Out",
        "Apr 03 - Clock In", "Apr 03 - Clock Out",
        "Apr 04 - Clock In", "Apr 04 - Clock Out",
        "Apr 05 - Clock In", "Apr 05 - Clock Out",
        "Apr 06 - Clock In", "Apr 06 - Clock Out",
        "Apr 15 - Clock In", "Apr 15 - Clock Out",
        "Apr 30 - Clock In", "Apr 30 - Clock Out",
    ])
    apr.append(_apr_row("Alpha Tech", "Neuron Deployments", {
        1: ("9:00 AM", "6:00 PM"),
        30: ("8:00 AM", "4:00 PM"),
    }))
    apr.append(_apr_row("Bravo Tech", "Neuron Deployments", {
        3: ("8:00 AM", "4:00 PM - lunch covered"),
    }))
    apr.append(_apr_row("Charlie Tech", "Neuron Deployments", {
        1: ("9:00:00 AM/ Bonita", "1:00 PM"),
    }))
    apr.append(_apr_row("Delta Tech", "Delivery / Transport", {
        2: ("7:00 AM", "3:00 PM"),
    }))
    apr.append(_apr_row("Echo Tech", "Neuron Deployments", {
        2: ("9:00 AM", "5:00 PM"),
    }))
    apr.append(_apr_row("Foxtrot Tech", "Neuron Deployments", {
        3: ("6:00 AM", "2:00 PM - Delivery / Transport"),
    }))
    apr.append(_apr_row("Golf Tech", "Neuron Deployments", {
        4: ("PTO", ""),
    }))
    apr.append(_apr_row("Hotel Tech", "Neuron Deployments", {
        15: ("6:00 AM", "11:00 PM"),
    }))
    apr.append(_apr_row("Yostinn Minaya", "Neuron Deployments", {
        1: ("9:00 AM", "6:00 PM"),
    }))
    apr.append(_apr_row("India Tech", "Delivery / Transport", {
        2: ("8:00 AM", "4:00 PM"),
    }))
    apr.append(_apr_row("Kilo Tech", "Neuron Deployments", {
        5: ("2:00 PM", "4:00 PM - client status call"),
    }))
    apr.append(_apr_row("Mixed Tech", "Neuron Deployments", {
        6: ("9:00 AM", "5:00 PM - inventory warehouse and configuration baseline"),
    }))

    # ── Worked Projects - April 2026 ─────────────────────────────────
    wapr = wb.create_sheet("Worked Projects - April 2026")
    wapr.append(["April 2026 - Worked Projects"])
    wapr.append(["Staff Name", "Default Project",
                 datetime(2026, 4, 2), datetime(2026, 4, 3)])
    wapr.append(["Delta Tech", "Delivery / Transport", "Neuron Deployments", ""])
    wapr.append(["Echo Tech", "Neuron Deployments", "Delivery / Transport", ""])

    # ── Assignments - April 2026 ─────────────────────────────────────
    aapr = wb.create_sheet("Assignments - April 2026")
    aapr.append(["Staff Name", datetime(2026, 4, 2)])
    aapr.append(["India Tech", "Neuron Deployments"])

    # ── Live - May 2026 ──────────────────────────────────────────────
    may = wb.create_sheet("Live - May 2026")
    may.append(["May 2026 - Attendance"])
    may.append([
        "Staff Name", "Project",
        "May 01 - Clock In", "May 01 - Clock Out",
        "May 02 - Clock In", "May 02 - Clock Out",
    ])
    may.append(["Alpha Tech", "Neuron Deployments", "8:00 AM", "4:00 PM", "", ""])
    may.append(["Bravo Tech", "Neuron Deployments", "", "", "9:00 AM", "5:00 PM"])

    wmay = wb.create_sheet("Worked Projects - May 2026")
    wmay.append(["May 2026 - Worked Projects"])
    wmay.append(["Staff Name", "Default Project", datetime(2026, 5, 1)])
    wmay.append(["Alpha Tech", "Neuron Deployments", ""])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_bonita_fixtures(base: Path) -> dict:
    base.mkdir(parents=True, exist_ok=True)
    roster = base / "bonita_roster.xlsx"
    write_bonita_roster(roster)
    return {"roster": roster}
