"""Synthetic roster workbook + malformed package builders for Neuron Track Hours tests.

Synthetic Neuron scope (deterministic):
  April: Alpha Tech 9h (Apr 01) + Beta Tech 4h (Apr 02, worked-project override) = 13h
  May:   Alpha Tech 16h (May 30) + 6h (May 31) = 22h (Go Live weekend)
  Gamma Tech is Delivery only -> excluded.
"""
from __future__ import annotations

import io
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook


def write_mini_roster(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # ── Live - April 2026 ───────────────────────────────────────────
    apr = wb.create_sheet("Live - April 2026")
    apr.append(["April 2026 - Attendance"])
    apr.append(["Staff Name", "Project",
                "Apr 01 - Clock In", "Apr 01 - Clock Out",
                "Apr 02 - Clock In", "Apr 02 - Clock Out"])
    apr.append(["Alpha Tech", "Neuron Deployments", "9:00 AM", "6:00 PM", "", ""])
    apr.append(["Beta Tech", "Delivery / Transport", "", "", "8:00:00 AM/ Bonita", "12:00 PM"])
    apr.append(["Gamma Tech", "Delivery / Transport", "9:00 AM", "5:00 PM", "", ""])

    # ── Worked Projects - April 2026 ────────────────────────────────
    wapr = wb.create_sheet("Worked Projects - April 2026")
    wapr.append(["April 2026 - Worked Projects"])
    wapr.append(["Staff Name", "Default Project",
                 datetime(2026, 4, 1), datetime(2026, 4, 2)])
    wapr.append(["Alpha Tech", "Neuron Deployments", "", ""])
    # Beta default is Delivery, but worked project on Apr 02 is Neuron
    wapr.append(["Beta Tech", "Delivery / Transport", "", "Neuron Deployments"])
    wapr.append(["Gamma Tech", "Delivery / Transport", "", ""])

    # ── Live - May 2026 (Go Live weekend cols only) ─────────────────
    may = wb.create_sheet("Live - May 2026")
    may.append(["May 2026 - Attendance"])
    may.append(["Staff Name", "Project",
                "May 30 - Clock In", "May 30 - Clock Out",
                "May 31 - Clock In", "May 31 - Clock Out"])
    may.append(["Alpha Tech", "Neuron Deployments", "12:00 AM", "4:00 PM", "12:00 AM", "6:00 AM"])
    may.append(["Gamma Tech", "Delivery / Transport", "9:00 AM", "5:00 PM", "", ""])

    wmay = wb.create_sheet("Worked Projects - May 2026")
    wmay.append(["May 2026 - Worked Projects"])
    wmay.append(["Staff Name", "Default Project",
                 datetime(2026, 5, 30), datetime(2026, 5, 31)])
    wmay.append(["Alpha Tech", "Neuron Deployments", "", ""])
    wmay.append(["Gamma Tech", "Delivery / Transport", "", ""])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_fixtures(base: Path) -> dict:
    base.mkdir(parents=True, exist_ok=True)
    roster = base / "mini_roster.xlsx"
    write_mini_roster(roster)
    return {"roster": roster}


def write_malformed_xlsx(path: Path) -> None:
    """A package that trips every stop-ship token: inlineStr, ns0:, calcChain.xml."""
    path.parent.mkdir(parents=True, exist_ok=True)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            "</Types>",
        )
        z.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets></workbook>',
        )
        z.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            'Target="worksheets/sheet1.xml"/></Relationships>',
        )
        z.writestr(
            "xl/worksheets/sheet1.xml",
            '<?xml version="1.0"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:ns0="http://example.com/ns0">'
            '<ns0:extLst/>'
            '<sheetData><row r="1"><c r="A1" t="inlineStr"><is><t>bad</t></is></c></row></sheetData></worksheet>',
        )
        z.writestr("xl/calcChain.xml", '<?xml version="1.0"?><calcChain/>')
    path.write_bytes(buf.getvalue())
