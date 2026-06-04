"""Sanitized fixtures for the 1 Marcus recon engine.

No real client data. Workbooks are generated at test time and gitignored.
"""
from __future__ import annotations

import io
import zipfile
from pathlib import Path

from openpyxl import Workbook

_EXT_LINK_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<externalLink xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    '<externalBook xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
    ' r:id="rId1"><sheetNames><sheetName val="5-07-2026 Part Numbers"/></sheetNames>'
    "</externalBook></externalLink>"
)
_EXT_LINK_RELS = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
    '<Relationship Id="rId1"'
    ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath"'
    ' Target="file:///C:/old/PartNumbers_Source.xlsx" TargetMode="External"/></Relationships>'
)
_CALC_CHAIN = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    '<c r="A1" i="1"/></calcChain>'
)


def _build_base(part_numbers_titles, *, with_external_formula: bool) -> bytes:
    """Build a minimal valid workbook with openpyxl and return its bytes."""
    wb = Workbook()
    pivot = wb.active
    pivot.title = "1M Recon Pivot Module"

    first_pn = part_numbers_titles[0]
    pivot["A1"] = "Recon totals"
    pivot["A2"] = f"=SUMIFS('{first_pn}'!$T$2:$T$10,'{first_pn}'!$Z$2:$Z$10,\"Include\")"
    pivot["A3"] = f"=COUNTIFS('{first_pn}'!$Z$2:$Z$10,\"Include\")"
    if with_external_formula:
        # Stored external-indexed reference Excel uses for external workbooks.
        pivot["A4"] = f"=[1]'{first_pn}'!$A$1"

    notes = wb.create_sheet("Notes")
    notes["A1"] = "Unrelated tab; must be preserved."

    for title in part_numbers_titles:
        pn = wb.create_sheet(title)
        pn["S1"] = "Site"
        pn["T1"] = "Qty"
        pn["Z1"] = "Decision"
        pn["S2"] = "OR"
        pn["T2"] = 5
        pn["Z2"] = "Include"

    readme = wb.create_sheet("README Integration")
    readme["A1"] = "Recon integration notes."

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _inject_defects(data: bytes, *, add_external: bool, add_calc_chain: bool) -> bytes:
    """Add external-link parts and/or calcChain to an openpyxl workbook."""
    with zipfile.ZipFile(io.BytesIO(data), "r") as zin:
        names = zin.namelist()
        parts = {n: zin.read(n) for n in names}

    if add_external:
        parts["xl/externalLinks/externalLink1.xml"] = _EXT_LINK_XML.encode("utf-8")
        parts["xl/externalLinks/_rels/externalLink1.xml.rels"] = _EXT_LINK_RELS.encode("utf-8")
        ct = parts["[Content_Types].xml"].decode("utf-8")
        if "externalLink" not in ct:
            ct = ct.replace(
                "</Types>",
                '<Override PartName="/xl/externalLinks/externalLink1.xml"'
                ' ContentType="application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.externalLink+xml"/></Types>',
            )
            parts["[Content_Types].xml"] = ct.encode("utf-8")
        rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8")
        rels = rels.replace(
            "</Relationships>",
            '<Relationship Id="rIdExt1"'
            ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink"'
            ' Target="externalLinks/externalLink1.xml"/></Relationships>',
        )
        parts["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")
        wb = parts["xl/workbook.xml"].decode("utf-8")
        if "<externalReferences" not in wb:
            wb = wb.replace(
                "</sheets>",
                '</sheets><externalReferences><externalReference r:id="rIdExt1"/>'
                "</externalReferences>",
            )
            parts["xl/workbook.xml"] = wb.encode("utf-8")

    if add_calc_chain:
        parts["xl/calcChain.xml"] = _CALC_CHAIN.encode("utf-8")
        ct = parts["[Content_Types].xml"].decode("utf-8")
        if "calcChain" not in ct:
            ct = ct.replace(
                "</Types>",
                '<Override PartName="/xl/calcChain.xml"'
                ' ContentType="application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.calcChain+xml"/></Types>',
            )
            parts["[Content_Types].xml"] = ct.encode("utf-8")
        rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8")
        if "calcChain" not in rels:
            rels = rels.replace(
                "</Relationships>",
                '<Relationship Id="rIdCalc"'
                ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/calcChain"'
                ' Target="calcChain.xml"/></Relationships>',
            )
            parts["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")

    order = list(parts.keys())
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in order:
            zout.writestr(n, parts[n])
    return buf.getvalue()


def make_stale_recon(path: str) -> str:
    """A recon workbook with a stale dated tab, external link, and calcChain."""
    data = _build_base(["5-07-2026 Part Numbers"], with_external_formula=True)
    data = _inject_defects(data, add_external=True, add_calc_chain=True)
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_bytes(data)
    return path


def make_ambiguous(path: str) -> str:
    """A recon workbook with two dated Part Numbers tabs (no filename date)."""
    data = _build_base(
        ["5-07-2026 Part Numbers", "5-21-2026 Part Numbers"], with_external_formula=False
    )
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_bytes(data)
    return path


def make_integrated_source(path: str) -> str:
    """Integrated workbook with Part Numbers helpers but no executive Visual column."""
    wb = Workbook()
    pivot = wb.active
    pivot.title = "1M Recon Pivot Module"
    pivot["A1"] = "Placeholder executive tab"
    pivot["A12"] = "Inventory Rollup by Item"
    pivot["B12"] = "Total Qty"
    # Deliberately missing Visual column on input.

    pn = wb.create_sheet("5-28-2026 Part Numbers")
    pn.append(
        [
            "Date Added",
            "Item Type",
            "Item Model/Brand",
            "Part / Model Number",
            "Category",
            "Description",
            "Quantity",
        ]
        + [None] * 11
        + ["PivotPartKey", "QtyNum", None, None, None, None, None, "IncludeFlag"]
    )
    pn.append([None, "TypeA", None, "PN-001", None, None, 5] + [None] * 11 + ["Alpha", 5, None, None, None, None, None, "Include"])
    pn.append([None, "TypeB", None, "PN-002", None, None, 3] + [None] * 11 + ["Beta", 3, None, None, None, None, None, "Include"])
    pn.append([None, "TypeA", None, "PN-003", None, None, 2] + [None] * 11 + ["Alpha", 2, None, None, None, None, None, "Include"])

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out)
