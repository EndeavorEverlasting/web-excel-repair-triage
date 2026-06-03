"""Synthetic roster log pairs for roster_log_compare tests."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

from tests.fixtures.admin_billing_summary.builders import write_roster


def _save(wb: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def build_identical_pair(base: Path) -> dict:
    left = base / "left.xlsx"
    right = base / "right.xlsx"
    write_roster(left)
    write_roster(right)
    return {"left": left, "right": right}


def build_newer_filename_same_content(base: Path) -> dict:
    pair = build_identical_pair(base)
    newer = base / "INTERNAL_Active_Roster_Log_2026-06-02.xlsx"
    older = base / "Active_Roster_Log_2026-05-19.xlsx"
    newer.write_bytes(pair["right"].read_bytes())
    older.write_bytes(pair["left"].read_bytes())
    return {"left": older, "right": newer}


def build_older_name_newer_content(base: Path) -> dict:
    pair = build_identical_pair(base)
    import openpyxl
    wb = openpyxl.load_workbook(pair["right"])
    may = wb["Live - May 2026"]
    may.append(["New Tech", "Neuron Deployments", "8:00 AM", "5:00 PM", "", ""])
    _save(wb, base / "newer_content.xlsx")
    wb.close()
    older = base / "Active_Roster_Log_2026-05-19.xlsx"
    newer = base / "INTERNAL_2026-06-02.xlsx"
    older.write_bytes(pair["left"].read_bytes())
    newer.write_bytes((base / "newer_content.xlsx").read_bytes())
    return {"left": older, "right": newer, "expect": "use_right"}


def build_punch_diff(base: Path) -> dict:
    import openpyxl
    write_roster(base / "left.xlsx")
    wb = openpyxl.load_workbook(base / "left.xlsx")
    may = wb["Live - May 2026"]
    may.cell(3, 3, "10:00 AM")
    _save(wb, base / "right.xlsx")
    wb.close()
    return {"left": base / "left.xlsx", "right": base / "right.xlsx", "expect": "manual_review_required"}


def build_increased_cf(base: Path) -> dict:
    pair = build_identical_pair(base)
    import openpyxl
    wb = openpyxl.load_workbook(pair["right"])
    ws = wb["Live - May 2026"]
    ws.conditional_formatting.add(
        "B3:F10",
        FormulaRule(formula=["TRUE"], fill=PatternFill("solid", fgColor="FFFF00")),
    )
    _save(wb, base / "right_cf.xlsx")
    wb.close()
    return {"left": pair["left"], "right": base / "right_cf.xlsx", "expect": "use_right"}


def build_missing_override(base: Path) -> dict:
    import openpyxl
    write_roster(base / "left.xlsx")
    wb = openpyxl.load_workbook(base / "left.xlsx")
    ws = wb["Assignments - April 2026"]
    for r in range(ws.max_row, 0, -1):
        v = ws.cell(r, 1).value
        if v and "override" in str(v).lower():
            ws.delete_rows(r, ws.max_row - r + 1)
            break
    _save(wb, base / "left_no_ov.xlsx")
    write_roster(base / "right.xlsx")
    wb.close()
    return {"left": base / "left_no_ov.xlsx", "right": base / "right.xlsx"}


def build_override_formula_refs(base: Path) -> dict:
    pair = build_identical_pair(base)
    import openpyxl
    wb = openpyxl.load_workbook(pair["right"])
    ws = wb["Assignments - May 2026"]
    ws.cell(3, 3, '=IFERROR(VLOOKUP(A3,$A$206:$C$505,3,FALSE),"")')
    _save(wb, base / "right_formula.xlsx")
    wb.close()
    return {"left": pair["left"], "right": base / "right_formula.xlsx"}


def build_static_expected_hours(base: Path) -> dict:
    pair = build_identical_pair(base)
    import openpyxl
    for label, path in (("left", pair["left"]), ("right", pair["right"])):
        wb = openpyxl.load_workbook(path)
        eh = wb.create_sheet("Expected Hours - May 2026")
        eh.append(["May 2026 Expected Hours"])
        eh.append(["Staff", datetime(2026, 5, 1)])
        eh.append(["Mensa Dee", 8.0])
        if label == "right":
            wb.create_sheet("Expected Hours Rules")
        _save(wb, path)
        wb.close()
    return {"left": pair["left"], "right": pair["right"]}


def build_fragmented_cf(base: Path) -> dict:
    pair = build_identical_pair(base)
    import openpyxl
    wb = openpyxl.load_workbook(pair["right"])
    ws = wb["Live - May 2026"]
    ws.conditional_formatting.add(
        "D17",
        FormulaRule(formula=["TRUE"], fill=PatternFill("solid", fgColor="FF0000")),
    )
    _save(wb, base / "right_frag.xlsx")
    wb.close()
    return {"left": pair["left"], "right": base / "right_frag.xlsx"}


def build_header_style_changed(base: Path) -> dict:
    pair = build_identical_pair(base)
    import openpyxl
    wb = openpyxl.load_workbook(pair["right"])
    ws = wb["Live - May 2026"]
    ws.cell(2, 1).fill = PatternFill("solid", fgColor="FF00FF")
    _save(wb, base / "right_hdr.xlsx")
    wb.close()
    return {"left": pair["left"], "right": base / "right_hdr.xlsx"}
