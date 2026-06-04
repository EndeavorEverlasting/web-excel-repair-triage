"""Synthetic roster workbooks for roster_log_review_queue tests."""
from __future__ import annotations

from pathlib import Path

from tests.fixtures.admin_billing_summary.builders import write_roster


def build_mini_roster(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    write_roster(path)
    return path


def build_roster_with_legacy_cf(path: Path) -> Path:
    """Roster with one pre-existing legacy CF block on May Live."""
    import openpyxl
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill

    build_mini_roster(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Live - May 2026"]
    ws.conditional_formatting.add(
        "B3:B10",
        FormulaRule(formula=["TRUE"], fill=PatternFill("solid", fgColor="FFFF00")),
    )
    wb.save(path)
    wb.close()
    return path
