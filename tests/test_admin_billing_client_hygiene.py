"""Client-facing billing exports must not expose internal timekeeping evidence."""
from __future__ import annotations

from pathlib import Path

import openpyxl

from tests.fixtures.admin_billing_summary.builders import build
from triage.admin_billing_summary.cli import run


def _all_text(path: str) -> str:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        values = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                values.extend(str(v) for v in row if v is not None)
        return "\n".join(values)
    finally:
        wb.close()


def test_client_export_excludes_employee_and_punch_detail(tmp_path):
    fixtures = build(Path(__file__).resolve().parent / "fixtures" / "admin_billing_summary")
    manifest = run(
        roster_log=str(fixtures["roster"]),
        out_dir=str(tmp_path),
        months=["2026-04"],
        websafe=True,
    )
    client = manifest["per_month"]["2026-04"]["outputs"]["client"]
    wb = openpyxl.load_workbook(client["workbook"], data_only=True, read_only=True)
    try:
        assert wb.sheetnames == [
            "Start Here",
            "Executive Dashboard",
            "Monthly Summary",
            "Project Summary",
            "Apr 26",
        ]
    finally:
        wb.close()

    text = _all_text(client["workbook"])
    for fixture_person in ("Mensa Dee", "Rao Tully", "Solo Vant"):
        assert fixture_person not in text
    for internal_label in (
        "Clock In",
        "Clock Out",
        "Review Net Hours",
        "Review row count",
        "Source roster",
        "Override > Worked > Assignment > Live default",
    ):
        assert internal_label not in text
    assert client["websafe_preflight_pass"] is True


def test_client_export_states_support_and_historical_period_boundary(tmp_path):
    fixtures = build(Path(__file__).resolve().parent / "fixtures" / "admin_billing_summary")
    manifest = run(
        roster_log=str(fixtures["roster"]),
        out_dir=str(tmp_path),
        months=["2026-04"],
        websafe=True,
    )
    client = manifest["per_month"]["2026-04"]["outputs"]["client"]
    text = _all_text(client["workbook"])
    assert "Client billing support copy" in text
    assert "not a standalone invoice or new billing request" in text
    assert "not reopened, rebilled, or superseded" in text
    assert "retained internally and excluded from this client copy" in text


def test_internal_export_retains_audit_detail(tmp_path):
    fixtures = build(Path(__file__).resolve().parent / "fixtures" / "admin_billing_summary")
    manifest = run(
        roster_log=str(fixtures["roster"]),
        out_dir=str(tmp_path),
        months=["2026-04"],
        websafe=True,
    )
    internal = manifest["per_month"]["2026-04"]["outputs"]["internal"]
    wb = openpyxl.load_workbook(internal["workbook"], read_only=True)
    try:
        assert "Tech Summary" in wb.sheetnames
        assert "Tech Project Summary" in wb.sheetnames
        assert "April Neuron Hours" in wb.sheetnames
        assert "Review Flags" in wb.sheetnames
    finally:
        wb.close()
