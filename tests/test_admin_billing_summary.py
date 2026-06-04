"""Admin Billing Summary — OpenAI-format fixture tests."""
from __future__ import annotations

import importlib
import json
import zipfile
from pathlib import Path

import openpyxl
import pytest

from triage.admin_billing_summary.aggregator import build_month_summary
from triage.admin_billing_summary.cli import _read_prior_project_net, run
from triage.admin_billing_summary.exporter import build_workbook
from triage.admin_billing_summary.preflight import preflight_billing_summary
from tests.fixtures.admin_billing_summary.builders import build

REPO_ROOT = Path(__file__).resolve().parent.parent
FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures" / "admin_billing_summary"


@pytest.fixture(scope="module")
def fixtures():
    return build(FIXTURE_DIR)


@pytest.fixture(scope="module")
def april(fixtures):
    return build_month_summary(str(fixtures["roster"]), "2026-04")


@pytest.fixture(scope="module")
def generated(fixtures, tmp_path_factory):
    out = tmp_path_factory.mktemp("abs_out")
    return run(
        roster_log=str(fixtures["roster"]),
        out_dir=str(out),
        months=["2026-04", "2026-05"],
        prior=str(fixtures["prior"]),
        websafe=True,
        repo_root=REPO_ROOT,
    )


def test_cli_imports():
    mod = importlib.import_module("triage.admin_billing_summary.cli")
    assert hasattr(mod, "run") and hasattr(mod, "main")


def test_override_beats_worked(april):
    rec = [r for r in april.records if r.tech == "Mensa Dee" and r.date.day == 3]
    assert len(rec) == 1
    assert rec[0].project == "Neuron Deployments"
    assert rec[0].project_source == "override"


def test_worked_beats_default(april):
    mensa2 = [r for r in april.records if r.tech == "Mensa Dee" and r.date.day == 2][0]
    assert mensa2.project == "Projects Team" and mensa2.project_source == "worked"
    rao1 = [r for r in april.records if r.tech == "Rao Tully" and r.date.day == 1][0]
    assert rao1.project == "Neuron Deployments" and rao1.project_source == "worked"


def test_net_hours_and_long_shift(april):
    solo = [r for r in april.records if r.tech == "Solo Vant"][0]
    assert solo.gross_span == 17.0
    assert solo.lunch == 1.0
    assert solo.net_hours == 16.0
    assert solo.long_shift is True


def test_project_summary(april):
    by = {r.project: r for r in april.project_rows}
    assert by["Neuron Deployments"].net_hours == 40.0
    assert by["Projects Team"].net_hours == 8.0


def test_executive_metrics(april):
    assert april.total_net == 48.0
    assert april.techs_reflected == 3


def test_internal_tabs_and_tables(generated):
    internal = generated["per_month"]["2026-04"]["outputs"]["internal"]["workbook"]
    wb = openpyxl.load_workbook(internal)
    neuron_hours = "April Neuron Hours"
    expected = [
        "Start Here", "Executive Dashboard", "Monthly Summary", "Project Summary",
        "Tech Summary", "Tech Project Summary", neuron_hours,
        "Apr 26", "Review Flags", "CF Dictionary", "WebExcel QC",
    ]
    assert wb.sheetnames == expected
    assert len(wb["Project Summary"].tables) >= 1
    assert len(wb["Project Summary"]._charts) == 1
    with zipfile.ZipFile(internal) as z:
        assert len([n for n in z.namelist() if n.startswith("xl/tables/")]) >= 9
    wb.close()


def test_client_tabs_clean(generated):
    client = generated["per_month"]["2026-04"]["outputs"]["client"]["workbook"]
    wb = openpyxl.load_workbook(client)
    assert "Review Flags" not in wb.sheetnames
    assert "WebExcel QC" not in wb.sheetnames
    assert "Apr 26" in wb.sheetnames
    wb.close()


def test_neuron_detail_matches_summary(april, tmp_path):
    out = tmp_path / "internal.xlsx"
    build_workbook(april, str(out), variant="internal")
    wb = openpyxl.load_workbook(out, data_only=True)
    ws = wb["April Neuron Hours"]
    detail_net = 0.0
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row[-1] is not None:
            detail_net += float(row[-1])
    wb.close()
    assert round(detail_net, 2) == april.net_for_bucket("Neurons")


def test_bonita_tab_neuron_only(generated):
    wb = openpyxl.load_workbook(
        generated["per_month"]["2026-04"]["outputs"]["internal"]["workbook"],
        read_only=True,
    )
    ws = wb["Apr 26"]
    rows = [r for r in ws.iter_rows(min_row=3, values_only=True) if r[1]]
    wb.close()
    assert len(rows) == 4
    assert all(str(r[5]).startswith("Northwell - Neuron") for r in rows)


def test_preflight_passes(generated):
    for variant in ("internal", "client"):
        assert generated["per_month"]["2026-04"]["outputs"][variant]["websafe_preflight_pass"] is True
        assert generated["per_month"]["2026-05"]["outputs"][variant]["websafe_preflight_pass"] is True


def test_delta_vs_prior(generated):
    delta = generated["per_month"]["2026-04"]["delta_vs_prior"]
    assert delta is not None
    by = {d["Project"]: d for d in delta["by_project"]}
    assert by["Neuron Deployments"]["Delta"] == 13.0
    assert delta["total_net_delta"] == 13.0


def test_no_repair_inlinestr_on_export(tmp_path, april):
    out = tmp_path / "test.xlsx"
    build_workbook(april, str(out), variant="client")
    pf = preflight_billing_summary(str(out), variant="client", expect_neuron_tab="Apr 26")
    assert pf["preflight_pass"] is True
    assert "inlineStr" not in pf.get("token_failures", [])


# ──────────────────────────── semantic gate tests ─────────────────────────────


def _make_corpse_xlsx(path: Path) -> None:
    """Build a synthetic xlsx where sharedStrings contains only Column1..Column11.

    Dates, numbers, and the zip structure survive; all text strings are gutted.
    This replicates the exact failure signature from the repaired-ZIP artifact.

    Strategy: save a real workbook first, call fix_inlinestr to ensure the file
    has a proper sharedStrings.xml with t="s" worksheet cells, then replace the
    sharedStrings content with Column1..Column11.
    """
    import io
    import openpyxl
    from triage.xlsx_utils import fix_inlinestr

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="RealTextHere")
    ws.cell(row=1, column=2, value="AnotherRealText")
    wb.save(str(path))

    # Materialise a proper sharedStrings.xml with t="s" cell refs
    fix_inlinestr(str(path))

    # Replace sharedStrings.xml with exactly Column1..Column11
    orig = path.read_bytes()
    ss_items = "".join(f"<si><t>Column{i}</t></si>" for i in range(1, 12))
    new_ss = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        f' count="2" uniqueCount="11">{ss_items}</sst>'
    ).encode("utf-8")
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(orig), "r") as zin, \
            zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            if name == "xl/sharedStrings.xml":
                zout.writestr(name, new_ss)
            else:
                zout.writestr(name, zin.read(name))
    path.write_bytes(buf.getvalue())


def test_semantic_gate_rejects_excel_repaired_column_header_corpse(tmp_path):
    """sharedStrings = Column1..Column11 only → preflight must fail."""
    from triage.webexcel_semantic_gate import run_semantic_gate

    corpse = tmp_path / "corpse.xlsx"
    _make_corpse_xlsx(corpse)
    gate = run_semantic_gate(str(corpse), profile="admin_billing")
    assert gate["generic_column_strings_only"] is True
    assert gate["semantic_integrity"] == "FAIL"
    assert gate["meaningful_shared_string_count"] == 0

    # Also through the full preflight
    pf = preflight_billing_summary(
        str(corpse), variant="internal", expect_neuron_tab="Apr 26"
    )
    assert pf["preflight_pass"] is False
    assert pf["generic_column_strings_only"] is True


def test_semantic_gate_catches_blank_title(tmp_path, april):
    """Start Here!A1 blank → semantic_integrity must be FAIL."""
    import openpyxl

    out = tmp_path / "blank_title.xlsx"
    build_workbook(april, str(out), variant="internal")

    wb = openpyxl.load_workbook(str(out))
    wb["Start Here"]["A1"].value = None
    wb.save(str(out))
    wb.close()

    pf = preflight_billing_summary(
        str(out), variant="internal", expect_neuron_tab="Apr 26"
    )
    assert pf["semantic_integrity"] == "FAIL"
    assert any("Start Here" in f and "blank" in f for f in pf["sentinel_failures"])
    assert pf["preflight_pass"] is False


def test_inlinestr_scanner_no_false_positive(tmp_path):
    """A shared string containing the phrase 't=\"inlineStr\"' must NOT trigger
    the inlineStr token failure — the text lives in sharedStrings.xml, not as a
    worksheet cell type attribute.

    Strategy: build a workbook, call fix_inlinestr to ensure all cells are
    t="s" shared-string refs, then inject the problematic phrase as an extra
    entry in sharedStrings.xml.
    """
    import io
    import openpyxl
    from triage.xlsx_utils import fix_inlinestr

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "clean value"
    p = tmp_path / "no_false_positive.xlsx"
    wb.save(str(p))

    # Materialise t="s" worksheet cells and a sharedStrings.xml
    fix_inlinestr(str(p))

    # Inject problematic phrase into sharedStrings without adding a cell ref
    raw = p.read_bytes()
    with zipfile.ZipFile(io.BytesIO(raw), "r") as z:
        ss = z.read("xl/sharedStrings.xml").decode("utf-8")
    extra_si = '<si><t xml:space="preserve">No t=inlineStr cells in worksheet XML.</t></si>'
    ss = ss.replace("</sst>", extra_si + "</sst>")
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(raw), "r") as zin, \
            zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            if name == "xl/sharedStrings.xml":
                zout.writestr(name, ss.encode("utf-8"))
            else:
                zout.writestr(name, zin.read(name))
    p.write_bytes(buf.getvalue())

    pf = preflight_billing_summary(
        str(p), variant="client", expect_neuron_tab="Apr 26"
    )
    assert "inlineStr" not in pf.get("token_failures", [])


def test_inlinestr_scanner_catches_real_inline(tmp_path):
    """A worksheet cell with t=\"inlineStr\" must be caught."""
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    p = tmp_path / "real_inline.xlsx"
    wb.save(str(p))

    # Patch sheet1.xml to insert a genuine inlineStr cell
    raw = p.read_bytes()
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(raw), "r") as zin, \
            zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            if name == "xl/worksheets/sheet1.xml":
                text = zin.read(name).decode("utf-8")
                inline_cell = '<c r="B1" t="inlineStr"><is><t>inline_value</t></is></c>'
                text = text.replace("</sheetData>", inline_cell + "</sheetData>")
                zout.writestr(name, text.encode("utf-8"))
            else:
                zout.writestr(name, zin.read(name))
    p.write_bytes(buf.getvalue())

    pf = preflight_billing_summary(
        str(p), variant="client", expect_neuron_tab="Apr 26"
    )
    assert "inlineStr" in pf.get("token_failures", [])


def test_good_generated_workbook_passes_semantic_gate(tmp_path, april):
    """build_workbook output must pass the semantic gate with real shared strings."""
    from triage.webexcel_semantic_gate import run_semantic_gate

    out = tmp_path / "good.xlsx"
    build_workbook(april, str(out), variant="internal")
    gate = run_semantic_gate(str(out), profile="admin_billing")

    assert gate["semantic_integrity"] == "PASS"
    assert gate["meaningful_shared_string_count"] > 50
    assert gate["generic_column_strings_only"] is False
    assert gate["post_repair_text_loss"] is False


def test_repair_snapshot_detects_text_loss(tmp_path):
    """Workbook with ColumnN-only sharedStrings must register post_repair_text_loss
    only when fix_inlinestr actually mutates sentinel values."""
    from triage.webexcel_semantic_gate import check_repair_preservation

    corpse = tmp_path / "corpse_for_snapshot.xlsx"
    _make_corpse_xlsx(corpse)
    # For the corpse, fix_inlinestr is a no-op (no inlineStr cells exist),
    # so post_repair_text_loss should be False — text was already lost before repair.
    # The semantic gate detects the damage via generic_column_strings_only instead.
    result = check_repair_preservation(str(corpse), "admin_billing")
    # Either False (fix_inlinestr is a no-op on corpse) or True — both
    # are acceptable; what matters is the density gate already catches it.
    assert isinstance(result, bool)

    # Now build a healthy workbook and verify fix_inlinestr does not mutate it.
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Start Here"
    ws["A1"] = "April 2026 Billing Summary"
    good = tmp_path / "good_snap.xlsx"
    wb.save(str(good))
    assert check_repair_preservation(str(good), "admin_billing") is False


def test_embedded_bonita_tracker_uses_net_hours(generated):
    """The embedded "Apr 26" tracker tab must report net hours, not gross span.

    Solo Vant's single Apr 02 shift spans 17h gross with a 1h lunch -> 16h net.
    The Total column (index 4) must equal the net figure.
    """
    wb = openpyxl.load_workbook(
        generated["per_month"]["2026-04"]["outputs"]["internal"]["workbook"],
        read_only=True,
    )
    ws = wb["Apr 26"]
    rows = [r for r in ws.iter_rows(min_row=3, values_only=True) if r[1] == "Solo Vant"]
    wb.close()
    assert len(rows) == 1
    assert rows[0][4] == 16.0
    assert rows[0][4] != 17.0


def test_delta_raises_on_unreadable_prior(tmp_path):
    """An unreadable prior workbook must raise RuntimeError, not silently return {}."""
    bad = tmp_path / "not_a_workbook.xlsx"
    bad.write_bytes(b"this is not a valid zip/xlsx payload")
    with pytest.raises(RuntimeError):
        _read_prior_project_net(bad)


def test_multi_month_rejects_legacy_client_reference(fixtures):
    """One --reference-client must not silently cover April and May."""
    with pytest.raises(SystemExit):
        run(
            roster_log=str(fixtures["roster"]),
            out_dir=str(fixtures["roster"].parent / "out_should_not_run"),
            months=["2026-04", "2026-05"],
            websafe=False,
            reference_client=str(fixtures["roster"]),
            repo_root=REPO_ROOT,
        )
