"""Artifact fingerprint and approved-reference comparison tests."""
from __future__ import annotations

import io
import json
import shutil
import zipfile
from pathlib import Path

import openpyxl
import pytest

from triage.admin_billing_summary.aggregator import build_month_summary
from triage.admin_billing_summary.exporter import build_workbook
from triage.nw_prj_neuron_track_hours.bonita_exporter import tab_name_for_month_key
from triage.artifact_compare import compare_artifacts
from triage.artifact_fingerprint import (
    canonical_package_sha256,
    fingerprint_file,
    raw_sha256,
    semantic_sha256,
)
from triage.artifact_profiles import load_profile, run_profile_checks
from tests.fixtures.admin_billing_summary.builders import build

REPO_ROOT = Path(__file__).resolve().parent.parent
FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures" / "admin_billing_summary"


@pytest.fixture(scope="module")
def april_workbook(tmp_path_factory):
    fixtures = build(FIXTURE_DIR)
    april = build_month_summary(str(fixtures["roster"]), "2026-04")
    out = tmp_path_factory.mktemp("artifact_ref") / "April_2026_Billing_Summary_Client.xlsx"
    build_workbook(april, str(out), variant="client")
    return out, tab_name_for_month_key("2026-04")


def _mutate_core_xml_only(src: Path, dst: Path) -> None:
    buf = io.BytesIO()
    new_core = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties">'
        "<dc:creator xmlns:dc='http://purl.org/dc/elements/1.1/'>MutatedAgent</dc:creator>"
        "</cp:coreProperties>"
    ).encode("utf-8")
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            if name == "docProps/core.xml":
                zout.writestr(name, new_core)
            else:
                zout.writestr(name, zin.read(name))
    dst.write_bytes(buf.getvalue())


def _make_corpse_xlsx(path: Path) -> None:
    from triage.xlsx_utils import fix_inlinestr

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Start Here"
    ws.cell(row=1, column=1, value="Title")
    wb.save(str(path))
    fix_inlinestr(str(path))
    orig = path.read_bytes()
    ss_items = "".join(f"<si><t>Column{i}</t></si>" for i in range(1, 12))
    new_ss = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        f' count="11" uniqueCount="11">{ss_items}</sst>'
    ).encode("utf-8")
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(orig), "r") as zin, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            if name == "xl/sharedStrings.xml":
                zout.writestr(name, new_ss)
            else:
                zout.writestr(name, zin.read(name))
    path.write_bytes(buf.getvalue())


def _add_hidden_reference_sheet(path: Path, marker: str, *, state: str = "hidden") -> None:
    from triage.xlsx_utils import fix_inlinestr

    wb = openpyxl.load_workbook(str(path))
    ws = wb.create_sheet("Reference Scratch")
    ws.sheet_state = state
    ws["A1"] = "Reference Scratch"
    ws["B2"] = marker
    wb.save(str(path))
    wb.close()
    fix_inlinestr(str(path))


def test_raw_differs_semantic_matches(april_workbook, tmp_path):
    ref_path, neuron_tab = april_workbook
    cand_path = tmp_path / "meta_mutated.xlsx"
    _mutate_core_xml_only(ref_path, cand_path)
    assert raw_sha256(ref_path) != raw_sha256(cand_path)
    assert canonical_package_sha256(ref_path) == canonical_package_sha256(cand_path)
    assert semantic_sha256(ref_path) == semantic_sha256(cand_path)
    report = compare_artifacts(
        str(ref_path), str(cand_path), "admin_billing_summary",
        expect_neuron_tab=neuron_tab,
    )
    assert report["semantic_sha_match"] is True
    assert report["compare_pass"] is True
    assert "raw_sha256_mismatch" in report["profile_warnings"]


def test_hidden_sheet_diff_does_not_poison_visible_semantic_hash(april_workbook, tmp_path):
    ref_path, neuron_tab = april_workbook
    ref_with_hidden = tmp_path / "ref_hidden.xlsx"
    cand_with_hidden = tmp_path / "cand_hidden.xlsx"
    shutil.copy2(ref_path, ref_with_hidden)
    shutil.copy2(ref_path, cand_with_hidden)

    _add_hidden_reference_sheet(ref_with_hidden, "alpha")
    _add_hidden_reference_sheet(cand_with_hidden, "bravo")

    assert semantic_sha256(ref_with_hidden) == semantic_sha256(cand_with_hidden)
    assert semantic_sha256(ref_with_hidden, include_hidden=True) != semantic_sha256(
        cand_with_hidden,
        include_hidden=True,
    )

    ref_fp = fingerprint_file(ref_with_hidden)
    cand_fp = fingerprint_file(cand_with_hidden)
    assert ref_fp.semantic_sha256 == cand_fp.semantic_sha256
    assert ref_fp.all_sheets_semantic_sha256 != cand_fp.all_sheets_semantic_sha256

    report = compare_artifacts(
        str(ref_with_hidden), str(cand_with_hidden), "admin_billing_summary",
        expect_neuron_tab=neuron_tab,
    )
    assert report["semantic_sha_match"] is True
    assert report["all_sheets_semantic_sha_match"] is False
    assert report["semantic_compare"] == "PASS"
    assert report["compare_pass"] is True


def test_column_corpse_fails_semantic_compare(april_workbook, tmp_path):
    ref_path, neuron_tab = april_workbook
    corpse = tmp_path / "corpse.xlsx"
    _make_corpse_xlsx(corpse)
    report = compare_artifacts(
        str(ref_path), str(corpse), "admin_billing_summary",
        expect_neuron_tab=neuron_tab,
    )
    assert report["semantic_compare"] == "FAIL"
    assert report["compare_pass"] is False
    assert any("generic" in f for f in report["profile_failures"])


def test_missing_sheet_fails_profile(tmp_path):
    p = tmp_path / "tiny.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Only"
    wb.save(str(p))
    prof = load_profile("internal_admin_log")
    res = run_profile_checks(str(p), prof)
    assert any("missing_sheet:Start Here" in f for f in res.failures)


def test_missing_required_header_bonita(tmp_path):
    p = tmp_path / "bad_bonita.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apr 26"
    ws.cell(row=1, column=2, value="WRONG")
    wb.create_sheet("May 26")
    wb.save(str(p))
    prof = load_profile("bonita_neuron_track_hours")
    res = run_profile_checks(str(p), prof)
    assert any("missing_header" in f for f in res.failures)


def test_neuron_total_change_fails_compare(april_workbook, tmp_path):
    ref_path, neuron_tab = april_workbook
    cand_path = tmp_path / "tweaked.xlsx"
    import shutil
    shutil.copy2(ref_path, cand_path)
    wb = openpyxl.load_workbook(str(cand_path))
    ws = wb["Executive Dashboard"]
    for row in ws.iter_rows(min_row=1, max_row=30):
        if row[0].value and str(row[0].value).strip() == "Neuron Net":
            row[1].value = 999.0
            break
    wb.save(str(cand_path))
    wb.close()
    report = compare_artifacts(
        str(ref_path), str(cand_path), "admin_billing_summary",
        expect_neuron_tab=neuron_tab,
    )
    assert report["semantic_compare"] == "FAIL"
    assert any("total_mismatch" in f for f in report["profile_failures"])


def test_approved_delta_allows_semantic_drift(april_workbook, tmp_path):
    ref_path, neuron_tab = april_workbook
    cand_path = tmp_path / "tweaked.xlsx"
    import shutil
    shutil.copy2(ref_path, cand_path)
    wb = openpyxl.load_workbook(str(cand_path))
    ws = wb["Executive Dashboard"]
    for row in ws.iter_rows(min_row=1, max_row=30):
        if row[0].value and str(row[0].value).strip() == "Neuron Net":
            row[1].value = 999.0
            break
    wb.save(str(cand_path))
    wb.close()
    cand_fp = fingerprint_file(cand_path)
    delta_path = tmp_path / "delta.json"
    delta_path.write_text(json.dumps({
        "allow_candidate_semantic_sha256": cand_fp.semantic_sha256,
        "reason": "test approved neuron net correction",
        "approved_utc": "2026-06-03T12:00:00Z",
        "scope": "test_fixture",
    }), encoding="utf-8")
    report = compare_artifacts(
        str(ref_path), str(cand_path), "admin_billing_summary",
        approved_delta=str(delta_path),
        expect_neuron_tab=neuron_tab,
    )
    assert report["semantic_sha_match"] is False
    assert report["semantic_compare"] == "PASS"
    assert "semantic_sha256_mismatch_approved_delta" in report["profile_warnings"]
    assert report["approved_delta_applied"] is True


def test_fingerprint_roundtrip_keys(april_workbook):
    ref_path, _ = april_workbook
    fp = fingerprint_file(ref_path)
    d = fp.to_dict()
    assert "raw_sha256" in d and "canonical_package_sha256" in d and "semantic_sha256" in d
    assert "all_sheets_semantic_sha256" in d


def test_load_profiles():
    for name in ("admin_billing_summary", "bonita_neuron_track_hours", "internal_admin_log"):
        prof = load_profile(name)
        assert prof.profile == name


def test_delta_missing_audit_fields_fails_semantic_compare(april_workbook, tmp_path):
    ref_path, neuron_tab = april_workbook
    cand_path = tmp_path / "tweaked2.xlsx"
    import shutil
    shutil.copy2(ref_path, cand_path)
    wb = openpyxl.load_workbook(str(cand_path))
    ws = wb["Executive Dashboard"]
    for row in ws.iter_rows(min_row=1, max_row=30):
        if row[0].value and str(row[0].value).strip() == "Neuron Net":
            row[1].value = 888.0
            break
    wb.save(str(cand_path))
    wb.close()
    cand_fp = fingerprint_file(cand_path)
    delta_path = tmp_path / "bad_delta.json"
    delta_path.write_text(json.dumps({
        "allow_candidate_semantic_sha256": cand_fp.semantic_sha256,
    }), encoding="utf-8")
    report = compare_artifacts(
        str(ref_path), str(cand_path), "admin_billing_summary",
        approved_delta=str(delta_path),
        expect_neuron_tab=neuron_tab,
    )
    assert report["semantic_compare"] == "FAIL"
    assert any("approved_delta_missing" in f for f in report["profile_failures"])


def test_forbidden_confidence_in_shared_strings_fails(generated_bonita_workbook):
    path = Path(generated_bonita_workbook)
    wb = openpyxl.load_workbook(str(path))
    wb["Apr 26"].cell(row=3, column=7, value="confidence heuristic leak")
    wb.save(str(path))
    wb.close()
    prof = load_profile("bonita_neuron_track_hours")
    res = run_profile_checks(str(path), prof)
    assert any(
        "forbidden_shared_string" in f or "forbidden_cell_text" in f
        for f in res.failures
    )


def test_required_nonblank_column_empty_fails(generated_bonita_workbook):
    """Row with hours but blank PROJECT must fail profile checks."""
    path = Path(generated_bonita_workbook)
    import shutil
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        bad = Path(td) / path.name
        shutil.copy2(path, bad)
        wb = openpyxl.load_workbook(str(bad))
        ws = wb["Apr 26"]
        col_proj = None
        col_total = None
        for c in range(1, (ws.max_column or 1) + 1):
            h = str(ws.cell(row=1, column=c).value or "").strip().upper()
            if h == "PROJECT":
                col_proj = c
            if h == "TOTAL":
                col_total = c
        assert col_proj and col_total
        for row in range(3, (ws.max_row or 3) + 1):
            if isinstance(ws.cell(row=row, column=col_total).value, (int, float)):
                ws.cell(row=row, column=col_proj).value = None
                break
        wb.save(str(bad))
        wb.close()
        prof = load_profile("bonita_neuron_track_hours")
        res = run_profile_checks(str(bad), prof)
        assert any("required_nonblank_column_empty" in f for f in res.failures)


@pytest.fixture(scope="module")
def generated_bonita_workbook(fixtures_bonita, tmp_path_factory):
    from tests.fixtures.nw_prj_neuron_track_hours.bonita_fixtures import build_bonita_fixtures
    from triage.nw_prj_neuron_track_hours.bonita_cli import run as bonita_run

    FIX = Path(__file__).resolve().parent / "fixtures" / "nw_prj_neuron_track_hours"
    fx = build_bonita_fixtures(FIX)
    out = tmp_path_factory.mktemp("bonita_nb")
    manifest = bonita_run(
        roster_log=str(fx["roster"]),
        out_dir=str(out),
        months=["2026-04", "2026-05"],
        websafe=True,
        repo_root=REPO_ROOT,
    )
    return manifest["outputs"]["workbook"]


@pytest.fixture(scope="module")
def fixtures_bonita():
    from tests.fixtures.nw_prj_neuron_track_hours.bonita_fixtures import build_bonita_fixtures
    FIX = Path(__file__).resolve().parent / "fixtures" / "nw_prj_neuron_track_hours"
    return build_bonita_fixtures(FIX)
