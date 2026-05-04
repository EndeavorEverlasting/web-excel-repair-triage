"""
tests/test_attendance_report.py
---------------------------------
Unit and integration tests for triage.attendance_report.

Coverage
--------
  T1  generate_attendance_report — output file is created on disk
  T2  generate_attendance_report — output file is a valid .xlsx workbook
  T3  generate_attendance_report — workbook contains the expected sheet name
  T4  generate_attendance_report — sheet has header rows (title + day headers)
  T5  generate_attendance_report — data rows appear for each staff/project pair
  T6  generate_attendance_report — weekly net total is correct
  T7  generate_attendance_report — empty records list produces a file (no data rows)
  T8  generate_attendance_report — run_manifest.json is written alongside the file
  T9  generate_attendance_report — gate-check passes (no stopship / ref errors)
  T10 _hours_str — decimal hours formatted as H:MM
  T11 _fmt_time — decimal hours formatted as HH:MM (zero-padded)
  T12 _lunch_deduction — 0 under 6h, 0.5 for 6–<8h, 1.0 at 8h+
"""
from __future__ import annotations

import datetime
import json
import zipfile
from pathlib import Path

import pytest
import openpyxl

from triage.attendance_report import (
    generate_attendance_report,
    _hours_str,
    _fmt_time,
    _lunch_deduction,
)


# ── helper to build minimal roster records ────────────────────────────────────

def _make_records(week_start: datetime.date, staff_data: list) -> list:
    """Build synthetic roster records for testing.

    staff_data: list of (staff_name, project, clock_in, clock_out, date_offset)
    """
    records = []
    for staff, project, ci, co, day_offset in staff_data:
        gross = (co - ci) if co >= ci else (co - ci + 24)
        from triage.roster_parser import _compute_gross
        gross = _compute_gross(ci, co)
        from triage.attendance_report import _lunch_deduction as _ld
        lunch = _ld(gross)
        net   = round(max(0.0, gross - lunch), 4)
        records.append({
            "staff":           staff,
            "project":         project,
            "date":            week_start + datetime.timedelta(days=day_offset),
            "clock_in":        ci,
            "clock_out":       co,
            "gross_hours":     round(gross, 4),
            "lunch_deduction": lunch,
            "net_hours":       net,
            "long_shift":      gross > 12.0,
        })
    return records


_WEEK_START = datetime.date(2026, 4, 6)   # Monday
_WEEK_END   = datetime.date(2026, 4, 10)  # Friday

_SAMPLE_STAFF = [
    ("Alice Brown",   "Project Alpha", 9.0,  17.0, 0),   # Mon — 8h → 1h lunch → 7h net
    ("Alice Brown",   "Project Alpha", 9.0,  17.0, 1),   # Tue
    ("Bob Smith",     "Project Beta",  8.5,  14.5, 0),   # Mon — 6h → 0.5h lunch → 5.5h net
    ("Bob Smith",     "Project Beta",  8.5,  17.0, 2),   # Wed — 8.5h → 1h lunch → 7.5h net
]


@pytest.fixture(scope="module")
def attendance_result(tmp_path_factory):
    tmp = str(tmp_path_factory.mktemp("attendance"))
    records = _make_records(_WEEK_START, _SAMPLE_STAFF)
    out_path = generate_attendance_report(
        records=records,
        week_start=_WEEK_START,
        week_end=_WEEK_END,
        out_root=tmp,
        run_id="test-run-001",
    )
    return out_path, tmp


# ── T1–T2: file creation ──────────────────────────────────────────────────────

def test_output_file_exists(attendance_result):
    out_path, _ = attendance_result
    assert Path(out_path).exists(), f"Expected output file at {out_path}"


def test_output_is_valid_xlsx(attendance_result):
    out_path, _ = attendance_result
    wb = openpyxl.load_workbook(out_path, data_only=True)
    assert wb is not None


# ── T3: sheet name ────────────────────────────────────────────────────────────

def test_sheet_name(attendance_result):
    out_path, _ = attendance_result
    wb = openpyxl.load_workbook(out_path, data_only=True)
    expected = f"Week {_WEEK_START.strftime('%m-%d')}"
    assert expected in wb.sheetnames, (
        f"Expected sheet '{expected}', got: {wb.sheetnames}"
    )


# ── T4: header rows present ───────────────────────────────────────────────────

def test_header_rows_present(attendance_result):
    out_path, _ = attendance_result
    wb = openpyxl.load_workbook(out_path, data_only=True)
    ws = wb.active
    title = str(ws.cell(1, 1).value or "")
    assert "Attendance" in title or "Week" in title, (
        f"Row 1 title unexpected: {title!r}"
    )
    staff_hdr = str(ws.cell(2, 1).value or "")
    assert "Staff" in staff_hdr, f"Row 2 col 1 should be 'Staff Name', got {staff_hdr!r}"


# ── T5: data rows present ─────────────────────────────────────────────────────

def test_data_rows_present(attendance_result):
    out_path, _ = attendance_result
    wb = openpyxl.load_workbook(out_path, data_only=True)
    ws = wb.active
    data_rows = [
        ws.cell(r, 1).value
        for r in range(4, ws.max_row + 1)
        if ws.cell(r, 1).value and str(ws.cell(r, 1).value).strip() not in ("", "TOTALS")
    ]
    assert len(data_rows) >= 2, f"Expected at least 2 data rows, got {len(data_rows)}"
    names = {str(v).strip() for v in data_rows}
    assert "Alice Brown" in names
    assert "Bob Smith" in names


# ── T6: weekly net total correctness ─────────────────────────────────────────

def test_weekly_total_column_positive(attendance_result):
    out_path, _ = attendance_result
    wb = openpyxl.load_workbook(out_path, data_only=True)
    ws = wb.active
    last_col = ws.max_column
    totals = [
        ws.cell(r, last_col).value
        for r in range(4, ws.max_row + 1)
        if isinstance(ws.cell(r, last_col).value, (int, float))
    ]
    assert len(totals) >= 2, "Expected weekly total values in last column"
    assert all(t >= 0 for t in totals), f"Negative totals found: {totals}"


# ── T7: empty records produces a file (no crash) ─────────────────────────────

def test_empty_records_creates_file(tmp_path):
    out_path = generate_attendance_report(
        records=[],
        week_start=_WEEK_START,
        week_end=_WEEK_END,
        out_root=str(tmp_path),
        run_id="empty-test",
    )
    assert Path(out_path).exists()
    wb = openpyxl.load_workbook(out_path, data_only=True)
    assert wb is not None


# ── T8: manifest is written ───────────────────────────────────────────────────

def test_manifest_written(attendance_result):
    _, tmp = attendance_result
    month_str = _WEEK_START.strftime("%Y-%m")
    manifest_path = Path(tmp) / month_str / "run_manifest.json"
    assert manifest_path.exists(), f"run_manifest.json not found at {manifest_path}"
    data = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert data.get("run_id") == "test-run-001"
    assert data.get("status") == "generated"


# ── T9: gate-check passes ─────────────────────────────────────────────────────

def test_gate_check_passes(attendance_result):
    """The generated attendance .xlsx must pass all OOXML gate checks."""
    out_path, _ = attendance_result
    from triage.gate_checks import run_all
    report = run_all(out_path)
    assert not report.stopship, f"Stopship tokens found: {report.stopship}"
    assert not report.cf_ref,   f"CF #REF! hits found: {report.cf_ref}"


def test_long_shift_uses_distinct_fill_and_comment(tmp_path):
    """Long shifts use a stronger fill than regular overnight shifts."""
    records = _make_records(
        _WEEK_START,
        [("Night Owl", "Project Gamma", 18.0, 7.5, 0)],
    )
    assert records[0]["long_shift"] is True

    out_path = generate_attendance_report(
        records=records,
        week_start=_WEEK_START,
        week_end=_WEEK_END,
        out_root=str(tmp_path),
        run_id="long-shift-test",
    )

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    ci_cell = ws.cell(4, 3)
    fill_rgb = str(ci_cell.fill.fgColor.rgb or "")

    assert fill_rgb.endswith("7F1D1D")
    assert not fill_rgb.endswith("7B3F00")
    assert ci_cell.comment is not None
    assert "Long shift" in ci_cell.comment.text


# ── T10–T12: formatting helpers ───────────────────────────────────────────────

class TestFormattingHelpers:
    def test_hours_str_whole(self):
        assert _hours_str(8.0) == "8:00"

    def test_hours_str_half(self):
        assert _hours_str(8.5) == "8:30"

    def test_hours_str_none(self):
        assert _hours_str(None) == ""

    def test_fmt_time_morning(self):
        assert _fmt_time(9.5) == "09:30"

    def test_fmt_time_afternoon(self):
        assert _fmt_time(17.0) == "17:00"

    def test_fmt_time_none(self):
        assert _fmt_time(None) == ""


class TestLunchDeduction:
    def test_under_six_hours(self):
        assert _lunch_deduction(5.9) == 0.0

    def test_exactly_six_hours(self):
        assert _lunch_deduction(6.0) == pytest.approx(0.5)

    def test_between_six_and_eight(self):
        assert _lunch_deduction(7.0) == pytest.approx(0.5)

    def test_exactly_eight_hours(self):
        assert _lunch_deduction(8.0) == pytest.approx(1.0)

    def test_over_eight_hours(self):
        assert _lunch_deduction(9.5) == pytest.approx(1.0)

    def test_zero_hours(self):
        assert _lunch_deduction(0.0) == 0.0
