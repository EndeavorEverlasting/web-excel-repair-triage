"""CLI-level E2E tests for billing context exporter (synthetic workbooks only)."""

from __future__ import annotations

import io
import json
import sys
import zipfile
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from tests.fixtures.billing_context.fixtures import write_all_fixtures
from triage.billing_context.cli import main
from triage.billing_context.exporters import LEADERSHIP_HEADERS


def _run_cli(
    out_dir: Path,
    inputs: dict[str, Path],
    *,
    html: bool = True,
    zip_bundle: bool = True,
    internal_xlsx: bool = True,
    include_tracker_import: bool = False,
) -> dict:
    argv = [
        "cli",
        "--track-hours",
        str(inputs["track_hours"]),
        "--april-context",
        str(inputs["april_context"]),
        "--roster-log",
        str(inputs["roster_log"]),
        "--admin-copy",
        str(inputs["admin_copy"]),
        "--dashboard",
        str(inputs["dashboard"]),
        "--out-dir",
        str(out_dir),
    ]
    if html:
        argv.append("--html")
    if zip_bundle:
        argv.append("--zip")
    if internal_xlsx:
        argv.append("--internal-xlsx")
    if include_tracker_import:
        argv.append("--include-tracker-import")

    buf = io.StringIO()
    with patch.object(sys, "argv", argv):
        with patch.object(sys, "stdout", buf):
            main()
    return json.loads(buf.getvalue())


@pytest.fixture()
def synthetic_inputs(tmp_path: Path) -> dict[str, Path]:
    return write_all_fixtures(tmp_path / "inputs")


def test_cli_e2e_produces_zip_and_manifest(synthetic_inputs, tmp_path: Path):
    payload = _run_cli(tmp_path / "outputs", synthetic_inputs)

    assert payload["entry_count"] >= 2
    assert payload["total_hours"] > 0
    assert all(m["exists"] for m in payload["manifest"])

    zip_path = Path(payload["outputs"]["zip_bundle"])
    assert zip_path.exists()
    with zipfile.ZipFile(zip_path) as zf:
        names = set(zf.namelist())
    assert "billing_context_dashboard.html" in names
    assert "billing_context_mismatches.csv" in names


def test_leadership_xlsx_has_clean_headers_only(synthetic_inputs, tmp_path: Path):
    payload = _run_cli(tmp_path / "outputs", synthetic_inputs)
    project_path = Path(payload["outputs"]["project_hours"])
    may_path = Path(payload["outputs"]["may_summary"])

    assert "Tracker Import" not in openpyxl.load_workbook(may_path, read_only=True).sheetnames

    for path in (project_path, may_path):
        wb = openpyxl.load_workbook(path, read_only=True)
        for ws in wb.worksheets:
            if ws.title in ("Admin Summary", "Work Context Summary", "Technician Summary", "Reporting Batch Summary"):
                continue
            if "Summary" in ws.title or ws.title.endswith("2026"):
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                if headers == LEADERSHIP_HEADERS:
                    for forbidden in ("Source Sheet", "Source Row", "Original Assignment", "Context Reason"):
                        assert forbidden not in headers
        wb.close()


def test_internal_xlsx_only_when_requested(synthetic_inputs, tmp_path: Path):
    payload = _run_cli(tmp_path / "outputs", synthetic_inputs, internal_xlsx=True)
    assert Path(payload["outputs"]["internal_detail"]).exists()

    payload_no = _run_cli(tmp_path / "outputs_no_internal", synthetic_inputs, internal_xlsx=False)
    assert "internal_detail" not in payload_no["outputs"]


def test_mismatch_csv_formula_prefix_neutralized(synthetic_inputs, tmp_path: Path):
    payload = _run_cli(tmp_path / "outputs", synthetic_inputs)
    csv_path = Path(payload["outputs"]["mismatches_csv"])
    text = csv_path.read_text(encoding="utf-8")
    for line in text.splitlines()[1:]:
        for field in line.split(","):
            if field.startswith(("=", "+", "-", "@")) and not field.startswith("'"):
                pytest.fail(f"Unneutralized CSV field: {field!r}")


def test_html_dashboard_has_escape_plumbing(synthetic_inputs, tmp_path: Path):
    payload = _run_cli(tmp_path / "outputs", synthetic_inputs)
    html = Path(payload["outputs"]["html_dashboard"]).read_text(encoding="utf-8")
    assert "function esc(" in html
    assert "https://" not in html
