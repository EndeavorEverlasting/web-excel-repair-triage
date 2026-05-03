"""
tests/test_billing_regression.py
---------------------------------
Structural regression test: compare key cells, headers, and section layout of
the generated billing workbook against the reference Agilant Admins template.

NOTE on record-count discrepancy
---------------------------------
The reference template was built from an earlier version of the roster file.
The current roster (Active_Roster_Log_5_1_2026_Billing_April_Pack_...) has 14
additional records (late additions by staff Alejandro Perales, Emmanuel Sanchez,
Geoff Gerber, Khadejah Harrison, Md Suhan Newaz, Valentyn Nykoliuk for dates
Apr 20–30) that were not present when the reference was created.  The reference
shows 89 rows / 722.92 gross; the current roster produces 103 rows / 826.97 gross.
This is a data-version difference, not a pipeline bug.  The tests below verify
that every structural property of the generated output matches the reference
template regardless of the roster-version difference.
"""
import os
import tempfile
import datetime
import pytest
import openpyxl

TEMPLATE = os.path.join(
    os.path.dirname(__file__),
    "..",
    "attached_assets",
    "Agilant_Admins_April_2026_Billing_Summary_PO176759_Neurons_1777807804119.xlsx",
)
ROSTER = os.path.join(
    os.path.dirname(__file__),
    "..",
    "attached_assets",
    "Active_Roster_Log_5_1_2026_Billing_April_Pack_(1)_1777807743057.xlsx",
)
TECH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "attached_assets",
    "NW_PRJ_Tech_hours_-_5-1-2026_1777807743058.xlsx",
)

pytestmark = pytest.mark.skipif(
    not os.path.exists(TEMPLATE) or not os.path.exists(ROSTER),
    reason="Reference assets not present",
)


# ── helpers ────────────────────────────────────────────────────────────────────

def _cell(ws, row, col):
    """Return stripped string value of a cell."""
    return str(ws.cell(row, col).value or "").strip()


def _find_row(ws, text, col=1, start=1):
    """Return row number where col has text, or None."""
    for r in range(start, ws.max_row + 1):
        if _cell(ws, r, col) == text:
            return r
    return None


def _build_generated(tmp_dir):
    """Build a fresh billing workbook from sample assets and return its path."""
    from triage.roster_parser import parse_roster
    from triage.tech_hours_parser import parse_tech_hours
    from triage.billing_summary_generator import generate_billing_summary

    malformed = []
    records = parse_roster(ROSTER, target_month="April 2026", malformed_out=malformed)

    tech_records = []
    if os.path.exists(TECH):
        tech_records = parse_tech_hours(
            path=TECH,
            sheet_name="Project Team",
            project="Projects Team",
            target_month="2026-04",
        )
    tech_keys = {(r["staff"], r["date"]) for r in tech_records}
    merged = []
    for r in records:
        if r["gross_hours"] <= 0:
            continue
        rec = dict(r)
        if (rec["staff"], rec["date"]) in tech_keys:
            rec["project"] = "Projects Team"
        merged.append(rec)

    return generate_billing_summary(
        records=merged,
        invoices=[],
        billing_month="2026-04",
        out_root=tmp_dir,
    )


# ── fixtures ───────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def ref_wb():
    return openpyxl.load_workbook(TEMPLATE, data_only=True)


@pytest.fixture(scope="module")
def gen_wb(tmp_path_factory):
    tmp = str(tmp_path_factory.mktemp("billing"))
    path = _build_generated(tmp)
    return openpyxl.load_workbook(path, data_only=True)


# ── tests: Billing Summary sheet ───────────────────────────────────────────────

class TestBillingSummarySheet:

    def test_sheet_exists(self, gen_wb):
        assert "Billing Summary - April 2026" in gen_wb.sheetnames, (
            "Missing sheet 'Billing Summary - April 2026'"
        )

    def test_title_row(self, gen_wb, ref_wb):
        ws_g = gen_wb["Billing Summary - April 2026"]
        ws_r = ref_wb["Billing Summary - April 2026"]
        assert _cell(ws_g, 1, 1) == _cell(ws_r, 1, 1), (
            f"Title mismatch: gen={_cell(ws_g, 1, 1)!r}, ref={_cell(ws_r, 1, 1)!r}"
        )

    def test_monthly_rollup_header(self, gen_wb, ref_wb):
        ws_g = gen_wb["Billing Summary - April 2026"]
        ws_r = ref_wb["Billing Summary - April 2026"]
        # Both should have "Monthly Rollup" in col 1
        row_g = _find_row(ws_g, "Monthly Rollup")
        row_r = _find_row(ws_r, "Monthly Rollup")
        assert row_g is not None, "Missing 'Monthly Rollup' in generated sheet"
        assert row_r is not None, "Missing 'Monthly Rollup' in reference sheet"

    def test_rollup_metrics_present(self, gen_wb):
        ws = gen_wb["Billing Summary - April 2026"]
        for label in ("Gross hours", "Lunch deducted", "Net billable hours"):
            assert _find_row(ws, label) is not None, (
                f"Missing rollup metric: {label!r}"
            )

    def test_hours_by_project_header(self, gen_wb):
        ws = gen_wb["Billing Summary - April 2026"]
        row = _find_row(ws, "Monthly Rollup")
        assert row is not None
        assert _cell(ws, row, 6) == "Hours by Project", (
            f"Expected 'Hours by Project' at R{row}C6"
        )

    def test_hours_by_project_column_labels(self, gen_wb, ref_wb):
        ws_g = gen_wb["Billing Summary - April 2026"]
        ws_r = ref_wb["Billing Summary - April 2026"]
        # "Project" header is in col 6 (not col 1) — use _find_row on col 6
        hdr_r = _find_row(ws_r, "Project", col=6)
        hdr_g = _find_row(ws_g, "Project", col=6)
        assert hdr_g is not None, "Missing 'Project' column label (col 6) in generated"
        ref_labels = [_cell(ws_r, hdr_r, c) for c in range(6, 12)]
        gen_labels = [_cell(ws_g, hdr_g, c) for c in range(6, 12)]
        assert gen_labels == ref_labels, (
            f"Hours-by-Project columns mismatch:\n  ref={ref_labels}\n  gen={gen_labels}"
        )

    def test_val_weekly_hours_present(self, gen_wb):
        ws = gen_wb["Billing Summary - April 2026"]
        row = _find_row(ws, "Val Weekly Hours")
        assert row is not None, "Missing 'Val Weekly Hours' section"
        assert _cell(ws, row, 9) == "Val Daily Detail", (
            f"Expected 'Val Daily Detail' at R{row}C9, got {_cell(ws, row, 9)!r}"
        )

    def test_daily_detail_header_columns(self, gen_wb, ref_wb):
        ws_g = gen_wb["Billing Summary - April 2026"]
        ws_r = ref_wb["Billing Summary - April 2026"]
        hdr_r = _find_row(ws_r, "Staff Name")
        hdr_g = _find_row(ws_g, "Staff Name")
        assert hdr_g is not None, "Missing daily detail 'Staff Name' header"
        ref_cols = [_cell(ws_r, hdr_r, c) for c in range(1, 11)]
        gen_cols = [_cell(ws_g, hdr_g, c) for c in range(1, 11)]
        assert gen_cols == ref_cols, (
            f"Daily-detail header columns mismatch:\n  ref={ref_cols}\n  gen={gen_cols}"
        )

    def test_daily_detail_has_data_rows(self, gen_wb):
        ws = gen_wb["Billing Summary - April 2026"]
        hdr = _find_row(ws, "Staff Name")
        assert hdr is not None
        data_rows = 0
        for r in range(hdr + 1, ws.max_row + 1):
            v = _cell(ws, r, 1)
            if v and v != "TOTAL":
                data_rows += 1
        assert data_rows > 0, "No data rows in daily billing detail"

    def test_gross_hours_positive(self, gen_wb):
        ws = gen_wb["Billing Summary - April 2026"]
        hdr = _find_row(ws, "Staff Name")
        gross_total = 0.0
        for r in range(hdr + 1, ws.max_row + 1):
            v = ws.cell(r, 7).value
            if isinstance(v, (int, float)):
                gross_total += v
        assert gross_total > 0, "Gross hours total must be positive"

    def test_net_le_gross(self, gen_wb):
        ws = gen_wb["Billing Summary - April 2026"]
        hdr = _find_row(ws, "Staff Name")
        gross_total = net_total = 0.0
        for r in range(hdr + 1, ws.max_row + 1):
            g = ws.cell(r, 7).value
            n = ws.cell(r, 9).value
            if isinstance(g, (int, float)):
                gross_total += g
            if isinstance(n, (int, float)):
                net_total += n
        assert net_total <= gross_total, (
            f"Net {net_total:.2f} > Gross {gross_total:.2f}"
        )

    def test_record_count_matches_current_roster(self, gen_wb):
        """Generated row count should match what the CURRENT roster produces.

        NOTE: the reference template shows 89 rows because it was built from an
        earlier roster version.  The current roster has 14 additional records
        (late entries for Apr 20–30).  This test asserts parity with the current
        roster, not the reference.  See module docstring for full context.
        """
        from triage.roster_parser import parse_roster
        malformed = []
        records = parse_roster(ROSTER, target_month="April 2026", malformed_out=malformed)
        valid = [r for r in records if r["gross_hours"] > 0]

        ws = gen_wb["Billing Summary - April 2026"]
        hdr = _find_row(ws, "Staff Name")
        data_rows = sum(
            1 for r in range(hdr + 1, ws.max_row + 1)
            if _cell(ws, r, 1) and _cell(ws, r, 1) != "TOTAL"
        )
        assert data_rows == len(valid), (
            f"Row count: generated={data_rows}, roster_valid={len(valid)}"
        )


# ── tests: Invoice Pivots sheet ────────────────────────────────────────────────

class TestInvoicePivotsSheet:

    def test_sheet_exists(self, gen_wb):
        assert "Invoice Pivots - Candidate" in gen_wb.sheetnames

    def test_monthly_invoice_totals_header(self, gen_wb):
        ws = gen_wb["Invoice Pivots - Candidate"]
        assert _cell(ws, 4, 1) == "Monthly Invoice Totals", (
            f"Expected 'Monthly Invoice Totals' at R4C1, got {_cell(ws, 4, 1)!r}"
        )

    def test_totals_by_category_header(self, gen_wb):
        ws = gen_wb["Invoice Pivots - Candidate"]
        assert _cell(ws, 4, 10) == "Totals by Category", (
            f"Expected 'Totals by Category' at R4C10, got {_cell(ws, 4, 10)!r}"
        )

    def test_monthly_totals_column_labels(self, gen_wb, ref_wb):
        ws_g = gen_wb["Invoice Pivots - Candidate"]
        ws_r = ref_wb["Invoice Pivots - Candidate"]
        ref_labels = [_cell(ws_r, 5, c) for c in range(1, 9)]
        gen_labels = [_cell(ws_g, 5, c) for c in range(1, 9)]
        assert gen_labels == ref_labels, (
            f"Monthly totals column labels mismatch:\n  ref={ref_labels}\n  gen={gen_labels}"
        )

    def test_category_column_labels(self, gen_wb, ref_wb):
        ws_g = gen_wb["Invoice Pivots - Candidate"]
        ws_r = ref_wb["Invoice Pivots - Candidate"]
        ref_labels = [_cell(ws_r, 5, c) for c in range(10, 15)]
        gen_labels = [_cell(ws_g, 5, c) for c in range(10, 15)]
        assert gen_labels == ref_labels, (
            f"Category column labels mismatch:\n  ref={ref_labels}\n  gen={gen_labels}"
        )

    def test_vendor_section_headers(self, gen_wb, ref_wb):
        ws_g = gen_wb["Invoice Pivots - Candidate"]
        ws_r = ref_wb["Invoice Pivots - Candidate"]
        # Find "Totals by PO Number" in reference
        po_row_r = None
        for r in range(5, ws_r.max_row + 1):
            if _cell(ws_r, r, 1) == "Totals by PO Number":
                po_row_r = r
                break
        assert po_row_r is not None, "Reference missing 'Totals by PO Number'"

        # Reference sub-header positions
        ref_sections = {
            "Totals by Project":         (po_row_r, 6),
            "Totals by Vendor / Crew":   (po_row_r, 10),
        }

        po_row_g = None
        for r in range(5, ws_g.max_row + 1):
            if _cell(ws_g, r, 1) == "Totals by PO Number":
                po_row_g = r
                break
        assert po_row_g is not None, "Generated output missing 'Totals by PO Number'"
        assert _cell(ws_g, po_row_g, 6)  == "Totals by Project", (
            f"Expected 'Totals by Project' at R{po_row_g}C6"
        )
        assert _cell(ws_g, po_row_g, 10) == "Totals by Vendor / Crew", (
            f"Expected 'Totals by Vendor / Crew' at R{po_row_g}C10"
        )

    def test_vendor_sub_header_columns(self, gen_wb, ref_wb):
        ws_g = gen_wb["Invoice Pivots - Candidate"]
        ws_r = ref_wb["Invoice Pivots - Candidate"]
        po_row_g = next(
            r for r in range(5, ws_g.max_row + 1)
            if _cell(ws_g, r, 1) == "Totals by PO Number"
        )
        sub_row = po_row_g + 1
        po_row_r = next(
            r for r in range(5, ws_r.max_row + 1)
            if _cell(ws_r, r, 1) == "Totals by PO Number"
        )
        sub_row_r = po_row_r + 1

        for col, label_key in [(10, "Vendor / Crew"), (12, "Trucking"), (14, "Total")]:
            ref_lbl = _cell(ws_r, sub_row_r, col)
            gen_lbl = _cell(ws_g, sub_row, col)
            assert gen_lbl == ref_lbl, (
                f"Vendor sub-header col {col}: ref={ref_lbl!r}, gen={gen_lbl!r}"
            )


# ── tests: strict malformed-row handling ───────────────────────────────────────

class TestMalformedRowHandling:

    def test_strict_mode_raises(self):
        """parse_roster without malformed_out must raise on the Patricia Marrero row."""
        from triage.roster_parser import parse_roster, RosterParseError
        with pytest.raises(RosterParseError, match="Patricia Marrero"):
            parse_roster(ROSTER, target_month="April 2026")

    def test_confirmed_exclusion_succeeds(self):
        """parse_roster with malformed_out collects the bad row and continues."""
        from triage.roster_parser import parse_roster
        malformed = []
        records = parse_roster(ROSTER, target_month="April 2026", malformed_out=malformed)
        assert len(malformed) == 1, f"Expected 1 malformed row, got {len(malformed)}"
        assert "Patricia Marrero" in malformed[0]
        assert len(records) == 103

    def test_excluded_row_not_in_billing(self):
        """The excluded malformed row must not appear in the generated billing summary."""
        from triage.roster_parser import parse_roster
        import datetime

        malformed = []
        records = parse_roster(ROSTER, target_month="April 2026", malformed_out=malformed)
        valid = [r for r in records if r["gross_hours"] > 0]

        malformed_keys = set()
        for m in malformed:
            if "Patricia Marrero" in m and "2026-04-30" in m:
                malformed_keys.add(("Patricia Marrero", datetime.date(2026, 4, 30)))

        for key in malformed_keys:
            hits = [r for r in valid if (r["staff"], r["date"]) == key]
            assert not hits, (
                f"Malformed row {key} appeared in valid records"
            )


# ── tests: data-version discrepancy documented ─────────────────────────────────

class TestDataVersionDiscrepancy:
    """
    Asserts that the 14-record gap between reference (89) and current roster
    (103) is exclusively explained by late-added roster entries, NOT by a
    pipeline bug.  Every row in the reference is present in our parse.
    """

    def test_all_reference_rows_present_in_parse(self):
        """No reference row should be missing from our parse."""
        from triage.roster_parser import parse_roster
        import datetime

        wb = openpyxl.load_workbook(TEMPLATE, data_only=True)
        ws = wb["Billing Summary - April 2026"]

        ref_keys = set()
        in_detail = False
        for r in range(1, ws.max_row + 1):
            v1 = str(ws.cell(r, 1).value or "").strip()
            if "Daily Billing Detail" in v1:
                in_detail = True
                continue
            if not in_detail:
                continue
            if v1 in ("Staff Name", "TOTAL", ""):
                continue
            v2 = ws.cell(r, 2).value
            if v2 is None:
                continue
            d = v2.date() if isinstance(v2, datetime.datetime) else v2
            ref_keys.add((v1, d))

        malformed = []
        records = parse_roster(ROSTER, target_month="April 2026", malformed_out=malformed)
        our_keys = {(r["staff"], r["date"]) for r in records}

        missing = ref_keys - our_keys
        assert not missing, (
            f"Reference rows missing from our parse: {sorted(missing)}"
        )

    def test_extra_rows_are_late_additions(self):
        """The 14 extra rows are all from the late-addition dates (Apr 20–30)."""
        from triage.roster_parser import parse_roster
        import datetime

        wb = openpyxl.load_workbook(TEMPLATE, data_only=True)
        ws = wb["Billing Summary - April 2026"]
        ref_keys = set()
        in_detail = False
        for r in range(1, ws.max_row + 1):
            v1 = str(ws.cell(r, 1).value or "").strip()
            if "Daily Billing Detail" in v1:
                in_detail = True
                continue
            if not in_detail:
                continue
            if v1 in ("Staff Name", "TOTAL", ""):
                continue
            v2 = ws.cell(r, 2).value
            if v2 is None:
                continue
            d = v2.date() if isinstance(v2, datetime.datetime) else v2
            ref_keys.add((v1, d))

        malformed = []
        records = parse_roster(ROSTER, target_month="April 2026", malformed_out=malformed)
        our_keys = {(r["staff"], r["date"]) for r in records}
        extra = our_keys - ref_keys

        # All 14 extra records should have dates >= April 20
        late_threshold = datetime.date(2026, 4, 20)
        early_extras = [
            (s, d) for s, d in extra
            if d < late_threshold
        ]
        assert not early_extras, (
            f"Extra records with dates before Apr 20 — unexpected: {early_extras}"
        )
        assert len(extra) == 14, (
            f"Expected exactly 14 extra (late-addition) records, got {len(extra)}"
        )


# ── tests: cross-month roster parsing ─────────────────────────────────────────

class TestCrossMonthRosterParsing:
    """Regression tests for cross-month attendance weeks.

    The Apr 27–May 1, 2026 week spans April and May.  The provided roster
    workbook contains a 'Live - May 2026' sheet, so parsing should succeed and
    return all records in scope.  When a month sheet is genuinely absent, a
    clear RosterParseError must be raised — no silent partial output.
    """

    def test_apr27_may1_complete_coverage(self):
        """Apr 27–May 1: both Live sheets present → parse succeeds.

        The May 2026 sheet exists in the workbook but has no staff entries for
        May 1 (no-one was rostered).  That is not a silent drop — it simply
        means no work records for that day.  The important assertion is that
        NO RosterParseError is raised and all April 27-30 records are present.
        """
        from triage.roster_parser import parse_roster
        malformed = []
        records = parse_roster(
            ROSTER,
            target_week_start=datetime.date(2026, 4, 27),
            target_week_end=datetime.date(2026, 5, 1),
            malformed_out=malformed,
        )
        april_recs = [r for r in records if r["date"].month == 4]
        may_recs   = [r for r in records if r["date"].month == 5]
        # April 27-30 should have records (Geoff Gerber, Khadejah Harrison)
        assert len(april_recs) > 0, (
            "Expected April 27-30 records in cross-month parse, got none"
        )
        # Each returned record must be within the requested week
        for r in records:
            assert datetime.date(2026, 4, 27) <= r["date"] <= datetime.date(2026, 5, 1), (
                f"Record date {r['date']} outside requested week Apr 27–May 1"
            )

    def test_apr27_may1_strict_mode_raises_on_malformed(self):
        """Cross-month strict mode: Patricia Marrero Apr 30 raises RosterParseError."""
        from triage.roster_parser import parse_roster, RosterParseError
        # Patricia Marrero has a malformed Apr 30 row; Apr 30 falls in the week.
        with pytest.raises(RosterParseError, match="(?i)(malformed|clock|patricia)"):
            parse_roster(
                ROSTER,
                target_week_start=datetime.date(2026, 4, 27),
                target_week_end=datetime.date(2026, 5, 1),
            )

    def test_missing_month_sheet_raises_explicitly(self, tmp_path):
        """Cross-month where a required Live sheet is absent must raise RosterParseError.

        Uses a synthetic minimal workbook that only contains 'Live - April 2026'.
        Parsing the cross-month week Apr 27–May 1 must raise a clear RosterParseError
        naming the missing 'Live - May 2026' sheet — not silently return April-only
        records.
        """
        import openpyxl as _xl
        from triage.roster_parser import parse_roster, RosterParseError

        # Build the minimal April-only workbook with a valid header structure
        wb = _xl.Workbook()
        ws = wb.active
        ws.title = "Live - April 2026"
        # Row 1: title; Row 2: Staff Name header + two date columns
        ws.append(["Attendance Log — April 2026"])
        ws.append(["Staff Name", "Project", "Apr 27 - Clock In", "Apr 27 - Clock Out"])
        ws.append(["Test Worker", "Project A", 9.0, 17.0])
        roster_path = str(tmp_path / "april_only_roster.xlsx")
        wb.save(roster_path)

        # Apr 27–May 1 requires both April and May Live sheets; May is absent.
        with pytest.raises(RosterParseError, match="(?i)(missing|cross.month|Live -)"):
            parse_roster(
                roster_path,
                target_week_start=datetime.date(2026, 4, 27),
                target_week_end=datetime.date(2026, 5, 1),
                malformed_out=[],
            )

    def test_no_silent_partial_output_on_missing_sheet(self, tmp_path):
        """When a required month sheet is missing, parse_roster must raise before
        returning any records — never silently return a partial list."""
        import openpyxl as _xl
        from triage.roster_parser import parse_roster, RosterParseError

        wb = _xl.Workbook()
        ws = wb.active
        ws.title = "Live - April 2026"
        ws.append(["Attendance Log"])
        ws.append(["Staff Name", "Project", "Apr 28 - Clock In", "Apr 28 - Clock Out"])
        ws.append(["Test Worker", "Project A", 9.0, 17.0])
        roster_path = str(tmp_path / "april_only_no_may.xlsx")
        wb.save(roster_path)

        result = []
        try:
            result = parse_roster(
                roster_path,
                target_week_start=datetime.date(2026, 4, 27),
                target_week_end=datetime.date(2026, 5, 1),
                malformed_out=[],
            )
        except RosterParseError:
            pass  # expected — confirmed error was raised
        assert len(result) == 0, (
            f"parse_roster must not return partial records when a sheet is missing "
            f"— got {len(result)} records despite the expected RosterParseError"
        )

    def test_wrong_year_sheet_not_accepted(self, tmp_path):
        """_find_live_sheet must NOT match 'Live - May 2025' when 'May 2026' is requested.

        Before the year-exact fix, the bare-month variant 'may' would match any
        May sheet regardless of year, silently returning wrong-year data.
        """
        import openpyxl as _xl
        from triage.roster_parser import parse_roster, RosterParseError

        # Workbook has only May 2025 Live sheet, not May 2026
        wb = _xl.Workbook()
        ws = wb.active
        ws.title = "Live - May 2025"
        ws.append(["Attendance Log — May 2025"])
        ws.append(["Staff Name", "Project", "May 27 - Clock In", "May 27 - Clock Out"])
        ws.append(["Test Worker", "Project A", 9.0, 17.0])
        roster_path = str(tmp_path / "may_2025_only.xlsx")
        wb.save(roster_path)

        # Requesting May 2026 — should raise because only May 2025 sheet exists
        with pytest.raises(RosterParseError, match="(?i)(matching|found|May 2026)"):
            parse_roster(roster_path, target_month="May 2026", malformed_out=[])

    def test_same_staff_multiple_projects_same_date(self, tmp_path):
        """Staff with two project rows on the same date must produce TWO records.

        Before the project-aware dedup fix, (staff, date) deduplication dropped
        the second project row, silently undercounting hours and corrupting the
        per-project breakdown.
        """
        import openpyxl as _xl
        from triage.roster_parser import parse_roster

        wb = _xl.Workbook()
        ws = wb.active
        ws.title = "Live - April 2026"
        ws.append(["Attendance Log — April 2026"])
        ws.append(["Staff Name", "Project", "Apr 15 - Clock In", "Apr 15 - Clock Out"])
        # Excel stores times as fractions of a day (9 AM = 9/24, 1 PM = 13/24).
        ws.append(["Jane Doe", "Project A", 9/24, 13/24])   # 4h gross
        ws.append(["Jane Doe", "Project B", 14/24, 17/24])  # 3h gross
        roster_path = str(tmp_path / "multi_project.xlsx")
        wb.save(roster_path)

        records = parse_roster(roster_path, target_month="April 2026", malformed_out=[])
        jane_recs = [r for r in records if r["staff"] == "Jane Doe"]
        assert len(jane_recs) == 2, (
            f"Expected 2 records for Jane Doe (one per project), got {len(jane_recs)}: "
            + str([(r['project'], r['clock_in'], r['clock_out']) for r in jane_recs])
        )
        projects = {r["project"] for r in jane_recs}
        assert projects == {"Project A", "Project B"}, (
            f"Expected projects {{'Project A', 'Project B'}}, got {projects}"
        )


# ── tests: invoice line-item categorization ────────────────────────────────────

class TestInvoiceCategorization:
    """Unit tests for _classify_line and _classify_invoice using real invoice assets."""

    def test_labor_prefix_wins_over_logistics(self):
        """'Labor: 3-person logistics team...' must classify as labor, not trucking."""
        from triage.invoice_parser import _classify_line
        desc = "Labor: 3-person logistics team(stock / staging / transport support)"
        assert _classify_line(desc) == "labor", (
            f"Expected 'labor', got {_classify_line(desc)!r} — logistics keyword "
            "must not override explicit 'Labor:' prefix"
        )

    def test_truck_logistics_support_is_trucking(self):
        """'Truck + logistics support...' must classify as trucking."""
        from triage.invoice_parser import _classify_line
        desc = "Truck + logistics support(weekend deployment / disposal runs)"
        assert _classify_line(desc) == "trucking"

    def test_courier_logistics_is_courier(self):
        """NYM courier line must classify as courier despite logistics keyword."""
        from triage.invoice_parser import _classify_line
        desc = "New York Minute courier & logistics support (NSUH pickup)"
        assert _classify_line(desc) == "courier"

    def test_plain_logistics_is_trucking(self):
        """Plain 'logistics support' (no explicit labor prefix) → trucking."""
        from triage.invoice_parser import _classify_line
        assert _classify_line("logistics support") == "trucking"

    def test_aaa_invoice_1_line_items(self):
        """AAA Feb-02 invoice: trucking line $300, labor line $696."""
        from triage.invoice_parser import parse_invoice
        p = next(
            f for f in [
                f"attached_assets/{f}" for f in os.listdir("attached_assets")
                if "AAA" in f and "2026-02-02" in f and f.endswith(".docx")
            ]
        )
        inv = parse_invoice(p)
        cats = {li["category"]: li["amount"] for li in inv["line_items"]}
        assert "trucking" in cats, f"Missing trucking line item; got {cats}"
        assert "labor" in cats, f"Missing labor line item; got {cats}"
        assert abs(cats["trucking"] - 300.0) < 0.01, (
            f"Trucking amount wrong: expected 300.0, got {cats['trucking']}"
        )
        assert abs(cats["labor"] - 696.0) < 0.01, (
            f"Labor amount wrong: expected 696.0, got {cats['labor']}"
        )

    def test_aaa_invoice_dominant_category_is_labor(self):
        """AAA Feb-02: labor ($696) > trucking ($300), so invoice category = labor."""
        from triage.invoice_parser import parse_invoice
        p = next(
            f for f in [
                f"attached_assets/{f}" for f in os.listdir("attached_assets")
                if "AAA" in f and "2026-02-02" in f and f.endswith(".docx")
            ]
        )
        inv = parse_invoice(p)
        assert inv["cost_category"] == "labor", (
            f"Expected invoice-level category 'labor' (dominant by amount), "
            f"got {inv['cost_category']!r}"
        )

    def test_nym_invoice_is_courier(self):
        """NYM invoice must classify entirely as courier."""
        from triage.invoice_parser import parse_invoice
        p = next(
            f for f in [
                f"attached_assets/{f}" for f in os.listdir("attached_assets")
                if "NYM" in f and f.endswith(".docx")
            ]
        )
        inv = parse_invoice(p)
        assert inv["cost_category"] == "courier"
        for li in inv["line_items"]:
            assert li["category"] == "courier", (
                f"NYM line item {li['description']!r} classified as "
                f"{li['category']!r}, expected 'courier'"
            )


# ── tests: invoice pivot totals with real invoices ────────────────────────────

class TestInvoicePivotTotals:
    """End-to-end tests: generate billing workbook WITH real invoices and
    verify that trucking, labor, courier amounts in the generated pivot are
    correct given the fixed line-item categorization."""

    @pytest.fixture(scope="class")
    def gen_wb_with_invoices(self, tmp_path_factory):
        from triage.roster_parser import parse_roster
        from triage.invoice_parser import parse_invoice
        from triage.billing_summary_generator import generate_billing_summary

        tmp = str(tmp_path_factory.mktemp("billing_inv"))

        malformed = []
        records = parse_roster(ROSTER, target_month="April 2026", malformed_out=malformed)
        merged = [r for r in records if r["gross_hours"] > 0]

        docx_files = sorted([
            f"attached_assets/{f}"
            for f in os.listdir("attached_assets")
            if f.endswith(".docx")
        ])
        invoices = [parse_invoice(p) for p in docx_files]

        path = generate_billing_summary(
            records=merged,
            invoices=invoices,
            billing_month="2026-04",
            out_root=tmp,
        )
        return openpyxl.load_workbook(path, data_only=True)

    def test_pivot_monthly_row_exists(self, gen_wb_with_invoices):
        ws = gen_wb_with_invoices["Invoice Pivots - Candidate"]
        data_row = 6
        service_month = ws.cell(data_row, 1).value
        assert service_month is not None, (
            "No data row at R6 in Monthly Invoice Totals section"
        )

    def _monthly_section_totals(self, ws, start_row=6):
        """Sum trucking/labor/courier/other columns in the Monthly Invoice Totals
        section only (stops at the TOTAL row; never crosses into PO/Vendor sections)."""
        cols = {"trucking": 3, "labor": 4, "courier": 5, "other": 6}
        totals = {k: 0.0 for k in cols}
        for r in range(start_row, ws.max_row + 1):
            label = str(ws.cell(r, 1).value or "").strip()
            # Stop at the TOTAL row — everything below belongs to PO/Vendor sections
            if label == "TOTAL":
                break
            # Skip empty label rows (blank separators between sections)
            if not label:
                continue
            for key, col in cols.items():
                v = ws.cell(r, col).value
                if isinstance(v, (int, float)):
                    totals[key] += v
        return totals

    def test_pivot_trucking_amount_correct(self, gen_wb_with_invoices):
        """Trucking total (monthly section, non-TOTAL rows) = $1500 from 3 AAA invoices."""
        ws = gen_wb_with_invoices["Invoice Pivots - Candidate"]
        totals = self._monthly_section_totals(ws)
        # AAA Feb-02 $300, Feb-09 $600, Feb-16 $600 = $1500
        assert abs(totals["trucking"] - 1500.0) < 0.50, (
            f"Trucking pivot total wrong: expected 1500.0, got {totals['trucking']:.2f}"
        )

    def test_pivot_labor_amount_correct(self, gen_wb_with_invoices):
        """Labor total (monthly section, non-TOTAL rows) = $3480 from 3 AAA invoices."""
        ws = gen_wb_with_invoices["Invoice Pivots - Candidate"]
        totals = self._monthly_section_totals(ws)
        # AAA Feb-02 $696, Feb-09 $1392, Feb-16 $1392 = $3480
        assert abs(totals["labor"] - 3480.0) < 0.50, (
            f"Labor pivot total wrong: expected 3480.0, got {totals['labor']:.2f}"
        )

    def test_pivot_courier_amount_correct(self, gen_wb_with_invoices):
        """Courier total (monthly section, non-TOTAL rows) = $634.50 from NYM invoice."""
        ws = gen_wb_with_invoices["Invoice Pivots - Candidate"]
        totals = self._monthly_section_totals(ws)
        assert abs(totals["courier"] - 634.5) < 0.50, (
            f"Courier pivot total wrong: expected 634.5, got {totals['courier']:.2f}"
        )

    def test_pivot_grand_total_correct(self, gen_wb_with_invoices):
        """Grand total (col 7 TOTAL row) = sum of all invoice amounts."""
        ws = gen_wb_with_invoices["Invoice Pivots - Candidate"]
        # Read the TOTAL row explicitly
        grand_total = None
        for r in range(6, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "").strip() == "TOTAL":
                v = ws.cell(r, 7).value
                if isinstance(v, (int, float)):
                    grand_total = v
                break
        assert grand_total is not None, "No TOTAL row found in Monthly Invoice Totals"
        expected = 996.0 + 1992.0 + 1992.0 + 634.5   # = 5614.5
        assert abs(grand_total - expected) < 1.0, (
            f"Grand total wrong: expected {expected:.2f}, got {grand_total:.2f}"
        )

    def test_vendor_section_aaa_trucking_correct(self, gen_wb_with_invoices):
        """Totals by Category (vendor section): AAA Disposal trucking = $1500."""
        ws = gen_wb_with_invoices["Invoice Pivots - Candidate"]
        # Vendor section is in col 10 (name), 12 (trucking), 13 (labor), 14 (total)
        for r in range(6, ws.max_row + 1):
            vendor_name = str(ws.cell(r, 10).value or "").strip()
            if "AAA" in vendor_name or "aaa" in vendor_name.lower():
                truck_val = ws.cell(r, 12).value
                labor_val = ws.cell(r, 13).value
                total_val = ws.cell(r, 14).value
                assert isinstance(truck_val, (int, float)), (
                    f"AAA trucking cell not numeric: {truck_val!r}"
                )
                assert abs(truck_val - 1500.0) < 0.50, (
                    f"AAA vendor trucking wrong: expected 1500.0, got {truck_val}"
                )
                assert abs(labor_val - 3480.0) < 0.50, (
                    f"AAA vendor labor wrong: expected 3480.0, got {labor_val}"
                )
                assert abs(total_val - 4980.0) < 0.50, (
                    f"AAA vendor total wrong: expected 4980.0, got {total_val}"
                )
                return
        pytest.fail("No AAA Disposal row found in vendor section (col 10)")
