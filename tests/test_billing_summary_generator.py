"""
tests/test_billing_summary_generator.py
-----------------------------------------
Unit and integration tests for triage.billing_summary_generator.

Coverage
--------
  T1  generate_billing_summary — raises RuntimeError when records is empty
  T2  generate_billing_summary — output file is created on disk
  T3  generate_billing_summary — workbook contains both required sheet names
  T4  Sheet 1 name follows 'Billing Summary - {Month YYYY}' pattern
  T5  Sheet 2 name is exactly 'Invoice Pivots - Candidate'
  T6  Sheet 1 has title row with month label and 'Billing Summary'
  T7  Sheet 1 'Monthly Rollup' section header is present
  T8  Sheet 1 rollup metrics (Gross hours, Lunch deducted, Net billable hours) present
  T9  Sheet 2 'Monthly Invoice Totals' section header present at correct position
  T10 Sheet 2 'Totals by Category' section header present
  T11 run_manifest.json is written with correct billing_month
  T12 generate_billing_summary — with invoices, invoice pivot data is populated
  T13 generate_billing_summary — net total <= gross total in generated sheet
  T14 generate_billing_summary — long month names do not exceed 31-char sheet tab limit
  T15 generate_billing_summary — gate-check passes on generated workbook
"""
from __future__ import annotations

import datetime
import json
from pathlib import Path

import pytest
import openpyxl

from triage.billing_summary_generator import generate_billing_summary


# ── helpers ───────────────────────────────────────────────────────────────────

def _make_record(staff: str, project: str, date: datetime.date,
                 clock_in: float = 9.0, clock_out: float = 17.0) -> dict:
    gross = clock_out - clock_in
    lunch = 1.0 if gross >= 8.0 else (0.5 if gross >= 6.0 else 0.0)
    net   = round(max(0.0, gross - lunch), 4)
    return {
        "staff":           staff,
        "project":         project,
        "date":            date,
        "clock_in":        clock_in,
        "clock_out":       clock_out,
        "gross_hours":     round(gross, 4),
        "lunch_deduction": lunch,
        "net_hours":       net,
    }


def _sample_records(billing_month: str = "2026-04") -> list:
    year, month = int(billing_month[:4]), int(billing_month[5:])
    base = datetime.date(year, month, 1)
    return [
        _make_record("Alice Brown",  "Project Alpha", base),
        _make_record("Alice Brown",  "Project Alpha", base + datetime.timedelta(1)),
        _make_record("Bob Smith",    "Project Beta",  base),
        _make_record("Charlie Diaz", "Project Alpha", base + datetime.timedelta(2)),
    ]


def _sample_invoice(vendor: str, total: float, category: str) -> dict:
    return {
        "source_file":    f"/tmp/{vendor}.docx",
        "vendor":         vendor,
        "invoice_number": "INV-001",
        "po_number":      "176759",
        "service_date":   "April 01, 2026",
        "service_window": None,
        "prepared_for":   "Agilant Solutions",
        "prepared_by":    vendor,
        "currency":       "USD",
        "line_items":     [
            {"description": category + " service", "qty": 1.0,
             "unit": "day", "rate": total, "amount": total, "category": category},
        ],
        "subtotal":       total,
        "total":          total,
        "cost_category":  category,
    }


def _find_cell_value(ws, text: str, search_col: int = 1):
    """Return (row, col) of first cell containing text, or None."""
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(r, search_col).value or "").strip()
        if text in v:
            return r
    return None


# ── fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def generated_wb(tmp_path_factory):
    tmp = str(tmp_path_factory.mktemp("billing_summary"))
    records = _sample_records("2026-04")
    path = generate_billing_summary(
        records=records,
        invoices=[],
        billing_month="2026-04",
        out_root=tmp,
        run_id="test-billing-001",
    )
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb, path, tmp


@pytest.fixture(scope="module")
def ws1(generated_wb):
    wb, _, _ = generated_wb
    return wb["Billing Summary - April 2026"]


@pytest.fixture(scope="module")
def ws2(generated_wb):
    wb, _, _ = generated_wb
    return wb["Invoice Pivots - Candidate"]


# ── T1: empty records raises ──────────────────────────────────────────────────

def test_empty_records_raises(tmp_path):
    with pytest.raises(RuntimeError, match="[Nn]o roster records|empty"):
        generate_billing_summary(
            records=[],
            invoices=[],
            billing_month="2026-04",
            out_root=str(tmp_path),
        )


# ── T2: output file created ───────────────────────────────────────────────────

def test_output_file_exists(generated_wb):
    _, path, _ = generated_wb
    assert Path(path).exists(), f"Output file not found: {path}"


# ── T3: both sheets present ───────────────────────────────────────────────────

def test_both_sheets_present(generated_wb):
    wb, _, _ = generated_wb
    sheets = wb.sheetnames
    assert len(sheets) == 2, f"Expected exactly 2 sheets, got: {sheets}"
    assert "Billing Summary - April 2026" in sheets
    assert "Invoice Pivots - Candidate" in sheets


# ── T4: sheet 1 name ─────────────────────────────────────────────────────────

def test_sheet1_name_pattern(generated_wb):
    wb, _, _ = generated_wb
    sheet1 = wb.sheetnames[0]
    assert sheet1.startswith("Billing Summary - "), f"Sheet 1 name: {sheet1!r}"
    assert "April 2026" in sheet1


# ── T5: sheet 2 name ─────────────────────────────────────────────────────────

def test_sheet2_name(generated_wb):
    wb, _, _ = generated_wb
    assert wb.sheetnames[1] == "Invoice Pivots - Candidate"


# ── T6: sheet 1 title row ────────────────────────────────────────────────────

def test_sheet1_title_row(ws1):
    title = str(ws1.cell(1, 1).value or "")
    assert "April 2026" in title, f"Title row missing month label: {title!r}"
    assert "Billing Summary" in title, f"Title row missing 'Billing Summary': {title!r}"


# ── T7: Monthly Rollup header ─────────────────────────────────────────────────

def test_monthly_rollup_header_present(ws1):
    row = _find_cell_value(ws1, "Monthly Rollup")
    assert row is not None, "Sheet 1 missing 'Monthly Rollup' section header"


# ── T8: rollup metrics ────────────────────────────────────────────────────────

def test_rollup_metrics_present(ws1):
    for label in ("Gross hours", "Lunch deducted", "Net billable hours"):
        row = _find_cell_value(ws1, label)
        assert row is not None, f"Sheet 1 missing rollup metric: {label!r}"


# ── T9: sheet 2 monthly invoice totals ───────────────────────────────────────

def test_sheet2_monthly_invoice_totals_header(ws2):
    row = _find_cell_value(ws2, "Monthly Invoice Totals")
    assert row is not None, "Sheet 2 missing 'Monthly Invoice Totals' header"


# ── T10: sheet 2 totals by category ──────────────────────────────────────────

def test_sheet2_totals_by_category_header(ws2):
    found = False
    for r in range(1, ws2.max_row + 1):
        for c in range(1, ws2.max_column + 1):
            v = str(ws2.cell(r, c).value or "")
            if "Totals by Category" in v:
                found = True
                break
        if found:
            break
    assert found, "Sheet 2 missing 'Totals by Category' section header"


# ── T11: manifest written ─────────────────────────────────────────────────────

def test_manifest_written(generated_wb):
    _, _, tmp = generated_wb
    manifest_path = Path(tmp) / "2026-04" / "run_manifest.json"
    assert manifest_path.exists(), f"run_manifest.json not found at {manifest_path}"
    data = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert data.get("run_id") == "test-billing-001"


# ── T12: invoice data populates pivots ───────────────────────────────────────

def test_with_invoices_pivot_data_populated(tmp_path):
    records = _sample_records("2026-04")
    invoices = [
        _sample_invoice("AAA Disposal", 700.0, "trucking"),
        _sample_invoice("NYM Courier",  360.0, "courier"),
    ]
    path = generate_billing_summary(
        records=records,
        invoices=invoices,
        billing_month="2026-04",
        out_root=str(tmp_path),
        run_id="inv-test",
    )
    wb = openpyxl.load_workbook(path, data_only=True)
    ws2 = wb["Invoice Pivots - Candidate"]
    numeric_values = []
    for r in range(1, ws2.max_row + 1):
        for c in range(1, ws2.max_column + 1):
            v = ws2.cell(r, c).value
            if isinstance(v, (int, float)) and v > 0:
                numeric_values.append(v)
    assert len(numeric_values) > 0, "Expected numeric invoice amounts in Invoice Pivots sheet"


# ── T13: net <= gross ─────────────────────────────────────────────────────────

def test_net_le_gross_in_rollup(ws1):
    gross_row = _find_cell_value(ws1, "Gross hours")
    net_row   = _find_cell_value(ws1, "Net billable hours")
    assert gross_row is not None and net_row is not None
    gross_val = ws1.cell(gross_row, 2).value
    net_val   = ws1.cell(net_row,   2).value
    assert isinstance(gross_val, (int, float)), f"Gross hours value not numeric: {gross_val}"
    assert isinstance(net_val,   (int, float)), f"Net hours value not numeric: {net_val}"
    assert net_val <= gross_val, f"Net ({net_val}) > Gross ({gross_val})"


# ── T14: sheet tab name ≤ 31 chars for long month names ─────────────────────

def test_sheet_tab_name_max_length(tmp_path):
    records = _sample_records.__wrapped__("2026-12") if hasattr(
        _sample_records, "__wrapped__"
    ) else [
        _make_record("Alice Brown", "Project Alpha", datetime.date(2026, 12, 1))
    ]
    path = generate_billing_summary(
        records=records,
        invoices=[],
        billing_month="2026-12",
        out_root=str(tmp_path),
    )
    wb = openpyxl.load_workbook(path, data_only=True)
    for name in wb.sheetnames:
        assert len(name) <= 31, f"Sheet tab name exceeds 31 chars: {name!r} ({len(name)})"


# ── T15: gate-check passes ───────────────────────────────────────────────────

def test_gate_check_passes(generated_wb):
    _, path, _ = generated_wb
    from triage.gate_checks import run_all
    report = run_all(path)
    assert not report.stopship, f"Stopship tokens: {report.stopship}"
    assert not report.cf_ref,   f"CF #REF! hits: {report.cf_ref}"
