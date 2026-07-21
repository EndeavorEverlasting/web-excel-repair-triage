"""Regression tests for the submitted Bonita Neuron Track Hours source hierarchy.

These tests encode the billing-submission lesson that made this path tedious:
Live tabs provide clocks, but monthly Assignments overrides and Worked Projects
determine whether a tech/date belongs to Neurons.
"""
from __future__ import annotations

from datetime import datetime

import openpyxl

from triage.nw_prj_neuron_track_hours.bonita_resolver import resolve_bonita_shifts


def _save_roster(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    live = wb.create_sheet("Live - April 2026")
    live.append(["April 2026 - Attendance"])
    live.append(["Staff Name", "Project", "Apr 23 - Clock In", "Apr 23 - Clock Out"])
    live.append(["Patricia Marrero", "Neuron Deployments", "9:00 AM / Neha", "5:00 PM / Neha"])
    live.append(["Zulu Tech", "Delivery / Transport", "8:00 AM", "4:00 PM"])
    live.append(["Echo Tech", "Neuron Deployments", "8:00 AM", "4:00 PM"])
    live.append(["Override Note Tech", "Delivery / Transport", "8:00 AM / Neha", "4:00 PM / Neha"])

    worked = wb.create_sheet("Worked Projects - April 2026")
    worked.append(["April 2026 - Worked Projects"])
    worked.append(["Staff Name", "Default Project", datetime(2026, 4, 23)])
    worked.append(["Patricia Marrero", "Neuron Deployments", "Neha / Other Project"])
    worked.append(["Zulu Tech", "Delivery / Transport", "Delivery / Transport"])
    worked.append(["Echo Tech", "Neuron Deployments", "Neha / Other Project"])
    worked.append(["Override Note Tech", "Delivery / Transport", "Delivery / Transport"])

    assign = wb.create_sheet("Assignments - April 2026")
    assign.append(["April 2026 - Project Assignments"])
    assign.append(["Staff Name", "Default Project", datetime(2026, 4, 23)])
    assign.append(["Patricia Marrero", "Neuron Deployments", "Neha / Other Project"])
    assign.append(["Zulu Tech", "Delivery / Transport", "Delivery / Transport"])
    assign.append(["Echo Tech", "Neuron Deployments", "Neuron Deployments"])
    assign.append(["Override Note Tech", "Delivery / Transport", "Delivery / Transport"])
    assign.append([])
    assign.append(["Overrides (only if different from Default Project)"])
    assign.append(["Override Staff Name", "Override Date", "Override Project", "Notes"])
    assign.append(["Zulu Tech", datetime(2026, 4, 23), "Neuron Deployments", "reviewed Neuron correction"])
    assign.append(["Override Note Tech", datetime(2026, 4, 23), "Neuron Deployments", "explicit Neuron override beats stale / Neha note"])

    wb.save(path)


def test_worked_projects_excludes_stale_live_neuron_default(tmp_path):
    roster = tmp_path / "roster.xlsx"
    _save_roster(roster)

    resolution = resolve_bonita_shifts(roster, ["2026-04"])

    assert all(s.tech != "Patricia Marrero" for s in resolution.shifts)
    assert all(s.tech != "Echo Tech" for s in resolution.shifts)


def test_assignments_override_table_can_include_neuron_day(tmp_path):
    roster = tmp_path / "roster.xlsx"
    _save_roster(roster)

    resolution = resolve_bonita_shifts(roster, ["2026-04"])
    zulu = [s for s in resolution.shifts if s.tech == "Zulu Tech"]

    assert len(zulu) == 1
    assert zulu[0].total_hours == 8.0
    assert zulu[0].project_name == "Northwell - Neurons"


def test_off_project_punch_note_excludes_unless_explicit_override(tmp_path):
    roster = tmp_path / "roster.xlsx"
    _save_roster(roster)

    resolution = resolve_bonita_shifts(roster, ["2026-04"])

    assert all(s.tech != "Patricia Marrero" for s in resolution.shifts)
    override_note = [s for s in resolution.shifts if s.tech == "Override Note Tech"]
    assert len(override_note) == 1
    assert override_note[0].total_hours == 8.0
    assert any(r.category == "off_project" and r.tech == "Patricia Marrero" for r in resolution.review)
