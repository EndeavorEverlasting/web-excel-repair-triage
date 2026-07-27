from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from triage.device_transfer_signoff import SignOffContractError, run


def _write_source(path: Path, count: int = 25) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Configs"
    ws.append([
        "Device Type",
        "Current Building",
        "Install Building",
        "Cybernet Hostname",
        "Cybernet Serial",
        "Neuron Hostname",
        "Neuron MAC",
        "Neuron S/N",
    ])
    for idx in range(1, count + 1):
        ws.append(["Cybernet", "STAGING", "EX", f"EX-CYB-{idx:03d}", f"CYB-SERIAL-{idx:03d}", "", "", ""])
    for idx in range(1, count + 1):
        ws.append(["Neuron", "STAGING", "EX", "", "", "", f"00AA00BB{idx:04d}", f"NEU-SERIAL-{idx:03d}"])
    wb.save(path)


def _config(serial_count: int = 25) -> dict:
    return {
        "artifact_family": "device_transfer_signoff",
        "site": {
            "name": "Example Hospital",
            "code": "EXAMPLE",
            "address": "100 Example Avenue, Example, NY 10000",
            "poc": "Example Receiver",
            "delivery_date": "2026-07-28",
            "delivery_time": "8:00 AM",
            "origin": "1 Marcus Ave / Staging",
            "prepared_by": "Example Coordinator",
            "signoff_id": "EXAMPLE-EDI-STOCK-20260728-001",
        },
        "source": {"sheet": None, "device_type_header": "Device Type"},
        "shipment": [
            {"item": "DIMs", "qty": 10},
            {"item": "Grey Cat5 Ethernet", "qty": 20},
            {"item": "Mice", "qty": serial_count},
            {"item": "Keyboards", "qty": serial_count},
            {"item": "Tap Badge Scanners", "qty": serial_count},
            {
                "item": "Neurons",
                "qty": serial_count,
                "serial_source": "Neuron",
                "serial_header": "Neuron S/N",
            },
            {
                "item": "Cybernets",
                "qty": serial_count,
                "serial_source": "Cybernet",
                "serial_header": "Cybernet Serial",
            },
        ],
    }


class DeviceTransferSignOffTests(unittest.TestCase):
    def test_generates_exact_shipment_and_serial_panels(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            source = root / "source.xlsx"
            config_path = root / "site.json"
            output = root / "out" / "Example_SignOff.xlsx"
            _write_source(source, 25)
            config = _config(25)
            config_path.write_text(json.dumps(config), encoding="utf-8")

            result = run(str(source), str(config_path), output=str(output))
            self.assertTrue(result.report["preflight_pass"])

            wb = load_workbook(output, data_only=True, read_only=True)
            try:
                ws = wb["Sign-Off"]
                items = [ws.cell(row, 2).value for row in range(9, 16)]
                quantities = [ws.cell(row, 3).value for row in range(9, 16)]
                self.assertEqual(items, [r["item"] for r in config["shipment"]])
                self.assertEqual(quantities, [r["qty"] for r in config["shipment"]])
                all_text = " ".join(
                    str(cell)
                    for row in ws.iter_rows(values_only=True)
                    for cell in row
                    if cell is not None
                )
                self.assertNotIn("Neuron Arm", all_text)
                self.assertNotIn("WC-0002-07", all_text)
                self.assertIn("CYB-SERIAL-025", all_text)
                self.assertIn("NEU-SERIAL-025", all_text)
            finally:
                wb.close()

    def test_serial_count_mismatch_fails_closed(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            source = root / "source.xlsx"
            config_path = root / "site.json"
            _write_source(source, 24)
            config_path.write_text(json.dumps(_config(25)), encoding="utf-8")
            with self.assertRaisesRegex(SignOffContractError, "quantity/serial mismatch"):
                run(str(source), str(config_path), output=str(root / "out.xlsx"))

    def test_missing_site_metadata_fails_closed(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            source = root / "source.xlsx"
            config_path = root / "site.json"
            _write_source(source, 25)
            config = _config(25)
            config["site"]["poc"] = ""
            config_path.write_text(json.dumps(config), encoding="utf-8")
            with self.assertRaisesRegex(SignOffContractError, "site.poc is required"):
                run(str(source), str(config_path), output=str(root / "out.xlsx"))

    def test_refuses_output_under_candidates(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            source = root / "source.xlsx"
            config_path = root / "site.json"
            _write_source(source, 25)
            config_path.write_text(json.dumps(_config(25)), encoding="utf-8")
            with self.assertRaisesRegex(SignOffContractError, "Candidates/ or Active"):
                run(
                    str(source),
                    str(config_path),
                    output=str(root / "Candidates" / "out.xlsx"),
                )


if __name__ == "__main__":
    unittest.main()
