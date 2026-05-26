from pathlib import Path

from openpyxl import Workbook

from scripts.inspect_admin_billing_workbook import inspect_workbook


def _save_workbook(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "01 Admin Summary"
    ws["A1"] = "Summary"
    wb.create_sheet("03 Friday Batches")
    tracker = wb.create_sheet("02 Tracker Import")
    tracker.append([
        "Date",
        "Technician",
        "Billing Category",
        "Hours",
        "Hours Status",
        "Admin Action",
        "Safe Billing Note",
    ])
    for row in rows:
        tracker.append(row)
    wb.save(path)


def test_inspector_passes_clean_admin_context(tmp_path):
    path = tmp_path / "clean.xlsx"
    _save_workbook(
        path,
        [
            ["2026-04-01", "Alejandro Perales", "Configuration", 8, "04/01 hours reviewed and carried in billing context.", "Keep", "Configuration support."],
            ["2026-04-01", "Khadejah Harrison", "Configuration", "", "04/01 confirmed OOO. No hours expected; reviewed and cleared.", "Keep", "Confirmed OOO for 04/01. No billable hours carried for this date."],
            ["2026-04-04", "Geoff Gerber", "Context", "", "04/04 reviewed as context-only. No hours carried; no billing action required.", "Keep", "Context-only entry for 04/04. No billable hours carried for this date."],
        ],
    )

    result = inspect_workbook(path)

    assert result.ok_to_transform
    assert result.total_carried_hours == 8.0
    assert result.numeric_hour_rows == 1
    assert result.blank_hour_rows == 2
    assert result.unresolved_review_rows == 0
    assert result.suspicious_language_hits == 0
    assert result.friday_batch_totals == {"2026-04-03": 8.0}


def test_inspector_flags_missing_required_tab(tmp_path):
    path = tmp_path / "missing_tab.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "02 Tracker Import"
    ws.append(["Date", "Technician", "Billing Category", "Hours"])
    wb.save(path)

    result = inspect_workbook(path)

    assert not result.ok_to_transform
    assert "01 Admin Summary" in result.required_admin_tabs_missing
    assert any(issue.issue_code == "missing_admin_tab" for issue in result.issues)


def test_inspector_flags_suspicious_language(tmp_path):
    path = tmp_path / "suspicious.xlsx"
    _save_workbook(
        path,
        [
            ["2026-04-01", "Tech", "Configuration", "", "Blank-hour OOO/context-only rows reviewed and cleared; no invented hours.", "Keep", "Context-only entry for 04/01. No billable hours carried for this date."],
        ],
    )

    result = inspect_workbook(path)

    assert not result.ok_to_transform
    assert result.suspicious_language_hits == 1
    assert any(issue.issue_code == "suspicious_language" for issue in result.issues)


def test_inspector_flags_ooo_rows_that_imply_work(tmp_path):
    path = tmp_path / "ooo_bad_note.xlsx"
    _save_workbook(
        path,
        [
            ["2026-04-01", "Khadejah Harrison", "Configuration", "", "04/01 confirmed OOO. No hours expected; reviewed and cleared.", "Keep", "Configuration support and QA readiness."],
        ],
    )

    result = inspect_workbook(path)

    assert not result.ok_to_transform
    assert result.ooo_work_language_hits == 1
    assert any(issue.issue_code == "ooo_implies_work_performed" for issue in result.issues)


def test_inspector_flags_unresolved_review_rows(tmp_path):
    path = tmp_path / "review.xlsx"
    _save_workbook(
        path,
        [
            ["2026-04-01", "Tech", "Configuration", 8, "REVIEW: confirm hours", "Review", "Configuration support."],
        ],
    )

    result = inspect_workbook(path)

    assert not result.ok_to_transform
    assert result.unresolved_review_rows == 1
    assert any(issue.issue_code == "unresolved_review_status" for issue in result.issues)
