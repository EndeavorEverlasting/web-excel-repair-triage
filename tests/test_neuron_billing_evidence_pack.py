"""Fixture-only tests for the Neuron billing evidence-pack generator."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from tests.fixtures.nw_prj_neuron_track_hours.bonita_fixtures import (
    build_bonita_fixtures,
)
from triage.nw_prj_neuron_track_hours.bonita_resolver import resolve_bonita_shifts
from triage.nw_prj_neuron_track_hours.evidence_pack_cli import run

MONTHS = ["2026-04", "2026-05"]


@pytest.fixture()
def roster(tmp_path: Path) -> Path:
    return build_bonita_fixtures(tmp_path / "fixtures")["roster"]


def _write_allocation_source(roster: Path, path: Path, *, omit_last: bool = False) -> None:
    resolution = resolve_bonita_shifts(str(roster), MONTHS)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Allocation Rules"
    ws.append([
        "DATE",
        "DAY",
        "TECH NAME",
        "START TIME",
        "END TIME",
        "TOTAL HOURS",
        "PROJECT NAME",
        "TASK / ASSIGNMENT TYPE",
        "SUPPORTING WORK / NOTES",
    ])
    non_deployment = [
        "Configurations",
        "Inventory Management",
        "Survey",
        "Ticket Forwarding",
    ]
    shifts = list(resolution.shifts)
    if omit_last:
        shifts = shifts[:-1]
    for index, shift in enumerate(shifts):
        assignment = (
            "Deployments"
            if shift.tech == "Delta Tech"
            else non_deployment[index % len(non_deployment)]
        )
        ws.append([
            shift.date,
            shift.day,
            shift.tech,
            shift.start_time,
            shift.end_time,
            shift.total_hours,
            shift.project_name,
            assignment,
            "Synthetic allocation label only",
        ])
    wb.save(path)
    wb.close()


def test_generates_expected_evidence_pack_from_roster(roster: Path, tmp_path: Path) -> None:
    manifest = run(
        roster_log=str(roster),
        out_dir=str(tmp_path / "out"),
        months=MONTHS,
        repo_root=Path(__file__).resolve().parents[1],
    )

    assert manifest["preflight_pass"] is True
    assert manifest["shift_count"] == 9
    assert manifest["grand_total_hours"] == 82.0
    assert manifest["daily_narrative_rows"] == 9
    assert manifest["event_rows"] == 9

    wb = openpyxl.load_workbook(manifest["outputs"]["workbook"], data_only=True)
    assert wb.sheetnames == [
        "April 2026",
        "May 2026",
        "Visual Summary",
        "Executive Dashboard",
        "Daily Narrative Log",
        "Event Log",
    ]
    assert "Task Summary" not in wb.sheetnames

    for sheet_name in ("April 2026", "May 2026"):
        for row in wb[sheet_name].iter_rows(min_row=5, values_only=True):
            if row[2]:
                assert row[8] == (
                    f"Assignment classification: {row[7]}. "
                    "No additional operational detail is asserted."
                )

    event = wb["Event Log"]
    actual_hours = [
        row[19]
        for row in event.iter_rows(min_row=6, values_only=True)
        if row[2] and isinstance(row[19], (int, float))
    ]
    assert len(actual_hours) == 9
    assert sum(actual_hours) == 82.0
    assert all(
        row[4] == "Not recorded in roster"
        for row in event.iter_rows(min_row=6, values_only=True)
        if row[2]
    )

    flat = [
        str(value)
        for ws in wb.worksheets
        for row in ws.iter_rows(values_only=True)
        for value in row
        if value is not None
    ]
    wb.close()
    assert not any("lunch covered" in value.casefold() for value in flat)
    assert not any("synthetic allocation label only" in value.casefold() for value in flat)


def test_allocation_source_overrides_labels_without_changing_hours(
    roster: Path, tmp_path: Path
) -> None:
    allocation = tmp_path / "allocation.xlsx"
    _write_allocation_source(roster, allocation)
    manifest = run(
        roster_log=str(roster),
        allocation_source=str(allocation),
        out_dir=str(tmp_path / "out"),
        months=MONTHS,
        repo_root=Path(__file__).resolve().parents[1],
    )

    assert manifest["preflight_pass"] is True
    assert manifest["allocation_overlay"]["matched"] == 9
    assert manifest["allocation_overlay"]["unmatched_shifts"] == 0
    assert manifest["grand_total_hours"] == 82.0

    wb = openpyxl.load_workbook(manifest["outputs"]["workbook"], data_only=True)
    deployment_people = []
    for sheet_name in ("April 2026", "May 2026"):
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=5, values_only=True):
            if len(row) >= 8 and row[7] == "Deployments":
                deployment_people.append(row[2])
    assert deployment_people == ["Delta Tech"]

    event = wb["Event Log"]
    deployment_event_people = [
        row[2]
        for row in event.iter_rows(min_row=6, values_only=True)
        if row[8] == "Deployments"
    ]
    wb.close()
    assert deployment_event_people == ["Delta Tech"]


def test_strict_allocation_source_fails_on_unmatched_shift(
    roster: Path, tmp_path: Path
) -> None:
    allocation = tmp_path / "allocation_missing_row.xlsx"
    _write_allocation_source(roster, allocation, omit_last=True)
    with pytest.raises(ValueError, match="did not reconcile every roster-derived shift"):
        run(
            roster_log=str(roster),
            allocation_source=str(allocation),
            out_dir=str(tmp_path / "out"),
            months=MONTHS,
            repo_root=Path(__file__).resolve().parents[1],
        )
