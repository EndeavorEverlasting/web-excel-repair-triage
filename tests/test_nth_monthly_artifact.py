from __future__ import annotations

from datetime import time
from pathlib import Path

import openpyxl

from triage.nth_month_readiness import inspect_month_readiness
from triage.nth_monthly_artifact import build_month_artifact, expected_sheets, workbook_name


def _write_ready_august(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Live - August 2026"
    ws.cell(2, 1, "Staff Name")
    ws.cell(2, 2, "Project")
    ws.cell(2, 3, "Aug 03 - Clock In")
    ws.cell(2, 4, "Aug 03 - Clock Out")
    ws.cell(3, 1, "Alpha Tech")
    ws.cell(3, 2, "Neuron Deployments")
    ws.cell(3, 3, time(8, 0))
    ws.cell(3, 4, time(17, 0))
    ws.cell(4, 1, "Other Tech")
    ws.cell(4, 2, "Projects Team")
    ws.cell(4, 3, time(8, 0))
    ws.cell(4, 4, time(17, 0))
    wb.save(path)
    wb.close()


def _write_presence_only_august(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "August 2026 - Attendance (Automated)"
    ws.cell(1, 1, "Staff Name")
    ws.cell(1, 2, "Project")
    ws.cell(1, 3, "2026-08-03")
    ws.cell(2, 1, "Alpha Tech")
    ws.cell(2, 2, "Neuron Deployments")
    ws.cell(2, 3, "Yes")
    wb.save(path)
    wb.close()


def test_presence_only_attendance_fails_closed(tmp_path):
    roster = tmp_path / "presence.xlsx"
    _write_presence_only_august(roster)
    result = inspect_month_readiness(roster, "2026-08")
    assert result.status == "NO_GO"
    assert result.attendance_authority == "not_proven"
    assert result.presence_only_sheets == ["August 2026 - Attendance (Automated)"]
    assert any("paid_hours_source_missing" in blocker for blocker in result.blockers)


def test_live_clock_pairs_make_august_ready(tmp_path):
    roster = tmp_path / "ready.xlsx"
    _write_ready_august(roster)
    result = inspect_month_readiness(roster, "2026-08")
    assert result.status == "READY"
    assert result.attendance_authority == "clock_in_out_live_sheet"
    assert result.live_sheet == "Live - August 2026"
    assert result.paired_clock_dates == 1


def test_august_build_is_dynamic_and_websafe(tmp_path):
    roster = tmp_path / "ready.xlsx"
    _write_ready_august(roster)
    out = tmp_path / "out"
    manifest = build_month_artifact(roster, "2026-08", out)

    assert manifest["month_label"] == "August 2026"
    assert manifest["hours"] == 9.0
    assert manifest["row_count"] == 1
    assert manifest["websafe_preflight_pass"] is True

    workbook = Path(manifest["workbook"])
    assert workbook.name == workbook_name("2026-08")
    assert workbook.exists()

    wb = openpyxl.load_workbook(workbook, read_only=True)
    try:
        assert wb.sheetnames == expected_sheets("August 2026")
        assert "August 2026 Neuron Hours" in wb.sheetnames
        assert "April Neuron Hours" not in wb.sheetnames
        assert "May Neuron Hours" not in wb.sheetnames
    finally:
        wb.close()
