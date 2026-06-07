"""NW PRJ Admin Log Project Team generator — fixture-only tests."""
from __future__ import annotations

import json
import zipfile
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from triage.artifact_fingerprint import raw_sha256
from triage.nw_prj_admin_log.cli import run
from triage.nw_prj_admin_log.grid import find_data_header_row, parse_date_columns
from tests.fixtures.nw_prj_admin_log.fixtures import build_fixtures

REPO_ROOT = Path(__file__).resolve().parent.parent
FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures" / "nw_prj_admin_log"
MONTHS = ["2026-04", "2026-05"]


@pytest.fixture(scope="module")
def fixtures():
    return build_fixtures(FIXTURE_DIR)


@pytest.fixture(scope="module")
def source_hashes(fixtures):
    return {
        "roster": raw_sha256(fixtures["roster"]),
        "donor": raw_sha256(fixtures["donor"]),
        "reference": raw_sha256(fixtures["reference"]),
    }


@pytest.fixture(scope="module")
def generated(fixtures, tmp_path_factory):
    out = tmp_path_factory.mktemp("admin_log_out")
    run_dir = out / "Outputs" / "nw_prj_admin_log" / "test_run"
    return run(
        roster_log=str(fixtures["roster"]),
        visual_donor=str(fixtures["donor"]),
        accepted_reference=str(fixtures["reference"]),
        out_dir=str(run_dir),
        months=MONTHS,
        websafe=True,
        repo_root=out,
    )


def test_exactly_one_visible_project_team_sheet(generated):
    wb = openpyxl.load_workbook(generated["delivery_workbook"])
    visible = [s.title for s in wb.worksheets if s.sheet_state == "visible"]
    assert visible == ["Project Team"]
    wb.close()


def test_column_a_hidden(generated):
    wb = openpyxl.load_workbook(generated["delivery_workbook"])
    ws = wb["Project Team"]
    assert ws.column_dimensions["A"].hidden is True
    wb.close()


def test_freeze_pane_c1(generated):
    wb = openpyxl.load_workbook(generated["delivery_workbook"])
    assert wb["Project Team"].freeze_panes == "C1"
    wb.close()


def test_april_1_through_5_present(generated):
    wb = openpyxl.load_workbook(generated["delivery_workbook"])
    ws = wb["Project Team"]
    hdr = find_data_header_row(ws)
    dates = set(parse_date_columns(ws, hdr).values())
    for day in range(1, 6):
        assert date(2026, 4, day) in dates
    wb.close()


def test_may_31_present(generated):
    wb = openpyxl.load_workbook(generated["delivery_workbook"])
    ws = wb["Project Team"]
    hdr = find_data_header_row(ws)
    dates = set(parse_date_columns(ws, hdr).values())
    assert date(2026, 5, 31) in dates
    wb.close()


def test_drawing_media_when_donor_has_logo(generated, fixtures):
    with zipfile.ZipFile(fixtures["donor"], "r") as z:
        donor_media = [n for n in z.namelist() if n.startswith("xl/media/")]
    if not donor_media:
        pytest.skip("donor fixture has no embedded media")
    with zipfile.ZipFile(generated["delivery_workbook"], "r") as z:
        cand_media = [n for n in z.namelist() if n.startswith("xl/media/")]
    assert cand_media, "expected media parts preserved from donor"


def test_forbidden_phrases_absent(generated):
    pf = generated["preflight"]
    assert pf.get("forbidden_visible_phrases") == []
    assert "forbidden_visible_text" not in pf.get("token_failures", [])


def test_source_files_unchanged(fixtures, generated, source_hashes):
    assert raw_sha256(fixtures["roster"]) == source_hashes["roster"]
    assert raw_sha256(fixtures["donor"]) == source_hashes["donor"]
    assert raw_sha256(fixtures["reference"]) == source_hashes["reference"]
    imm = generated["manifest"]["source_immutability"]
    assert imm["pass"] is True


def test_no_stopship_tokens(generated):
    pf = generated["preflight"]
    assert "inlineStr" not in pf.get("token_failures", [])
    for tok in ("ns0:", "_xlfn.", "_xludf.", "_xlpm."):
        assert not any(tok in t for t in pf.get("token_failures", []))
    assert not any("formula_error" in t for t in pf.get("token_failures", []))


def test_sharedstrings_invariant(generated):
    pf = generated["preflight"]
    assert pf.get("sharedstrings_count_ok") is True


def test_manifest_provenance_fields(generated):
    m = generated["manifest"]
    assert m.get("run_id")
    assert m.get("source_emulator_path")
    assert m.get("source_raw_sha256")
    assert m.get("output_layout_version") == 1
    assert Path(m["outputs"]["delivery_workbook"]).is_file()
    side = Path(m["outputs"]["preflight_json"])
    assert side.is_file()
    assert json.loads(side.read_text(encoding="utf-8")).get("preflight_pass") is True
