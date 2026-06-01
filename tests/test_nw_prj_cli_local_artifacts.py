"""End-to-end smoke tests for NW PRJ local artifact CLI pipeline."""
from __future__ import annotations

import json
import zipfile
from pathlib import Path

import openpyxl
import pytest

from triage.nw_prj_cli import run as cli_run


def _write_roster(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Live Projects"
    ws.append(["tech", "date", "project", "worked project", "in", "out", "hours"])
    ws.append(["Alice", "2026-04-03", "ProjA", "ProjA", "08:00", "16:00", 8.0])
    ws.append(["Bob",   "2026-04-04", "ProjB", "ProjB", "08:00", "12:00", 4.0])
    ws.append(["Rich Perez", "2026-04-05", "ProjC", "ProjC", "08:00", "12:00", 4.0])
    ws.append(["Alice", "2026-05-02", "ProjA", "ProjA", "09:00", "17:00", 8.0])
    wb.create_sheet("Worked Projects")
    wb["Worked Projects"].append(["tech", "date", "project", "in", "out", "hours"])
    wb.save(path)


def _write_admin(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Team"
    ws.append(["tech", "date", "hours"])
    ws.append(["Alice", "2026-04-03", 8.0])
    ws.append(["Bob",   "2026-04-04", 4.0])
    ws.append(["Rich Perez", "2026-04-05", 8.0])  # admin says 8, roster says 4
    ws.append(["Alice", "2026-05-02", 8.0])
    wb.save(path)


@pytest.fixture
def synthetic_inputs(tmp_path: Path) -> dict:
    roster = tmp_path / "roster.xlsx"
    admin = tmp_path / "admin.xlsx"
    _write_roster(roster)
    _write_admin(admin)
    return {"roster": roster, "admin": admin, "out": tmp_path / "out"}


def test_cli_pipeline_generates_all_artifacts(synthetic_inputs):
    out_dir = synthetic_inputs["out"]
    manifest = cli_run(
        roster_log=str(synthetic_inputs["roster"]),
        admin_folder=None,
        admin_april=str(synthetic_inputs["admin"]),
        admin_may=str(synthetic_inputs["admin"]),
        out_dir=str(out_dir),
        months=["2026-04", "2026-05"],
        webexcel=True,
        zip_output=True,
    )
    # Expected workbooks
    for fname in (
        "NW_PRJ_April_2026_Billing_Summary_WEBSAFE.xlsx",
        "NW_PRJ_May_2026_Billing_Summary_WEBSAFE.xlsx",
        "Neuron_Track_Hours_April_May_2026_WEBSAFE.xlsx",
    ):
        assert (out_dir / fname).exists(), f"missing {fname}"

    # Expected reports
    for fname in (
        "nw_prj_admin_project_team_control_records.csv",
        "nw_prj_roster_work_records_april_may.csv",
        "nw_prj_reconciliation_queue.csv",
        "nw_prj_workbook_preflight_report.json",
        "nw_prj_artifact_scan_report.json",
    ):
        assert (out_dir / fname).exists(), f"missing {fname}"

    # Outer zip
    zip_path = Path(manifest["outer_zip"])
    assert zip_path.exists()
    with zipfile.ZipFile(zip_path) as z:
        names = z.namelist()
        assert "NW_PRJ_April_2026_Billing_Summary_WEBSAFE.xlsx" in names
        assert "Neuron_Track_Hours_April_May_2026_WEBSAFE.xlsx" in names

    # Manifest shape
    assert manifest["counts"]["admin_records"] >= 4
    assert manifest["counts"]["roster_records"] >= 4
    for art in manifest["artifacts"]:
        assert "artifact_name" in art
        assert "webexcel_preflight_pass" in art
        assert "has_filters" in art
        assert "has_frozen_header" in art


def test_billing_summary_has_filters_and_frozen(synthetic_inputs):
    out_dir = synthetic_inputs["out"]
    cli_run(
        roster_log=str(synthetic_inputs["roster"]),
        admin_folder=None,
        admin_april=str(synthetic_inputs["admin"]),
        admin_may=str(synthetic_inputs["admin"]),
        out_dir=str(out_dir),
        months=["2026-04"],
        webexcel=True,
        zip_output=False,
    )
    wb = openpyxl.load_workbook(out_dir / "NW_PRJ_April_2026_Billing_Summary_WEBSAFE.xlsx")
    ws = wb.active
    assert ws.freeze_panes == "A5"
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.startswith("A4")


def test_neuron_tabs_present(synthetic_inputs):
    out_dir = synthetic_inputs["out"]
    cli_run(
        roster_log=str(synthetic_inputs["roster"]),
        admin_folder=None,
        admin_april=str(synthetic_inputs["admin"]),
        admin_may=str(synthetic_inputs["admin"]),
        out_dir=str(out_dir),
        months=["2026-04", "2026-05"],
        webexcel=False,
        zip_output=False,
    )
    wb = openpyxl.load_workbook(out_dir / "Neuron_Track_Hours_April_May_2026_WEBSAFE.xlsx")
    names = wb.sheetnames
    for expected in ("Summary", "April 2026", "May 2026", "Go Live Weekend Support", "CF Dictionary", "WebExcel QC"):
        assert expected in names, f"missing tab {expected}"


def test_preflight_passes_on_generated_billing_summary(synthetic_inputs):
    """Generated billing summary must pass webexcel preflight (no inlineStr)."""
    from triage.webexcel_preflight import run_preflight

    out_dir = synthetic_inputs["out"]
    cli_run(
        roster_log=str(synthetic_inputs["roster"]),
        admin_folder=None,
        admin_april=str(synthetic_inputs["admin"]),
        admin_may=str(synthetic_inputs["admin"]),
        out_dir=str(out_dir),
        months=["2026-04"],
        webexcel=True,
        zip_output=False,
    )
    rpt = run_preflight(str(out_dir / "NW_PRJ_April_2026_Billing_Summary_WEBSAFE.xlsx"))
    assert rpt.webexcel_preflight_pass, (
        f"Preflight failed — token_failures={rpt.token_failures} gate_failures={rpt.gate_failures}"
    )
    assert "inlineStr" not in rpt.token_failures


def test_rich_guard_preserves_admin_hours(synthetic_inputs):
    """Rich's admin says 8h, roster says 4h. Resolved hours must stay at admin's 8."""
    from triage.nw_prj_admin_reader import NwPrjAdminReader
    from triage.nw_prj_classifier import NwPrjClassifier
    from triage.nw_prj_roster_reader import NwPrjRosterReader

    admin_records = NwPrjAdminReader(str(synthetic_inputs["admin"])).read_records()
    roster_records = NwPrjRosterReader(str(synthetic_inputs["roster"])).read_all_records()
    results = NwPrjClassifier().classify(admin_records, roster_records)

    rich_rows = [r for r in results if r.tech == "Rich Perez"]
    assert rich_rows, "Rich row missing from classifier output"
    rich = rich_rows[0]
    assert rich.resolved_hours == 8.0
    assert rich.reason_code.startswith("RICH_")
