"""NW PRJ dashboard v6 contract, gates, generator, and comparator tests."""
from __future__ import annotations

import io
import zipfile
from pathlib import Path

import openpyxl
import pytest

from triage import gate_checks
from triage.nw_prj_artifact_compare import CompareInputs, compare_artifacts
from triage.nw_prj_config import is_repair_filename
from triage.nw_prj_dashboard_validator import review_status_bucket
from triage.nw_prj_dashboard_generator import GenerateInputs, generate_dashboard
from triage.nw_prj_dashboard_validator import (
    check_cf_dictionary_exists,
    validate_nw_prj_dashboard,
)


def _minimal_xlsx(
    sheet_names: list[str] | None = None,
    cf_block: str = "",
    formula: str = "",
    table_parts: list[tuple[str, str]] | None = None,
) -> bytes:
    sheet_names = sheet_names or ["Sheet1"]
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(
            "[Content_Types].xml",
            b'<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            b'<Default Extension="xml" ContentType="application/xml"/>'
            b'<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            b"</Types>",
        )
        sheets_xml = "".join(
            f'<sheet name="{n}" sheetId="{i+1}" r:id="rId{i+1}"/>' for i, n in enumerate(sheet_names)
        )
        z.writestr(
            "xl/workbook.xml",
            (
                '<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                f"<sheets>{sheets_xml}</sheets></workbook>"
            ).encode(),
        )
        rels = "".join(
            f'<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            f'Target="worksheets/sheet{i+1}.xml" Id="rId{i+1}"/>'
            for i in range(len(sheet_names))
        )
        z.writestr(
            "xl/_rels/workbook.xml.rels",
            (
                '<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                f"{rels}</Relationships>"
            ).encode(),
        )
        fxml = f"<c r='A1'><f>{formula}</f></c>" if formula else ""
        cfxml = cf_block or ""
        for i in range(len(sheet_names)):
            z.writestr(
                f"xl/worksheets/sheet{i+1}.xml",
                (
                    '<?xml version="1.0"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                    f"<sheetData><row r='1'>{fxml}</row></sheetData>{cfxml}</worksheet>"
                ).encode(),
            )
        for i, (tname, tpart) in enumerate(table_parts or []):
            z.writestr(
                tpart,
                f'<?xml version="1.0"?><table xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                f'displayName="{tname}" name="{tname}" ref="A1:B2"/>'.encode(),
            )
    return buf.getvalue()


def _write_prior_dashboard(path: Path, rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Active_Admin_Targets"
    headers = list(rows[0].keys()) if rows else ["Review Status", "Tech", "Date", "Edit Sheet", "Edit Row"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    wb.create_sheet("CF_Dictionary")
    wb["CF_Dictionary"].append(["Rule ID"])
    wb.save(path)


# ── config / filename ──


def test_no_repair_chain_filename():
    assert is_repair_filename("Deprecated_repaired_foo.xlsx")
    assert is_repair_filename("repaired_bar.xlsx")
    assert not is_repair_filename("NW_PRJ_Tech_Roster_Dashboard_v6_7_OK_WEBSAFE.xlsx")


# ── gate checks ──


def test_no_web_excel_bad_tokens(tmp_path: Path):
    p = tmp_path / "bad.xlsx"
    p.write_bytes(_minimal_xlsx(formula='SUM(_xlfn.CONCAT(A1))'))
    rpt = gate_checks.run_all(str(p))
    assert rpt.stopship


def test_no_rc_formula_refs(tmp_path: Path):
    cf = '<conditionalFormatting sqref="A2:A10"><cfRule type="expression"><formula>RC2="AMBER"</formula></cfRule></conditionalFormatting>'
    p = tmp_path / "rc.xlsx"
    p.write_bytes(_minimal_xlsx(cf_block=cf))
    rpt = gate_checks.run_all(str(p))
    assert rpt.rc_formula_refs


def test_unique_table_names(tmp_path: Path):
    p = tmp_path / "dup.xlsx"
    p.write_bytes(
        _minimal_xlsx(
            table_parts=[
                ("tblA", "xl/tables/table1.xml"),
                ("tblA", "xl/tables/table2.xml"),
            ]
        )
    )
    rpt = gate_checks.run_all(str(p))
    assert rpt.duplicate_table_names


def test_xml_parts_parse(tmp_path: Path):
    p = tmp_path / "ok.xlsx"
    p.write_bytes(_minimal_xlsx(sheet_names=["CF_Dictionary", "Active_Admin_Targets"]))
    rpt = gate_checks.run_all(str(p))
    assert not rpt.xml_wellformed


# ── profile validator ──


def test_cf_dictionary_exists(tmp_path: Path):
    p = tmp_path / "nodict.xlsx"
    p.write_bytes(_minimal_xlsx(sheet_names=["Active_Admin_Targets"]))
    assert check_cf_dictionary_exists(str(p))

    p2 = tmp_path / "dict.xlsx"
    p2.write_bytes(_minimal_xlsx(sheet_names=["CF_Dictionary", "Active_Admin_Targets"]))
    assert not check_cf_dictionary_exists(str(p2))


def test_column_a_override_wins(tmp_path: Path):
    cf = '<conditionalFormatting sqref="A2:A10"><cfRule><formula>SEARCH("AMBER",RC2)</formula></cfRule></conditionalFormatting>'
    p = tmp_path / "badcf.xlsx"
    p.write_bytes(_minimal_xlsx(sheet_names=["Active_Admin_Targets"], cf_block=cf))
    val = validate_nw_prj_dashboard(str(p))
    assert any("column_a" in f for f in val.failures)


# ── carry-forward / compare ──


def test_manual_status_carryforward(tmp_path: Path):
    dash = tmp_path / "prior.xlsx"
    _write_prior_dashboard(
        dash,
        [
            {
                "Review Status": "Done",
                "Tech": "Alice",
                "Date": "2026-05-01",
                "Edit Sheet": "Hours",
                "Edit Row": "5",
                "Manual Note / Resolution Note": "confirmed",
                "Work Queue Status": "AMBER",
            }
        ],
    )
    scratch = tmp_path / "scratch.xlsx"
    openpyxl.Workbook().save(scratch)
    rpt = compare_artifacts(
        CompareInputs(dashboard_path=str(dash), admin_scratch_path=str(scratch))
    )
    assert rpt.archive_rows
    assert rpt.archive_rows[0]["Review Status"] == "Done"
    assert rpt.archive_rows[0]["Manual Note / Resolution Note"] == "confirmed"


def test_gray_rows_archived(tmp_path: Path):
    dash = tmp_path / "gray.xlsx"
    _write_prior_dashboard(
        dash,
        [
            {
                "Review Status": "Skipped/Gray",
                "Tech": "Bob",
                "Date": "2026-05-02",
                "Edit Sheet": "Hours",
                "Edit Row": "3",
            }
        ],
    )
    scratch = tmp_path / "scratch2.xlsx"
    openpyxl.Workbook().save(scratch)
    rpt = compare_artifacts(
        CompareInputs(dashboard_path=str(dash), admin_scratch_path=str(scratch))
    )
    assert all(
        review_status_bucket(r.get("Review Status", "")) == "skipped_gray"
        for r in rpt.archive_rows
    )
    assert not rpt.active_rows


def test_admin_scratch_targeting(tmp_path: Path):
    scratch = tmp_path / "admin_scratch.xlsx"
    openpyxl.Workbook().save(scratch)
    rpt = compare_artifacts(CompareInputs(admin_scratch_path=str(scratch)))
    assert rpt.admin_authority == "manual_admin_scratch"
    assert rpt.inputs["admin_scratch"]


def test_rich_guard_does_not_downgrade(tmp_path: Path):
    dash = tmp_path / "rich.xlsx"
    _write_prior_dashboard(
        dash,
        [
            {
                "Review Status": "",
                "Tech": "Rich",
                "Date": "2026-05-03",
                "Edit Sheet": "H",
                "Edit Row": "1",
                "Current Admin Value": "8",
                "Roster Latest Hours": "4",
            }
        ],
    )
    scratch = tmp_path / "scratch_rich.xlsx"
    openpyxl.Workbook().save(scratch)
    rpt = compare_artifacts(
        CompareInputs(dashboard_path=str(dash), admin_scratch_path=str(scratch))
    )
    assert rpt.rich_guard_rows
    assert "Preserve admin" in rpt.rich_guard_rows[0]["Action Needed"]
    assert float(rpt.rich_guard_rows[0]["Current Admin Value"]) >= 8


def test_partial_hours_are_review_not_error():
    from triage.nw_prj_artifact_compare import _partial_hours

    assert _partial_hours(4.5)
    assert not _partial_hours(8)
    assert not _partial_hours(0)


def test_total_formula_protection(tmp_path: Path):
    scratch = tmp_path / "adm.xlsx"
    openpyxl.Workbook().save(scratch)
    res = generate_dashboard(
        GenerateInputs(
            admin_scratch_path=str(scratch),
            descriptor="TEST",
            version_minor="9",
            out_dir=str(tmp_path),
        )
    )
    wb = openpyxl.load_workbook(res.output_path, data_only=True)
    text = wb["Start Here"]["A1"].value or ""
    assert "Total" in text
    assert res.web_excel_safe


def test_review_status_bucket():
    assert review_status_bucket("Done") == "resolved_green"
    assert review_status_bucket("Skipped/Gray") == "skipped_gray"
