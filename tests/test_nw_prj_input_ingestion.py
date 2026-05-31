"""
Contract tests for NW PRJ input ingestion (admin scratch + roster readers).

These tests are xfail-strict against ``NotImplementedError`` because the
readers are scaffolded for feature/nw-prj-ingest-admin-roster-rows and the
ingestion PR has not landed yet. When implementation arrives, the xfails
flip to passes automatically and the strict flag fails the suite if a test
unexpectedly passes for the wrong reason.

The body of each test describes the binding contract for the reader so the
implementation PR has a clear target.
"""
from __future__ import annotations

from pathlib import Path

import pytest

from triage.nw_prj_admin_scratch_reader import (
    AdminScratchEvidence,
    read_admin_scratch,
    read_official_admin,
)
from triage.nw_prj_roster_reader import (
    RosterEvidence,
    read_roster_log,
    split_note_bearing_punch,
)


# ── shape ──


def test_admin_scratch_evidence_is_frozen_dataclass():
    ev = AdminScratchEvidence(
        tech="Alice",
        date="2026-05-01",
        source_workbook="scratch.xlsx",
        source_sheet="Hours",
        source_row=5,
    )
    with pytest.raises(Exception):
        ev.tech = "Bob"  # type: ignore[misc]


def test_roster_evidence_is_frozen_dataclass():
    ev = RosterEvidence(
        tech="Alice",
        date="2026-05-01",
        project="NW PRJ",
        clock_in_raw=None,
        clock_out_raw=None,
        clock_in_hours=None,
        clock_out_hours=None,
        gross_hours=None,
        lunch_deduction=None,
        net_hours=None,
        long_shift=False,
    )
    with pytest.raises(Exception):
        ev.tech = "Bob"  # type: ignore[misc]


# ── admin scratch reader contract ──


@pytest.mark.xfail(raises=NotImplementedError, strict=True, reason="ingestion PR pending")
def test_read_admin_scratch_returns_evidence_list(tmp_path: Path):
    import openpyxl

    p = tmp_path / "scratch.xlsx"
    openpyxl.Workbook().save(p)
    result = read_admin_scratch(p)
    assert isinstance(result, list)
    assert all(isinstance(r, AdminScratchEvidence) for r in result)


@pytest.mark.xfail(raises=NotImplementedError, strict=True, reason="ingestion PR pending")
def test_read_official_admin_returns_evidence_list(tmp_path: Path):
    import openpyxl

    p = tmp_path / "official.xlsx"
    openpyxl.Workbook().save(p)
    result = read_official_admin(p)
    assert isinstance(result, list)
    assert all(isinstance(r, AdminScratchEvidence) for r in result)


def test_read_admin_scratch_missing_file_raises(tmp_path: Path):
    with pytest.raises((FileNotFoundError, NotImplementedError)):
        read_admin_scratch(tmp_path / "does_not_exist.xlsx")


# ── roster reader contract ──


@pytest.mark.xfail(raises=NotImplementedError, strict=True, reason="ingestion PR pending")
def test_read_roster_log_returns_evidence_list(tmp_path: Path):
    import openpyxl

    p = tmp_path / "roster.xlsx"
    openpyxl.Workbook().save(p)
    result = read_roster_log(p)
    assert isinstance(result, list)
    assert all(isinstance(r, RosterEvidence) for r in result)


def test_split_note_bearing_punch_separates_time_and_note():
    time_text, note_text = split_note_bearing_punch("9:28:00 AM/ Bonita")
    assert time_text.strip() == "9:28:00 AM"
    assert note_text == "Bonita"


def test_split_note_bearing_punch_pure_time_has_empty_note():
    time_text, note_text = split_note_bearing_punch("9:28:00 AM")
    assert time_text == "9:28:00 AM"
    assert note_text == ""


# ── readers stay dumb ──


def _imported_modules(module) -> set[str]:
    import ast

    src = Path(module.__file__).read_text(encoding="utf-8")
    tree = ast.parse(src)
    names: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                names.add(alias.name)
        elif isinstance(node, ast.ImportFrom):
            if node.module:
                names.add(node.module)
    return names


def test_admin_scratch_reader_does_not_import_classifier():
    """Reader must not import the classifier or status taxonomy.

    Readers extract evidence only. Authority and classification belong to
    ``triage.nw_prj_target_classifier``. Importing the classifier from a
    reader would invert the dependency direction.
    """
    import triage.nw_prj_admin_scratch_reader as mod

    imports = _imported_modules(mod)
    assert "triage.nw_prj_target_classifier" not in imports
    assert "triage.nw_prj_config" not in imports
    assert "triage.nw_prj_dashboard_validator" not in imports


def test_roster_reader_does_not_import_classifier():
    import triage.nw_prj_roster_reader as mod

    imports = _imported_modules(mod)
    assert "triage.nw_prj_target_classifier" not in imports
    assert "triage.nw_prj_config" not in imports
    assert "triage.nw_prj_dashboard_validator" not in imports
