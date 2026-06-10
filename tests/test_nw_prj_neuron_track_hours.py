"""NW PRJ Neuron Track Hours engine tests — real implementation, no xfails."""
from __future__ import annotations

import io
import shutil
import zipfile
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from triage.webexcel_semantic_gate import run_semantic_gate

from triage.nw_prj_neuron_track_hours.classifier import (
    flag_missing_roster,
    rich_guard_review,
)
from triage.nw_prj_neuron_track_hours.cli import REFERENCE_TARGETS, build_report, run
from triage.nw_prj_neuron_track_hours.exporter import EXPECTED_SHEETS
from triage.nw_prj_neuron_track_hours.preflight import run_preflight
from triage.nw_prj_neuron_track_hours.reader import split_note_bearing_punch
from tests.fixtures.nw_prj_neuron_track_hours.fixtures import (
    build_fixtures,
    write_malformed_xlsx,
)

REPO_ROOT = Path(__file__).resolve().parent.parent
FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures" / "nw_prj_neuron_track_hours"
REAL_ROSTER = (
    REPO_ROOT
    / "Candidates/attendacne artifacts 6-1-2026/INTERNAL_May_Billing_Active_Roster_Log_2026-06-01-update so that partial hours are flagged before submission.xlsx"
)


@pytest.fixture(scope="module")
def fixtures():
    return build_fixtures(FIXTURE_DIR)


@pytest.fixture(scope="module")
def generated(fixtures, tmp_path_factory):
    root = tmp_path_factory.mktemp("nth_repo")
    out = root / "Outputs" / "nw_prj_neuron_track_hours" / "test_run"
    out.mkdir(parents=True)
    manifest = run(
        roster_log=str(fixtures["roster"]),
        out_dir=str(out),
        months=["2026-04", "2026-05"],
        websafe=True,
        zip_output=True,
        repo_root=root,
    )
    return manifest


# ── CLI / workbook structure ───────────────────────────────────────

def test_neuron_track_cli_generates_workbook(generated):
    wb_path = Path(generated["outputs"]["workbook"])
    assert wb_path.exists()
    assert Path(generated["outputs"]["zip"]).exists()
    assert Path(generated["outputs"]["reconciliation_json"]).exists()
    assert Path(generated["outputs"]["review_queue_csv"]).exists()


def test_neuron_track_has_april_and_may_tabs(generated):
    wb = openpyxl.load_workbook(generated["outputs"]["workbook"], read_only=True)
    for sheet in EXPECTED_SHEETS:
        assert sheet in wb.sheetnames, f"missing sheet {sheet}"
    assert "April Neuron Hours" in wb.sheetnames
    assert "May Neuron Hours" in wb.sheetnames
    wb.close()


def test_neuron_track_go_live_weekend_support_visible(generated):
    wb = openpyxl.load_workbook(generated["outputs"]["workbook"], read_only=True)
    ws = wb["Go Live Weekend"]
    vals = [tuple(r) for r in ws.iter_rows(values_only=True)]
    flat = [str(c) for row in vals for c in row if c is not None]
    assert any("2026-05-30" in s for s in flat)
    assert any("2026-05-31" in s for s in flat)
    wb.close()
    assert generated["totals"]["go_live_rows"] == 2
    assert generated["totals"]["go_live_hours"] == 22.0


def test_synthetic_totals(generated):
    # April: 9 (Alpha) + 4 (Beta override) = 13 ; May go-live = 22
    assert generated["totals"]["april"] == 13.0
    assert generated["totals"]["may"] == 22.0
    assert generated["totals"]["total"] == 35.0


def test_workbook_has_filters_and_frozen_headers(generated):
    pf = run_preflight(generated["outputs"]["workbook"], expected_sheets=EXPECTED_SHEETS)
    assert pf.has_filters
    assert pf.has_frozen_header


def test_workbook_has_cf_dictionary(generated):
    pf = run_preflight(generated["outputs"]["workbook"], expected_sheets=EXPECTED_SHEETS)
    assert pf.has_cf_dictionary
    assert pf.has_conditional_formatting


def test_workbook_has_dropdown_validations(generated):
    pf = run_preflight(generated["outputs"]["workbook"], expected_sheets=EXPECTED_SHEETS)
    assert pf.has_dropdowns


def test_workbook_preflight_pass(generated):
    assert generated["websafe_preflight_pass"] is True


# ── Classifier units ───────────────────────────────────────────────

def test_rich_guard_preserves_full_day_review():
    flag = rich_guard_review("Asher Ali", date(2026, 4, 1), roster_hours=4.0, admin_hours=8.0)
    assert flag is not None
    assert flag.severity == "PURPLE"
    assert flag.gross_hours == 8.0  # admin full day preserved, not downgraded
    # No protection when roster already meets/exceeds admin
    assert rich_guard_review("Asher Ali", date(2026, 4, 1), roster_hours=9.0, admin_hours=8.0) is None


def test_pinned_names_do_not_become_missing_roster_failures():
    flags = flag_missing_roster(
        expected_techs=["Alpha Tech", "Pinned Tech"],
        present_techs=["Alpha Tech"],
        pinned_techs=["Pinned Tech"],
    )
    assert flags == []
    # Without pinning, the absent tech IS flagged
    flags2 = flag_missing_roster(
        expected_techs=["Alpha Tech", "Ghost Tech"],
        present_techs=["Alpha Tech"],
        pinned_techs=[],
    )
    assert len(flags2) == 1
    assert flags2[0].severity == "RED"


def test_note_bearing_punch_parses_time_and_preserves_note():
    hours, note = split_note_bearing_punch("9:28:00 AM/ Bonita")
    assert abs(hours - (9 + 28 / 60)) < 1e-6
    assert note == "Bonita"
    hours2, note2 = split_note_bearing_punch("5:00 PM")
    assert hours2 == 17.0
    assert note2 == ""


# ── Preflight rejection ─────────────────────────────────────────────

def test_webexcel_preflight_rejects_inlineStr_ns0_calcChain(tmp_path):
    bad = tmp_path / "bad.xlsx"
    write_malformed_xlsx(bad)
    pf = run_preflight(str(bad))
    assert "inlineStr" in pf.token_failures
    assert "ns0:" in pf.token_failures
    assert "calcChain.xml" in pf.token_failures
    assert pf.preflight_pass is False


# ── Real roster regression ──────────────────────────────────────────

@pytest.mark.skipif(not REAL_ROSTER.exists(), reason="Real roster log not present")
def test_neuron_track_totals_match_reference_targets(tmp_path):
    report = build_report(str(REAL_ROSTER), ["2026-04", "2026-05"])
    assert abs(report.month_total("April") - REFERENCE_TARGETS["april"]) <= 0.05
    assert abs(report.month_total("May") - REFERENCE_TARGETS["may"]) <= 0.05
    assert abs(report.grand_total() - REFERENCE_TARGETS["total"]) <= 0.05
    assert len(report.go_live_rows()) == REFERENCE_TARGETS["go_live_rows"]
    assert abs(report.go_live_hours() - REFERENCE_TARGETS["go_live_hours"]) <= 0.05


# ── sharedStrings invariant (Web Excel repair guard) ───────────────

def test_missing_sharedstrings_with_live_refs_fails_preflight(tmp_path):
    """Worksheet t="s" refs with no sharedStrings.xml must fail preflight.

    This is the exact signature that triggers Excel-for-Web "repair", which
    silently strips text. The count guard must catch the missing part rather
    than passing it through.
    """
    from triage.xlsx_utils import fix_inlinestr

    src = tmp_path / "with_ss.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "RealNeuronText"
    ws["B1"] = "AnotherRealString"
    wb.save(str(src))
    # openpyxl writes inlineStr; convert to shared-string t="s" refs + a real
    # sharedStrings.xml so that dropping the part leaves live refs behind.
    fix_inlinestr(str(src))

    raw = src.read_bytes()
    stripped = tmp_path / "no_ss.xlsx"
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(raw), "r") as zin, \
            zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            if name == "xl/sharedStrings.xml":
                continue  # drop the table while leaving t="s" refs behind
            zout.writestr(name, zin.read(name))
    stripped.write_bytes(buf.getvalue())

    pf = run_preflight(str(stripped))
    assert pf.sharedstrings_count_ok is False
    assert pf.preflight_pass is False


# ── neuron_track dashboard structural sentinels ────────────────────

def test_neuron_track_dashboard_passes_sentinels(generated):
    gate = run_semantic_gate(generated["outputs"]["workbook"], profile="neuron_track")
    assert gate["semantic_integrity"] == "PASS"
    assert gate["sentinel_failures"] == []


def test_neuron_track_blank_start_here_fails_sentinels(generated, tmp_path):
    damaged = tmp_path / "blank_start.xlsx"
    shutil.copy(generated["outputs"]["workbook"], damaged)
    wb = openpyxl.load_workbook(str(damaged))
    start = next(n for n in wb.sheetnames if "start here" in n.lower())
    wb[start]["A1"].value = None
    wb.save(str(damaged))
    wb.close()

    gate = run_semantic_gate(str(damaged), profile="neuron_track")
    assert gate["semantic_integrity"] == "FAIL"
    assert any("Start Here" in f and "blank" in f for f in gate["sentinel_failures"])
