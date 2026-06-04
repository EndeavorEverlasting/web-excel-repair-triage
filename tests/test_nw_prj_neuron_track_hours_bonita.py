"""Bonita-format Neuron Track Hours generator — fixture-only tests (13 cases)."""
from __future__ import annotations

import importlib
from pathlib import Path

import openpyxl
import pytest

from triage.nw_prj_neuron_track_hours.bonita_cli import run
from triage.nw_prj_neuron_track_hours.bonita_resolver import resolve_bonita_shifts
from tests.fixtures.nw_prj_neuron_track_hours.bonita_fixtures import build_bonita_fixtures

REPO_ROOT = Path(__file__).resolve().parent.parent
FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures" / "nw_prj_neuron_track_hours"
MONTHS = ["2026-04", "2026-05"]


@pytest.fixture(scope="module")
def fixtures():
    return build_bonita_fixtures(FIXTURE_DIR)


@pytest.fixture(scope="module")
def resolution(fixtures):
    return resolve_bonita_shifts(str(fixtures["roster"]), MONTHS)


@pytest.fixture(scope="module")
def generated(fixtures, tmp_path_factory):
    out = tmp_path_factory.mktemp("bonita_out")
    return run(
        roster_log=str(fixtures["roster"]),
        out_dir=str(out),
        months=MONTHS,
        websafe=True,
        repo_root=REPO_ROOT,
    )


# 1 ─ CLI import smoke ───────────────────────────────────────────────
def test_cli_imports():
    mod = importlib.import_module("triage.nw_prj_neuron_track_hours.bonita_cli")
    assert hasattr(mod, "run")
    assert hasattr(mod, "main")


# 2 ─ Exactly Apr 26 + May 26 tabs ───────────────────────────────────
def test_exactly_two_named_tabs(generated):
    wb = openpyxl.load_workbook(generated["outputs"]["workbook"], read_only=True)
    assert wb.sheetnames == ["Apr 26", "May 26"]
    wb.close()


# 3 ─ April spans full month, not Apr 1-4 ────────────────────────────
def test_april_spans_full_month(resolution):
    april_days = {s.date.day for s in resolution.shifts_for_month("April")}
    assert max(april_days) == 30, f"April should reach day 30, saw {sorted(april_days)}"
    assert any(d > 4 for d in april_days)


# 4 ─ Note-bearing punches parse start/end; notes absent from cells ──
def test_note_bearing_punch_parsed_and_notes_not_in_workbook(resolution, generated):
    bravo = [s for s in resolution.shifts_for_month("April") if s.tech == "Bravo Tech"]
    assert len(bravo) == 1
    assert bravo[0].clock_in == "8:00 AM"
    assert bravo[0].clock_out == "4:00 PM"
    assert bravo[0].total_hours == 8.0
    # The note text must never appear in any workbook cell.
    wb = openpyxl.load_workbook(generated["outputs"]["workbook"], read_only=True)
    flat = [
        str(c)
        for ws in wb.worksheets
        for row in ws.iter_rows(values_only=True)
        for c in row
        if c is not None
    ]
    wb.close()
    assert not any("covered" in s.lower() for s in flat)
    assert not any("bonita" in s.lower() for s in flat)


# 5 ─ Non-work markers skipped -> review queue ───────────────────────
def test_non_work_marker_skipped_to_review(resolution):
    golf_shifts = [s for s in resolution.shifts if s.tech == "Golf Tech"]
    assert golf_shifts == []
    markers = [r for r in resolution.review if r.category == "non_work_marker"]
    assert any(r.tech == "Golf Tech" and "PTO" in r.note.upper() for r in markers)


# 6b ─ One workbook row per resolved shift (not collapsed aggregate) ─────────
def test_workbook_row_per_shift_not_collapsed(resolution, generated):
    wb = openpyxl.load_workbook(generated["outputs"]["workbook"], read_only=True)
    for month_name, tab in (("April", "Apr 26"), ("May", "May 26")):
        expected = len(resolution.shifts_for_month(month_name))
        ws = wb[tab]
        col_tech = col_total = None
        for c in range(1, (ws.max_column or 1) + 1):
            h = str(ws.cell(row=1, column=c).value or "").strip().upper()
            if h == "TECH":
                col_tech = c
            if h == "TOTAL":
                col_total = c
        assert col_tech and col_total
        rows = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if len(row) < col_total:
                continue
            hours = row[col_total - 1]
            tech = row[col_tech - 1] if len(row) >= col_tech else None
            if isinstance(hours, (int, float)) and float(hours) > 0 and tech:
                rows += 1
        assert rows == expected, f"{tab}: expected {expected} shift rows, saw {rows}"
    wb.close()


def test_assignment_column_not_uniform_generic_label(generated):
    wb = openpyxl.load_workbook(generated["outputs"]["workbook"], read_only=True)
    seen: set[str] = set()
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        col_assign = None
        for c in range(1, (ws.max_column or 1) + 1):
            if str(ws.cell(row=1, column=c).value or "").strip().upper() == "ASSIGNMENT":
                col_assign = c
                break
        if not col_assign:
            continue
        for row in ws.iter_rows(min_row=3, values_only=True):
            if len(row) < col_assign:
                continue
            val = row[col_assign - 1]
            if val is not None and str(val).strip():
                seen.add(str(val).strip())
    wb.close()
    assert seen, "expected assignment values in workbook"
    assert not all(v.lower() == "generic" for v in seen)


# 6 ─ Worked-project override beats default ──────────────────────────
def test_worked_project_override_beats_default(resolution):
    # Delta default Delivery, worked-project override -> Neuron: counted.
    delta = [s for s in resolution.shifts if s.tech == "Delta Tech"]
    assert len(delta) == 1 and delta[0].total_hours == 8.0
    # Echo default Neuron, worked-project override -> Delivery: excluded.
    echo = [s for s in resolution.shifts if s.tech == "Echo Tech"]
    assert echo == []


# 7 ─ Approved Assignments override beats default ────────────────────
def test_assignments_override_beats_default(resolution):
    india = [s for s in resolution.shifts if s.tech == "India Tech"]
    assert len(india) == 1
    assert india[0].total_hours == 8.0
    assert india[0].project_name == "Northwell - Neurons"


# 8 ─ Long shift included but flagged in review ──────────────────────
def test_long_shift_included_and_flagged(resolution):
    hotel = [s for s in resolution.shifts if s.tech == "Hotel Tech"]
    assert len(hotel) == 1
    assert hotel[0].total_hours == 17.0
    assert hotel[0].long_shift is True
    assert any(r.category == "long_shift" and r.tech == "Hotel Tech"
               for r in resolution.review)


# 9 ─ Excluded names never counted ───────────────────────────────────
def test_excluded_name_never_counted(resolution):
    assert all(s.tech != "Yostinn Minaya" for s in resolution.shifts)
    assert any(r.category == "excluded_name" and r.tech == "Yostinn Minaya"
               for r in resolution.review)


# 10 ─ No blank start/end/total on populated rows ────────────────────
def test_no_blank_cells_on_populated_rows(generated):
    wb = openpyxl.load_workbook(generated["outputs"]["workbook"], read_only=True)
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        for row in rows:
            if not any(c is not None for c in row):
                continue
            # cols: DATE, TECH, START, END, TOTAL, PROJECT, ASSIGNMENT
            for idx in (1, 2, 3, 4, 5, 6):
                assert row[idx] not in (None, ""), f"blank col {idx} in {ws.title}: {row}"
    wb.close()


# 11 ─ Manifest has per-month row counts + total hours ───────────────
def test_manifest_has_per_month_counts_and_totals(generated):
    pm = generated["per_month"]
    assert pm["2026-04"]["row_count"] == 7
    assert pm["2026-04"]["total_hours"] == 66.0
    assert pm["2026-05"]["row_count"] == 2
    assert pm["2026-05"]["total_hours"] == 16.0
    assert generated["grand_total_hours"] == 82.0


# 12 ─ Preflight passes (no inlineStr/ns0/calcChain/external links) ──
def test_preflight_passes(generated):
    assert generated["websafe_preflight_pass"] is True
    import json
    pf = json.loads(Path(generated["outputs"]["preflight_json"]).read_text(encoding="utf-8"))
    assert pf["token_failures"] == []
    assert pf["has_calc_chain"] is False
    assert pf["has_external_links"] is False


# 13 ─ / Bonita off-project punch excluded from totals but in review ─
def test_off_project_bonita_excluded_but_in_review(resolution):
    assert all(s.tech != "Charlie Tech" for s in resolution.shifts)
    off = [r for r in resolution.review if r.category == "off_project"]
    assert any(r.tech == "Charlie Tech" and "bonita" in r.note.lower() for r in off)


# 14 ─ sharedStrings count invariant (the Excel-for-Web repair guard) ─
def test_sharedstrings_count_invariant(generated):
    import json
    pf = json.loads(Path(generated["outputs"]["preflight_json"]).read_text(encoding="utf-8"))
    assert pf["sharedstrings_count_ok"] is True
    assert pf["sharedstrings_declared_count"] == pf["sharedstrings_actual_refs"]


# 15 ─ Tracker uses real time cells (h:mm AM/PM), numeric totals, bold ─
def test_tracker_time_values_and_number_formats(generated):
    import datetime
    wb = openpyxl.load_workbook(generated["outputs"]["workbook"])
    ws = wb["Apr 26"]
    assert ws.cell(1, 1).value in (None, "")   # column A header is blank
    assert ws.cell(3, 3).value.__class__ is datetime.time
    assert ws.cell(3, 4).value.__class__ is datetime.time
    assert ws.cell(3, 3).number_format == "h:mm AM/PM"
    assert ws.cell(3, 5).number_format == "0.00"
    assert isinstance(ws.cell(3, 5).value, (int, float))
    assert ws.cell(3, 2).font.bold is True
    wb.close()


# 16 ─ Approved Overrides sub-table beats a non-Neuron worked project ─
def test_override_beats_worked_project_in_bonita_resolver(tmp_path):
    """A reviewed Assignments *Overrides* entry must outrank Worked Projects.

    Zulu defaults to Delivery and the Worked Projects sheet also says Delivery
    (non-Neuron -> would be excluded). A reviewed override to Neuron Deployments
    must win, so the shift is counted.
    """
    from datetime import datetime

    roster = tmp_path / "override_roster.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    live = wb.create_sheet("Live - April 2026")
    live.append(["April 2026 - Attendance"])
    live.append(["Staff Name", "Project",
                 "Apr 02 - Clock In", "Apr 02 - Clock Out"])
    live.append(["Zulu Tech", "Delivery / Transport", "8:00 AM", "4:00 PM"])

    worked = wb.create_sheet("Worked Projects - April 2026")
    worked.append(["April 2026 - Worked Projects"])
    worked.append(["Staff Name", "Default Project", datetime(2026, 4, 2)])
    worked.append(["Zulu Tech", "Delivery / Transport", "Delivery / Transport"])

    assign = wb.create_sheet("Assignments - April 2026")
    assign.append(["April 2026 - Project Assignments"])
    assign.append(["Staff Name", "Default Project", datetime(2026, 4, 2)])
    assign.append(["Overrides (only if different from Default Project)"])
    assign.append(["Override Staff Name", "Override Date", "Override Project", "Notes"])
    assign.append(["Zulu Tech", datetime(2026, 4, 2), "Neuron Deployments",
                   "Reviewed: neuron confirmed"])
    wb.save(str(roster))

    resolution = resolve_bonita_shifts(str(roster), ["2026-04"])
    zulu = [s for s in resolution.shifts if s.tech == "Zulu Tech"]
    assert len(zulu) == 1
    assert zulu[0].total_hours == 8.0
    assert zulu[0].project_name == "Northwell - Neurons"
