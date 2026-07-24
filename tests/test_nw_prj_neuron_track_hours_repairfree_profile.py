"""Repair-free Bonita-format Neuron Track Hours profile gate tests."""
from __future__ import annotations

import importlib
import json
import zipfile
from pathlib import Path

import openpyxl
import pytest

from triage.nw_prj_neuron_track_hours.bonita_cli import run
from triage.nw_prj_neuron_track_hours.repairfree_profile_gate import (
    load_repairfree_profile,
    run_repairfree_profile_gate,
)
from tests.fixtures.nw_prj_neuron_track_hours.bonita_fixtures import build_bonita_fixtures

REPO_ROOT = Path(__file__).resolve().parent.parent
FIXTURE_DIR = REPO_ROOT / "tests" / "fixtures" / "nw_prj_neuron_track_hours"
PROFILE_PATH = REPO_ROOT / "tests" / "fixtures" / "nw_prj_neuron_track_hours_repairfree" / "golden_profile.json"
MONTHS = ["2026-04", "2026-05"]


@pytest.fixture(scope="module")
def fixtures():
    return build_bonita_fixtures(FIXTURE_DIR)


@pytest.fixture(scope="module")
def generated(fixtures, tmp_path_factory):
    out = tmp_path_factory.mktemp("repairfree_out")
    return run(
        roster_log=str(fixtures["roster"]),
        out_dir=str(out),
        months=MONTHS,
        websafe=True,
        reference_profile=str(PROFILE_PATH),
        repo_root=REPO_ROOT,
    )


def test_cli_imports():
    mod = importlib.import_module("triage.nw_prj_neuron_track_hours.bonita_cli")
    assert hasattr(mod, "run")


def test_golden_profile_json_valid():
    profile = load_repairfree_profile(PROFILE_PATH)
    assert profile["profile"] == "neuron_track_hours_repairfree_golden"
    assert "Apr 26" in profile["required_sheets"]
    assert profile["header_row2"][0] == "DATE"


def test_generated_workbook_passes_profile_gate(generated):
    wb_path = generated["outputs"]["workbook"]
    gate = run_repairfree_profile_gate(wb_path, PROFILE_PATH)
    assert gate["profile_pass"] is True, gate.get("profile_failures")


def test_preflight_includes_repairfree_profile(generated):
    pf = generated["preflight_data"]
    assert pf.get("repairfree_profile_pass") is True


def test_month_sheet_headers_match_profile(generated):
    profile = load_repairfree_profile(PROFILE_PATH)
    wb = openpyxl.load_workbook(generated["outputs"]["workbook"], read_only=True)
    ws = wb["Apr 26"]
    assert ws.cell(1, 2).value == profile["header_row1"][1]
    assert ws.cell(2, 1).value == "DATE"
    wb.close()


def test_month_sheet_uses_columns_a_through_g_only(generated):
    wb = openpyxl.load_workbook(generated["outputs"]["workbook"], read_only=True)
    ws = wb["Apr 26"]
    assert ws.max_column <= 7
    wb.close()


def test_no_formulas_in_month_sheets(generated):
    wb = openpyxl.load_workbook(generated["outputs"]["workbook"], data_only=False)
    for name in ("Apr 26", "May 26"):
        ws = wb[name]
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
            for cell in row:
                assert getattr(cell, "data_type", None) != "f"
    wb.close()


def test_package_has_no_calc_chain_or_external_links(generated):
    path = generated["outputs"]["workbook"]
    with zipfile.ZipFile(path, "r") as z:
        names = z.namelist()
    assert "xl/calcChain.xml" not in names
    assert not any("externalLink" in n for n in names)


def test_websafe_alias_emitted(generated, tmp_path_factory):
    out = Path(generated["outputs"]["workbook"]).parent
    alias = out / "Neuron_Track_Hours_April_May_2026_WEBSAFE.xlsx"
    assert alias.is_file()


def test_sidecars_emitted(generated):
    out = Path(generated["outputs"]["workbook"]).parent
    assert (out / "Neuron_Track_Hours_April_May_2026_manifest.json").is_file()
    assert (out / "Neuron_Track_Hours_April_May_2026_review_queue.csv").is_file()
    assert (out / "Neuron_Track_Hours_April_May_2026_preflight.json").is_file()
