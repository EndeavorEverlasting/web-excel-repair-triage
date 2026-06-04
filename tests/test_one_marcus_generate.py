"""Tests for the One Marcus clean-render generator."""
from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

from tests.fixtures.one_marcus_recon import fixtures as fx
from triage.one_marcus_recon.exporter import run_generate
from triage.one_marcus_recon.operational_checks import run_operational_checks
from triage.one_marcus_recon.style_pass import apply_style_pass

READY_WB = Path("Candidates/inventory recon/1M_Recon_READY.xlsx")


@pytest.fixture
def integrated_source(tmp_path):
    return fx.make_integrated_source(str(tmp_path / "integrated_source.xlsx"))


@pytest.fixture
def generated_output(integrated_source, tmp_path):
    out = tmp_path / "1M_Recon_READY.xlsx"
    result = run_generate(integrated_source, output_path=str(out))
    assert result.report.webexcel_preflight_pass
    assert result.report.operational_pass
    return str(out)


def test_generate_creates_required_sheets(integrated_source, tmp_path):
    out = tmp_path / "out.xlsx"
    result = run_generate(integrated_source, output_path=str(out))
    assert result.report.mode == "generate"
    assert result.report.operational_pass
    with zipfile.ZipFile(out) as z:
        wb = z.read("xl/workbook.xml").decode("utf-8")
    assert "Part Numbers" in wb
    assert "1M Recon Pivot Module" in wb


def test_visual_column_after_total_qty(generated_output):
    wb = openpyxl.load_workbook(generated_output, data_only=False, read_only=True)
    try:
        pivot = wb["1M Recon Pivot Module"]
        headers = [pivot.cell(12, c).value for c in range(1, 8)]
        assert headers[0] == "Inventory Rollup by Item"
        assert headers[1] == "Total Qty"
        assert headers[2] == "Visual"
        formula = pivot.cell(13, 3).value
        assert isinstance(formula, str) and "REPT" in formula.upper()
    finally:
        wb.close()


def test_no_forbidden_workbook_text(generated_output):
    ops = run_operational_checks(generated_output)
    forbidden = [f for f in ops.failures if "websafe" in f or "%20" in f]
    assert not forbidden


def test_style_pass_preserves_formulas(generated_output, tmp_path):
    copy = tmp_path / "styled.xlsx"
    copy.write_bytes(Path(generated_output).read_bytes())
    before, after = apply_style_pass(str(copy))
    assert before == after


def test_operational_checks_fail_without_visual(tmp_path, integrated_source):
    out = tmp_path / "bad.xlsx"
    run_generate(integrated_source, output_path=str(out))
    wb = openpyxl.load_workbook(out)
    ws = wb["1M Recon Pivot Module"]
    ws.cell(12, 3, "NotVisual")
    ws.cell(13, 3, "")
    wb.save(out)
    wb.close()
    ops = run_operational_checks(str(out))
    assert not ops.operational_pass


@pytest.mark.skipif(not READY_WB.exists(), reason="private operator reference workbook not present")
def test_generate_from_operator_reference(tmp_path):
    out = tmp_path / "1M_Recon_READY_generated.xlsx"
    result = run_generate(str(READY_WB), output_path=str(out))
    assert result.report.webexcel_preflight_pass
    assert result.report.operational_pass
    assert result.report.rollup_key_count >= 1
