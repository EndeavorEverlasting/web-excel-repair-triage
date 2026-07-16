from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection

from triage.prompt_kit_v33_artifact_contract import validate_artifact


def _make_valid_artifact(path: Path) -> None:
    workbook = Workbook()
    library = workbook.active
    library.title = "Prompt_Library"
    library.append(["↓ Bottom", "Seq", "Prompt ID"] + [None] * 11 + ["Copy-Safe Sheet", "↓ Bottom"])
    prompt_ids = ("P07", "P45", "P46", "P47")
    fill = PatternFill("solid", fgColor="FFF2CC")
    for index, prompt_id in enumerate(prompt_ids, start=2):
        sheet_name = f"{prompt_id}_COPY_SAFE"
        values = [None, prompt_id[1:], prompt_id] + [f"value-{prompt_id}"] * 11 + [sheet_name, None]
        library.append(values)
        for column in range(1, 17):
            library.cell(row=index, column=column).fill = fill
        prompt_cell = library.cell(row=index, column=3)
        prompt_cell.font = Font(bold=True, underline="single")
        if prompt_id == "P07":
            prompt_cell.value = f'=HYPERLINK("#\'{sheet_name}\'!A1:A3","{prompt_id}")'
        else:
            prompt_cell.hyperlink = f"#'{sheet_name}'!A1:A3"
        ws = workbook.create_sheet(sheet_name)
        ws.column_dimensions["A"].width = 110
        ws["A1"] = f"{prompt_id} first line"
        ws["A2"] = ""
        ws["A3"] = "last line"
        back = f"#'Prompt_Library'!A{index}:P{index}"
        for coordinate in ("B1", "E1", "B3", "E3"):
            ws[coordinate] = "Prompt Library"
            ws[coordinate].hyperlink = back
        ws.protection.sheet = True
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True)
    footer_row = len(prompt_ids) + 2
    library.append(["↑ Top", "End of Prompt Library"] + [None] * 13 + ["↑ Top"])
    library["A1"].hyperlink = f"#'Prompt_Library'!A{footer_row}"
    library["P1"].hyperlink = f"#'Prompt_Library'!P{footer_row}"
    library[f"A{footer_row}"].hyperlink = "#'Prompt_Library'!A1"
    library[f"P{footer_row}"].hyperlink = "#'Prompt_Library'!P1"
    library.protection.sheet = True
    for row in library.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

    discovery = workbook.create_sheet("Opportunity_Discovery")
    discovery.protection.sheet = True
    for row in discovery.iter_rows(min_row=1, max_row=100, min_col=1, max_col=18):
        for cell in row:
            cell.protection = Protection(locked=False)
    discovery["S101"] = "locked"
    discovery["S101"].protection = Protection(locked=True)

    for sheet_name in ("Prompt_Library", "Opportunity_Discovery", "P07_COPY_SAFE", "P46_COPY_SAFE"):
        workbook[sheet_name].sheet_properties.tabColor = "FFF2CC"
    workbook.save(path)


def test_valid_artifact_passes(tmp_path: Path) -> None:
    artifact = tmp_path / "AI_Harness_Prompt_Kit_v33.xlsx"
    _make_valid_artifact(artifact)
    result = validate_artifact(artifact)
    assert result.passed is True
    assert result.prompt_count == 4
    assert result.findings == ()
    assert len(result.sha256) == 64


def test_contract_rejects_single_cell_link_wrong_tab_and_unlocked_cell(tmp_path: Path) -> None:
    artifact = tmp_path / "AI_Harness_Prompt_Kit_v33.xlsx"
    _make_valid_artifact(artifact)
    workbook = load_workbook(artifact)
    library = workbook["Prompt_Library"]
    library["C3"].hyperlink = "#'P45_COPY_SAFE'!A1"
    library["C3"].protection = Protection(locked=False)
    workbook["P46_COPY_SAFE"].sheet_properties.tabColor = "FF0000"
    workbook.save(artifact)

    findings = validate_artifact(artifact).findings
    assert any("P45 forward link must select full A1:A<n> range" in item for item in findings)
    assert any("P46_COPY_SAFE tab color" in item for item in findings)
    assert any("cell outside editable range is unlocked: Prompt_Library!C3" in item for item in findings)


def test_contract_rejects_missing_required_prompt(tmp_path: Path) -> None:
    artifact = tmp_path / "AI_Harness_Prompt_Kit_v33.xlsx"
    _make_valid_artifact(artifact)
    workbook = load_workbook(artifact)
    library = workbook["Prompt_Library"]
    library["B5"] = ""
    library["C5"] = ""
    workbook.save(artifact)
    findings = validate_artifact(artifact).findings
    assert "missing required prompt: P47" in findings
