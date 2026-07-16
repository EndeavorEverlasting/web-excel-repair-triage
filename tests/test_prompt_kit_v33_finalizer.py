from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from triage.prompt_kit_v33_finalizer import finalize_workbook


def _make_source(path: Path) -> None:
    wb = Workbook()
    library = wb.active
    library.title = "Prompt_Library"
    headers = [
        "↓ Bottom",
        "Seq",
        "Prompt ID",
        "Prompt Type",
        "Prompt Class",
        "Sprint Path Role",
        "Use For Progress?",
        "Prompt Name",
        "Use This When",
        "Inspect First",
        "Expected Output",
        "Next Step",
        "Proof / Acceptance Gate",
        "Color",
        "Copy-Safe Sheet",
        "↓ Bottom",
    ]
    library.append(headers)
    for number in range(45):
        prompt_id = f"P{number:02d}"
        library.append([
            None,
            f"{number:02d}",
            prompt_id,
            "BUILD",
            "TEST",
            "role",
            "YES",
            f"Prompt {prompt_id}",
            "when",
            "inspect",
            "output",
            "next",
            "gate",
            "Slate",
            f"{prompt_id}_COPY_SAFE",
            None,
        ])
        ws = wb.create_sheet(f"{prompt_id}_COPY_SAFE")
        ws["A1"] = f"{prompt_id} body"
        ws["A2"] = "second line"
    library.append(["↑ Top", "End of Prompt Library · P00–P44"] + [None] * 13 + ["↑ Top"])
    discovery = wb.create_sheet("Opportunity_Discovery")
    discovery["A1"] = "editable"
    discovery["S101"] = "locked"
    wb.save(path)


def _target(cell) -> str:
    return cell.hyperlink.target if cell.hyperlink else ""


def test_finalizer_adds_gnhf_prompts_range_links_colors_and_protection(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    report = tmp_path / "report.json"
    _make_source(source)

    result = finalize_workbook(source, output)
    report.write_text(json.dumps(result.to_dict(), indent=2), encoding="utf-8")

    wb = load_workbook(output)
    library = wb["Prompt_Library"]
    assert result.prompt_ids == ("P02", "P45", "P46", "P47", "P48", "P49")

    for prompt_id, expected_sheet in {
        "P02": "P02_COPY_SAFE",
        "P45": "P45_COPY_SAFE",
        "P46": "P46_COPY_SAFE",
        "P47": "P47_COPY_SAFE",
        "P48": "P48_COPY_SAFE",
        "P49": "P49_COPY_SAFE",
    }.items():
        row = result.library_rows[prompt_id]
        prompt_range = result.prompt_ranges[prompt_id]
        assert _target(library.cell(row=row, column=3)) == f"#'{expected_sheet}'!{prompt_range}"
        ws = wb[expected_sheet]
        expected_backlink = f"#'Prompt_Library'!A{row}:P{row}"
        last_row = int(prompt_range.split(":A", 1)[1])
        assert _target(ws["B1"]) == expected_backlink
        assert _target(ws["E1"]) == expected_backlink
        assert _target(ws[f"B{last_row}"]) == expected_backlink
        assert _target(ws[f"E{last_row}"]) == expected_backlink
        assert ws.protection.sheet is True
        assert ws["A1"].protection.locked is True

    footer_row = max(result.library_rows.values()) + 1
    assert _target(library["A1"]) == f"#'Prompt_Library'!A{footer_row}"
    assert _target(library["P1"]) == f"#'Prompt_Library'!P{footer_row}"
    assert _target(library[f"A{footer_row}"]) == "#'Prompt_Library'!A1"
    assert _target(library[f"P{footer_row}"]) == "#'Prompt_Library'!P1"

    cream = "00FFF2CC"
    assert library.sheet_properties.tabColor.rgb == cream
    assert wb["Opportunity_Discovery"].sheet_properties.tabColor.rgb == cream
    assert wb["P07_COPY_SAFE"].sheet_properties.tabColor.rgb == cream
    assert wb["P46_COPY_SAFE"].sheet_properties.tabColor.rgb == cream

    assert library.protection.sheet is True
    discovery = wb["Opportunity_Discovery"]
    assert discovery.protection.sheet is True
    assert discovery["A1"].protection.locked is False
    assert discovery["R100"].protection.locked is False
    assert discovery["S101"].protection.locked is True
    assert wb.security.lockStructure is True

    p02_text = "\n".join(
        wb["P02_COPY_SAFE"].cell(row=row, column=1).value or ""
        for row in range(1, int(result.prompt_ranges["P02"].split(":A", 1)[1]) + 1)
    )
    assert "Modify tracked files" in p02_text
    assert "Open or update the intended pull request" in p02_text

    p45_text = "\n".join(
        wb["P45_COPY_SAFE"].cell(row=row, column=1).value or ""
        for row in range(1, int(result.prompt_ranges["P45"].split(":A", 1)[1]) + 1)
    )
    assert "compiled-gnhf-prompt-result v1" in p45_text
    assert "DO NOT EXECUTE THE SPRINT" in p45_text

    p46_text = "\n".join(
        wb["P46_COPY_SAFE"].cell(row=row, column=1).value or ""
        for row in range(1, int(result.prompt_ranges["P46"].split(":A", 1)[1]) + 1)
    )
    assert p46_text.startswith("gnhf `")
    assert "Build or repair the repo-local AI harness" in p46_text
    assert "--max-iterations 8" in p46_text
    assert "--max-tokens 800000" in p46_text
    assert "--prevent-sleep on" in p46_text

    p47_text = "\n".join(
        wb["P47_COPY_SAFE"].cell(row=row, column=1).value or ""
        for row in range(1, int(result.prompt_ranges["P47"].split(":A", 1)[1]) + 1)
    )
    assert p47_text.startswith("gnhf `")
    assert "Execute one repo-local harness workflow" in p47_text
    assert "A successful process exit without the required artifact and commit is failure" in p47_text

    p48_text = "\n".join(
        wb["P48_COPY_SAFE"].cell(row=row, column=1).value or ""
        for row in range(1, int(result.prompt_ranges["P48"].split(":A", 1)[1]) + 1)
    )
    assert "Invoke-ChatGPTDesktopGnhfSprint.ps1" in p48_text
    assert "Do not promise unconditional failed-worktree preservation" in p48_text

    p49_text = "\n".join(
        wb["P49_COPY_SAFE"].cell(row=row, column=1).value or ""
        for row in range(1, int(result.prompt_ranges["P49"].split(":A", 1)[1]) + 1)
    )
    assert "Start-TmuxGnhfWorkspaceSetup.ps1" in p49_text
    assert "-Mode Plan" in p49_text
    assert "-Mode Apply" in p49_text


def test_finalizer_is_idempotent_for_prompt_rows(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    _make_source(source)
    first_result = finalize_workbook(source, first)
    second_result = finalize_workbook(first, second)
    assert first_result.library_rows == second_result.library_rows
    wb = load_workbook(second)
    library = wb["Prompt_Library"]
    ids = [library.cell(row=row, column=3).value for row in range(2, library.max_row)]
    for prompt_id in ("P02", "P45", "P46", "P47", "P48", "P49"):
        assert ids.count(prompt_id) == 1


def test_self_service_wrappers_remain_bounded() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    generate_ps1 = (repo_root / "scripts" / "Generate-AIPromptKitV33.ps1").read_text(encoding="utf-8")
    generate_cmd = (repo_root / "scripts" / "Generate-AIPromptKitV33.cmd").read_text(encoding="utf-8")
    finalize_ps1 = (repo_root / "scripts" / "Finalize-AIPromptKitV33.ps1").read_text(encoding="utf-8")
    finalize_cmd = (repo_root / "scripts" / "Finalize-AIPromptKitV33.cmd").read_text(encoding="utf-8")
    assert "triage.prompt_kit_v33_generator" in generate_ps1
    assert '".xlsx", ".zip"' in generate_ps1
    assert "Output must not overwrite the source workbook" in generate_ps1
    assert "pwsh -NoLogo -NoProfile" in generate_cmd
    assert "Generate-AIPromptKitV33.ps1" in generate_cmd
    assert "pause" in generate_cmd
    assert "Generate-AIPromptKitV33.ps1" in finalize_ps1
    assert "compatibility shim" in finalize_ps1
    assert "Generate-AIPromptKitV33.cmd" in finalize_cmd
