from __future__ import annotations

import hashlib
import json
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection

from tests.test_prompt_kit_operability_contract import _valid_prompt_command
from triage.prompt_kit_v33_finalizer import PALETTE
from triage.prompt_kit_v33_generator import REPORT_NAMES, generate_v33
from triage.prompt_kit_v33_prompt_contract import DEFAULT_SPEC_PATH


HEADERS = [
    "↓ Bottom",
    "Seq",
    "Prompt ID",
    "Prompt Type",
    "Prompt Class",
    "Sprint Path Role",
    "Use For Progress?",
    "Prompt Name",
    "Use This When",
    "Inspect First",
    "Expected Output",
    "Next Step",
    "Proof / Acceptance Gate",
    "Color",
    "Copy-Safe Sheet",
    "↓ Bottom",
]


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _prompt_lines(prompt_id: str) -> list[str]:
    number = int(prompt_id[1:])
    if 26 <= number <= 36:
        return _valid_prompt_command(prompt_id)
    return [f"{prompt_id} CHAT PROMPT", "Body"]


def _make_source(path: Path) -> None:
    spec = json.loads(DEFAULT_SPEC_PATH.read_text(encoding="utf-8"))
    excluded = {f"P{number:02d}_COPY_SAFE" for number in range(45, 50)}
    source_order = [name for name in spec["sheet_order"] if name not in excluded]
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in source_order:
        workbook.create_sheet(sheet_name)

    library = workbook["Prompt_Library"]
    for column, value in enumerate(HEADERS, start=1):
        library.cell(row=1, column=column).value = value
    library.freeze_panes = "C2"

    fill_color, font_color = PALETTE["Night"]
    for number in range(45):
        prompt_id = f"P{number:02d}"
        sheet_name = f"{prompt_id}_COPY_SAFE"
        row = number + 2
        values = [
            None,
            f"{number:02d}",
            prompt_id,
            "TYPE",
            "CLASS",
            "ROLE",
            "YES",
            f"Name {prompt_id}",
            f"Use {prompt_id}",
            f"Inspect {prompt_id}",
            f"Output {prompt_id}",
            "Next",
            "Gate",
            "Night",
            sheet_name,
            None,
        ]
        for column, value in enumerate(values, start=1):
            cell = library.cell(row=row, column=column)
            cell.value = value
            cell.fill = PatternFill("solid", fgColor=fill_color)
            size = 28 if column == 3 else 12 if column == 8 else 10
            bold = column in {2, 3, 4, 5, 7, 8, 14, 15}
            cell.font = Font(name="Aptos", size=size, bold=bold, color=font_color)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.protection = Protection(locked=True)
        lines = _prompt_lines(prompt_id)
        target = f"#'{sheet_name}'!A1:A{len(lines)}"
        for column in (3, 15):
            library.cell(row=row, column=column).hyperlink = target
            library.cell(row=row, column=column).font = Font(
                name="Aptos",
                size=28 if column == 3 else 10,
                bold=True,
                underline="single",
                color=font_color,
            )

        sheet = workbook[sheet_name]
        sheet.column_dimensions["A"].width = 110
        for line_number, line in enumerate(lines, start=1):
            sheet.cell(row=line_number, column=1).value = line
        last_row = len(lines)
        back = f"#'Prompt_Library'!A{row}:P{row}"
        for coordinate in ("B1", "E1", f"B{last_row}", f"E{last_row}"):
            sheet[coordinate] = "Prompt Library"
            sheet[coordinate].hyperlink = back
        sheet["C1"] = f"Copy A1:A{last_row} only"
        sheet[f"C{last_row}"] = f"Copy A1:A{last_row} only"

    footer_row = 47
    library[f"A{footer_row}"] = "↑ Top"
    library[f"P{footer_row}"] = "↑ Top"
    library["A1"].hyperlink = f"#'Prompt_Library'!A{footer_row}"
    library["P1"].hyperlink = f"#'Prompt_Library'!P{footer_row}"
    library[f"A{footer_row}"].hyperlink = "#'Prompt_Library'!A1"
    library[f"P{footer_row}"].hyperlink = "#'Prompt_Library'!P1"

    opportunity = workbook["Opportunity_Discovery"]
    opportunity["A1"] = "Editable"
    opportunity["S101"] = "Locked"
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True)
        sheet.protection.sheet = True
    for row in opportunity.iter_rows(min_row=1, max_row=100, min_col=1, max_col=18):
        for cell in row:
            cell.protection = Protection(locked=False)
    workbook.security.lockStructure = True
    workbook.save(path)


def test_generator_uses_canonical_finalizer_and_writes_reports(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _make_source(source)
    before = _sha256(source)
    output_dir = tmp_path / "out"

    result = generate_v33(source, output_dir)

    workbook = Path(result["workbook"])
    assert workbook.exists()
    assert _sha256(source) == before == result["source_sha256"]
    assert result["source_immutable"] is True
    assert result["prompt_ids"] == [f"P{number:02d}" for number in range(50)]
    assert result["protected_sheets"] == "all"
    assert result["editable_ranges"] == {"Opportunity_Discovery": "A1:R100"}
    assert all(status == "PASS" for status in result["validators"].values())
    assert (output_dir / "AI_Harness_Prompt_Kit_v33_manifest.json").exists()
    assert all((output_dir / filename).exists() for filename in REPORT_NAMES.values())
    assert not (Path(__file__).parents[1] / "triage" / "prompt_kit_v33_ooxml.py").exists()

    generated = load_workbook(workbook)
    assert generated.sheetnames[-4:-2] == ["P48_COPY_SAFE", "P49_COPY_SAFE"]
    assert generated.security.lockStructure is True
    assert generated["Opportunity_Discovery"].protection.sheet is True
    assert generated["Opportunity_Discovery"]["A1"].protection.locked is False
    assert generated["Opportunity_Discovery"]["S101"].protection.locked is True


def test_bundle_input_preserves_support_files_and_existing_outputs_are_backed_up(tmp_path: Path) -> None:
    source_workbook = tmp_path / "source.xlsx"
    _make_source(source_workbook)
    source_bundle = tmp_path / "source-bundle.zip"
    with zipfile.ZipFile(source_bundle, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.write(source_workbook, "source/AI_Harness_Prompt_Kit_v33.xlsx")
        archive.writestr("AI_Harness_Prompt_Kit_v33_Night_Shift_Quickstart.md", "quickstart")

    output_dir = tmp_path / "out"
    first = generate_v33(source_bundle, output_dir)
    first_hash = first["workbook_sha256"]
    second = generate_v33(source_bundle, output_dir)
    assert second["workbook_sha256"] == first_hash
    assert second["backup_directory"]
    assert Path(second["backup_directory"]).is_dir()
    assert second["support_files"][0]["path"].endswith("Night_Shift_Quickstart.md")

    with zipfile.ZipFile(second["bundle"]) as archive:
        names = set(archive.namelist())
        assert "AI_Harness_Prompt_Kit_v33.xlsx" in names
        assert "AI_Harness_Prompt_Kit_v33_manifest.json" in names
        assert "AI_Harness_Prompt_Kit_v33_Night_Shift_Quickstart.md" in names
        assert set(REPORT_NAMES.values()).issubset(names)


def test_generator_refuses_source_overwrite_and_unsafe_bundle_members(tmp_path: Path) -> None:
    source = tmp_path / "AI_Harness_Prompt_Kit_v33.xlsx"
    _make_source(source)
    with pytest.raises(ValueError, match="output must not overwrite"):
        generate_v33(source, tmp_path)

    unsafe = tmp_path / "unsafe.zip"
    with zipfile.ZipFile(unsafe, "w") as archive:
        archive.write(source, "source.xlsx")
        archive.writestr("../escape.md", "no")
    with pytest.raises(ValueError, match="unsafe member path"):
        generate_v33(unsafe, tmp_path / "unsafe-out")
