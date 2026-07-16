from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection

from triage.prompt_kit_v33_layout_finalizer import DEFAULT_SPEC_PATH, canonicalize_layout


def _load_spec() -> dict:
    return json.loads(DEFAULT_SPEC_PATH.read_text(encoding="utf-8"))


def _make_pre_layout_artifact(path: Path) -> None:
    spec = _load_spec()
    order = list(spec["sheet_order"])
    wb = Workbook()
    library = wb.active
    library.title = "Prompt_Library"
    library.append(["↓ Bottom", "Seq", "Prompt ID"] + [None] * 11 + ["Copy-Safe Sheet", "↓ Bottom"])

    for number in range(48):
        prompt_id = f"P{number:02d}"
        sheet_name = f"{prompt_id}_COPY_SAFE"
        row = number + 2
        values = [None, f"{number:02d}", prompt_id] + [f"value-{prompt_id}"] * 11 + [sheet_name, None]
        library.append(values)
        prompt_cell = library.cell(row=row, column=3)
        if prompt_id == "P07":
            prompt_cell.value = f'=HYPERLINK("#\'{sheet_name}\'!A1:A3","{prompt_id}")'
        else:
            prompt_cell.hyperlink = f"#'{sheet_name}'!A1:A3"
        if number >= 45:
            fill = PatternFill("solid", fgColor="FFF2CC")
            for column in range(1, 17):
                library.cell(row=row, column=column).fill = fill
            prompt_cell.font = Font(bold=True, underline="single")

    footer_row = 50
    library.append(["↑ Top", "End of Prompt Library"] + [None] * 13 + ["↑ Top"])
    library["A1"].hyperlink = f"#'Prompt_Library'!A{footer_row}"
    library["P1"].hyperlink = f"#'Prompt_Library'!P{footer_row}"
    library[f"A{footer_row}"].hyperlink = "#'Prompt_Library'!A1"
    library[f"P{footer_row}"].hyperlink = "#'Prompt_Library'!P1"

    # Deliberately wrong order. The canonicalizer owns the final sequence.
    for sheet_name in reversed([name for name in order if name != "Prompt_Library"]):
        wb.create_sheet(sheet_name)

    for number in range(48):
        prompt_id = f"P{number:02d}"
        sheet_name = f"{prompt_id}_COPY_SAFE"
        row = number + 2
        ws = wb[sheet_name]
        ws.column_dimensions["A"].width = 110
        ws["A1"] = f"{prompt_id} first line"
        ws["A2"] = ""
        ws["A3"] = "last line"
        back = f"#'Prompt_Library'!A{row}:P{row}"
        for coordinate in ("B1", "E1", "B3", "E3"):
            ws[coordinate] = "Prompt Library"
            ws[coordinate].hyperlink = back
        ws["C1"] = "Copy A1:A3 only"
        ws["C3"] = "Copy A1:A3 only"

    # Seed manual drift that must be overwritten.
    wb["START_HERE"].sheet_properties.tabColor = "FF0000"
    wb["Opportunity_Discovery"].protection.sheet = True
    wb["START_HERE"]["A1"].protection = Protection(locked=False)
    wb.save(path)


def _target(cell) -> str:
    return cell.hyperlink.target if cell.hyperlink else ""


def _assert_color(actual, expected: dict) -> None:
    if "rgb" in expected:
        assert actual.type == "rgb"
        assert str(actual.rgb).upper()[-6:] == str(expected["rgb"]).upper()[-6:]
    else:
        assert actual.type == "theme"
        assert int(actual.theme) == int(expected["theme"])
        assert abs(float(actual.tint or 0.0) - float(expected.get("tint", 0.0))) < 1e-12


def test_layout_finalizer_codifies_range_links_order_colors_and_protection(tmp_path: Path) -> None:
    artifact = tmp_path / "AI_Harness_Prompt_Kit_v33.xlsx"
    _make_pre_layout_artifact(artifact)
    result = canonicalize_layout(artifact)
    spec = _load_spec()
    wb = load_workbook(artifact, data_only=False)

    assert wb.sheetnames == spec["sheet_order"]
    assert wb.sheetnames.index("P45_COPY_SAFE") == wb.sheetnames.index("P44_COPY_SAFE") + 1
    assert result.prompt_ranges["P00"] == "A1:A3"
    assert result.prompt_ranges["P47"] == "A1:A3"

    for prompt_id, prompt_range in result.prompt_ranges.items():
        ws = wb[f"{prompt_id}_COPY_SAFE"]
        expected = f"#'{ws.title}'!{prompt_range}"
        assert ws["C1"].value == f"Copy {prompt_range} only"
        assert ws["C3"].value == f"Copy {prompt_range} only"
        assert _target(ws["C1"]) == expected
        assert _target(ws["C3"]) == expected
        assert ws["C1"].font.underline == "single"

    for sheet_name, color_spec in spec["tab_colors"].items():
        _assert_color(wb[sheet_name].sheet_properties.tabColor, color_spec)
    for ws in wb.worksheets:
        if ws.title not in spec["tab_colors"]:
            assert ws.sheet_properties.tabColor is None

    assert wb["Opportunity_Discovery"].protection.sheet is False
    for ws in wb.worksheets:
        if ws.title != "Opportunity_Discovery":
            assert ws.protection.sheet is True
            for row in ws.iter_rows():
                for cell in row:
                    assert cell.protection.locked is True


def test_layout_finalizer_is_idempotent(tmp_path: Path) -> None:
    artifact = tmp_path / "AI_Harness_Prompt_Kit_v33.xlsx"
    _make_pre_layout_artifact(artifact)
    first = canonicalize_layout(artifact)
    second = canonicalize_layout(artifact)
    assert first.prompt_ranges == second.prompt_ranges
    assert first.sheet_order == second.sheet_order
    assert first.colored_tabs == second.colored_tabs
