"""Tests for roster log review queue XML graft engine."""
from __future__ import annotations

import io
import json
import zipfile
from pathlib import Path
from unittest import mock

import openpyxl
import pytest

from tests.fixtures.roster_log_review_queue.builders import (
    build_mini_roster,
    build_roster_with_legacy_cf,
)
from triage.roster_log_review_queue.priority_allocator import (
    count_cf_groups,
    load_cf_markers,
)
from triage.roster_log_review_queue.run import run


def _marker_present(xlsx_bytes: bytes, sheet: str) -> bool:
    markers = load_cf_markers()
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as z:
        from triage.xlsx_utils import read_text, sheet_name_map

        part = [p for p, n in sheet_name_map(z).items() if n == sheet][0]
        xml = read_text(z, part)
    return any(m in xml for m in markers)


def test_blank_mode_builds_review_first_shell(tmp_path: Path) -> None:
    out = tmp_path / "blank.xlsx"
    prov = tmp_path / "blank.provenance.json"

    result = run(
        mode="blank",
        output_path=str(out),
        provenance_out=str(prov),
        months=["2026-04", "2026-05"],
    )

    assert result.preflight_pass, result.errors
    assert result.review_queue_rows == 0
    wb = openpyxl.load_workbook(out, read_only=True)
    assert wb.sheetnames[:4] == [
        "Review Dashboard",
        "Review Queue",
        "Review Rules",
        "CF Dictionary",
    ]
    assert wb.sheetnames[4:] == ["Live - April 2026", "Live - May 2026"]
    assert wb["Review Queue"]["A1"].value == "Review ID"
    assert wb["Review Rules"].max_row == 8
    assert wb["CF Dictionary"].max_row == 7
    assert wb["Live - April 2026"]["C2"].value == "Apr 01 - Clock In"
    assert wb["Live - April 2026"]["D2"].value == "Apr 01 - Clock Out"
    assert wb["Live - April 2026"].max_column == 62
    assert wb["Live - May 2026"].max_column == 64
    wb.close()

    data = json.loads(prov.read_text(encoding="utf-8"))
    assert data["mode"] == "blank"
    assert data["input_workbook"] == "<generated blank roster shell>"
    assert data["repair_safety"]["openpyxl_save_used"] is True
    assert data["verification"]["review_dashboard_first"] is True
    assert data["verification"]["all_live_tabs_patched"] is True
    assert data["verification"]["live_tabs_patched_count"] == 2


def test_blank_mode_requires_months(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="--months is required for blank mode"):
        run(mode="blank", output_path=str(tmp_path / "blank.xlsx"))


def test_live_cf_patcher_adds_markers(tmp_path: Path) -> None:
    src = build_mini_roster(tmp_path / "roster.xlsx")
    before_may = count_cf_groups(src.read_bytes(), "Live - May 2026")
    out = tmp_path / "out.xlsx"

    result = run(
        mode="live-cf-only",
        input_path=str(src),
        output_path=str(out),
    )
    assert result.preflight_pass, result.errors

    data = out.read_bytes()
    assert _marker_present(data, "Live - May 2026")
    assert _marker_present(data, "Live - April 2026")

    after_may = count_cf_groups(data, "Live - May 2026")
    assert after_may > before_may
    may_stats = result.live_cf_stats["Live - May 2026"]
    assert may_stats.new_cf_groups >= 2
    assert may_stats.new_cf_rules >= 7


def test_legacy_cf_preserved(tmp_path: Path) -> None:
    src = build_roster_with_legacy_cf(tmp_path / "roster.xlsx")
    before = count_cf_groups(src.read_bytes(), "Live - May 2026")
    out = tmp_path / "out.xlsx"
    run(mode="live-cf-only", input_path=str(src), output_path=str(out))
    after = count_cf_groups(out.read_bytes(), "Live - May 2026")
    assert after > before


def test_no_openpyxl_save_on_graft_path(tmp_path: Path) -> None:
    src = build_mini_roster(tmp_path / "roster.xlsx")
    out = tmp_path / "out.xlsx"

    with mock.patch("openpyxl.workbook.workbook.Workbook.save") as save_mock:
        run(mode="live-cf-only", input_path=str(src), output_path=str(out))
    save_mock.assert_not_called()


def test_skip_already_patched_sheet(tmp_path: Path) -> None:
    src = build_mini_roster(tmp_path / "roster.xlsx")
    mid = tmp_path / "once.xlsx"
    run(mode="live-cf-only", input_path=str(src), output_path=str(mid))
    groups_once = count_cf_groups(mid.read_bytes(), "Live - May 2026")

    out = tmp_path / "twice.xlsx"
    run(mode="live-cf-only", input_path=str(mid), output_path=str(out))
    groups_twice = count_cf_groups(out.read_bytes(), "Live - May 2026")
    assert groups_twice == groups_once


def test_provenance_shape(tmp_path: Path) -> None:
    src = build_mini_roster(tmp_path / "roster.xlsx")
    out = tmp_path / "out.xlsx"
    prov = tmp_path / "prov.json"
    run(
        mode="live-cf-only",
        input_path=str(src),
        output_path=str(out),
        provenance_out=str(prov),
    )
    data = json.loads(prov.read_text(encoding="utf-8"))
    for key in (
        "generated_at",
        "method",
        "repair_safety",
        "verification",
        "live_cf",
        "live_cf_counts_after",
    ):
        assert key in data
    assert data["repair_safety"]["openpyxl_save_used"] is False
