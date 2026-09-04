"""Roster Log V2 normalized-allocation and Web Excel regressions."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from triage.roster_log_v2.builder import build_roster_workbook
from triage.roster_log_v2.schema import normalize_state, reconcile_state
from triage.web_excel_compatibility_rules import inspect_web_excel_package


def _base_state():
    return {
        "schema_version": "roster-log-v2/v1",
        "projects": ["Northwell", "H&H"],
        "workstreams": ["Project Delivery", "Management"],
        "attendance": [
            {
                "date": "2026-09-01",
                "staff": "Operator",
                "clock_in": "08:00",
                "clock_out": "17:00",
                "paid_hours": 8,
                "default_project": "Northwell",
            }
        ],
        "allocations": [],
    }


def test_single_project_day_is_default() -> None:
    state = normalize_state(_base_state())
    assert len(state["allocations"]) == 1
    assert state["allocations"][0]["project"] == "Northwell"
    assert state["allocations"][0]["hours"] == 8
    rec = reconcile_state(state)[0]
    assert rec.mode == "SINGLE"
    assert rec.project_count == 1
    assert rec.reconciled


def test_multi_project_day_is_normal_when_hours_reconcile() -> None:
    state = _base_state()
    state["allocations"] = [
        {"allocation_id": "A1", "date": "2026-09-01", "staff": "Operator", "project": "H&H", "workstream": "Management", "hours": 6.4},
        {"allocation_id": "A2", "date": "2026-09-01", "staff": "Operator", "project": "Northwell", "workstream": "Project Delivery", "hours": 1.6},
    ]
    rec = reconcile_state(state)[0]
    assert rec.mode == "MULTI"
    assert rec.project_count == 2
    assert rec.allocated_hours == 8
    assert rec.variance == 0
    assert rec.reconciled


def test_operator_can_call_whole_day_one_project() -> None:
    state = _base_state()
    state["allocations"] = [
        {"allocation_id": "A1", "date": "2026-09-01", "staff": "Operator", "project": "Northwell", "workstream": "Project Delivery", "hours": 8}
    ]
    rec = reconcile_state(state)[0]
    assert rec.mode == "SINGLE"
    assert rec.reconciled


def test_only_variance_is_reconciliation_failure() -> None:
    state = _base_state()
    state["allocations"] = [
        {"allocation_id": "A1", "date": "2026-09-01", "staff": "Operator", "project": "Northwell", "hours": 5},
        {"allocation_id": "A2", "date": "2026-09-01", "staff": "Operator", "project": "H&H", "hours": 2},
    ]
    rec = reconcile_state(state)[0]
    assert rec.mode == "MULTI"
    assert not rec.reconciled
    assert rec.variance == 1


def test_allocation_without_attendance_is_rejected() -> None:
    state = _base_state()
    state["allocations"] = [
        {"date": "2026-09-02", "staff": "Operator", "project": "Northwell", "hours": 8}
    ]
    with pytest.raises(ValueError, match="allocation without attendance day"):
        normalize_state(state)


def test_generated_workbook_is_webexcel_safe_and_has_v2_contract(tmp_path: Path) -> None:
    state = _base_state()
    state["allocations"] = [
        {"allocation_id": "A1", "date": "2026-09-01", "staff": "Operator", "project": "H&H", "hours": 6.4},
        {"allocation_id": "A2", "date": "2026-09-01", "staff": "Operator", "project": "Northwell", "hours": 1.6},
    ]
    out = tmp_path / "Roster_Log_V2.xlsx"
    result = build_roster_workbook(state, out, require_reconciled=True)
    assert result["preflight"]["preflight_pass"]
    assert inspect_web_excel_package(out) == []

    wb = openpyxl.load_workbook(out, data_only=False)
    try:
        assert wb.sheetnames == ["Dashboard", "Attendance", "Project Allocations", "Dictionaries", "Review Queue", "Read Me"]
        assert wb["Attendance"]["G2"].value.startswith("=IF(COUNTIFS")
        assert wb["Attendance"]["H2"].value.startswith("=SUMIFS")
        assert wb["Review Queue"].max_row == 1
        readme = " ".join(str(wb["Read Me"].cell(r, 1).value or "") for r in range(1, wb["Read Me"].max_row + 1))
        assert "Multi-project days are supported" in readme
        assert "does not manufacture an 80/20 split" in readme
    finally:
        wb.close()


def test_unreconciled_workbook_routes_variance_to_review(tmp_path: Path) -> None:
    state = _base_state()
    state["allocations"] = [
        {"allocation_id": "A1", "date": "2026-09-01", "staff": "Operator", "project": "Northwell", "hours": 7}
    ]
    out = tmp_path / "draft.xlsx"
    result = build_roster_workbook(state, out)
    assert result["unreconciled_days"] == 1
    wb = openpyxl.load_workbook(out, data_only=False)
    try:
        assert wb["Review Queue"]["C2"].value == "ALLOCATION_VARIANCE"
        assert wb["Review Queue"]["F2"].value == 1
    finally:
        wb.close()


def test_local_web_app_supports_cache_add_project_and_exports() -> None:
    root = Path(__file__).resolve().parents[1]
    html = (root / "web" / "roster-log-v2" / "index.html").read_text(encoding="utf-8")
    js = (root / "web" / "roster-log-v2" / "app.js").read_text(encoding="utf-8")
    assert "Add project" in html
    assert "Use one project for whole day" in html
    assert "localStorage" in js
    assert "roster-log-v2-state-v1" in js
    assert "exportAttendance" in js
    assert "exportAllocations" in js
    assert "fetch(" not in js
