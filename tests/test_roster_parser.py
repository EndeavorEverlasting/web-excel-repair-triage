"""
tests/test_roster_parser.py
----------------------------
Unit and integration tests for triage.roster_parser.

Coverage
--------
  T1  _time_to_hours — standard datetime.time input
  T2  _time_to_hours — datetime.datetime input
  T3  _time_to_hours — Excel float fraction input
  T4  _time_to_hours — None input
  T5  _time_to_hours — annotated string "9:28:00 AM/ Bonita"  (real-file quirk)
  T6  _time_to_hours — plain AM/PM string "5:00:00 PM"
  T7  _time_to_hours — 24-h string without AM/PM "17:00:00"
  T8  _time_to_hours — HH:MM string without seconds "9:28 AM/note"
  T9  _time_to_hours — midnight AM edge case "12:00:00 AM"
  T10 _time_to_hours — noon PM edge case "12:00:00 PM"
  T11 _time_to_hours — unrecognisable string returns None
  T12 _time_to_hours — invalid int (e.g. 8) returns None (not a fraction)
  T13 parse_roster — smoke test against live April 2026 roster file
  T14 parse_roster — Patricia Marrero Apr 30 record parsed (was the bug)
  T15 parse_roster — zero malformed rows on April 2026 file

Assignments / per-day project override coverage
  A1  _load_assignments — None worksheet returns empty dict (safe no-op)
  A2  _load_assignments — main table populates lookup for worked days
  A3  _load_assignments — overrides section wins over main-table entry
  A4  _find_assignments_sheet — returns correct sheet when present
  A5  _find_assignments_sheet — returns None when absent (non-fatal)
  A6  parse_roster (synthetic) — assignments override replaces Live project
  A7  parse_roster (synthetic) — days with different projects produce
      separate records each with the correct project label
  A8  parse_roster (synthetic) — no Assignments sheet → Live project used
  A9  parse_roster (live) — Cyen Heyliger Apr 6 gets 'Projects Team'
      from the Overrides sub-table (not her default project)
  A10 parse_roster (live) — Alejandro Perales all records use
      'Neuron Deployments' as confirmed by the Assignments main table
"""
from __future__ import annotations

import datetime
import os
from pathlib import Path

import pytest

from triage.roster_parser import (
    _time_to_hours,
    _compute_gross,
    _find_assignments_sheet,
    _load_assignments,
    parse_roster,
    RosterParseError,
)

_ROSTER = Path(__file__).parent.parent / "attached_assets" / (
    "Active_Roster_Log_5_1_2026_Billing_April_Pack_(1)_1777807743057.xlsx"
)
_ROSTER_PRESENT = _ROSTER.exists()

# ── _time_to_hours unit tests ─────────────────────────────────────────────────

class TestTimeToHours:
    def test_datetime_time(self):
        assert _time_to_hours(datetime.time(9, 30)) == pytest.approx(9.5)

    def test_datetime_time_with_seconds(self):
        assert _time_to_hours(datetime.time(9, 28, 0)) == pytest.approx(9 + 28/60, rel=1e-6)

    def test_datetime_datetime(self):
        dt = datetime.datetime(2026, 4, 1, 17, 0, 0)
        assert _time_to_hours(dt) == pytest.approx(17.0)

    def test_excel_float_fraction(self):
        # 0.5 = noon in Excel's serial-time system
        assert _time_to_hours(0.5) == pytest.approx(12.0)

    def test_none_returns_none(self):
        assert _time_to_hours(None) is None

    # ── string branch (new in this fix) ──────────────────────────────────────

    def test_string_annotated_am_time(self):
        """Real-file value: '9:28:00 AM/ Bonita' — appended note must be ignored."""
        result = _time_to_hours("9:28:00 AM/ Bonita")
        assert result == pytest.approx(9 + 28/60, rel=1e-6)

    def test_string_pm_time(self):
        assert _time_to_hours("5:00:00 PM") == pytest.approx(17.0)

    def test_string_24h_no_ampm(self):
        assert _time_to_hours("17:00:00") == pytest.approx(17.0)

    def test_string_hhmm_no_seconds(self):
        result = _time_to_hours("9:28 AM/note")
        assert result == pytest.approx(9 + 28/60, rel=1e-6)

    def test_string_midnight_am(self):
        """12:xx AM should be treated as 0:xx (midnight)."""
        assert _time_to_hours("12:00:00 AM") == pytest.approx(0.0)

    def test_string_noon_pm(self):
        """12:xx PM should stay 12:xx (not add 12)."""
        assert _time_to_hours("12:00:00 PM") == pytest.approx(12.0)

    def test_string_no_time_returns_none(self):
        assert _time_to_hours("N/A") is None

    def test_string_empty_returns_none(self):
        assert _time_to_hours("") is None

    def test_large_int_returns_none(self):
        """Integers ≥2 are not Excel time fractions; should return None."""
        assert _time_to_hours(8) is None

    def test_string_dash_note(self):
        """Dash-separated annotation like '8:30 AM - note here'."""
        assert _time_to_hours("8:30 AM - note here") == pytest.approx(8.5)


class TestComputeGross:
    def test_compute_gross_same_day_shift(self):
        assert _compute_gross(9.0, 17.0) == pytest.approx(8.0)

    def test_compute_gross_overnight_shift(self):
        assert _compute_gross(23.0, 7.0) == pytest.approx(8.0)

    def test_compute_gross_midnight_exact_clock_out(self):
        assert _compute_gross(22.0, 0.0) == pytest.approx(2.0)

    def test_compute_gross_missing_clock_returns_zero(self):
        assert _compute_gross(None, 17.0) == pytest.approx(0.0)
        assert _compute_gross(9.0, None) == pytest.approx(0.0)


# ── Integration: parse_roster against the live April 2026 file ───────────────

@pytest.mark.skipif(not _ROSTER_PRESENT, reason="Live roster asset not present")
class TestParseRosterApril2026:
    @pytest.fixture(scope="class")
    def april_records(self):
        malformed: list[str] = []
        records = parse_roster(
            str(_ROSTER),
            target_month="April 2026",
            malformed_out=malformed,
        )
        return records, malformed

    def test_non_empty(self, april_records):
        records, _ = april_records
        assert len(records) > 0, "Expected at least one parsed record"

    def test_expected_staff_count(self, april_records):
        records, _ = april_records
        staffs = {r["staff"] for r in records}
        assert len(staffs) == 12

    def test_known_staff_present(self, april_records):
        records, _ = april_records
        staffs = {r["staff"] for r in records}
        for name in ("Alejandro Perales", "Geoff Gerber", "Patricia Marrero"):
            assert name in staffs, f"Expected staff member '{name}' not found"

    def test_zero_malformed_rows(self, april_records):
        """No rows should be dropped as malformed after the string-time fix."""
        _, malformed = april_records
        assert malformed == [], f"Unexpected malformed rows: {malformed}"

    def test_patricia_marrero_apr30_parsed(self, april_records):
        """Patricia Marrero's Apr 30 record was the triggering bug — verify present."""
        records, _ = april_records
        target = datetime.date(2026, 4, 30)
        pm_apr30 = [
            r for r in records
            if r["staff"] == "Patricia Marrero" and r["date"] == target
        ]
        assert len(pm_apr30) == 1, "Patricia Marrero Apr 30 record missing"
        rec = pm_apr30[0]
        assert rec["clock_in"]  == pytest.approx(9 + 28/60, rel=1e-4)
        assert rec["clock_out"] == pytest.approx(18.0)
        assert rec["gross_hours"] == pytest.approx(8.5333, rel=1e-3)
        assert rec["net_hours"]   == pytest.approx(7.5333, rel=1e-3)

    def test_record_schema(self, april_records):
        """Every record has the required keys with correct types."""
        records, _ = april_records
        required_keys = {
            "staff", "project", "date",
            "clock_in", "clock_out",
            "gross_hours", "lunch_deduction", "net_hours",
            "long_shift",
        }
        for rec in records:
            assert required_keys == set(rec.keys()), f"Unexpected record keys: {rec.keys()}"
            assert isinstance(rec["staff"],  str)
            assert isinstance(rec["date"],   datetime.date)
            assert isinstance(rec["net_hours"], float)

    def test_all_dates_in_april(self, april_records):
        records, _ = april_records
        for rec in records:
            assert rec["date"].month == 4, (
                f"Expected April date, got {rec['date']} for {rec['staff']}"
            )
            assert rec["date"].year == 2026

    def test_net_hours_positive(self, april_records):
        records, _ = april_records
        for rec in records:
            assert rec["net_hours"] >= 0.0, (
                f"Negative net hours for {rec['staff']} on {rec['date']}"
            )


# ── Synthetic workbook tests: missing sheet, missing columns, lunch deduction ──

class TestRosterParserSynthetic:
    """Tests using minimal in-memory workbooks — no live asset required."""

    def _build_wb(self, tmp_path, sheet_name: str, headers: list, rows: list):
        """Create a minimal roster workbook with the given sheet/header/row layout."""
        import openpyxl as _xl
        wb = _xl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Attendance Log — Test"])      # row 1: title
        ws.append(headers)                         # row 2: headers
        for row in rows:
            ws.append(row)                         # row 3+: data
        path = str(tmp_path / "roster.xlsx")
        wb.save(path)
        return path

    def test_missing_live_sheet_raises(self, tmp_path):
        """Workbook with no 'Live - ...' sheet must raise RosterParseError."""
        import openpyxl as _xl
        wb = _xl.Workbook()
        wb.active.title = "Summary"
        path = str(tmp_path / "no_live.xlsx")
        wb.save(path)
        with pytest.raises(RosterParseError, match="[Ll]ive"):
            parse_roster(path)

    def test_missing_staff_column_raises(self, tmp_path):
        """Sheet that has no 'Staff Name' column must raise RosterParseError."""
        path = self._build_wb(
            tmp_path,
            sheet_name="Live - April 2026",
            headers=["Apr 01 - Clock In", "Apr 01 - Clock Out"],
            rows=[[datetime.time(9, 0), datetime.time(17, 0)]],
        )
        with pytest.raises(RosterParseError, match="[Ss]taff"):
            parse_roster(path)

    def test_missing_date_columns_raises(self, tmp_path):
        """Sheet with only Staff Name column and no date columns raises RosterParseError."""
        path = self._build_wb(
            tmp_path,
            sheet_name="Live - April 2026",
            headers=["Staff Name", "Project"],
            rows=[["Test Worker", "Alpha"]],
        )
        with pytest.raises(RosterParseError, match="[Dd]ate"):
            parse_roster(path)

    def test_target_month_not_found_raises(self, tmp_path):
        """Requesting a month that has no matching Live sheet raises RosterParseError."""
        path = self._build_wb(
            tmp_path,
            sheet_name="Live - April 2026",
            headers=["Staff Name", "Project", "Apr 01 - Clock In", "Apr 01 - Clock Out"],
            rows=[["Test Worker", "Alpha", datetime.time(9, 0), datetime.time(17, 0)]],
        )
        with pytest.raises(RosterParseError, match="[Ll]ive|[Mm]atching"):
            parse_roster(path, target_month="March 2026")

    def test_happy_path_parses_record(self, tmp_path):
        """Valid workbook with one staff member produces one record."""
        path = self._build_wb(
            tmp_path,
            sheet_name="Live - April 2026",
            headers=["Staff Name", "Project", "Apr 01 - Clock In", "Apr 01 - Clock Out"],
            rows=[["Jane Doe", "Alpha", datetime.time(9, 0), datetime.time(17, 0)]],
        )
        records = parse_roster(path, target_month="April 2026")
        assert len(records) == 1
        rec = records[0]
        assert rec["staff"] == "Jane Doe"
        assert rec["project"] == "Alpha"
        assert rec["gross_hours"] == pytest.approx(8.0)

    def test_parse_roster_collects_overnight_records(self, tmp_path):
        """Overnight shifts remain normal records and are copied to overnight_out."""
        path = self._build_wb(
            tmp_path,
            sheet_name="Live - April 2026",
            headers=["Staff Name", "Project", "Apr 01 - Clock In", "Apr 01 - Clock Out"],
            rows=[["Night Worker", "Ops", datetime.time(23, 0), datetime.time(7, 0)]],
        )
        overnight: list[dict] = []
        records = parse_roster(
            path,
            target_month="April 2026",
            overnight_out=overnight,
        )

        assert len(records) == 1
        rec = records[0]
        assert rec["gross_hours"] == pytest.approx(8.0)
        assert rec["long_shift"] is False

        assert len(overnight) == 1
        assert overnight[0]["staff"] == "Night Worker"
        assert overnight[0]["overnight"] is True
        assert overnight[0]["long_shift"] is False

    def test_parse_roster_flags_long_overnight_shift(self, tmp_path):
        """Overnight shifts above the threshold are flagged for extra review."""
        path = self._build_wb(
            tmp_path,
            sheet_name="Live - April 2026",
            headers=["Staff Name", "Project", "Apr 01 - Clock In", "Apr 01 - Clock Out"],
            rows=[["Long Night Worker", "Ops", datetime.time(18, 0), datetime.time(7, 30)]],
        )
        overnight: list[dict] = []
        records = parse_roster(
            path,
            target_month="April 2026",
            overnight_out=overnight,
            long_shift_threshold_hours=12.0,
        )

        assert len(records) == 1
        assert records[0]["gross_hours"] == pytest.approx(13.5)
        assert records[0]["long_shift"] is True
        assert overnight[0]["long_shift"] is True

    def test_parse_roster_long_shift_threshold_is_configurable(self, tmp_path):
        """The long-shift threshold can be raised for special cases."""
        path = self._build_wb(
            tmp_path,
            sheet_name="Live - April 2026",
            headers=["Staff Name", "Project", "Apr 01 - Clock In", "Apr 01 - Clock Out"],
            rows=[["Long But Allowed", "Ops", datetime.time(18, 0), datetime.time(7, 30)]],
        )
        records = parse_roster(path, target_month="April 2026", long_shift_threshold_hours=14.0)
        assert records[0]["gross_hours"] == pytest.approx(13.5)
        assert records[0]["long_shift"] is False

    def test_lunch_deduction_under_six_hours(self, tmp_path):
        """Shifts under 6h receive no lunch deduction."""
        path = self._build_wb(
            tmp_path,
            sheet_name="Live - April 2026",
            headers=["Staff Name", "Project", "Apr 01 - Clock In", "Apr 01 - Clock Out"],
            rows=[["Short Worker", "Beta", datetime.time(9, 0), datetime.time(13, 0)]],
        )
        records = parse_roster(path, target_month="April 2026")
        assert len(records) == 1
        assert records[0]["lunch_deduction"] == pytest.approx(0.0)
        assert records[0]["net_hours"] == pytest.approx(4.0)

    def test_lunch_deduction_six_to_eight_hours(self, tmp_path):
        """Shifts between 6h and <8h receive 0.5h lunch deduction."""
        path = self._build_wb(
            tmp_path,
            sheet_name="Live - April 2026",
            headers=["Staff Name", "Project", "Apr 01 - Clock In", "Apr 01 - Clock Out"],
            rows=[["Mid Worker", "Beta", datetime.time(9, 0), datetime.time(15, 0)]],
        )
        records = parse_roster(path, target_month="April 2026")
        assert len(records) == 1
        assert records[0]["lunch_deduction"] == pytest.approx(0.5)
        assert records[0]["net_hours"] == pytest.approx(5.5)

    def test_lunch_deduction_eight_or_more_hours(self, tmp_path):
        """Shifts of 8h+ receive 1.0h lunch deduction."""
        path = self._build_wb(
            tmp_path,
            sheet_name="Live - April 2026",
            headers=["Staff Name", "Project", "Apr 01 - Clock In", "Apr 01 - Clock Out"],
            rows=[["Full Worker", "Alpha", datetime.time(9, 0), datetime.time(17, 0)]],
        )
        records = parse_roster(path, target_month="April 2026")
        assert len(records) == 1
        assert records[0]["lunch_deduction"] == pytest.approx(1.0)
        assert records[0]["net_hours"] == pytest.approx(7.0)

    def test_both_clocks_missing_skips_row(self, tmp_path):
        """A row with no clock-in and no clock-out is silently skipped."""
        path = self._build_wb(
            tmp_path,
            sheet_name="Live - April 2026",
            headers=["Staff Name", "Project", "Apr 01 - Clock In", "Apr 01 - Clock Out"],
            rows=[["Ghost Worker", "Alpha", None, None]],
        )
        records = parse_roster(path, target_month="April 2026")
        assert records == []

    def test_one_clock_missing_strict_raises(self, tmp_path):
        """One clock value missing in strict mode (no malformed_out) raises."""
        path = self._build_wb(
            tmp_path,
            sheet_name="Live - April 2026",
            headers=["Staff Name", "Project", "Apr 01 - Clock In", "Apr 01 - Clock Out"],
            rows=[["Half Worker", "Alpha", datetime.time(9, 0), None]],
        )
        with pytest.raises(RosterParseError, match="[Mm]alformed|[Bb]lank|[Cc]lock"):
            parse_roster(path, target_month="April 2026")

    def test_one_clock_missing_collect_mode_warns(self, tmp_path):
        """One clock value missing in collect mode appends warning and skips row."""
        path = self._build_wb(
            tmp_path,
            sheet_name="Live - April 2026",
            headers=["Staff Name", "Project", "Apr 01 - Clock In", "Apr 01 - Clock Out"],
            rows=[["Half Worker", "Alpha", datetime.time(9, 0), None]],
        )
        malformed = []
        records = parse_roster(path, target_month="April 2026", malformed_out=malformed)
        assert records == []
        assert len(malformed) == 1
        assert "Half Worker" in malformed[0]


# ── Assignments / per-day project override tests ──────────────────────────────

class TestLoadAssignments:
    """A1–A3: unit tests for _load_assignments and _find_assignments_sheet."""

    def test_none_worksheet_returns_empty_dict(self):
        """A1: _load_assignments(None) must return an empty dict without error."""
        result = _load_assignments(None)
        assert result == {}

    def _build_assignments_wb(self, tmp_path, main_rows, override_rows=None):
        """
        Build a minimal workbook with:
          Live - April 2026   (one clock pair so parse_roster works)
          Assignments - April 2026  (main table + optional overrides sub-table)
        """
        import openpyxl as _xl
        wb = _xl.Workbook()

        # Live sheet
        live_ws = wb.active
        live_ws.title = "Live - April 2026"
        live_ws.append(["Attendance Log"])
        live_ws.append(["Staff Name", "Project", "Apr 01 - Clock In", "Apr 01 - Clock Out",
                         "Apr 02 - Clock In", "Apr 02 - Clock Out"])
        for r in main_rows:
            live_ws.append([r[0], r[1],
                             datetime.time(9, 0), datetime.time(17, 0),
                             datetime.time(9, 0), datetime.time(17, 0)])

        # Assignments sheet
        asn_ws = wb.create_sheet("Assignments - April 2026")
        asn_ws.append(["April 2026 - Project Assignments"])
        asn_ws.append(["Staff Name", "Default Project",
                        datetime.datetime(2026, 4, 1),
                        datetime.datetime(2026, 4, 2)])
        for r in main_rows:
            asn_ws.append([r[0], r[1], r[2], r[3]])

        if override_rows:
            asn_ws.append(["Overrides (only if different from Default Project)"])
            asn_ws.append(["Override Staff Name", "Override Date", "Override Project", "Notes"])
            for ov in override_rows:
                asn_ws.append(ov)

        path = str(tmp_path / "roster.xlsx")
        wb.save(path)
        return path

    def test_main_table_populates_lookup(self, tmp_path):
        """A2: per-day project values in the main table are loaded into the lookup."""
        path = self._build_assignments_wb(
            tmp_path,
            main_rows=[["Jane Doe", "Alpha", "Alpha", "Alpha"]],
        )
        import openpyxl as _xl
        wb = _xl.load_workbook(path, data_only=True)
        asn_ws = wb["Assignments - April 2026"]
        lookup = _load_assignments(asn_ws)
        assert lookup.get((datetime.date(2026, 4, 1), "Jane Doe")) == "Alpha"
        assert lookup.get((datetime.date(2026, 4, 2), "Jane Doe")) == "Alpha"

    def test_overrides_win_over_main_table(self, tmp_path):
        """A3: an Overrides sub-table entry overrides the main-table project for that day."""
        path = self._build_assignments_wb(
            tmp_path,
            main_rows=[["Jane Doe", "Alpha", "Alpha", "Alpha"]],
            override_rows=[
                ["Jane Doe", datetime.datetime(2026, 4, 2), "Beta", "switched"],
            ],
        )
        import openpyxl as _xl
        wb = _xl.load_workbook(path, data_only=True)
        asn_ws = wb["Assignments - April 2026"]
        lookup = _load_assignments(asn_ws)
        assert lookup.get((datetime.date(2026, 4, 1), "Jane Doe")) == "Alpha"
        assert lookup.get((datetime.date(2026, 4, 2), "Jane Doe")) == "Beta"


class TestFindAssignmentsSheet:
    """A4–A5: unit tests for _find_assignments_sheet."""

    def _make_wb_with_sheet(self, tmp_path, sheet_name):
        import openpyxl as _xl
        wb = _xl.Workbook()
        wb.active.title = sheet_name
        path = str(tmp_path / "wb.xlsx")
        wb.save(path)
        return path

    def test_finds_matching_sheet(self, tmp_path):
        """A4: _find_assignments_sheet returns the correct sheet when present."""
        import openpyxl as _xl
        wb = _xl.Workbook()
        wb.active.title = "Summary"
        wb.create_sheet("Assignments - April 2026")
        path = str(tmp_path / "wb.xlsx")
        wb.save(path)
        wb2 = _xl.load_workbook(path)
        ws = _find_assignments_sheet(wb2, "April 2026")
        assert ws is not None
        assert ws.title == "Assignments - April 2026"

    def test_returns_none_when_absent(self, tmp_path):
        """A5: _find_assignments_sheet returns None (no error) when sheet is missing."""
        import openpyxl as _xl
        wb = _xl.Workbook()
        wb.active.title = "Live - April 2026"
        path = str(tmp_path / "wb.xlsx")
        wb.save(path)
        wb2 = _xl.load_workbook(path)
        ws = _find_assignments_sheet(wb2, "April 2026")
        assert ws is None


class TestAssignmentsIntegrationSynthetic:
    """A6–A8: integration tests with synthetic workbooks."""

    def _build_wb(self, tmp_path, live_rows, asn_rows=None, override_rows=None):
        """
        Build a workbook with a Live sheet and optional Assignments sheet.
        live_rows: list of [name, live_project, in_apr1, out_apr1, in_apr2, out_apr2]
        asn_rows:  list of [name, default_project, asn_project_apr1, asn_project_apr2]
        """
        import openpyxl as _xl
        wb = _xl.Workbook()

        live_ws = wb.active
        live_ws.title = "Live - April 2026"
        live_ws.append(["Attendance Log"])
        live_ws.append(["Staff Name", "Project",
                         "Apr 01 - Clock In", "Apr 01 - Clock Out",
                         "Apr 02 - Clock In", "Apr 02 - Clock Out"])
        for r in live_rows:
            live_ws.append(r)

        if asn_rows is not None:
            asn_ws = wb.create_sheet("Assignments - April 2026")
            asn_ws.append(["April 2026 - Project Assignments"])
            asn_ws.append(["Staff Name", "Default Project",
                            datetime.datetime(2026, 4, 1),
                            datetime.datetime(2026, 4, 2)])
            for r in asn_rows:
                asn_ws.append(r)
            if override_rows:
                asn_ws.append(["Overrides (only if different from Default Project)"])
                asn_ws.append(["Override Staff Name", "Override Date",
                                "Override Project", "Notes"])
                for ov in override_rows:
                    asn_ws.append(ov)

        path = str(tmp_path / "roster.xlsx")
        wb.save(path)
        return path

    def test_assignments_project_overrides_live_project(self, tmp_path):
        """A6: per-day project from Assignments sheet replaces the Live sheet project."""
        path = self._build_wb(
            tmp_path,
            live_rows=[["Jane Doe", "OldProject",
                         datetime.time(9, 0), datetime.time(17, 0),
                         datetime.time(9, 0), datetime.time(17, 0)]],
            asn_rows=[["Jane Doe", "NewProject", "NewProject", "NewProject"]],
        )
        records = parse_roster(path, target_month="April 2026")
        assert all(r["project"] == "NewProject" for r in records), (
            f"Expected 'NewProject' for all records, got: {[r['project'] for r in records]}"
        )

    def test_different_projects_per_day_produce_separate_records(self, tmp_path):
        """A7: days with different Assignments projects each get the correct label."""
        path = self._build_wb(
            tmp_path,
            live_rows=[["Sam Smith", "DefaultProject",
                         datetime.time(9, 0), datetime.time(17, 0),
                         datetime.time(9, 0), datetime.time(17, 0)]],
            asn_rows=[["Sam Smith", "DefaultProject", "ProjectA", "ProjectB"]],
        )
        records = parse_roster(path, target_month="April 2026")
        assert len(records) == 2
        by_date = {r["date"]: r["project"] for r in records}
        assert by_date[datetime.date(2026, 4, 1)] == "ProjectA"
        assert by_date[datetime.date(2026, 4, 2)] == "ProjectB"

    def test_no_assignments_sheet_falls_back_to_live_project(self, tmp_path):
        """A8: when the Assignments sheet is absent the Live project column is used."""
        path = self._build_wb(
            tmp_path,
            live_rows=[["Jane Doe", "FallbackProject",
                         datetime.time(9, 0), datetime.time(17, 0),
                         None, None]],
            asn_rows=None,
        )
        records = parse_roster(path, target_month="April 2026")
        assert len(records) == 1
        assert records[0]["project"] == "FallbackProject"


@pytest.mark.skipif(not _ROSTER_PRESENT, reason="Live roster asset not present")
class TestAssignmentsIntegrationLive:
    """A9–A10: integration tests against the real April 2026 workbook."""

    @pytest.fixture(scope="class")
    def april_records(self):
        malformed: list[str] = []
        records = parse_roster(
            str(_ROSTER),
            target_month="April 2026",
            malformed_out=malformed,
        )
        return records

    def test_cyen_heyliger_apr6_is_projects_team(self, april_records):
        """A9: Cyen Heyliger's Apr 6 record uses 'Projects Team' from the Overrides table."""
        target = datetime.date(2026, 4, 6)
        matches = [
            r for r in april_records
            if r["staff"] == "Cyen Heyliger" and r["date"] == target
        ]
        assert len(matches) == 1, (
            f"Expected exactly one record for Cyen Heyliger on Apr 6, got {len(matches)}"
        )
        assert matches[0]["project"] == "Projects Team", (
            f"Expected 'Projects Team', got '{matches[0]['project']}'"
        )

    def test_cyen_heyliger_apr2_is_bonita(self, april_records):
        """A9b: Cyen Heyliger's Apr 2 record uses 'Bonita' from the Assignments main table."""
        target = datetime.date(2026, 4, 2)
        matches = [
            r for r in april_records
            if r["staff"] == "Cyen Heyliger" and r["date"] == target
        ]
        assert len(matches) == 1
        assert matches[0]["project"] == "Bonita", (
            f"Expected 'Bonita', got '{matches[0]['project']}'"
        )

    def test_alejandro_perales_all_records_neuron_deployments(self, april_records):
        """A10: Alejandro Perales's Assignments entries are all 'Neuron Deployments'."""
        alejandro = [r for r in april_records if r["staff"] == "Alejandro Perales"]
        assert len(alejandro) > 0, "No records found for Alejandro Perales"
        for rec in alejandro:
            assert rec["project"] == "Neuron Deployments", (
                f"Expected 'Neuron Deployments' on {rec['date']}, "
                f"got '{rec['project']}'"
            )
