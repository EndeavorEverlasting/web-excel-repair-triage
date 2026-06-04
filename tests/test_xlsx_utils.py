from __future__ import annotations

import io
import zipfile

import openpyxl

from triage.xlsx_utils import fix_inlinestr, sheet_name_map, sheet_index_map


def _make_xlsx_with_absolute_rels_targets() -> bytes:
    """Create a tiny .xlsx where workbook.xml.rels uses absolute Targets (leading '/').

    Real-world example: Deprecated/readme/README_*.xlsx uses Target="/xl/worksheets/sheet1.xml".
    """
    workbook_xml = b"""<?xml version=\"1.0\"?>
<workbook xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\">
  <sheets>
    <sheet xmlns:r=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships\" name=\"README\" sheetId=\"1\" r:id=\"rId1\"/>
    <sheet xmlns:r=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships\" name=\"ERROR_LOG\" sheetId=\"2\" r:id=\"rId2\"/>
  </sheets>
</workbook>"""

    # Note: Target appears before Id, and Target is absolute (leading slash).
    rels_xml = b"""<?xml version=\"1.0\"?>
<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">
  <Relationship Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet\" Target=\"/xl/worksheets/sheet1.xml\" Id=\"rId1\"/>
  <Relationship Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet\" Target=\"/xl/worksheets/sheet2.xml\" Id=\"rId2\"/>
</Relationships>"""

    sheet1 = b"<?xml version=\"1.0\"?><worksheet xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\"><sheetData/></worksheet>"
    sheet2 = b"<?xml version=\"1.0\"?><worksheet xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\"><sheetData/></worksheet>"

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("xl/workbook.xml", workbook_xml)
        z.writestr("xl/_rels/workbook.xml.rels", rels_xml)
        z.writestr("xl/worksheets/sheet1.xml", sheet1)
        z.writestr("xl/worksheets/sheet2.xml", sheet2)
    return buf.getvalue()


def test_sheet_name_map_handles_absolute_rels_targets():
    data = _make_xlsx_with_absolute_rels_targets()
    with zipfile.ZipFile(io.BytesIO(data), "r") as z:
        m = sheet_name_map(z)
    assert m["xl/worksheets/sheet1.xml"] == "README"
    assert m["xl/worksheets/sheet2.xml"] == "ERROR_LOG"


def test_sheet_index_map_handles_absolute_rels_targets():
    data = _make_xlsx_with_absolute_rels_targets()
    with zipfile.ZipFile(io.BytesIO(data), "r") as z:
        m = sheet_index_map(z)
    assert m["xl/worksheets/sheet1.xml"] == 0
    assert m["xl/worksheets/sheet2.xml"] == 1


def test_fix_inlinestr_no_double_escape(tmp_path):
    """Re-running fix_inlinestr must not double-escape XML entities.

    A value containing ``&``, ``<`` and ``>`` is stored as escaped entities in
    sharedStrings.xml. When fix_inlinestr re-reads that table it must unescape
    before re-indexing, otherwise a second pass turns ``&amp;`` into
    ``&amp;amp;`` and the round-tripped cell value drifts from the original.
    """
    original = "R&D <Imaging> \"baseline\" & sign-off"
    p = tmp_path / "entities.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = original
    wb.save(str(p))

    fix_inlinestr(str(p))
    fix_inlinestr(str(p))  # second pass must be idempotent

    wb2 = openpyxl.load_workbook(str(p))
    value = wb2["Sheet1"]["A1"].value
    wb2.close()
    assert value == original
