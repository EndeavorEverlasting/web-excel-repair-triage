import openpyxl
from openpyxl.utils import get_column_letter

class TrackerTooling:
    def __init__(self, template_path=None):
        self.template_path = template_path

    def correct_dupdeployed_formula(self, file_path, output_path=None):
        """
        Corrects the DupDeployed formula to ensure that a duplicate is only flagged
        if the same identifier is Deployed="YES" in a DIFFERENT location.
        Locations are derived from Current Building (G), Install Building (H), Area/Unit/Dept (I), Room (J), and Bay (K).
        Identifiers checked: V, W, X, Y, AB, AH, AP, AQ.
        """
        if output_path is None:
            output_path = file_path

        wb = openpyxl.load_workbook(file_path)
        if 'Deployments' not in wb.sheetnames:
            print("Error: 'Deployments' sheet not found.")
            return False

        ws = wb['Deployments']
        max_row = ws.max_row

        # Columns mapped
        # B = Deployed
        # F = DupDeployed
        # G = Current Building, H = Install Building, I = Area/Unit/Dept, J = Room, K = Bay
        # ID columns: V, W, X, Y, AB, AH, AP, AQ

        id_cols = ['V', 'W', 'X', 'Y', 'AB', 'AH', 'AP', 'AQ']

        for row in range(2, max_row + 1):
            # Construct the complex OR condition
            conditions = []
            for c in id_cols:
                # If MAC address, strip colons and dashes
                if c in ['Y', 'AH']:
                    val_expr = f'UPPER(SUBSTITUTE(SUBSTITUTE(TRIM(${c}{row}),":",""),"-",""))'
                    range_expr = f'UPPER(SUBSTITUTE(SUBSTITUTE(TRIM(${c}$2:${c}${max_row}),":",""),"-",""))'
                else:
                    val_expr = f'UPPER(TRIM(${c}{row}))'
                    range_expr = f'UPPER(TRIM(${c}$2:${c}${max_row}))'

                cond = (
                    f'AND(${c}{row}<>"", LOWER(TRIM(${c}{row}))<>"n/a", LOWER(TRIM(${c}{row}))<>"na", '
                    f'SUMPRODUCT(--({range_expr}={val_expr}), '
                    f'--(UPPER(TRIM($B$2:$B${max_row}))="YES"), '
                    f'--(UPPER(TRIM($G$2:$G${max_row})&TRIM($H$2:$H${max_row})&TRIM($I$2:$I${max_row})&TRIM($J$2:$J${max_row})&TRIM($K$2:$K${max_row}))<>UPPER(TRIM($G{row})&TRIM($H{row})&TRIM($I{row})&TRIM($J{row})&TRIM($K{row}))))>0)'
                )
                conditions.append(cond)

            or_statement = "OR(" + ", ".join(conditions) + ")"

            formula = f'=IF(UPPER(TRIM($B{row}))<>"YES","", IF({or_statement}, "Yes", "No"))'
            ws[f'F{row}'] = formula

        wb.save(output_path)
        print(f"Successfully updated DupDeployed formula and saved to {output_path}")
        return True

    def validate_peripherals_sites(self, file_path, output_path=None):
        """
        Validates that 'PERIPHERALS' are only deployed in allowed sites:
        valley stream, forest hills, syosset, plainview.
        Creates/updates a 'PeripheralsAllowedSite' column.
        """
        import re
        if output_path is None:
            output_path = file_path

        wb = openpyxl.load_workbook(file_path)
        if 'Deployments' not in wb.sheetnames:
            print("Error: 'Deployments' sheet not found.")
            return False

        ws = wb['Deployments']
        max_row = ws.max_row

        # Find where to put the new validation flag (append to the end if not exists)
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        if 'PeripheralsAllowedSite' not in headers:
            col_idx = ws.max_column + 1
            ws.cell(1, col_idx, 'PeripheralsAllowedSite')
        else:
            col_idx = headers['PeripheralsAllowedSite']

        allowed_patterns = [
            re.compile(r'valley\s+stream', re.IGNORECASE),
            re.compile(r'forest\s+hills', re.IGNORECASE),
            re.compile(r'syosset', re.IGNORECASE),
            re.compile(r'plainview', re.IGNORECASE)
        ]

        for row in range(2, max_row + 1):
            device_type = str(ws.cell(row, 1).value or '').strip().upper()
            if device_type == 'PERIPHERALS':
                # Check site columns G, H, I (7, 8, 9)
                site_parts = [
                    str(ws.cell(row, 7).value or '').strip(),
                    str(ws.cell(row, 8).value or '').strip(),
                    str(ws.cell(row, 9).value or '').strip()
                ]
                blob = ' '.join(filter(None, site_parts))

                is_allowed = False
                if blob:
                    for pat in allowed_patterns:
                        if pat.search(blob):
                            is_allowed = True
                            break

                ws.cell(row, col_idx, is_allowed)
            else:
                ws.cell(row, col_idx, None)

        wb.save(output_path)
        print(f"Successfully validated Peripherals sites and saved to {output_path}")
        return True

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Tracker Tooling")
    parser.add_argument("--fix-duplicates", dest="fix_dup", help="Path to tracker to fix DupDeployed formula")
    parser.add_argument("--validate-sites", dest="val_sites", help="Path to tracker to validate PeripheralsAllowedSite")
    parser.add_argument("--out", help="Output path")
    args = parser.parse_args()

    tool = TrackerTooling()
    if args.fix_dup:
        tool.correct_dupdeployed_formula(args.fix_dup, args.out)
    if args.val_sites:
        tool.validate_peripherals_sites(args.val_sites, args.out)
