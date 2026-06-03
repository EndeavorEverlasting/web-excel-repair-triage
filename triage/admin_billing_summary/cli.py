"""CLI for OpenAI-format admin billing summaries (Internal + Client per month).

    python -m triage.admin_billing_summary.cli \\
        --roster-log <roster.xlsx> --months 2026-04 2026-05 \\
        --out-dir Outputs/admin_billing_summary_2026_06_02 \\
        --prior "<April copy>.xlsx" --websafe
"""
from __future__ import annotations

import argparse
import csv
import datetime as _dt
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from triage.admin_billing_summary.aggregator import build_month_summary
from triage.admin_billing_summary.exporter import build_workbook
from triage.admin_billing_summary.models import MonthSummary
from triage.admin_billing_summary.preflight import preflight_billing_summary
from triage.nw_prj_neuron_track_hours.bonita_exporter import tab_name_for_month_key
from triage.artifact_compare import compare_artifacts
from triage.sidecar_html.adapters import admin_billing_sections
from triage.sidecar_html.portal import build_run_portal

DEFAULT_MONTHS = ["2026-04", "2026-05"]


def _resolve(p: Optional[str], base: Path) -> Optional[Path]:
    if not p:
        return None
    pp = Path(p)
    return pp if pp.is_absolute() else (base / pp).resolve()


def _month_stem(month_key: str) -> str:
    y, m = month_key.split("-")
    from calendar import month_name
    return f"{month_name[int(m)]}_{y}"


def _workbook_name(month_key: str, variant: str) -> str:
    return f"{_month_stem(month_key)}_Billing_Summary_{variant.title()}.xlsx"


def _read_prior_project_net(prior_path: Path) -> Dict[str, float]:
    import openpyxl
    out: Dict[str, float] = {}
    try:
        wb = openpyxl.load_workbook(str(prior_path), data_only=True, read_only=True)
    except Exception as exc:
        raise RuntimeError(
            f"Cannot read prior workbook for delta: {prior_path} — {exc}"
        ) from exc
    try:
        if "Project Summary" not in wb.sheetnames:
            return out
        ws = wb["Project Summary"]
        rows = list(ws.iter_rows(values_only=True))
        hdr_idx = None
        for i, row in enumerate(rows):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if "project" in cells and "net hours" in cells:
                hdr_idx = i
                break
            if row and str(row[0]).strip().lower() == "project":
                hdr_idx = i
                break
        if hdr_idx is None:
            return out
        header = [str(c).strip() if c is not None else "" for c in rows[hdr_idx]]
        try:
            proj_col = header.index("Project")
            net_col = header.index("Net Hours")
        except ValueError:
            net_col = len(header) - 1
            proj_col = 0
        for row in rows[hdr_idx + 1:]:
            if not row or row[proj_col] is None:
                continue
            proj = str(row[proj_col]).strip()
            val = row[net_col] if net_col < len(row) else None
            if isinstance(val, (int, float)):
                out[proj] = round(float(val), 2)
    finally:
        wb.close()
    return out


def _delta(summary: MonthSummary, prior_path: Path) -> Dict[str, Any]:
    prior = _read_prior_project_net(prior_path)
    current = {r.project: round(r.net_hours, 2) for r in summary.project_rows}
    keys = sorted(set(prior) | set(current))
    rows = []
    for k in keys:
        pv = prior.get(k)
        cv = current.get(k)
        delta = round((cv or 0.0) - (pv or 0.0), 2)
        rows.append({"Project": k, "Prior Net": pv, "Current Net": cv, "Delta": delta})
    return {
        "prior_file": str(prior_path),
        "prior_total_net": round(sum(prior.values()), 2),
        "current_total_net": summary.total_net,
        "total_net_delta": round(summary.total_net - sum(prior.values()), 2),
        "by_project": rows,
    }


def _write_review_csv(path: Path, summary: MonthSummary) -> None:
    cols = ["Category", "Date", "Day", "Tech", "Project", "Project Source",
            "Clock In", "Clock Out", "Gross Span", "Net Hours", "Note"]
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in summary.records:
            cat = None
            if r.long_shift:
                cat = "long_shift"
            elif r.project_source == "override":
                cat = "override_applied"
            elif r.project == "Unassigned / Review":
                cat = "unassigned"
            if not cat:
                continue
            row = {
                "Category": cat, "Date": r.date.isoformat(), "Day": r.day, "Tech": r.tech,
                "Project": r.project, "Project Source": r.project_source,
                "Clock In": r.clock_in, "Clock Out": r.clock_out,
                "Gross Span": round(r.gross_span, 2), "Net Hours": round(r.net_hours, 2),
                "Note": r.note,
            }
            for k, v in list(row.items()):
                if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
                    row[k] = "'" + v
            w.writerow(row)
        for m in summary.malformed:
            w.writerow({"Category": "malformed", "Note": m})


def run(
    roster_log: str,
    out_dir: str,
    months: Optional[List[str]] = None,
    prior: Optional[str] = None,
    websafe: bool = True,
    repo_root: Optional[Path] = None,
    reference: Optional[str] = None,
    artifact_profile: str = "admin_billing_summary",
    approved_delta: Optional[str] = None,
) -> Dict[str, Any]:
    root = repo_root or Path(__file__).resolve().parent.parent.parent
    months = months or DEFAULT_MONTHS
    roster_path = _resolve(roster_log, root)
    if roster_path is None or not roster_path.exists():
        raise FileNotFoundError(f"roster-log not found: {roster_path}")
    out = _resolve(out_dir, root) or (root / "Outputs")
    out.mkdir(parents=True, exist_ok=True)
    prior_path = _resolve(prior, root)
    reference_path = _resolve(reference, root)
    delta_path = _resolve(approved_delta, root)
    generated_utc = _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%d %H:%M")

    months_out: Dict[str, Any] = {}
    for mk in months:
        summary = build_month_summary(str(roster_path), mk)
        bonita_tab = tab_name_for_month_key(mk)
        outputs: Dict[str, Any] = {}

        for variant in ("internal", "client"):
            xlsx_path = out / _workbook_name(mk, variant)
            build_workbook(
                summary,
                str(xlsx_path),
                variant=variant,
                roster_name=roster_path.name,
                generated_utc=generated_utc,
            )
            preflight = (
                preflight_billing_summary(
                    str(xlsx_path),
                    variant=variant,
                    expect_neuron_tab=bonita_tab,
                )
                if websafe
                else {}
            )
            preflight_path = out / f"{xlsx_path.stem}_preflight.json"
            preflight_path.write_text(
                json.dumps(preflight, indent=2, default=str), encoding="utf-8"
            )
            variant_out: Dict[str, Any] = {
                "workbook": str(xlsx_path),
                "preflight_json": str(preflight_path),
                "websafe_preflight_pass": bool(preflight.get("preflight_pass")) if websafe else None,
                "native_table_count": preflight.get("native_table_count"),
            }
            if (
                reference_path
                and reference_path.is_file()
                and variant == "client"
            ):
                cmp_path = out / f"{xlsx_path.stem}_artifact_compare.json"
                cmp_report = compare_artifacts(
                    str(reference_path),
                    str(xlsx_path),
                    artifact_profile,
                    approved_delta=str(delta_path) if delta_path else None,
                    expect_neuron_tab=bonita_tab,
                    variant=variant,
                )
                cmp_path.write_text(
                    json.dumps(cmp_report, indent=2, default=str), encoding="utf-8"
                )
                variant_out["artifact_compare_json"] = str(cmp_path)
                variant_out["artifact_compare_pass"] = bool(cmp_report.get("compare_pass"))
            outputs[variant] = variant_out

        review_path = out / f"{_month_stem(mk)}_Billing_Summary_review_queue.csv"
        _write_review_csv(review_path, summary)

        delta = None
        if prior_path and prior_path.exists() and mk.endswith("-04"):
            delta = _delta(summary, prior_path)
            (out / f"{_month_stem(mk)}_Billing_Summary_Internal_delta.json").write_text(
                json.dumps(delta, indent=2, default=str), encoding="utf-8"
            )

        neuron_net_detail = round(sum(r.net_hours for r in summary.neuron_records()), 2)
        months_out[mk] = {
            "month_name": summary.month_name,
            "bonita_tab": bonita_tab,
            "review_queue_csv": str(review_path),
            "row_count": len(summary.records),
            "total_net": summary.total_net,
            "neuron_net": summary.net_for_bucket("Neurons"),
            "neuron_detail_net": neuron_net_detail,
            "outputs": outputs,
            "warnings": summary.warnings,
            "malformed_count": len(summary.malformed),
            "delta_vs_prior": delta,
        }

    manifest = {
        "engine": "triage.admin_billing_summary.cli",
        "format": "openai_native_tables_v1",
        "generated_utc": _dt.datetime.now(_dt.timezone.utc).isoformat(),
        "roster_log": str(roster_path),
        "prior": str(prior_path) if prior_path else "",
        "reference": str(reference_path) if reference_path else "",
        "artifact_profile": artifact_profile if reference_path else "",
        "months": months,
        "per_month": months_out,
    }
    manifest_path = out / "admin_billing_summary_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, default=str), encoding="utf-8")
    manifest["manifest_path"] = str(manifest_path)

    portal_path = build_run_portal(
        out,
        title="Admin Billing Summary — Run Review",
        subtitle=f"Roster: {roster_path.name}",
        sections=admin_billing_sections(manifest, out),
    )
    manifest["html_portal"] = str(portal_path)
    manifest_path.write_text(json.dumps(manifest, indent=2, default=str), encoding="utf-8")
    return manifest


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(prog="triage.admin_billing_summary.cli")
    ap.add_argument("--roster-log", required=True)
    ap.add_argument("--months", nargs="+", default=DEFAULT_MONTHS)
    ap.add_argument("--out-dir", default="Outputs/admin_billing_summary_2026_06_02")
    ap.add_argument("--prior")
    ap.add_argument(
        "--reference",
        help="Approved reference Client workbook for artifact_compare",
    )
    ap.add_argument(
        "--artifact-profile",
        default="admin_billing_summary",
        help="configs/artifact_profiles/<name>.json",
    )
    ap.add_argument("--approved-delta", help="JSON allowlist for semantic hash drift")
    ap.add_argument("--websafe", action="store_true", default=True)
    ap.add_argument("--no-websafe", action="store_false", dest="websafe")
    args = ap.parse_args(argv)
    manifest = run(
        roster_log=args.roster_log,
        out_dir=args.out_dir,
        months=args.months,
        prior=args.prior,
        websafe=args.websafe,
        reference=args.reference,
        artifact_profile=args.artifact_profile,
        approved_delta=args.approved_delta,
    )
    print(json.dumps(manifest, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
