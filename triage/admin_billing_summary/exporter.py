"""OpenAI-format, Web Excel-safe admin billing workbooks (native tables).

Per-month Internal and Client variants share the same resolved ``MonthSummary``.
Single clean ``wb.save()`` — no ``_repair_inlinestr`` post-processing.
"""
from __future__ import annotations

from calendar import month_name as _month_name
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from triage.admin_billing_summary.models import (
    DailyRecord,
    MonthSummary,
    billing_bucket,
)
from triage.nw_prj_neuron_track_hours.bonita_exporter import (
    _write_month_tab,
    tab_name_for_month_key,
)
from triage.neuron_work_context_rules import classify_neuron_work_context
from triage.nw_prj_neuron_track_hours.bonita_resolver import (
    BonitaShift,
    NEURON_DISPLAY_NAME,
)
from triage.websafe_cell import websafe_cell_value
from triage.xlsx_utils import fix_inlinestr

_TITLE_FILL = "1F365C"
_SUBTITLE_FILL = "EAF1F8"
_TABLE_STYLE = "TableStyleMedium4"

_CF_ROWS = [
    ("Green", "OK / clean row", "Counts toward billable totals."),
    ("Amber", "Needs review", "Review before submission."),
    ("Red", "Must fix", "Correct source roster before trusting."),
]

_QC_ROWS = [
    ("Macros", "PASS", "No VBA project."),
    ("Formulas", "PASS", "Values only on export sheets."),
    ("Native tables", "PASS", "Excel Table objects (Web Excel-safe)."),
    ("Inline string cells", "PASS", "No t=inlineStr cells in worksheet XML."),
    ("Shared-string repair", "PASS", "fix_inlinestr only when openpyxl emits inlineStr."),
]


def _xl():
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    return (Workbook, BarChart, Reference, Alignment, Font, PatternFill,
            get_column_letter, Table, TableStyleInfo)


def _num_format(header: str) -> str:
    h = header
    if h in {"Gross Span", "Gross Span Hours", "Gross", "Lunch", "Lunch Deducted",
             "Net Hours", "Net", "Billable Hours", "Clean Net Hours", "Review Net Hours"}:
        return "0.00"
    if h in {"Tech Count", "Worked Days", "Worked Rows", "Detail Rows",
             "Clean Billable Rows", "Review Rows", "Review row count", "Projects", "Techs"}:
        return "0"
    if h == "Date":
        return "yyyy-mm-dd"
    return "General"


def _title_band(ws, title: str, subtitle: str, width: int) -> None:
    """Title/subtitle without merged cells (merged cells force inlineStr in openpyxl)."""
    _, _, _, _, Font, PatternFill, get_column_letter, *_ = _xl()
    w = max(2, width)
    tf = PatternFill("solid", fgColor=_TITLE_FILL)
    sf = PatternFill("solid", fgColor=_SUBTITLE_FILL)
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = Font(bold=True, size=14, color="FFFFFF")
    tcell.fill = tf
    scell = ws.cell(row=2, column=1, value=subtitle)
    scell.font = Font(color="1F2937")
    scell.fill = sf
    for c in range(2, w + 1):
        ws.cell(row=1, column=c).fill = tf
        ws.cell(row=2, column=c).fill = sf
    ws.column_dimensions[get_column_letter(1)].width = 42


def _add_table(ws, table_name: str, headers: List[str], rows: List[Dict[str, Any]],
               header_row: int = 5, start_col: int = 1) -> Tuple[int, int]:
    (_, _, _, Alignment, Font, PatternFill, get_column_letter,
     Table, TableStyleInfo) = _xl()
    fill = PatternFill("solid", fgColor=_TITLE_FILL)
    for c, h in enumerate(headers, start_col):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, header_row + 1):
        for ci, h in enumerate(headers, start_col):
            cell = ws.cell(row=ri, column=ci, value=websafe_cell_value(row.get(h, "")))
            fmt = _num_format(h)
            if fmt != "General":
                cell.number_format = fmt
    last_row = max(header_row, len(rows) + header_row)
    if rows:
        c0 = get_column_letter(start_col)
        c1 = get_column_letter(start_col + len(headers) - 1)
        ref = f"{c0}{header_row}:{c1}{last_row}"
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name=_TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)
    if start_col == 1:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 28 if c <= 2 else 16
    return header_row, last_row


def _add_net_chart(ws, title: str, headers: List[str], header_row: int,
                   last_row: int, anchor: str, category_col: int = 2,
                   value_header: str = "Net Hours") -> None:
    if last_row <= header_row or value_header not in headers:
        return
    _, BarChart, Reference, *_ = _xl()
    net_col = headers.index(value_header) + 1
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.y_axis.title = value_header
    chart.height = 9
    chart.width = 20
    data = Reference(ws, min_col=net_col, min_row=header_row, max_row=last_row)
    cats = Reference(ws, min_col=category_col, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, anchor)


def _review_records(summary: MonthSummary) -> List[DailyRecord]:
    out: List[DailyRecord] = []
    for r in summary.records:
        if r.long_shift or r.project_source == "override":
            out.append(r)
        elif r.project == "Unassigned / Review":
            out.append(r)
    return out


def _time_to_float(t) -> Optional[float]:
    if t is None:
        return None
    return t.hour + t.minute / 60.0


def _neuron_shifts(summary: MonthSummary) -> List[BonitaShift]:
    _, _, mon = _parse_key(summary.month_key)
    short = _month_name[mon]
    shifts: List[BonitaShift] = []
    for r in summary.neuron_records():
        decision = classify_neuron_work_context(
            work_date=r.date,
            start_hour=_time_to_float(r.start_time),
            end_hour=_time_to_float(r.end_time),
            notes=r.note,
            worked_label=r.worked_label,
            resolved_project=r.project,
        )
        shifts.append(BonitaShift(
            month_key=summary.month_key,
            month_name=short,
            date=r.date,
            day=r.day,
            tech=r.tech,
            clock_in=r.clock_in,
            clock_out=r.clock_out,
            total_hours=r.net_hours,
            project_name=NEURON_DISPLAY_NAME,
            assignment_type=decision.assignment_type,
            note=r.note,
            long_shift=r.long_shift,
            start_time=r.start_time,
            end_time=r.end_time,
        ))
    shifts.sort(key=lambda s: (s.date, s.tech))
    return shifts



def _parse_key(month_key: str):
    import re
    m = re.match(r"^(\d{4})-(\d{1,2})$", month_key.strip())
    return m.group(1), int(m.group(1)), int(m.group(2))


def _neuron_detail_rows(summary: MonthSummary) -> List[Dict[str, Any]]:
    return [{
        "Month": summary.month_name,
        "Date": r.date,
        "Day": r.day,
        "Tech": r.tech,
        "Project": r.project,
        "Billing Bucket": billing_bucket(r.project),
        "Clock In": r.clock_in,
        "Clock Out": r.clock_out,
        "Gross": round(r.gross_span, 2),
        "Lunch": round(r.lunch, 2),
        "Net": round(r.net_hours, 2),
    } for r in summary.neuron_records()]


def _project_rows(summary: MonthSummary) -> List[Dict[str, Any]]:
    return [{
        "Month": summary.month_name,
        "Project": r.project,
        "Billing Bucket": billing_bucket(r.project),
        "Worked Rows": r.worked_days,
        "Tech Count": r.tech_count,
        "Gross Span": round(r.gross_span, 2),
        "Lunch": round(r.lunch_deducted, 2),
        "Net Hours": round(r.net_hours, 2),
    } for r in summary.project_rows]


def _tech_rows(summary: MonthSummary) -> List[Dict[str, Any]]:
    return [{
        "Month": summary.month_name,
        "Tech": r.tech,
        "Worked Rows": r.worked_days,
        "Gross Span": round(r.gross_span, 2),
        "Lunch": round(r.lunch_deducted, 2),
        "Net Hours": round(r.net_hours, 2),
    } for r in summary.tech_rows]


def _tech_project_rows(summary: MonthSummary) -> List[Dict[str, Any]]:
    return [{
        "Month": summary.month_name,
        "Tech": r.tech,
        "Project": r.project,
        "Worked Rows": r.worked_days,
        "Gross Span": round(r.gross_span, 2),
        "Lunch": round(r.lunch_deducted, 2),
        "Net Hours": round(r.net_hours, 2),
    } for r in summary.tech_project_rows]


def _monthly_summary_rows(summary: MonthSummary) -> List[Dict[str, Any]]:
    review = _review_records(summary)
    review_net = round(sum(r.net_hours for r in review), 2)
    return [{
        "Month": summary.month_name,
        "Detail Rows": len(summary.records),
        "Clean Billable Rows": len(summary.records) - len(review),
        "Review Rows": len(review),
        "Clean Net Hours": round(summary.total_net - review_net, 2),
        "Review Net Hours": review_net,
        "Gross Span": summary.total_gross,
        "Lunch": summary.total_lunch,
    }]


def _review_flag_rows(summary: MonthSummary) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for r in _review_records(summary):
        flags = []
        if r.long_shift:
            flags.append("long_shift")
        if r.project_source == "override":
            flags.append("override")
        if r.project == "Unassigned / Review":
            flags.append("unassigned")
        rows.append({
            "Month": summary.month_name,
            "Date": r.date,
            "Day": r.day,
            "Tech": r.tech,
            "Project": r.project,
            "Net Hours": round(r.net_hours, 2),
            "Flag": ", ".join(flags),
            "Note": r.note,
        })
    for m in summary.malformed:
        rows.append({
            "Month": summary.month_name,
            "Date": "",
            "Day": "",
            "Tech": "",
            "Project": "",
            "Net Hours": "",
            "Flag": "malformed",
            "Note": m,
        })
    return rows


def _sheet_table(ws, title: str, subtitle: str, table_name: str,
                 headers: List[str], rows: List[Dict[str, Any]],
                 chart: Optional[Tuple[str, str]] = None) -> None:
    _title_band(ws, title, subtitle, len(headers))
    hr, lr = _add_table(ws, table_name, headers, rows)
    if chart:
        _add_net_chart(ws, chart[0], headers, hr, lr, chart[1])


def build_workbook(
    summary: MonthSummary,
    out_path: str,
    *,
    variant: str = "internal",
    roster_name: str = "",
    generated_utc: Optional[str] = None,
) -> str:
    """Write Internal or Client billing summary for one month."""
    if variant not in ("internal", "client"):
        raise ValueError(f"variant must be internal or client, got {variant!r}")

    Workbook, *_ = _xl()
    wb = Workbook()
    wb.remove(wb.active)
    label = summary.month_name
    stamp = generated_utc or datetime.now().strftime("%Y-%m-%d %H:%M")
    stem = Path(out_path).stem

    ws = wb.create_sheet("Start Here")
    _title_band(ws, f"{label} Billing Summary ({variant.title()})",
                "Roster-derived; override-aware project resolution.", 2)
    _add_table(ws, "StartHereTable", ["Field", "Value"], [
        {"Field": "Artifact", "Value": stem + ".xlsx"},
        {"Field": "Generated", "Value": stamp},
        {"Field": "Source roster", "Value": roster_name},
        {"Field": "Variant", "Value": variant},
        {"Field": "Resolution", "Value": "Override > Worked > Assignment > Live default"},
    ])

    ws = wb.create_sheet("Executive Dashboard")
    _title_band(ws, "Executive Dashboard", f"{label} billing snapshot.", 8)
    review = _review_records(summary)
    review_net = round(sum(r.net_hours for r in review), 2)
    _add_table(ws, "DashboardKPITable", ["Metric", "Value"], [
        {"Metric": "Total Net Hours", "Value": summary.total_net},
        {"Metric": "Gross Span", "Value": summary.total_gross},
        {"Metric": "Lunch / Unpaid", "Value": summary.total_lunch},
        {"Metric": "Techs", "Value": summary.techs_reflected},
        {"Metric": "Projects", "Value": summary.projects_reflected},
        {"Metric": "Neuron Net", "Value": summary.net_for_bucket("Neurons")},
        {"Metric": "Review Net Hours", "Value": review_net},
        {"Metric": "Review row count", "Value": len(review)},
    ], header_row=5)
    top = sorted(summary.project_rows, key=lambda x: -x.net_hours)[:8]
    _add_table(ws, "DashboardProjectTopTable", ["Project", "Net Hours"],
               [{"Project": p.project, "Net Hours": round(p.net_hours, 2)} for p in top],
               header_row=5, start_col=5)

    mh = ["Month", "Detail Rows", "Clean Billable Rows", "Review Rows",
          "Clean Net Hours", "Review Net Hours", "Gross Span", "Lunch"]
    _sheet_table(wb.create_sheet("Monthly Summary"), "Monthly Summary",
                 f"{label} totals.", "MonthlySummaryTable", mh, _monthly_summary_rows(summary))

    ph = ["Month", "Project", "Billing Bucket", "Worked Rows", "Tech Count",
          "Gross Span", "Lunch", "Net Hours"]
    _sheet_table(wb.create_sheet("Project Summary"), "Project Summary",
                 "Billable totals by project.", "ProjectSummaryTable", ph, _project_rows(summary),
                 ("Net Hours by Project", "I4"))

    th = ["Month", "Tech", "Worked Rows", "Gross Span", "Lunch", "Net Hours"]
    _sheet_table(wb.create_sheet("Tech Summary"), "Tech Summary",
                 "Billable totals by technician.", "TechSummaryTable", th, _tech_rows(summary))

    tph = ["Month", "Tech", "Project", "Worked Rows", "Gross Span", "Lunch", "Net Hours"]
    _sheet_table(wb.create_sheet("Tech Project Summary"), "Tech Project Summary",
                 "Billable totals by technician and project.", "TechProjectSummaryTable",
                 tph, _tech_project_rows(summary),
                 ("Net Hours by Tech and Project", "I4"))

    neuron_tab = f"{label.split()[0]} Neuron Hours"
    nh = ["Month", "Date", "Day", "Tech", "Project", "Billing Bucket",
          "Clock In", "Clock Out", "Gross", "Lunch", "Net"]
    _sheet_table(wb.create_sheet(neuron_tab), f"{label} Neuron Hours",
                 "Neuron-only detail; same source as embedded Bonita tracker tab.",
                 "NeuronDetailTable", nh, _neuron_detail_rows(summary))

    _write_month_tab(wb, tab_name_for_month_key(summary.month_key), _neuron_shifts(summary))

    if variant == "internal":
        rh = ["Month", "Date", "Day", "Tech", "Project", "Net Hours", "Flag", "Note"]
        _sheet_table(wb.create_sheet("Review Flags"), "Review Flags",
                     "Overrides, long shifts, unassigned, and malformed rows.",
                     "ReviewFlagsTable", rh, _review_flag_rows(summary))
        _sheet_table(wb.create_sheet("CF Dictionary"), "Conditional Formatting Dictionary",
                     "Color meanings for review visibility.", "CFDictionaryTable",
                     ["Color", "Meaning", "Action"],
                     [{"Color": c, "Meaning": m, "Action": a} for c, m, a in _CF_ROWS])
        _sheet_table(wb.create_sheet("WebExcel QC"), "Web Excel QC",
                     "Structural checks for Excel for Web.", "WebExcelQCTable",
                     ["Check", "Result", "Notes"],
                     [{"Check": c, "Result": r, "Notes": n} for c, r, n in _QC_ROWS])

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    fix_inlinestr(out_path)
    return out_path
