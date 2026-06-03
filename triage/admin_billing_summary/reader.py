"""Resolve the Active Roster Log into multi-project daily records.

Project resolution precedence (matches the documented hierarchy and the
April "My Preferred Format" build, which extracted from Worked Projects):

    Assignments Overrides sub-table  (Richard-reviewed authority)
    > Worked Projects - {Month} cell
    > Assignments - {Month} main-table cell
    > Live default Project column

Reuses proven helpers: time/lunch math from ``roster_parser`` and the
Worked-Projects lookup + clock formatting from the Neuron engine.
"""
from __future__ import annotations

import datetime as _dt
from calendar import month_name as _month_name
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from triage.admin_billing_summary.models import DailyRecord
from triage.nw_prj_neuron_track_hours.reader import (
    _decimal_to_time,
    _format_clock,
    _worked_project_lookup,
)
from triage.roster_parser import (
    RosterParseError,
    _compute_gross,
    _extract_year_from_sheet,
    _find_assignments_sheet,
    _find_header_row,
    _find_live_sheet,
    _is_overnight,
    _load_assignments,
    _lunch_deduction,
    _parse_date_header,
    _time_to_hours,
)
from triage.nw_prj_neuron_track_hours.reader import split_note_bearing_punch

LONG_SHIFT_HOURS = 12.0


def _clean_project(value: Any) -> str:
    """Normalize a project cell; treat blanks and placeholder '0' as empty."""
    s = str(value).strip() if value is not None else ""
    return "" if s in ("", "0", "None") else s


def _month_label(month_key: str) -> Tuple[str, int, int]:
    from triage.month_validation import validate_month_key
    try:
        year, mon = validate_month_key(month_key)
    except ValueError as exc:
        raise RosterParseError(str(exc)) from exc
    return f"{_month_name[mon]} {year}", year, mon


def _find_worked_sheet(wb, label: str):
    target = f"worked projects - {label}".lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return wb[name]
    month_word = label.split()[0].lower()
    for name in wb.sheetnames:
        low = name.strip().lower()
        if low.startswith("worked projects") and month_word in low and label.split()[-1] in name:
            return wb[name]
    return None


def _load_overrides(ws) -> Dict[Tuple[date, str], Tuple[str, str]]:
    """Parse ONLY the Assignments 'Overrides' sub-table.

    Returns {(date, staff): (project, note)} for explicit reviewed overrides.
    """
    out: Dict[Tuple[date, str], Tuple[str, str]] = {}
    if ws is None:
        return out
    max_row = ws.max_row
    overrides_start: Optional[int] = None
    for r in range(1, max_row + 1):
        v = ws.cell(r, 1).value
        if v and "override" in str(v).strip().lower():
            overrides_start = r
            break
    if overrides_start is None:
        return out
    hdr: Optional[int] = None
    for r in range(overrides_start, min(overrides_start + 5, max_row + 1)):
        v = ws.cell(r, 1).value
        if not v:
            continue
        s = str(v).strip().lower()
        if "override staff" in s:
            hdr = r
            break
        if s in ("staff name", "staff") and r > 1:
            prev = ws.cell(r - 1, 1).value
            if prev and "override" in str(prev).strip().lower():
                hdr = r
                break
    if hdr is None:
        return out
    for r in range(hdr + 1, max_row + 1):
        staff = ws.cell(r, 1).value
        dval = ws.cell(r, 2).value
        proj = ws.cell(r, 3).value
        note = ws.cell(r, 4).value
        if not staff or not dval or not proj:
            continue
        if isinstance(dval, _dt.datetime):
            d = dval.date()
        elif isinstance(dval, _dt.date):
            d = dval
        else:
            continue
        out[(d, str(staff).strip())] = (str(proj).strip(), str(note).strip() if note else "")
    return out


def read_month(roster_path: str | Path, month_key: str) -> Tuple[List[DailyRecord], List[str], List[str]]:
    """Return (records, warnings, malformed) for one month."""
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover
        raise RosterParseError("openpyxl is required: pip install openpyxl") from e
    p = Path(roster_path)
    if not p.exists():
        raise RosterParseError(f"Roster file not found: {roster_path}")
    label, year, mon = _month_label(month_key)
    wb = openpyxl.load_workbook(str(p), data_only=True, read_only=False)

    warnings: List[str] = []
    malformed: List[str] = []
    try:
        live_ws = _find_live_sheet(wb, label)
        worked = _worked_project_lookup(_find_worked_sheet(wb, label))
        asn_ws = _find_assignments_sheet(wb, label)
        assignments = _load_assignments(asn_ws)            # main + overrides merged
        overrides = _load_overrides(asn_ws)                # explicit reviewed overrides
        if not worked:
            warnings.append(f"worked_projects_missing_or_empty:{label}")
        if asn_ws is None:
            warnings.append(f"assignments_missing:{label}")

        year_hint = _extract_year_from_sheet(live_ws.title)
        hdr_row = _find_header_row(live_ws)
        if hdr_row is None:
            raise RosterParseError(f"No header row in {live_ws.title}")
        headers = [live_ws.cell(hdr_row, c).value for c in range(1, live_ws.max_column + 1)]

        idx_staff, idx_project = None, None
        for i, h in enumerate(headers):
            hs = str(h).lower().strip() if h else ""
            if idx_staff is None and ("staff" in hs or "name" in hs):
                idx_staff = i
            elif idx_project is None and ("project" in hs or "team" in hs or "bucket" in hs):
                idx_project = i
        if idx_staff is None:
            raise RosterParseError(f"No Staff Name column in {live_ws.title}")

        date_to_cols: Dict[date, Dict[str, int]] = {}
        for i, h in enumerate(headers):
            if h is None:
                continue
            res = _parse_date_header(str(h), year_hint)
            if res:
                d, direction = res
                date_to_cols.setdefault(d, {})[direction] = i
        if not date_to_cols:
            raise RosterParseError(f"No date columns in {live_ws.title}")

        records: List[DailyRecord] = []
        for r in range(hdr_row + 1, live_ws.max_row + 1):
            row = [live_ws.cell(r, c).value for c in range(1, live_ws.max_column + 1)]
            staff_val = row[idx_staff] if idx_staff < len(row) else None
            if not staff_val or str(staff_val).strip() in ("", "None", "0"):
                continue
            if isinstance(staff_val, (int, float)):
                continue
            staff = str(staff_val).strip()
            live_default = ""
            if idx_project is not None and idx_project < len(row) and row[idx_project] is not None:
                live_default = str(row[idx_project]).strip()
            if live_default in ("0", ""):
                live_default = ""

            for d, dirs in sorted(date_to_cols.items()):
                in_val = row[dirs["in"]] if "in" in dirs and dirs["in"] < len(row) else None
                out_val = row[dirs["out"]] if "out" in dirs and dirs["out"] < len(row) else None
                ci, note_in = split_note_bearing_punch(in_val)
                co, note_out = split_note_bearing_punch(out_val)
                if ci is None and co is None:
                    continue
                if ci is None or co is None:
                    miss = "Clock In" if ci is None else "Clock Out"
                    malformed.append(
                        f"{staff} on {d.isoformat()}: {miss} blank while other present - row excluded"
                    )
                    continue

                worked_label = _clean_project(worked.get((staff, d), ""))
                asn_label = _clean_project(assignments.get((d, staff), ""))
                override = overrides.get((d, staff))
                override_proj = _clean_project(override[0]) if override else ""
                note = " ".join(n for n in (note_in, note_out) if n).strip()
                if override_proj:
                    project, project_source = override_proj, "override"
                    if override[1]:
                        note = (note + " | " + override[1]).strip(" |")
                elif worked_label:
                    project, project_source = worked_label, "worked"
                elif asn_label:
                    project, project_source = asn_label, "assignment"
                else:
                    project, project_source = (live_default or "Unassigned / Review"), "live_default"

                gross = _compute_gross(ci, co)
                lunch = _lunch_deduction(gross)
                net = round(max(0.0, gross - lunch), 4)
                records.append(DailyRecord(
                    date=d,
                    day=d.strftime("%a"),
                    tech=staff,
                    project=project,
                    project_source=project_source,
                    clock_in=_format_clock(ci),
                    clock_out=_format_clock(co),
                    gross_span=gross,
                    lunch=lunch,
                    net_hours=net,
                    long_shift=gross > LONG_SHIFT_HOURS,
                    note=note,
                    worked_label=worked_label,
                    start_time=_decimal_to_time(ci),
                    end_time=_decimal_to_time(co),
                ))
    finally:
        wb.close()

    records.sort(key=lambda r: (r.date, r.tech, r.project))
    return records, warnings, malformed
