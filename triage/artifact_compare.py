"""Compare generated XLSX artifacts against approved references.

This is intentionally lightweight. It provides a first repo-native comparison
surface without committing private approved workbooks.
"""
from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any


def raw_sha256(path: str | Path) -> str:
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


def workbook_visible_payload(path: str | Path) -> dict[str, Any]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        sheets = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                values = [str(value) if value is not None else "" for value in row]
                if any(values):
                    rows.append(values)
            sheets.append({"title": ws.title, "rows": rows})
        return {"sheet_order": wb.sheetnames, "sheets": sheets}
    finally:
        wb.close()


def semantic_sha256(path: str | Path) -> str:
    payload = workbook_visible_payload(path)
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def compare_workbooks(reference: str | Path, candidate: str | Path) -> dict[str, Any]:
    ref_raw = raw_sha256(reference)
    cand_raw = raw_sha256(candidate)
    ref_semantic = semantic_sha256(reference)
    cand_semantic = semantic_sha256(candidate)
    semantic_match = ref_semantic == cand_semantic
    return {
        "reference": str(reference),
        "candidate": str(candidate),
        "raw_sha_match": ref_raw == cand_raw,
        "semantic_match": semantic_match,
        "comparison_status": "PASS" if semantic_match else "FAIL",
        "reference_raw_sha256": ref_raw,
        "candidate_raw_sha256": cand_raw,
        "reference_semantic_sha256": ref_semantic,
        "candidate_semantic_sha256": cand_semantic,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Compare generated workbook against approved reference")
    parser.add_argument("--reference", required=True)
    parser.add_argument("--candidate", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--profile", default="")
    args = parser.parse_args(argv)

    result = compare_workbooks(args.reference, args.candidate)
    if args.profile:
        result["profile"] = args.profile
    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    Path(args.out).write_text(json.dumps(result, indent=2), encoding="utf-8")
    return 0 if result["comparison_status"] == "PASS" else 1


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
