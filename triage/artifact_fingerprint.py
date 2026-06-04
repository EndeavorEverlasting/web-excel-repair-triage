"""Three-layer fingerprinting for .xlsx submission artifacts.

Layers
------
1. ``raw_sha256`` — exact file bytes (brittle for OOXML).
2. ``canonical_package_sha256`` — normalized ZIP parts, volatile metadata stripped.
3. ``semantic_sha256`` — stable JSON of sheet meaning (values, formulas, layout signals).

Distinct from ``triage.nw_prj_artifact_compare`` (dashboard row reconciliation).
"""
from __future__ import annotations

import hashlib
import json
import posixpath
import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List

from triage.xlsx_utils import get_attr, read_bytes, read_text, sheet_name_map, table_parts

_VOLATILE_PARTS = frozenset({
    "docProps/core.xml",
    "docProps/app.xml",
    "xl/calcChain.xml",
})

_CALCPR_RE = re.compile(r"<calcPr\b[^>]*/>", re.IGNORECASE)
_WS_RE = re.compile(r">\s+<")


def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def raw_sha256(path: str | Path) -> str:
    """SHA-256 of the full file on disk."""
    return _sha256_bytes(Path(path).read_bytes())


def _normalize_part_content(name: str, raw: bytes) -> bytes:
    if not name.lower().endswith(".xml"):
        return raw
    text = raw.decode("utf-8", errors="ignore")
    if name == "xl/workbook.xml":
        text = _CALCPR_RE.sub("", text)
    text = _WS_RE.sub("><", text)
    return text.encode("utf-8")


def canonical_package_sha256(path: str | Path) -> str:
    """Hash OOXML package after dropping volatile parts and light XML normalization."""
    chunks: List[bytes] = []
    with zipfile.ZipFile(path, "r") as z:
        for name in sorted(z.namelist()):
            if name.endswith("/") or name in _VOLATILE_PARTS:
                continue
            raw = z.read(name)
            norm = _normalize_part_content(name, raw)
            chunks.append(name.encode("utf-8") + b"\0" + norm)
    return _sha256_bytes(b"".join(chunks))


def _stable_value(val: Any) -> Any:
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float, Decimal)):
        return round(float(val), 10)
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, time):
        return val.isoformat()
    s = str(val).strip()
    return s


def _sheet_xml_features(z: zipfile.ZipFile, sheet_part: str) -> Dict[str, Any]:
    try:
        raw = read_bytes(z, sheet_part)
    except KeyError:
        return {}
    text = raw.decode("utf-8", errors="ignore")
    frozen = bool(re.search(r"<pane\b", text))
    autofilter = re.findall(r'<autoFilter\b[^>]*\bref="([^"]+)"', text)
    return {
        "frozen_pane": frozen,
        "autofilter_refs": sorted(autofilter),
    }


def _rels_part_for(source_part: str) -> str:
    return posixpath.join(
        posixpath.dirname(source_part),
        "_rels",
        posixpath.basename(source_part) + ".rels",
    )


def _resolve_rel_target(source_part: str, target: str) -> str:
    t = (target or "").strip().replace("\\", "/")
    while t.startswith("/"):
        t = t[1:]
    if t.startswith("xl/"):
        return t
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), t))


def _related_parts(
    z: zipfile.ZipFile,
    source_part: str,
    *,
    rel_type_tail: str = "",
    target_prefix: str = "",
) -> List[str]:
    try:
        rels = read_text(z, _rels_part_for(source_part))
    except KeyError:
        return []

    found: set[str] = set()
    for m in re.finditer(r"<Relationship\b[^>]*>", rels):
        frag = m.group(0)
        rel_type = get_attr(frag, "Type") or ""
        target = get_attr(frag, "Target") or ""
        resolved = _resolve_rel_target(source_part, target)
        if rel_type_tail and not rel_type.endswith(rel_type_tail):
            continue
        if target_prefix and not resolved.startswith(target_prefix):
            continue
        found.add(resolved)
    return sorted(found)


def _table_features(z: zipfile.ZipFile, sheet_parts: List[str] | None = None) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    table_part_names = table_parts(z) if sheet_parts is None else sorted({
        table_part
        for sheet_part in sheet_parts
        for table_part in _related_parts(
            z,
            sheet_part,
            rel_type_tail="/table",
            target_prefix="xl/tables/",
        )
    })
    for part in table_part_names:
        try:
            text = read_text(z, part)
        except KeyError:
            continue
        name_m = re.search(r'\bname="([^"]+)"', text)
        ref_m = re.search(r'\bref="([^"]+)"', text)
        out.append({
            "part": part,
            "name": name_m.group(1) if name_m else "",
            "ref": ref_m.group(1) if ref_m else "",
        })
    return sorted(out, key=lambda x: (x["name"], x["ref"]))


def _chart_count(z: zipfile.ZipFile, sheet_parts: List[str] | None = None) -> int:
    if sheet_parts is None:
        return sum(1 for n in z.namelist() if n.startswith("xl/charts/chart") and n.endswith(".xml"))

    drawing_parts = {
        drawing_part
        for sheet_part in sheet_parts
        for drawing_part in _related_parts(
            z,
            sheet_part,
            rel_type_tail="/drawing",
            target_prefix="xl/drawings/",
        )
    }
    chart_parts = {
        chart_part
        for drawing_part in drawing_parts
        for chart_part in _related_parts(
            z,
            drawing_part,
            rel_type_tail="/chart",
            target_prefix="xl/charts/",
        )
    }
    return len(chart_parts)


def _semantic_workbook_model(path: str | Path, *, include_hidden: bool = False) -> Dict[str, Any]:
    """Build canonical semantic document (before hashing).

    Omits filesystem path so copies and metadata-only ZIP edits hash identically.
    """
    p = Path(path)
    model: Dict[str, Any] = {"sheets": []}

    try:
        import openpyxl
    except ImportError as exc:
        model["error"] = f"openpyxl_not_installed:{exc}"
        return model

    wb = openpyxl.load_workbook(str(p), data_only=False, read_only=True)
    try:
        included_sheet_names: List[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if not include_hidden and getattr(ws, "sheet_state", "visible") in {"hidden", "veryHidden"}:
                continue
            included_sheet_names.append(sheet_name)
        included_name_set = set(included_sheet_names)

        sheets_out: List[Dict[str, Any]] = []
        with zipfile.ZipFile(p, "r") as z:
            part_to_name = sheet_name_map(z)
            sheet_parts_sorted = sorted(
                (
                    (part, name)
                    for part, name in part_to_name.items()
                    if name in included_name_set
                ),
                key=lambda kv: kv[1],
            )
            included_sheet_parts = [part for part, _ in sheet_parts_sorted]
            model["sheet_order"] = [name for _, name in sheet_parts_sorted]
            model["table_features"] = _table_features(z, included_sheet_parts)
            model["chart_count"] = _chart_count(z, included_sheet_parts)
            name_to_part = {v: k for k, v in part_to_name.items()}
            for sheet_name in included_sheet_names:
                ws = wb[sheet_name]
                cells: Dict[str, Any] = {}
                for row in ws.iter_rows(
                    min_row=1,
                    max_row=ws.max_row or 1,
                    min_col=1,
                    max_col=ws.max_column or 1,
                ):
                    for cell in row:
                        if cell.value is None and not cell.data_type == "f":
                            if cell.number_format in (None, "General"):
                                continue
                        addr = f"{sheet_name}!{cell.coordinate}"
                        entry: Dict[str, Any] = {"v": _stable_value(cell.value)}
                        if cell.data_type == "f" and cell.value is not None:
                            entry["f"] = str(cell.value)
                        nf = cell.number_format
                        if nf and nf != "General":
                            entry["nf"] = nf
                        cells[addr] = entry
                part = name_to_part.get(sheet_name, "")
                xml_feat = _sheet_xml_features(z, part) if part else {}
                sheets_out.append({
                    "name": sheet_name,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "cells": cells,
                    **xml_feat,
                })
        model["sheets"] = sheets_out
    finally:
        wb.close()

    return model


def _semantic_model_sha256(model: Dict[str, Any]) -> str:
    payload = json.dumps(model, sort_keys=True, separators=(",", ":"), default=str)
    return _sha256_bytes(payload.encode("utf-8"))


def semantic_sha256(path: str | Path, *, include_hidden: bool = False) -> str:
    """Hash canonical semantic workbook JSON.

    By default only visible worksheets contribute to the semantic hash.
    Set ``include_hidden=True`` for an explicit all-sheets semantic hash.
    """
    return _semantic_model_sha256(_semantic_workbook_model(path, include_hidden=include_hidden))


def per_sheet_semantic_detail(path: str | Path, *, include_hidden: bool = False) -> Dict[str, Any]:
    """Per-sheet breakdown for compare reports (not used for pass/fail hash)."""
    model = _semantic_workbook_model(path, include_hidden=include_hidden)
    sheets: Dict[str, Any] = {}
    for sh in model.get("sheets") or []:
        name = sh.get("name", "")
        cells = sh.get("cells") or {}
        cell_payload = json.dumps(cells, sort_keys=True, separators=(",", ":"), default=str)
        sheets[name] = {
            "cell_value_hash": _sha256_bytes(cell_payload.encode("utf-8")),
            "cell_count": len(cells),
            "chart_count": model.get("chart_count", 0),
            "frozen_pane": sh.get("frozen_pane", False),
            "autofilter_refs": sh.get("autofilter_refs", []),
        }
    return sheets


@dataclass
class ArtifactFingerprint:
    path: str
    raw_sha256: str
    canonical_package_sha256: str
    semantic_sha256: str
    all_sheets_semantic_sha256: str
    sheet_order: List[str] = field(default_factory=list)
    sheets: Dict[str, Any] = field(default_factory=dict)
    chart_count: int = 0
    table_count: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "path": self.path,
            "raw_sha256": self.raw_sha256,
            "canonical_package_sha256": self.canonical_package_sha256,
            "semantic_sha256": self.semantic_sha256,
            "all_sheets_semantic_sha256": self.all_sheets_semantic_sha256,
            "sheet_order": list(self.sheet_order),
            "chart_count": self.chart_count,
            "table_count": self.table_count,
            "sheets": dict(self.sheets),
        }


def fingerprint_file(path: str | Path) -> ArtifactFingerprint:
    """Compute all three fingerprint layers for one workbook."""
    p = Path(path)
    model = _semantic_workbook_model(p, include_hidden=False)
    all_sheets_model = _semantic_workbook_model(p, include_hidden=True)
    sem_hash = _semantic_model_sha256(model)
    all_sheets_sem_hash = _semantic_model_sha256(all_sheets_model)
    sheets_detail: Dict[str, Any] = {}
    for sh in model.get("sheets") or []:
        name = sh.get("name", "")
        cells = sh.get("cells") or {}
        cell_payload = json.dumps(cells, sort_keys=True, separators=(",", ":"), default=str)
        sheets_detail[name] = {
            "cell_value_hash": _sha256_bytes(cell_payload.encode("utf-8")),
            "cell_count": len(cells),
            "frozen_pane": sh.get("frozen_pane", False),
            "autofilter_refs": sh.get("autofilter_refs", []),
        }
    return ArtifactFingerprint(
        path=str(p.resolve()),
        raw_sha256=raw_sha256(p),
        canonical_package_sha256=canonical_package_sha256(p),
        semantic_sha256=sem_hash,
        all_sheets_semantic_sha256=all_sheets_sem_hash,
        sheet_order=list(model.get("sheet_order") or []),
        sheets=sheets_detail,
        chart_count=int(model.get("chart_count") or 0),
        table_count=len(model.get("table_features") or []),
    )
