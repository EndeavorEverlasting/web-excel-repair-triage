"""Reusable workbook formatting/profile helpers for submission artifacts.

The goal is to preserve the formatting posture of the approved generated outputs
without committing private workbook artifacts. Engines should load a profile and
apply the same simple, Web-Excel-friendly conventions every time:

- readable title and header rows
- frozen panes and autofilters
- conservative column widths
- values-only submission workbooks
- internal review/audit content kept out of client-facing tabs
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable
import json

HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"
SUBTITLE_FILL = "EAF1F8"
TITLE_FONT_SIZE = 14
HEADER_FONT_SIZE = 11
BODY_FONT_SIZE = 10
DEFAULT_ROW_HEIGHT = 18
TITLE_ROW_HEIGHT = 24
HEADER_ROW_HEIGHT = 22


@dataclass(frozen=True)
class ArtifactFormatProfile:
    """Workbook format/profile contract loaded from JSON."""

    name: str
    sheets: tuple[str, ...]
    submission_safe: bool
    forbidden_text: tuple[str, ...]
    required_headers: tuple[str, ...]
    frozen_panes: str = "A3"
    autofilter: bool = True
    title_row_height: int = TITLE_ROW_HEIGHT
    header_row_height: int = HEADER_ROW_HEIGHT


def load_profile(path: str | Path) -> ArtifactFormatProfile:
    """Load an artifact format profile from JSON."""

    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    return ArtifactFormatProfile(
        name=payload["name"],
        sheets=tuple(payload.get("sheets", [])),
        submission_safe=bool(payload.get("submission_safe", False)),
        forbidden_text=tuple(payload.get("forbidden_text", [])),
        required_headers=tuple(payload.get("required_headers", [])),
        frozen_panes=payload.get("frozen_panes", "A3"),
        autofilter=bool(payload.get("autofilter", True)),
        title_row_height=int(payload.get("title_row_height", TITLE_ROW_HEIGHT)),
        header_row_height=int(payload.get("header_row_height", HEADER_ROW_HEIGHT)),
    )


def apply_submission_sheet_format(
    ws: Any,
    header_row: int,
    last_row: int | None = None,
    freeze_cell: str | None = None,
    enable_filter: bool = True,
) -> None:
    """Apply the repo-standard submission sheet formatting to an openpyxl sheet.

    This intentionally avoids fragile features. No pivots, no macros, no external
    links, no volatile formulas. It is designed for Excel desktop and Excel for
    Web compatibility.
    """

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True, color=HEADER_FONT, size=HEADER_FONT_SIZE)
    body_font = Font(size=BODY_FONT_SIZE)

    ws.freeze_panes = freeze_cell or ws.cell(header_row + 1, 1).coordinate
    ws.row_dimensions[header_row].height = HEADER_ROW_HEIGHT

    max_col = max(1, ws.max_column)
    max_row = max(header_row, last_row or ws.max_row)

    for cell in ws[header_row]:
        if cell.column <= max_col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=header_row + 1, max_row=max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letter]:
            if cell.value not in (None, ""):
                max_len = max(max_len, min(len(str(cell.value)), 45))
        ws.column_dimensions[letter].width = max_len + 2

    if enable_filter:
        last_col = get_column_letter(max_col)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{max_row}"


def assert_no_forbidden_submission_text(values: Iterable[object], forbidden: Iterable[str]) -> None:
    """Fail if internal/review-only text leaks into a submission workbook."""

    forbidden_lower = [token.lower() for token in forbidden]
    for value in values:
        if value is None:
            continue
        text = str(value).lower()
        hit = next((token for token in forbidden_lower if token and token in text), None)
        if hit:
            raise ValueError(f"forbidden submission text leaked: {hit}")
