"""JSON-driven artifact profiles for stop-ship checks and compare totals."""
from __future__ import annotations

import json
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

from triage.webexcel_semantic_gate import (
    _GENERIC_RE,
    extract_shared_strings,
    run_semantic_gate,
)

_REPO_ROOT = Path(__file__).resolve().parent.parent
_PROFILES_DIR = _REPO_ROOT / "configs" / "artifact_profiles"

_GATE_PROFILE_ALIAS = {
    "admin_billing_summary": "admin_billing",
    "bonita_neuron_track_hours": "bonita",
    "internal_admin_log": "admin_billing",
}


@dataclass
class ArtifactProfile:
    profile: str
    required_sheets: List[str] = field(default_factory=list)
    required_headers: List[str] = field(default_factory=list)
    required_semantic_cells: List[Dict[str, Any]] = field(default_factory=list)
    required_nonblank_columns: List[str] = field(default_factory=list)
    forbidden_values: List[str] = field(default_factory=list)
    compare_totals: bool = False
    totals: Dict[str, Any] = field(default_factory=dict)
    fail_on_canonical_mismatch: bool = False
    variant: Optional[str] = None

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "ArtifactProfile":
        return cls(
            profile=str(data.get("profile", "")),
            required_sheets=list(data.get("required_sheets") or []),
            required_headers=list(data.get("required_headers") or []),
            required_semantic_cells=list(data.get("required_semantic_cells") or []),
            required_nonblank_columns=list(data.get("required_nonblank_columns") or []),
            forbidden_values=list(data.get("forbidden_values") or []),
            compare_totals=bool(data.get("compare_totals")),
            totals=dict(data.get("totals") or {}),
            fail_on_canonical_mismatch=bool(data.get("fail_on_canonical_mismatch")),
            variant=data.get("variant"),
        )


@dataclass
class ProfileResult:
    failures: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    totals: Dict[str, Any] = field(default_factory=dict)

    @property
    def passed(self) -> bool:
        return not self.failures


def load_profile(name: str) -> ArtifactProfile:
    path = _PROFILES_DIR / f"{name}.json"
    if not path.is_file():
        raise FileNotFoundError(f"artifact profile not found: {path}")
    data = json.loads(path.read_text(encoding="utf-8"))
    return ArtifactProfile.from_dict(data)


def resolve_profile_sheets(profile: ArtifactProfile, *, expect_neuron_tab: Optional[str] = None) -> List[str]:
    out: List[str] = []
    for s in profile.required_sheets:
        if s == "{{expect_neuron_tab}}" and expect_neuron_tab:
            out.append(expect_neuron_tab)
        else:
            out.append(s)
    return out


def gate_profile_for(profile_name: str) -> str:
    return _GATE_PROFILE_ALIAS.get(profile_name, profile_name)


def _workbook_tabs(path: str) -> List[str]:
    try:
        with zipfile.ZipFile(path, "r") as z:
            if "xl/workbook.xml" not in z.namelist():
                return []
            text = z.read("xl/workbook.xml").decode("utf-8", errors="ignore")
            return re.findall(r'<sheet[^>]*name="([^"]+)"', text)
    except zipfile.BadZipFile:
        return []


def _check_forbidden_strings(path: str, forbidden: List[str]) -> List[str]:
    failures: List[str] = []
    try:
        with zipfile.ZipFile(path, "r") as z:
            if "xl/sharedStrings.xml" in z.namelist():
                strings = extract_shared_strings(z.read("xl/sharedStrings.xml"))
                for s in strings:
                    for bad in forbidden:
                        if bad.lower() in s.lower():
                            failures.append(f"forbidden_shared_string:{bad!r}")
                    if _GENERIC_RE.match(s.strip()):
                        failures.append(f"generic_column_string:{s}")
    except zipfile.BadZipFile:
        failures.append("bad_zip")
    return failures


def _check_required_headers(path: str, headers: List[str], data_tabs: Optional[List[str]] = None) -> List[str]:
    if not headers:
        return []
    try:
        import openpyxl
    except ImportError:
        return ["openpyxl_not_installed"]

    failures: List[str] = []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        tabs = data_tabs or wb.sheetnames
        for tab in tabs:
            if tab not in wb.sheetnames:
                continue
            ws = wb[tab]
            found: set[str] = set()
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                for c in row:
                    if c is not None:
                        found.add(str(c).strip().upper())
            for h in headers:
                if h.upper() not in found:
                    failures.append(f"missing_header:{tab}:{h}")
    finally:
        wb.close()
    return failures


def _check_semantic_cells(path: str, cells: List[Dict[str, Any]]) -> List[str]:
    if not cells:
        return []
    try:
        import openpyxl
    except ImportError:
        return ["openpyxl_not_installed"]

    failures: List[str] = []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for spec in cells:
            sheet = spec.get("sheet", "")
            cell = spec.get("cell", "")
            if not sheet or not cell:
                continue
            if sheet not in wb.sheetnames:
                failures.append(f"missing_sheet:{sheet}")
                continue
            val = wb[sheet][cell].value
            check = spec.get("check", "contains")
            expected = spec.get("contains") or spec.get("equals") or spec.get("expected")
            if check == "nonblank":
                if val is None or str(val).strip() == "":
                    failures.append(f"{sheet}!{cell} is blank")
            elif check == "equals":
                if str(val or "").strip() != str(expected or ""):
                    failures.append(f"{sheet}!{cell} expected {expected!r}, got {val!r}")
            else:
                if expected and str(expected).lower() not in str(val or "").lower():
                    failures.append(f"{sheet}!{cell} does not contain {expected!r}")
    finally:
        wb.close()
    return failures


def _extract_admin_totals(path: str) -> Dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        return {}
    out: Dict[str, Any] = {}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if "Executive Dashboard" in wb.sheetnames:
            ws = wb["Executive Dashboard"]
            for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                if not row or row[0] is None:
                    continue
                label = str(row[0]).strip()
                if label == "Total Net Hours" and len(row) > 1:
                    out["total_net"] = round(float(row[1] or 0), 2)
                if label == "Neuron Net" and len(row) > 1:
                    out["neuron_net"] = round(float(row[1] or 0), 2)
    finally:
        wb.close()
    return out


def _extract_bonita_totals(path: str, tabs: List[str]) -> Dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        return {}
    meta = {"CF Dictionary", "CF_Dictionary", "WebExcel QC", "Review Flags"}
    data_tabs = [t for t in tabs if t not in meta]
    out: Dict[str, Any] = {"per_tab_hours": {}, "grand_total_hours": 0.0}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        total = 0.0
        for tab in data_tabs:
            if tab not in wb.sheetnames:
                continue
            ws = wb[tab]
            col_total = None
            for c in range(1, (ws.max_column or 1) + 1):
                if str(ws.cell(row=1, column=c).value or "").strip().upper() == "TOTAL":
                    col_total = c
                    break
            if col_total is None:
                continue
            tab_sum = 0.0
            for row in ws.iter_rows(min_row=3, values_only=True):
                if len(row) < col_total:
                    continue
                v = row[col_total - 1]
                if isinstance(v, (int, float)):
                    tab_sum += float(v)
            tab_sum = round(tab_sum, 2)
            out["per_tab_hours"][tab] = tab_sum
            total += tab_sum
        out["grand_total_hours"] = round(total, 2)
    finally:
        wb.close()
    return out


def extract_profile_totals(path: str, profile: ArtifactProfile) -> Dict[str, Any]:
    tabs = _workbook_tabs(path)
    if profile.profile == "bonita_neuron_track_hours":
        return _extract_bonita_totals(path, tabs)
    return _extract_admin_totals(path)


def run_profile_checks(
    path: str,
    profile: ArtifactProfile,
    *,
    expect_neuron_tab: Optional[str] = None,
    reference_totals: Optional[Dict[str, Any]] = None,
) -> ProfileResult:
    """Run profile stop-ship rules on *path* (candidate or reference)."""
    res = ProfileResult()
    p = Path(path)
    if not p.is_file():
        res.failures.append("file_not_found")
        return res

    tabs = _workbook_tabs(str(p))
    required = resolve_profile_sheets(profile, expect_neuron_tab=expect_neuron_tab)
    for sheet in required:
        if sheet not in tabs:
            res.failures.append(f"missing_sheet:{sheet}")

    gate_name = gate_profile_for(profile.profile)
    gate = run_semantic_gate(str(p), profile=gate_name)
    if gate.get("semantic_integrity") != "PASS":
        for f in gate.get("sentinel_failures") or []:
            res.failures.append(f"semantic_gate:{f}")
    if gate.get("generic_column_strings_only"):
        res.failures.append("generic_column_strings_only")

    res.failures.extend(_check_forbidden_strings(str(p), profile.forbidden_values))
    if profile.required_headers:
        meta = {"CF Dictionary", "CF_Dictionary", "WebExcel QC", "Review Flags"}
        data_tabs = [t for t in tabs if t not in meta] if profile.profile == "bonita_neuron_track_hours" else None
        res.failures.extend(_check_required_headers(str(p), profile.required_headers, data_tabs))

    res.failures.extend(_check_semantic_cells(str(p), profile.required_semantic_cells))

    res.totals = extract_profile_totals(str(p), profile)
    if profile.compare_totals and reference_totals:
        ref = reference_totals
        cur = res.totals
        for key in ("total_net", "neuron_net", "grand_total_hours"):
            if key in ref and key in cur and ref[key] != cur[key]:
                res.failures.append(f"total_mismatch:{key}:{ref[key]}!={cur[key]}")
        ref_tabs = ref.get("per_tab_hours") or {}
        cur_tabs = cur.get("per_tab_hours") or {}
        for tab, rv in ref_tabs.items():
            cv = cur_tabs.get(tab)
            if cv is not None and rv != cv:
                res.failures.append(f"tab_hours_mismatch:{tab}:{rv}!={cv}")

    return res
