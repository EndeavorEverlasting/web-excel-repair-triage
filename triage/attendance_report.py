"""
triage/attendance_report.py
---------------------------
Build a weekly attendance .xlsx report from parsed roster records.

Output: billing_runs/YYYY-MM/attendance/attendance_week_{YYYY-MM-DD}.xlsx

Columns:
  Staff Name | Project | Mon CI | Mon CO | Mon Net | ... Fri CI | Fri CO | Fri Net | Weekly Total
"""
from __future__ import annotations

import json
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional


WEEKDAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def _hours_str(h: Optional[float]) -> str:
    if h is None:
        return ""
    hh = int(h)
    mm = int(round((h - hh) * 60))
    return f"{hh}:{mm:02d}"


def _fmt_time(decimal_hours: Optional[float]) -> str:
    """Format a decimal time (e.g. 8.5 → '08:30')."""
    if decimal_hours is None:
        return ""
    hh = int(decimal_hours)
    mm = int(round((decimal_hours - hh) * 60))
    return f"{hh:02d}:{mm:02d}"


def _lunch_deduction(gross: float) -> float:
    if gross >= 8.0:
        return 1.0
    if gross >= 6.0:
        return 0.5
    return 0.0


def _make_overnight_comment(rec: Dict[str, Any], Comment):
    """Build an openpyxl Comment for an overnight-shift cell."""
    ci_s = _fmt_time(rec.get("clock_in"))
    co_s = _fmt_time(rec.get("clock_out"))
    text = (
        f"Overnight shift\n"
        f"Clock-in:  {ci_s}\n"
        f"Clock-out: {co_s}\n"
        f"Gross hrs: {rec.get('gross_hours', 0):.2f}  — review recommended"
    )
    return Comment(text, "Attendance Report")


def _make_long_shift_comment(rec: Dict[str, Any], Comment):
    """Build an openpyxl Comment for a suspiciously long shift cell."""
    ci_s = _fmt_time(rec.get("clock_in"))
    co_s = _fmt_time(rec.get("clock_out"))
    text = (
        f"Long shift - possible data error\n"
        f"Clock-in:  {ci_s}\n"
        f"Clock-out: {co_s}\n"
        f"Gross hrs: {rec.get('gross_hours', 0):.2f}\n"
        f"Review the source roster before billing."
    )
    return Comment(text, "Attendance Report")


def generate_attendance_report(
    records: List[Dict[str, Any]],
    week_start: date,
    week_end: date,
    out_root: str = "billing_runs",
    run_id: Optional[str] = None,
    input_paths: Optional[List[str]] = None,
) -> str:
    """
    Generate an .xlsx attendance report for the given week.

    Parameters
    ----------
    records    : output of roster_parser.parse_roster()
    week_start : Monday of the target week
    week_end   : Friday (or end) of the target week
    out_root   : root folder for output (default 'billing_runs')
    run_id     : optional identifier for the manifest

    Returns
    -------
    Absolute path to the generated .xlsx file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.comments import Comment
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    # Filter records to this week
    week_records = [r for r in records if week_start <= r["date"] <= week_end]

    # Build: {(staff, project) → {date → record}}
    staff_project_map: Dict[tuple, Dict[date, Dict]] = {}
    for rec in week_records:
        key = (rec["staff"], rec["project"])
        staff_project_map.setdefault(key, {})[rec["date"]] = rec

    # Output path
    month_str = week_start.strftime("%Y-%m")
    out_dir   = Path(out_root) / month_str / "attendance"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path  = out_dir / f"attendance_week_{week_start.isoformat()}.xlsx"

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Week {week_start.strftime('%m-%d')}"

    # Styles
    hdr_fill       = PatternFill("solid", fgColor="1A5C38")
    date_fill      = PatternFill("solid", fgColor="0D3320")
    overnight_fill = PatternFill("solid", fgColor="7B3F00")  # dark amber — overnight rows
    long_shift_fill = PatternFill("solid", fgColor="7F1D1D")  # dark red — suspicious duration
    bold_font = Font(bold=True, color="FFFFFF", size=11)
    bold_dark = Font(bold=True, color="FFFFFF", size=10)
    normal    = Font(size=10)
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center")
    thin      = Side(style="thin", color="444444")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title
    num_days = (week_end - week_start).days + 1
    total_cols = 2 + num_days * 3 + 1  # Staff | Project | (CI+CO+Net)*days | Weekly Total

    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = (
        f"Weekly Attendance Report — Week of {week_start.strftime('%B %d, %Y')} "
        f"to {week_end.strftime('%B %d, %Y')}"
    )
    title_cell.font  = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill  = PatternFill("solid", fgColor="0E4C2F")
    title_cell.alignment = center

    # Row 2: Day headers (merged over 3 sub-columns each)
    ws.row_dimensions[2].height = 20
    ws.cell(row=2, column=1, value="Staff Name").font  = bold_font
    ws.cell(row=2, column=1).fill      = hdr_fill
    ws.cell(row=2, column=1).alignment = center
    ws.cell(row=2, column=2, value="Project").font     = bold_font
    ws.cell(row=2, column=2).fill      = hdr_fill
    ws.cell(row=2, column=2).alignment = center

    day_dates: List[date] = []
    col = 3
    for d in range(num_days):
        day_date = week_start + timedelta(days=d)
        day_dates.append(day_date)
        label = day_date.strftime("%a %m/%d")
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        cell = ws.cell(row=2, column=col, value=label)
        cell.font      = bold_font
        cell.fill      = date_fill
        cell.alignment = center
        col += 3

    ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
    total_hdr = ws.cell(row=2, column=col, value="Weekly Total\nNet Hrs")
    total_hdr.font      = bold_font
    total_hdr.fill      = hdr_fill
    total_hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row 3: Sub-headers (CI / CO / Net)
    ws.row_dimensions[3].height = 16
    ws.cell(row=3, column=1).fill = hdr_fill
    ws.cell(row=3, column=2).fill = hdr_fill
    sub_col = 3
    for _ in day_dates:
        for lbl in ("In", "Out", "Net"):
            c = ws.cell(row=3, column=sub_col, value=lbl)
            c.font      = bold_dark
            c.fill      = hdr_fill
            c.alignment = center
            sub_col += 1

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 22
    ws.column_dimensions[get_column_letter(2)].width = 18
    for c in range(3, total_cols):
        ws.column_dimensions[get_column_letter(c)].width = 8
    ws.column_dimensions[get_column_letter(total_cols)].width = 12

    # Data rows
    data_row = 4
    alt_fill = PatternFill("solid", fgColor="0A1A10")

    for (staff, project), day_map in sorted(staff_project_map.items()):
        row_fill = alt_fill if data_row % 2 == 0 else None
        weekly_net = 0.0

        r_staff = ws.cell(row=data_row, column=1, value=staff)
        r_staff.font      = Font(size=10, bold=True)
        r_staff.alignment = left
        if row_fill:
            r_staff.fill = row_fill

        r_proj = ws.cell(row=data_row, column=2, value=project)
        r_proj.font      = normal
        r_proj.alignment = left
        if row_fill:
            r_proj.fill = row_fill

        d_col = 3
        for day_date in day_dates:
            rec = day_map.get(day_date)
            if rec:
                ci         = _fmt_time(rec["clock_in"])
                co         = _fmt_time(rec["clock_out"])
                net        = rec["net_hours"]
                weekly_net += net
                net_s      = f"{net:.2f}"
                ci_h       = rec.get("clock_in")
                co_h       = rec.get("clock_out")
                is_night   = (ci_h is not None and co_h is not None and co_h < ci_h)
                is_long    = bool(rec.get("long_shift"))
            else:
                ci = co = net_s = ""
                is_night = False
                is_long = False

            for val in (ci, co, net_s):
                c = ws.cell(row=data_row, column=d_col, value=val)
                c.font      = normal
                c.alignment = center
                if is_long:
                    c.fill    = long_shift_fill
                    c.comment = _make_long_shift_comment(rec, Comment)
                elif is_night:
                    c.fill    = overnight_fill
                    c.comment = _make_overnight_comment(rec, Comment)
                elif row_fill:
                    c.fill = row_fill
                d_col += 1

        total_cell = ws.cell(row=data_row, column=total_cols, value=round(weekly_net, 2))
        total_cell.font      = Font(size=10, bold=True)
        total_cell.alignment = center
        if row_fill:
            total_cell.fill = row_fill

        data_row += 1

    # Totals row
    if data_row > 4:
        ws.row_dimensions[data_row].height = 18
        tot_label = ws.cell(row=data_row, column=1, value="TOTALS")
        tot_label.font      = Font(bold=True, color="FFFFFF", size=10)
        tot_label.fill      = hdr_fill
        tot_label.alignment = left

        grand_total = sum(
            sum(r["net_hours"] for r in day_map.values())
            for day_map in staff_project_map.values()
        )
        gt_cell = ws.cell(row=data_row, column=total_cols, value=round(grand_total, 2))
        gt_cell.font      = Font(bold=True, color="FFFFFF", size=10)
        gt_cell.fill      = hdr_fill
        gt_cell.alignment = center

    # Freeze panes
    ws.freeze_panes = "C4"

    wb.save(str(out_path))

    import time as _ts
    _write_manifest(
        out_root   = out_root,
        month_str  = month_str,
        run_id     = run_id or f"attendance-{week_start.isoformat()}-{int(_ts.time())}",
        inputs     = input_paths or [],
        outputs    = [str(out_path)],
        status     = "generated",
        meta       = {
            "week_start": week_start.isoformat(),
            "week_end":   week_end.isoformat(),
            "staff_count": len(staff_project_map),
            "record_count": len(week_records),
        },
    )

    return str(out_path)


def update_attendance_manifest_status(
    out_root: str,
    month_str: str,
    status: str,
    gate_status: str = "",
    failures: Optional[List[str]] = None,
    warnings: Optional[List[str]] = None,
) -> None:
    """Update existing attendance manifest after gate validation."""
    import time as _time, json as _json
    manifest_path = Path(out_root) / month_str / "run_manifest.json"
    if not manifest_path.exists():
        return
    try:
        data = _json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:
        data = {}
    data["status"]        = status
    data["gate_status"]   = gate_status
    data["gate_failures"] = failures or []
    data["gate_warnings"] = warnings or []
    data["validated_at"]  = _time.strftime("%Y-%m-%dT%H:%M:%SZ")
    manifest_path.write_text(_json.dumps(data, indent=2, default=str), encoding="utf-8")


def _write_manifest(
    out_root: str,
    month_str: str,
    run_id: str,
    inputs: List[str],
    outputs: List[str],
    status: str,
    meta: Dict[str, Any],
) -> None:
    import time
    run_dir = Path(out_root) / month_str
    run_dir.mkdir(parents=True, exist_ok=True)
    manifest = {
        "run_id":    run_id,
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "status":    status,
        "inputs":    inputs,
        "outputs":   outputs,
        **meta,
    }
    (run_dir / "run_manifest.json").write_text(
        json.dumps(manifest, indent=2, default=str), encoding="utf-8"
    )
