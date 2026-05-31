from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from triage.admin_billing_context_rules import contains_suspicious_language

from .models import Mismatch, WorkEntry
from .reconcile import friday_batch_key

HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"
ERROR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A")
FORBIDDEN_TEXT_TOKENS = ("...", "…", "TBD", "TODO", "context goes here")

LEADERSHIP_HEADERS = ["Date", "Tech", "Hours", "Work Context"]

INTERNAL_HEADERS = [
    "Date",
    "Tech",
    "Hours",
    "Work Context",
    "Context Reason",
    "Original Assignment",
    "Start Time",
    "End Time",
    "Source Sheet",
    "Source Row",
]


def safe_csv_value(value: object) -> str:
    text = "" if value is None else str(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def autosize(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 70))
        ws.column_dimensions[letter].width = max_len + 2


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.alignment = Alignment(horizontal="center")


def write_rows(ws, headers: list[str], rows: list[dict]) -> None:
    ws.append(headers)
    style_header(ws)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    autosize(ws)


def entries_to_leadership_rows(entries: list[WorkEntry]) -> list[dict]:
    return [
        {
            "Date": e.work_date.isoformat(),
            "Tech": e.tech,
            "Hours": round(e.hours, 2),
            "Work Context": e.work_context,
        }
        for e in entries
    ]


def entries_to_internal_rows(entries: list[WorkEntry]) -> list[dict]:
    rows = []
    for e in entries:
        assignment = e.original_assignment
        if is_placeholder_empty(assignment):
            assignment = ""
        rows.append(
            {
                "Date": e.work_date.isoformat(),
                "Tech": e.tech,
                "Hours": round(e.hours, 2),
                "Work Context": e.work_context,
                "Context Reason": e.context_reason,
                "Original Assignment": assignment,
                "Start Time": e.start_time.isoformat() if e.start_time else "",
                "End Time": e.end_time.isoformat() if e.end_time else "",
                "Source Sheet": e.sheet_name,
                "Source Row": e.row_number,
            }
        )
    return rows


def is_placeholder_empty(assignment: str) -> bool:
    from .context_rules import is_placeholder_assignment

    return not assignment or is_placeholder_assignment(assignment)


def summarize_by_context(entries: list[WorkEntry]) -> list[dict]:
    totals: dict[str, float] = defaultdict(float)
    for e in entries:
        totals[e.work_context] += e.hours
    return [
        {"Work Context": k, "Hours": round(v, 2)}
        for k, v in sorted(totals.items(), key=lambda item: item[1], reverse=True)
    ]


def summarize_by_tech(entries: list[WorkEntry]) -> list[dict]:
    totals: dict[str, float] = defaultdict(float)
    for e in entries:
        totals[e.tech] += e.hours
    return [
        {"Tech": k, "Hours": round(v, 2)}
        for k, v in sorted(totals.items(), key=lambda item: item[1], reverse=True)
    ]


def summarize_by_batch(entries: list[WorkEntry]) -> list[dict]:
    totals: dict[str, float] = defaultdict(float)
    for e in entries:
        totals[friday_batch_key(e.work_date)] += e.hours
    return [
        {"Reporting Batch Friday": k, "Hours": round(v, 2)}
        for k, v in sorted(totals.items())
    ]


def add_bar_chart(ws, title: str, category_col: int, value_col: int, anchor: str) -> None:
    if ws.max_row < 2:
        return
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "Hours"
    chart.x_axis.title = ws.cell(1, category_col).value or "Category"
    data = Reference(ws, min_col=value_col, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=category_col, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 14
    ws.add_chart(chart, anchor)


def add_line_chart(ws, title: str, category_col: int, value_col: int, anchor: str) -> None:
    if ws.max_row < 2:
        return
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Hours"
    chart.x_axis.title = "Batch"
    data = Reference(ws, min_col=value_col, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=category_col, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 14
    ws.add_chart(chart, anchor)


def _scan_cells(path: str | Path, data_only: bool) -> list[tuple[str, str, str]]:
    hits: list[tuple[str, str, str]] = []
    wb = load_workbook(path, data_only=data_only)
    for ws in wb.worksheets:
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value or "")
                if any(tok in val for tok in ERROR_TOKENS):
                    hits.append((ws.title, cell.coordinate, val))
    wb.close()
    return hits


def scan_workbook_errors(path: str | Path) -> list[tuple[str, str, str]]:
    hits = _scan_cells(path, data_only=True)
    if not hits:
        hits = _scan_cells(path, data_only=False)
    return hits


def scan_forbidden_text(path: str | Path) -> list[str]:
    issues: list[str] = []
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue
        for row in ws.iter_rows():
            for cell in row:
                value = str(cell.value or "")
                lower = value.lower()
                for token in FORBIDDEN_TEXT_TOKENS:
                    if token.lower() in lower:
                        issues.append(f"{ws.title}!{cell.coordinate}: forbidden token {token!r}")
    wb.close()
    return issues


def scan_leadership_language(path: str | Path) -> list[str]:
    issues: list[str] = []
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if contains_suspicious_language(cell.value):
                    issues.append(f"{ws.title}!{cell.coordinate}: {cell.value}")
    wb.close()
    return issues


def export_neuron_project_hours(entries: list[WorkEntry], out_path: str) -> str:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for month, title in [(4, "April 2026"), (5, "May 2026")]:
        month_entries = [e for e in entries if e.work_date.month == month]

        ws = wb.create_sheet(title)
        write_rows(ws, LEADERSHIP_HEADERS, entries_to_leadership_rows(month_entries))

        context_ws = wb.create_sheet(f"{title} Context Summary")
        write_rows(context_ws, ["Work Context", "Hours"], summarize_by_context(month_entries))
        add_bar_chart(context_ws, f"{title} Hours by Work Context", 1, 2, "D2")

        tech_ws = wb.create_sheet(f"{title} Tech Summary")
        write_rows(tech_ws, ["Tech", "Hours"], summarize_by_tech(month_entries))
        add_bar_chart(tech_ws, f"{title} Top Technician Hours", 1, 2, "D2")

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def export_monthly_summary(
    entries: list[WorkEntry],
    month: int,
    out_path: str,
    *,
    include_tracker_import: bool = False,
) -> str:
    month_labels = {4: "April 2026", 5: "May 2026"}
    if month not in month_labels:
        raise ValueError(f"Unsupported month for billing summary: {month}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Admin Summary"

    month_entries = [e for e in entries if e.work_date.month == month]
    total_hours = round(sum(e.hours for e in month_entries), 2)

    ws["A1"] = "Billing Summary"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "Month"
    ws["B3"] = month_labels[month]
    ws["A4"] = "Total Hours"
    ws["B4"] = total_hours
    ws["A5"] = "Total Rows"
    ws["B5"] = len(month_entries)

    context_ws = wb.create_sheet("Work Context Summary")
    write_rows(context_ws, ["Work Context", "Hours"], summarize_by_context(month_entries))
    add_bar_chart(context_ws, "Hours by Work Context", 1, 2, "D2")

    tech_ws = wb.create_sheet("Technician Summary")
    write_rows(tech_ws, ["Tech", "Hours"], summarize_by_tech(month_entries))
    add_bar_chart(tech_ws, "Top Technician Hours", 1, 2, "D2")

    batch_ws = wb.create_sheet("Reporting Batch Summary")
    write_rows(batch_ws, ["Reporting Batch Friday", "Hours"], summarize_by_batch(month_entries))
    add_line_chart(batch_ws, "Hours by Reporting Batch", 1, 2, "D2")

    if include_tracker_import:
        detail_ws = wb.create_sheet("Tracker Import")
        write_rows(detail_ws, INTERNAL_HEADERS, entries_to_internal_rows(month_entries))

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def export_internal_detail(entries: list[WorkEntry], out_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Internal Detail"
    write_rows(ws, INTERNAL_HEADERS, entries_to_internal_rows(entries))
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def export_mismatches(mismatches: list[Mismatch], out_json: str, out_csv: str) -> None:
    rows = [m.to_dict() for m in mismatches]
    Path(out_json).parent.mkdir(parents=True, exist_ok=True)
    Path(out_csv).parent.mkdir(parents=True, exist_ok=True)

    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2)

    fieldnames = [
        "severity",
        "mismatch_type",
        "tech",
        "work_date",
        "source_a",
        "source_b",
        "source_a_value",
        "source_b_value",
        "recommendation",
        "leadership_safe",
    ]
    with open(out_csv, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: safe_csv_value(row.get(k, "")) for k in fieldnames})


def build_output_manifest(outputs: dict[str, str]) -> list[dict]:
    manifest: list[dict] = []
    for name, path in outputs.items():
        p = Path(path)
        manifest.append(
            {
                "name": name,
                "path": str(p.resolve()),
                "exists": p.exists(),
                "bytes": p.stat().st_size if p.exists() else 0,
            }
        )
    return manifest


def create_zip_bundle(paths: list[str], zip_path: str) -> str:
    import zipfile

    outp = Path(zip_path)
    outp.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(outp, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in paths:
            p = Path(path)
            if p.exists():
                zf.write(p, p.name)
    return str(outp)
