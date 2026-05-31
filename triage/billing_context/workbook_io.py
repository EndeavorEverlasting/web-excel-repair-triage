from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def load_xlsx(path: str | Path, data_only: bool = True):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Workbook not found: {p}")
    return load_workbook(p, data_only=data_only)


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\n", " ").replace("\r", " ")


def header_map(ws: Worksheet, header_row: int = 1) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[header_row]:
        name = normalize_header(cell.value)
        if name:
            headers[name] = cell.column
    return headers


_HEADER_HINTS: tuple[str, ...] = (
    "Tech", "Technician", "Name", "Resource", "Staff", "Employee",
    "Date", "Work Date", "Day",
    "Hours", "Total Hours", "Net Hours", "Total", "Duration",
    "In", "Out", "Start", "End", "Start Time", "End Time", "Clock In", "Clock Out",
    "Assignment", "Assignment Type", "Project", "Project Name", "Task", "Work Context",
    "Work / Context",
)


def auto_detect_header_row(ws: Worksheet, max_rows: int = 10) -> int | None:
    """Scan the first *max_rows* rows for the one with the most recognizable header keywords.

    Returns 1-based row index, or None if no recognizable headers found.
    """
    best_row: int = 1
    best_score: int = 0
    for r in range(1, min(max_rows, ws.max_row) + 1):
        score = 0
        for cell in ws[r]:
            norm = normalize_header(cell.value)
            for hint in _HEADER_HINTS:
                if hint.lower() == norm.lower():
                    score += 1
                    break
        if score > best_score:
            best_score = score
            best_row = r
    return best_row if best_score >= 2 else None


def iter_dict_rows(ws: Worksheet, header_row: int | None = None) -> Iterable[dict[str, Any]]:
    if header_row is None:
        detected = auto_detect_header_row(ws)
        header_row = detected if detected is not None else 1
    headers = header_map(ws, header_row)
    reverse = {col: name for name, col in headers.items()}

    for r in range(header_row + 1, ws.max_row + 1):
        row: dict[str, Any] = {}
        empty = True
        for col, name in reverse.items():
            value = ws.cell(r, col).value
            if value not in (None, ""):
                empty = False
            row[name] = value
        if not empty:
            row["_row_number"] = r
            yield row


def fuzzy_get(row: dict[str, Any], *candidates: str, default: Any = None) -> Any:
    """Case-insensitive key lookup: return first matching value from *row*."""
    lower_keys = {k.lower(): k for k in row if not k.startswith("_")}
    for cand in candidates:
        key = lower_keys.get(cand.lower())
        if key is not None:
            val = row[key]
            if val not in (None, ""):
                return val
    return default


def safe_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return default


def sheet_names(path: str | Path) -> list[str]:
    wb = load_xlsx(path)
    names = list(wb.sheetnames)
    wb.close()
    return names
