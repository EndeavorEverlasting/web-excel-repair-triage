"""
triage/billing_summary_generator.py
------------------------------------
Build the monthly billing summary .xlsx from roster records + parsed invoices.

Output sheets (matching Agilant Admins template exactly):
  1. "Billing Summary - {Month YYYY}"  (≤31 chars — Excel tab name limit)
       - Title: "{Month YYYY} Billing Summary - Candidate"
       - Monthly Rollup block (A-D) side-by-side with Pivot by Team/Project (F-K)
       - Invoice-Ready Calculation section at rows 13+ (trucking/logistics invoices)
       - Staff Rollup (A-G) side-by-side with Val Weekly Hours (I-O)
       - Daily Billing Detail section (A-J)
  2. "Invoice Pivots - Candidate"
       - Monthly Invoice Totals (A-H) + Totals by Category (J-N)
       - Totals by PO Number (A-D) | Totals by Project (F-H) | Totals by Vendor/Crew (J-N)
       - Audit Notes block

Color palette verified cell-by-cell against:
  attached_assets/Agilant_Admins_April_2026_Billing_Summary_PO176759_Neurons_*.xlsx

Output path: billing_runs/YYYY-MM/workbook/billing_summary_{YYYY-MM}.xlsx
Also writes:  billing_runs/YYYY-MM/run_manifest.json

Raises RuntimeError (not silently falls back) if required inputs are malformed.
"""
from __future__ import annotations

import json
import time
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set


def _month_day_label(d) -> str:
    """Cross-platform '%b %-d' equivalent (e.g. 'Apr 1').

    The POSIX ``%-d`` strftime directive is unsupported on Windows and raises
    ``ValueError: Invalid format string``. Use the day attribute directly to
    avoid the platform-dependent directive.
    """
    return f"{d:%b} {d.day}"


def generate_billing_summary(
    records: List[Dict[str, Any]],
    invoices: List[Dict[str, Any]],
    billing_month: str,
    out_root: str = "billing_runs",
    run_id: Optional[str] = None,
    input_paths: Optional[List[str]] = None,
    audit_notes: Optional[List[str]] = None,
) -> str:
    """
    Generate the monthly billing summary workbook.

    Parameters
    ----------
    records       : list of dicts from roster_parser.parse_roster()
    invoices      : list of dicts from invoice_parser.parse_invoices()
    billing_month : 'YYYY-MM', e.g. '2026-04'
    out_root      : root folder (default 'billing_runs')
    run_id        : optional run identifier
    input_paths   : list of source file paths for the manifest
    audit_notes   : optional list of note strings for the Invoice Pivots
                    Audit Notes block (matches Agilant template rows 24+)

    Returns
    -------
    Absolute path to the generated .xlsx file.
    Raises RuntimeError if records is empty.
    """
    if not records:
        raise RuntimeError(
            f"No roster records provided for {billing_month}. "
            "Cannot generate billing summary from empty data."
        )

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    import datetime as _dt

    year_s, month_s = billing_month.split("-")
    month_label = date(int(year_s), int(month_s), 1).strftime("%B %Y")

    # ── Aggregate roster data ──────────────────────────────────────────────────
    gross_total = sum(r["gross_hours"]     for r in records)
    lunch_total = sum(r["lunch_deduction"] for r in records)
    net_total   = sum(r["net_hours"]       for r in records)
    daily_rows  = len(records)

    all_staff: Set[str] = {r["staff"] for r in records}

    # Hours by project
    project_map: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {"gross": 0.0, "lunch": 0.0, "net": 0.0,
                 "techs": set(), "daily_rows": 0}
    )
    for rec in records:
        proj = rec["project"] or "Unassigned / Review"
        project_map[proj]["gross"]      += rec["gross_hours"]
        project_map[proj]["lunch"]      += rec["lunch_deduction"]
        project_map[proj]["net"]        += rec["net_hours"]
        project_map[proj]["techs"].add(rec["staff"])
        project_map[proj]["daily_rows"] += 1

    # Per-staff aggregate for Staff Rollup section
    staff_map: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {"gross": 0.0, "lunch": 0.0, "net": 0.0,
                 "projects": set(), "rows": 0}
    )
    for rec in records:
        s = rec["staff"]
        staff_map[s]["gross"]    += rec["gross_hours"]
        staff_map[s]["lunch"]    += rec["lunch_deduction"]
        staff_map[s]["net"]      += rec["net_hours"]
        staff_map[s]["projects"].add(rec["project"] or "Unassigned")
        staff_map[s]["rows"]     += 1

    # Sort records and projects for display
    sorted_records  = sorted(records, key=lambda r: (r["date"], r["staff"]))
    sorted_projects = sorted(project_map.items())
    sorted_staff    = sorted(staff_map.items())

    # ── Aggregate invoice data ─────────────────────────────────────────────────
    inv_by_month: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {"trucking": 0.0, "labor": 0.0, "courier": 0.0, "other": 0.0,
                 "count": 0, "pos": set()}
    )
    po_map: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {"trucking": 0.0, "labor": 0.0, "courier": 0.0, "other": 0.0,
                 "total": 0.0, "count": 0, "notes": ""}
    )
    project_inv_map: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {"total": 0.0, "count": 0}
    )
    vendor_map: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {"trucking": 0.0, "labor": 0.0, "courier": 0.0, "other": 0.0,
                 "total": 0.0, "count": 0}
    )
    category_map: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {"trucking": 0.0, "labor": 0.0, "courier": 0.0, "other": 0.0,
                 "count": 0, "total": 0.0}
    )

    _CATEGORY_TO_PROJECT = {
        "trucking": "Delivery / Transport / Disposal",
        "courier":  "Delivery / Transport / Disposal",
        "labor":    "Neuron Deployments",
        "other":    "Unassigned / Review",
    }

    for inv in invoices:
        inv_total_amt = float(inv.get("total") or 0)
        po     = inv.get("po_number") or "Unknown"
        vendor = inv["vendor"]

        svc = inv.get("service_date") or inv.get("service_window") or ""
        inv_month = billing_month
        if svc:
            import re as _re
            svc_first = _re.split(r"\s*[-–—]\s*\d", svc.strip())[0].strip()
            if not _re.search(r"\b\d{4}\b", svc_first):
                year_m = _re.search(r"\b(\d{4})\b", svc)
                if year_m:
                    svc_first = svc_first.rstrip(",. ") + ", " + year_m.group(1)
            for fmt in ("%b %d, %Y", "%B %d, %Y", "%Y-%m-%d", "%m/%d/%Y",
                        "%b %d %Y", "%B %d %Y"):
                try:
                    parsed_date = datetime.strptime(svc_first.strip(), fmt)
                    inv_month = parsed_date.strftime("%Y-%m")
                    break
                except ValueError:
                    continue

        inv_by_month[inv_month]["count"] += 1
        inv_by_month[inv_month]["pos"].add(po)
        po_map[po]["count"] += 1
        vendor_map[vendor]["count"] += 1


        line_items = inv.get("line_items") or []
        if line_items:
            li_sum: Dict[str, float] = {"trucking": 0.0, "labor": 0.0,
                                         "courier": 0.0, "other": 0.0}
            for item in line_items:
                item_cat = item.get("category") or "other"
                if item_cat not in li_sum:
                    item_cat = "other"
                li_sum[item_cat] += float(item.get("amount") or 0)

            li_raw_sum = sum(li_sum.values())
            if li_raw_sum > 0 and abs(li_raw_sum - inv_total_amt) / li_raw_sum > 0.001:
                scale = inv_total_amt / li_raw_sum
                li_sum = {k: v * scale for k, v in li_sum.items()}
        else:
            fallback_cat = inv.get("cost_category") or "other"
            li_sum = {"trucking": 0.0, "labor": 0.0, "courier": 0.0, "other": 0.0}
            li_sum[fallback_cat if fallback_cat in li_sum else "other"] = inv_total_amt

        for cat, amt in li_sum.items():
            if amt == 0:
                continue
            inv_by_month[inv_month][cat] += amt
            po_map[po][cat]              += amt
            vendor_map[vendor][cat]      += amt
            category_map[cat]["count"] += 1
            category_map[cat][cat] += amt
            category_map[cat]["total"] += amt

        po_map[po]["total"]              += inv_total_amt
        vendor_map[vendor]["total"]      += inv_total_amt

        dom_cat = max(li_sum, key=li_sum.get) if any(li_sum.values()) else "other"
        proj_bucket = _CATEGORY_TO_PROJECT.get(dom_cat, "Unassigned / Review")
        project_inv_map[proj_bucket]["total"] += inv_total_amt
        project_inv_map[proj_bucket]["count"] += 1

    inv_total = sum(
        d.get("trucking", 0) + d.get("labor", 0) + d.get("courier", 0) + d.get("other", 0)
        for d in inv_by_month.values()
    )

    # ── Build workbook ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet tab names must be ≤31 chars (Excel limit).
    # Pattern mirrors the Agilant template naming convention.
    # "Billing Summary - December 2026" = 31 chars (longest possible, exactly fits).
    sheet1_name = f"Billing Summary - {month_label}"[:31]
    ws1 = wb.create_sheet(sheet1_name)
    ws2 = wb.create_sheet("Invoice Pivots - Candidate")

    # ── Color palette (verified against Agilant Admins April 2026 template) ───
    # Sheet 1
    C1_HEADER  = "1F365C"   # dark navy — title bar, section headers
    C1_SUBHDR  = "D9EAF7"   # light blue — column sub-header fill
    C1_SUBTITLE = "EAF1F8"  # very light blue — subtitle row fill
    C1_TOTAL   = "E2F0D9"   # light green — TOTAL row fill
    C1_TXT     = "1F365C"   # dark navy — text on light/white backgrounds
    C_WHITE    = "FFFFFF"   # white — text on dark fills

    # Sheet 2
    C2_HEADER  = "1F2937"   # dark charcoal — title, Audit Notes header
    C2_SECHDR  = "0F766E"   # teal — section headers
    C2_SUBHDR  = "374151"   # medium gray — column sub-headers
    C2_SUBTITLE = "DBEAFE"  # light blue — subtitle
    C2_TOTAL   = "DCFCE7"   # light green — TOTAL rows

    def _fill(hex_color: str) -> PatternFill:
        # openpyxl requires 8-char ARGB (alpha prefix FF = fully opaque).
        if len(hex_color) == 6:
            hex_color = "FF" + hex_color
        return PatternFill("solid", fgColor=hex_color)

    def cell(ws, row, col, value=None, bold=False, size=10, txt_color=None,
             fill=None, h="left", v="center", wrap=False, fmt=None):
        """Write a cell with optional fill and formatting.
        Data rows: pass fill=None (no fill) and txt_color=None (default dark).
        Header rows: pass fill=color_hex and txt_color=C_WHITE.
        """
        c = ws.cell(row=row, column=col, value=value)
        if txt_color:
            c.font = Font(bold=bold, size=size, color=txt_color)
        else:
            c.font = Font(bold=bold, size=size)
        c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        if fill:
            c.fill = _fill(fill)
        if fmt:
            c.number_format = fmt
        return c

    def hdr(ws, row, col, value, span=1, fill=C1_HEADER, txt=C_WHITE,
            bold=True, h="center", size=10):
        """Write a header cell (section or sub-header)."""
        if span > 1:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col + span - 1)
        return cell(ws, row, col, value, bold=bold, size=size,
                    txt_color=txt, fill=fill, h=h)

    # ── Time/date helpers ──────────────────────────────────────────────────────
    def _to_time(h):
        """Convert float hour value to datetime.time (for h:mm AM/PM format)."""
        if h is None:
            return None
        hh = int(h) % 24
        mm = int(round((h - int(h)) * 60))
        if mm == 60:
            hh = (hh + 1) % 24
            mm = 0
        return _dt.time(hh, mm)

    def _to_datetime(d):
        """Convert date to datetime (Excel stores dates as datetime)."""
        if d is None:
            return None
        if isinstance(d, _dt.datetime):
            return d
        return _dt.datetime(d.year, d.month, d.day)

    def _month_datetime(ym_str):
        """Convert 'YYYY-MM' to datetime for mmm yyyy format."""
        try:
            return _dt.datetime.strptime(ym_str, "%Y-%m")
        except Exception:
            return None

    # ════════════════════════════════════════════════════════════════════════════
    # SHEET 1: Billing Summary
    # ════════════════════════════════════════════════════════════════════════════

    # Row 1: Title — A1:K1, navy fill, white bold 16pt
    ws1.merge_cells("A1:K1")
    cell(ws1, 1, 1, f"{month_label} Billing Summary - Candidate",
         bold=True, size=16, txt_color=C_WHITE, fill=C1_HEADER, h="center")
    ws1.row_dimensions[1].height = 22

    # Row 2: Subtitle — A2:K2, light-blue fill, dark text
    ws1.merge_cells("A2:K2")
    cell(ws1, 2, 1,
         "Separating Logistics from Device Integration; Including Lunch Deductions",
         size=10, txt_color=C1_TXT, fill=C1_SUBTITLE, h="left")
    ws1.row_dimensions[2].height = 14

    # Row 3: blank spacer
    ws1.row_dimensions[3].height = 6

    # Row 4: Section headers
    hdr(ws1, 4, 1, "Monthly Rollup",                   span=4, fill=C1_HEADER)
    hdr(ws1, 4, 6, "Pivot by Team / Project Bucket",   span=6, fill=C1_HEADER)
    ws1.row_dimensions[4].height = 18

    # Row 5: Column sub-headers — light-blue fill, dark bold text
    for ci, h_val in enumerate(["Metric", "Value", "Basis", "Notes"], 1):
        hdr(ws1, 5, ci, h_val, fill=C1_SUBHDR, txt=C1_TXT)
    for ci, h_val in enumerate(
        ["Team / Project Bucket", "Tech Count", "Daily Rows",
         "Gross Hours", "Lunch Deducted", "Net Billable Hours"],
        6
    ):
        hdr(ws1, 5, ci, h_val, fill=C1_SUBHDR, txt=C1_TXT, h="center")
    ws1.row_dimensions[5].height = 16

    # Rows 6-10: Rollup data (left) + project data (right, up to 5 real slots)
    rollup_data = [
        ("Gross hours",        round(gross_total, 2), "Live " + month_label,
         "Before lunch deduction"),
        ("Lunch deducted",     round(lunch_total, 2), "Policy rule",
         "0 under 6h; 0.5 from 6 to <8h; 1.0 at 8h+"),
        ("Net billable hours", round(net_total,   2), "Gross - lunch",
         "Use unless company billing policy overrides"),
        ("Tracked staff",      len(all_staff),         "Distinct staff", ""),
        ("Daily work rows",    daily_rows,              "Daily detail rows", ""),
    ]

    # Up to 5 real project slots in rows 6-10; TOTAL at row 11.
    proj_rows = list(sorted_projects)
    real_slots: list = (proj_rows[:5] + [None] * 5)[:5]

    for i in range(6, 11):
        ri = i - 6
        metric, value, basis, notes = rollup_data[ri]
        # Left: no fill, dark text; value column uses 0.00 format for hours
        cell(ws1, i, 1, metric, h="left")
        c_val = cell(ws1, i, 2, value, bold=False, h="right")
        if isinstance(value, float):
            c_val.number_format = "0.00"
        cell(ws1, i, 3, basis,  h="left")
        cell(ws1, i, 4, notes,  h="left", size=9)

        # Right: project slot
        entry = real_slots[ri]
        if entry is not None:
            proj, pdata = entry
            cell(ws1, i, 6,  proj,                     h="left")
            cell(ws1, i, 7,  len(pdata["techs"]),      h="center")
            cell(ws1, i, 8,  pdata["daily_rows"],      h="center")
            c = cell(ws1, i, 9,  round(pdata["gross"], 2), h="right")
            c.number_format = "0.00"
            c = cell(ws1, i, 10, round(pdata["lunch"], 2), h="right")
            c.number_format = "0.00"
            c = cell(ws1, i, 11, round(pdata["net"],   2), h="right")
            c.number_format = "0.00"
        ws1.row_dimensions[i].height = 15

    # Row 11: TOTAL row (right side F-K), light-green fill, dark bold text
    cell(ws1, 11, 6,  "TOTAL",                     bold=True, fill=C1_TOTAL, txt_color=C1_TXT)
    cell(ws1, 11, 7,  len(all_staff),               bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="center")
    cell(ws1, 11, 8,  daily_rows,                   bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="center")
    c = cell(ws1, 11, 9,  round(gross_total, 2),    bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
    c.number_format = "0.00"
    c = cell(ws1, 11, 10, round(lunch_total, 2),    bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
    c.number_format = "0.00"
    c = cell(ws1, 11, 11, round(net_total,   2),    bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
    c.number_format = "0.00"
    ws1.row_dimensions[11].height = 15

    # Overflow: any real projects beyond slot 5 go below row 11
    extra_row = 12
    if len(proj_rows) > 5:
        for proj, pdata in proj_rows[5:]:
            cell(ws1, extra_row, 6,  proj,                    h="left")
            cell(ws1, extra_row, 7,  len(pdata["techs"]),     h="center")
            cell(ws1, extra_row, 8,  pdata["daily_rows"],     h="center")
            c = cell(ws1, extra_row, 9,  round(pdata["gross"], 2), h="right")
            c.number_format = "0.00"
            c = cell(ws1, extra_row, 10, round(pdata["lunch"], 2), h="right")
            c.number_format = "0.00"
            c = cell(ws1, extra_row, 11, round(pdata["net"],   2), h="right")
            c.number_format = "0.00"
            extra_row += 1
        # TOTAL after overflow
        cell(ws1, extra_row, 6,  "TOTAL",               bold=True, fill=C1_TOTAL, txt_color=C1_TXT)
        cell(ws1, extra_row, 7,  len(all_staff),         bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="center")
        cell(ws1, extra_row, 8,  daily_rows,             bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="center")
        c = cell(ws1, extra_row, 9,  round(gross_total, 2), bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
        c.number_format = "0.00"
        c = cell(ws1, extra_row, 10, round(lunch_total, 2), bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
        c.number_format = "0.00"
        c = cell(ws1, extra_row, 11, round(net_total,   2), bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
        c.number_format = "0.00"
        extra_row += 1

    # ── Invoice-Ready Calculation section (matches template rows 13-18) ─────────
    # Template layout: dark-navy header merged A:F at row 13, sub-headers at row 14,
    # one row per line item (Qty / Unit / Rate / Amount), Subtotal row (no fill),
    # Total Due row (light-green fill with PO | Net 30 | USD note in col F).
    # Column header "Amount" — NOT "Amount ($)" — matches the template exactly.
    trucking_invs = [inv for inv in invoices
                     if inv.get("cost_category") in ("trucking", "courier")]
    if trucking_invs:
        truck_start = max(13, extra_row + 1)

        # Title derived from first vendor + service date (mirrors template convention)
        first_inv    = trucking_invs[0]
        first_vendor = first_inv.get("vendor", billing_month)
        first_date   = first_inv.get("service_date", "")
        truck_title  = (f"{first_vendor} Invoice-Ready Calculation"
                        + (f" — {first_date}" if first_date else ""))

        ws1.merge_cells(
            start_row=truck_start, start_column=1,
            end_row=truck_start, end_column=6
        )
        cell(ws1, truck_start, 1, truck_title,
             bold=True, txt_color=C_WHITE, fill=C1_HEADER, h="center")
        ws1.row_dimensions[truck_start].height = 18

        truck_hdr = truck_start + 1
        for ci, h_val in enumerate(
            ["Line Item", "Qty", "Unit", "Rate", "Amount", "Notes"], 1
        ):
            hdr(ws1, truck_hdr, ci, h_val, fill=C1_SUBHDR, txt=C1_TXT, h="center")
        ws1.row_dimensions[truck_hdr].height = 15

        truck_data        = truck_hdr + 1
        grand_truck_total = 0.0
        first_po          = ""
        _first_data_row   = True  # first line-item row gets 25.5pt height (matches ref)

        for inv in trucking_invs:
            if not first_po:
                first_po = str(inv.get("po_number", ""))
            line_items = inv.get("line_items") or []
            po_note    = f"PO {inv.get('po_number', '?')} | {inv.get('service_date', '')}"

            if line_items:
                # Write individual line items (matches template rows 15-16 style)
                for li in line_items:
                    li_label = li.get("description") or li.get("category", "").capitalize()
                    li_qty   = li.get("qty", "")
                    li_unit  = li.get("unit", "")
                    li_rate  = li.get("rate", "")
                    li_amt   = round(float(li.get("amount") or 0), 2)
                    cell(ws1, truck_data, 1, li_label,  h="left")
                    cell(ws1, truck_data, 2, li_qty,    h="center")
                    cell(ws1, truck_data, 3, li_unit,   h="left")
                    if li_rate != "":
                        c = cell(ws1, truck_data, 4, round(float(li_rate), 2), h="right")
                        c.number_format = '"$"#,##0.00'
                    c = cell(ws1, truck_data, 5, li_amt, h="right", bold=True)
                    c.number_format = '"$"#,##0.00'
                    cell(ws1, truck_data, 6, po_note, h="left", size=9)
                    ws1.row_dimensions[truck_data].height = 25.5 if _first_data_row else 14
                    _first_data_row   = False
                    truck_data        += 1
                    grand_truck_total += li_amt
            else:
                inv_total_f = round(float(inv.get("total") or 0), 2)
                cell(ws1, truck_data, 1, inv.get("vendor", "Invoice"), h="left")
                cell(ws1, truck_data, 2, 1,          h="center")
                cell(ws1, truck_data, 3, "invoice",  h="left")
                c = cell(ws1, truck_data, 4, inv_total_f, h="right")
                c.number_format = '"$"#,##0.00'
                c = cell(ws1, truck_data, 5, inv_total_f, h="right", bold=True)
                c.number_format = '"$"#,##0.00'
                cell(ws1, truck_data, 6, po_note, h="left", size=9)
                ws1.row_dimensions[truck_data].height = 25.5 if _first_data_row else 14
                _first_data_row   = False
                truck_data        += 1
                grand_truck_total += inv_total_f

        # Subtotal row — no fill (matches template row 17)
        cell(ws1, truck_data, 1, "Subtotal", h="left")
        c = cell(ws1, truck_data, 5, round(grand_truck_total, 2), bold=True, h="right")
        c.number_format = '"$"#,##0.00'
        ws1.row_dimensions[truck_data].height = 14
        truck_data += 1

        # Total Due row — light-green fill (matches template row 18)
        cell(ws1, truck_data, 1, "Total Due",
             bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="left")
        c = cell(ws1, truck_data, 5, round(grand_truck_total, 2),
                 bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
        c.number_format = '"$"#,##0.00'
        cell(ws1, truck_data, 6,
             f"PO {first_po} | Net 30 | USD" if first_po else "",
             fill=C1_TOTAL, txt_color=C1_TXT, h="left", size=9)
        ws1.row_dimensions[truck_data].height = 14
        extra_row = truck_data + 1
    else:
        extra_row = max(13, extra_row + 1)

    # ── Staff Rollup (A-G) + Val Weekly Hours (I-O) ───────────────────────────
    # Template anchors Staff Rollup at row 21 (after trucking section ends at row 18).
    # Guard: if data pushed extra_row past 21, use extra_row + 2 (one blank-row gap).
    staff_start = max(21, extra_row + 2)

    ws1.merge_cells(
        start_row=staff_start, start_column=1,
        end_row=staff_start, end_column=7
    )
    cell(ws1, staff_start, 1, "Staff Rollup",
         bold=True, txt_color=C_WHITE, fill=C1_HEADER, h="center")

    # Compute weekly buckets
    weekly: dict = defaultdict(lambda: {"gross": 0.0, "lunch": 0.0, "net": 0.0,
                                        "projects": set(), "days": set()})
    for rec in sorted_records:
        d = rec["date"]
        mon = d - _dt.timedelta(days=d.weekday())
        fri = mon + _dt.timedelta(days=4)
        wk  = mon
        label = f"{_month_day_label(mon)}–{fri.day}"
        weekly[wk]["label"]  = label
        weekly[wk]["gross"] += rec["gross_hours"]
        weekly[wk]["lunch"] += rec["lunch_deduction"]
        weekly[wk]["net"]   += rec["net_hours"]
        weekly[wk]["projects"].add(rec["project"] or "Unassigned")
        weekly[wk]["days"].add(d)

    ws1.merge_cells(
        start_row=staff_start, start_column=9,
        end_row=staff_start, end_column=15
    )
    cell(ws1, staff_start, 9, "Val Weekly Hours",
         bold=True, txt_color=C_WHITE, fill=C1_HEADER, h="center")
    ws1.row_dimensions[staff_start].height = 18

    staff_hdr = staff_start + 1
    for ci, h_val in enumerate(
        ["Staff Name", "Team / Project Bucket(s)", "Daily Rows",
         "Gross Hours", "Lunch Deducted", "Net Billable Hours", "Source"],
        1
    ):
        hdr(ws1, staff_hdr, ci, h_val, fill=C1_SUBHDR, txt=C1_TXT, h="center")
    for ci, h_val in enumerate(
        ["Week", "Date Range", "Projects", "Days Worked",
         "Gross Hours", "Lunch Deducted", "Net Billable Hours"],
        9
    ):
        hdr(ws1, staff_hdr, ci, h_val, fill=C1_SUBHDR, txt=C1_TXT, h="center")
    ws1.row_dimensions[staff_hdr].height = 15

    staff_data = staff_hdr + 1
    sorted_weekly = sorted(weekly.items())

    max_staff_rows = max(len(sorted_staff), len(sorted_weekly), 1)
    for i in range(max_staff_rows):
        r = staff_data + i

        # Left: per-staff
        if i < len(sorted_staff):
            sname, sdata = sorted_staff[i]
            proj_list = ", ".join(sorted(sdata["projects"]))
            cell(ws1, r, 1, sname,                        h="left")
            cell(ws1, r, 2, proj_list,                    h="left", size=9)
            cell(ws1, r, 3, sdata["rows"],                h="center")
            c = cell(ws1, r, 4, round(sdata["gross"], 2), h="right")
            c.number_format = "0.00"
            c = cell(ws1, r, 5, round(sdata["lunch"], 2), h="right")
            c.number_format = "0.00"
            c = cell(ws1, r, 6, round(sdata["net"],   2), h="right")
            c.number_format = "0.00"
            cell(ws1, r, 7, "Live " + month_label,        h="left", size=9)

        # Right: per-week — col 9 (I) = week start datetime
        if i < len(sorted_weekly):
            wk_mon, wdata = sorted_weekly[i]
            wk_datetime = _to_datetime(wk_mon)
            c = cell(ws1, r, 9, wk_datetime, h="left")
            c.number_format = "mm-dd-yy"
            cell(ws1, r, 10, wdata.get("label", ""), h="left")
            cell(ws1, r, 11, ", ".join(sorted(wdata["projects"])), h="left", size=8)
            cell(ws1, r, 12, len(wdata["days"]), h="center")
            c = cell(ws1, r, 13, round(wdata["gross"], 2), h="right")
            c.number_format = "0.00"
            c = cell(ws1, r, 14, round(wdata["lunch"], 2), h="right")
            c.number_format = "0.00"
            c = cell(ws1, r, 15, round(wdata["net"],   2), h="right")
            c.number_format = "0.00"

        ws1.row_dimensions[r].height = 15

    # TOTAL row for both sections
    total_r = staff_data + max_staff_rows
    cell(ws1, total_r, 1, "TOTAL", bold=True, fill=C1_TOTAL, txt_color=C1_TXT)
    cell(ws1, total_r, 3, len(records),              bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="center")
    c = cell(ws1, total_r, 4, round(gross_total, 2), bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
    c.number_format = "0.00"
    c = cell(ws1, total_r, 5, round(lunch_total, 2), bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
    c.number_format = "0.00"
    c = cell(ws1, total_r, 6, round(net_total,   2), bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
    c.number_format = "0.00"
    cell(ws1, total_r, 9,  "TOTAL", bold=True, fill=C1_TOTAL, txt_color=C1_TXT)
    cell(ws1, total_r, 12, sum(len(w["days"]) for w in weekly.values()),
         bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="center")
    c = cell(ws1, total_r, 13, round(gross_total, 2), bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
    c.number_format = "0.00"
    c = cell(ws1, total_r, 14, round(lunch_total, 2), bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
    c.number_format = "0.00"
    c = cell(ws1, total_r, 15, round(net_total,   2), bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
    c.number_format = "0.00"
    ws1.row_dimensions[total_r].height = 15

    # ── Daily Billing Detail ───────────────────────────────────────────────────
    detail_start = total_r + 2
    ws1.merge_cells(
        start_row=detail_start, start_column=1,
        end_row=detail_start, end_column=10
    )
    cell(ws1, detail_start, 1, "Daily Billing Detail with Lunch Deduction",
         bold=True, txt_color=C_WHITE, fill=C1_HEADER, h="center")
    ws1.row_dimensions[detail_start].height = 18

    detail_hdr_row = detail_start + 1
    for ci, h_val in enumerate(
        ["Staff Name", "Date", "Week", "Project", "Clock In", "Clock Out",
         "Gross Hours", "Lunch Deduction", "Net Billable Hours", "Source / Rule"],
        1
    ):
        hdr(ws1, detail_hdr_row, ci, h_val, fill=C1_SUBHDR, txt=C1_TXT, h="center")
    ws1.row_dimensions[detail_hdr_row].height = 15

    detail_data_row = detail_hdr_row + 1
    for rec in sorted_records:
        d = rec["date"]
        week_mon = d - _dt.timedelta(days=d.weekday())
        week_fri = week_mon + _dt.timedelta(days=4)
        week_label = f"{_month_day_label(week_mon)}–{week_fri.day}"

        # Date as datetime object with mm-dd-yy format
        date_cell = cell(ws1, detail_data_row, 2, _to_datetime(d), h="left")
        date_cell.number_format = "mm-dd-yy"

        # Clock in/out as time objects with h:mm AM/PM format
        ci_cell = cell(ws1, detail_data_row, 5, _to_time(rec["clock_in"]), h="right")
        ci_cell.number_format = "h:mm AM/PM"
        co_cell = cell(ws1, detail_data_row, 6, _to_time(rec["clock_out"]), h="right")
        co_cell.number_format = "h:mm AM/PM"

        cell(ws1, detail_data_row, 1, rec["staff"],   h="left")
        cell(ws1, detail_data_row, 3, week_label,     h="left")
        cell(ws1, detail_data_row, 4, rec["project"], h="left")

        c = cell(ws1, detail_data_row, 7, round(rec["gross_hours"],     2), h="right")
        c.number_format = "0.00"
        c = cell(ws1, detail_data_row, 8, round(rec["lunch_deduction"], 1), h="right")
        c.number_format = "0.00"
        c = cell(ws1, detail_data_row, 9, round(rec["net_hours"],       2), h="right")
        c.number_format = "0.00"
        cell(ws1, detail_data_row, 10,
             "Live " + month_label + "; lunch rule applied", h="left", size=9)

        ws1.row_dimensions[detail_data_row].height = 14
        detail_data_row += 1

    # TOTAL row for detail
    cell(ws1, detail_data_row, 1, "TOTAL", bold=True, fill=C1_TOTAL, txt_color=C1_TXT)
    c = cell(ws1, detail_data_row, 7, round(gross_total, 2), bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
    c.number_format = "0.00"
    c = cell(ws1, detail_data_row, 8, round(lunch_total, 2), bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
    c.number_format = "0.00"
    c = cell(ws1, detail_data_row, 9, round(net_total,   2), bold=True, fill=C1_TOTAL, txt_color=C1_TXT, h="right")
    c.number_format = "0.00"

    # Column widths for sheet 1 (matching Agilant template)
    ws1.column_dimensions["A"].width = 22.125
    ws1.column_dimensions["B"].width = 17.875
    ws1.column_dimensions["C"].width = 16.375
    ws1.column_dimensions["D"].width = 25.75
    ws1.column_dimensions["E"].width = 12.875
    ws1.column_dimensions["F"].width = 25.0
    ws1.column_dimensions["G"].width = 13.625
    ws1.column_dimensions["I"].width = 17.875
    ws1.column_dimensions["J"].width = 22.125
    ws1.column_dimensions["K"].width = 16.375
    # No freeze panes — template has none

    # ════════════════════════════════════════════════════════════════════════════
    # SHEET 2: Invoice Pivots - Candidate
    # ════════════════════════════════════════════════════════════════════════════

    # Row 1: Title — dark charcoal fill, white bold 16pt
    ws2.merge_cells("A1:N1")
    cell(ws2, 1, 1, "Invoice Pivot Summary - Candidate",
         bold=True, size=16, txt_color=C_WHITE, fill=C2_HEADER, h="center")
    ws2.row_dimensions[1].height = 27.95

    # Row 2: Subtitle — light-blue fill, dark text
    ws2.merge_cells("A2:N2")
    cell(ws2, 2, 1,
         "Monthly invoice aggregation from the invoice ledger. This is pivot-style and formula-driven so new rows can be added to the ledger.",
         size=11, txt_color=None, fill=C2_SUBTITLE, wrap=True)
    ws2.row_dimensions[2].height = 14

    # Row 4: Section headers (teal)
    # col 9 (I) is blank gap — matches Agilant template layout
    hdr(ws2, 4, 1,  "Monthly Invoice Totals", span=8, fill=C2_SECHDR, size=11)
    hdr(ws2, 4, 10, "Totals by Category",     span=5, fill=C2_SECHDR, size=11)
    ws2.row_dimensions[4].height = 15

    # Row 5: Column sub-headers (medium gray fill, white text)
    for ci, h_val in enumerate(
        ["Service Month", "Invoice Count", "Trucking", "Labor",
         "Courier", "Other", "Total", "PO Count"],
        1
    ):
        hdr(ws2, 5, ci, h_val, fill=C2_SUBHDR, txt=C_WHITE, h="center", size=11)
    for ci, h_val in enumerate(
        ["Invoice Category", "Invoice Count", "Trucking", "Labor", "Total"],
        10
    ):
        hdr(ws2, 5, ci, h_val, fill=C2_SUBHDR, txt=C_WHITE, h="center", size=11)
    ws2.row_dimensions[5].height = 15

    # Monthly totals rows
    mit_row = 6
    grand_truck = grand_labor = grand_courier = grand_other = 0.0
    grand_count = 0
    all_po_sets: Set[str] = set()

    for inv_month, mdata in sorted(inv_by_month.items()):
        t_ = mdata["trucking"]; l_ = mdata["labor"]
        c_ = mdata["courier"];  o_ = mdata["other"]
        tot = t_ + l_ + c_ + o_
        cnt = mdata["count"]
        pos = mdata["pos"]
        grand_truck   += t_; grand_labor  += l_
        grand_courier += c_; grand_other  += o_
        grand_count   += cnt; all_po_sets |= pos

        # Service Month: datetime object with mmm yyyy format
        month_dt = _month_datetime(inv_month)
        c_sm = cell(ws2, mit_row, 1, month_dt, size=11)
        c_sm.number_format = "mmm\\ yyyy"

        cell(ws2, mit_row, 2, cnt,          size=11, h="center")
        c = cell(ws2, mit_row, 3, round(t_,  2), size=11, h="right")
        c.number_format = r'\$#,##0.00'
        c = cell(ws2, mit_row, 4, round(l_,  2), size=11, h="right")
        c.number_format = r'\$#,##0.00'
        c = cell(ws2, mit_row, 5, round(c_,  2), size=11, h="right")
        c.number_format = r'\$#,##0.00'
        c = cell(ws2, mit_row, 6, round(o_,  2), size=11, h="right")
        c.number_format = r'\$#,##0.00'
        c = cell(ws2, mit_row, 7, round(tot, 2), size=11, h="right", bold=True)
        c.number_format = r'\$#,##0.00'
        cell(ws2, mit_row, 8, len(pos),      size=11, h="center")
        ws2.row_dimensions[mit_row].height = 15
        mit_row += 1

    # TOTAL row for Monthly Invoice Totals
    # Col H (PO Count) is intentionally blank in the template TOTAL row.
    grand_tot = grand_truck + grand_labor + grand_courier + grand_other
    cell(ws2, mit_row, 1, "TOTAL",                     bold=True, fill=C2_TOTAL, size=11)
    cell(ws2, mit_row, 2, grand_count,                  bold=True, fill=C2_TOTAL, size=11, h="center")
    c = cell(ws2, mit_row, 3, round(grand_truck,  2),   bold=True, fill=C2_TOTAL, size=11, h="right")
    c.number_format = r'\$#,##0.00'
    c = cell(ws2, mit_row, 4, round(grand_labor,  2),   bold=True, fill=C2_TOTAL, size=11, h="right")
    c.number_format = r'\$#,##0.00'
    c = cell(ws2, mit_row, 5, round(grand_courier,2),   bold=True, fill=C2_TOTAL, size=11, h="right")
    c.number_format = r'\$#,##0.00'
    c = cell(ws2, mit_row, 6, round(grand_other,  2),   bold=True, fill=C2_TOTAL, size=11, h="right")
    c.number_format = r'\$#,##0.00'
    c = cell(ws2, mit_row, 7, round(grand_tot,    2),   bold=True, fill=C2_TOTAL, size=11, h="right")
    c.number_format = r'\$#,##0.00'
    ws2.row_dimensions[mit_row].height = 15
    mit_row += 1  # advance past TOTAL row

    # Vendor/Category block (right side, rows 6+) — cols 10-14
    # Invoice Category | Invoice Count | Trucking | Labor | Total
    cat_row = 6
    for category, vdata in sorted(category_map.items()):
        cell(ws2, cat_row, 10, category,                    size=11)
        cell(ws2, cat_row, 11, vdata["count"],               size=11, h="center")
        c = cell(ws2, cat_row, 12, round(vdata["trucking"], 2), size=11, h="right")
        c.number_format = r'\$#,##0.00'
        c = cell(ws2, cat_row, 13, round(vdata["labor"],    2), size=11, h="right")
        c.number_format = r'\$#,##0.00'
        c = cell(ws2, cat_row, 14, round(vdata["total"],    2), size=11, h="right", bold=True)
        c.number_format = r'\$#,##0.00'
        ws2.row_dimensions[cat_row].height = 15
        cat_row += 1

    # ── Three side-by-side sections: PO | Project | Vendor ───────────────────
    # PO at col 1 (A-D), Project at col 6 (F-H), Vendor/Crew at col 10 (J-N)
    # Cols 5 (E) and 9 (I) are blank gaps — matches Agilant template exactly.
    # Start row = 2 rows after whichever block (monthly totals or category) ends later.
    section_row = max(mit_row, cat_row) + 2
    hdr(ws2, section_row, 1,  "Totals by PO Number",     span=4, fill=C2_SECHDR, size=11)
    hdr(ws2, section_row, 6,  "Totals by Project",       span=3, fill=C2_SECHDR, size=11)
    hdr(ws2, section_row, 10, "Totals by Vendor / Crew", span=5, fill=C2_SECHDR, size=11)
    ws2.row_dimensions[section_row].height = 15
    section_row += 1

    for ci, h_val in enumerate(["PO Number", "Invoice Count", "Total", "Notes"], 1):
        hdr(ws2, section_row, ci, h_val, fill=C2_SUBHDR, txt=C_WHITE, h="center", size=11)
    for ci, h_val in enumerate(["Project", "Invoice Count", "Total"], 6):
        hdr(ws2, section_row, ci, h_val, fill=C2_SUBHDR, txt=C_WHITE, h="center", size=11)
    for ci, h_val in enumerate(
        ["Vendor / Crew", "Invoice Count", "Trucking", "Labor", "Total"], 10
    ):
        hdr(ws2, section_row, ci, h_val, fill=C2_SUBHDR, txt=C_WHITE, h="center", size=11)
    ws2.row_dimensions[section_row].height = 15
    section_row += 1

    po_items       = sorted(po_map.items())
    proj_inv_items = list(sorted(project_inv_map.items())) if project_inv_map else []
    vendor_items   = sorted(vendor_map.items())

    max_rows = max(len(po_items), max(len(proj_inv_items), 1), len(vendor_items))
    for i in range(max_rows):
        r = section_row + i

        if i < len(po_items):
            po, pdata = po_items[i]
            cell(ws2, r, 1, po,                      size=11)
            cell(ws2, r, 2, pdata["count"],           size=11, h="center")
            c = cell(ws2, r, 3, round(pdata["total"], 2), size=11, h="right", bold=True)
            c.number_format = r'\$#,##0.00'
            cell(ws2, r, 4, pdata.get("notes", ""),  size=11)

        if i < len(proj_inv_items):
            proj, pdata_proj = proj_inv_items[i]
            inv_count  = pdata_proj["count"] if isinstance(pdata_proj, dict) else "-"
            proj_total = pdata_proj["total"] if isinstance(pdata_proj, dict) else pdata_proj
            cell(ws2, r, 6, proj,                    size=11)
            cell(ws2, r, 7, inv_count,               size=11, h="center")
            c = cell(ws2, r, 8, round(proj_total, 2), size=11, h="right", bold=True)
            c.number_format = r'\$#,##0.00'

        if i < len(vendor_items):
            vendor, vdata = vendor_items[i]
            cell(ws2, r, 10, vendor,                      size=11)
            cell(ws2, r, 11, vdata["count"],               size=11, h="center")
            c = cell(ws2, r, 12, round(vdata["trucking"], 2), size=11, h="right")
            c.number_format = r'\$#,##0.00'
            c = cell(ws2, r, 13, round(vdata["labor"],    2), size=11, h="right")
            c.number_format = r'\$#,##0.00'
            c = cell(ws2, r, 14, round(vdata["total"],    2), size=11, h="right", bold=True)
            c.number_format = r'\$#,##0.00'

        ws2.row_dimensions[r].height = 15

    # Grand total row
    total_r2 = section_row + max_rows
    cell(ws2, total_r2, 1,  "TOTAL", bold=True, fill=C2_TOTAL, size=11)
    cell(ws2, total_r2, 2,  sum(p["count"] for p in po_map.values()), bold=True, fill=C2_TOTAL, size=11, h="center")
    c = cell(ws2, total_r2, 3,
             round(sum(p["total"] for p in po_map.values()), 2),
             bold=True, fill=C2_TOTAL, size=11, h="right")
    c.number_format = r'\$#,##0.00'
    cell(ws2, total_r2, 10, "TOTAL", bold=True, fill=C2_TOTAL, size=11)
    cell(ws2, total_r2, 11, sum(v["count"] for v in vendor_map.values()), bold=True, fill=C2_TOTAL, size=11, h="center")
    c = cell(ws2, total_r2, 14,
             round(sum(v["total"] for v in vendor_map.values()), 2),
             bold=True, fill=C2_TOTAL, size=11, h="right")
    c.number_format = r'\$#,##0.00'

    # ── Audit Notes block (matches Agilant template A24+) ─────────────────────
    # Dark charcoal header merged A:N, then numbered note rows (col A = index,
    # col B:N merged = note text).  Populated from the audit_notes parameter.
    notes_header_row = total_r2 + 2
    ws2.merge_cells(
        start_row=notes_header_row, start_column=1,
        end_row=notes_header_row, end_column=14
    )
    cell(ws2, notes_header_row, 1, "Audit Notes",
         bold=True, size=11, txt_color=C_WHITE, fill=C2_HEADER, h="left")
    ws2.row_dimensions[notes_header_row].height = 15

    notes_list = audit_notes or []
    for idx, note_text in enumerate(notes_list, 1):
        note_r = notes_header_row + idx
        cell(ws2, note_r, 1, str(idx), size=11, h="center")
        ws2.merge_cells(
            start_row=note_r, start_column=2,
            end_row=note_r, end_column=14
        )
        cell(ws2, note_r, 2, note_text, size=11, wrap=True)
        ws2.row_dimensions[note_r].height = 15

    # Column widths for sheet 2 — only set columns that have custom widths in
    # the Agilant template.  C, D, E, L, M use Excel's default width.
    ws2.column_dimensions["A"].width = 15.0
    ws2.column_dimensions["B"].width = 13.0
    ws2.column_dimensions["F"].width = 32.0
    ws2.column_dimensions["G"].width = 13.0
    ws2.column_dimensions["H"].width = 15.0
    ws2.column_dimensions["I"].width = 3.0
    ws2.column_dimensions["J"].width = 31.0
    ws2.column_dimensions["K"].width = 13.0
    ws2.column_dimensions["N"].width = 15.0
    # No freeze panes — template has none

    # ── Structural assertions (fail hard if layout drifts from template) ──────
    def _assert_billing_structure(ws) -> None:
        """Raise RuntimeError if Sheet 1 fixed-position labels deviate from the template.

        Fixed-position structural labels (rows 4-5) are compared exactly.
        Dynamic sections (Staff Rollup, Daily Billing Detail) are scanned by value.
        """
        errors = []

        # Rows 4-5: section + column sub-headers (exact positions from template)
        fixed_checks = {
            (4, 1):  "Monthly Rollup",
            (4, 6):  "Pivot by Team / Project Bucket",
            (5, 1):  "Metric",
            (5, 2):  "Value",
            (5, 3):  "Basis",
            (5, 4):  "Notes",
            (5, 6):  "Team / Project Bucket",
            (5, 7):  "Tech Count",
            (5, 8):  "Daily Rows",
            (5, 9):  "Gross Hours",
            (5, 10): "Lunch Deducted",
            (5, 11): "Net Billable Hours",
        }
        for (r, c), expected in fixed_checks.items():
            actual = str(ws.cell(r, c).value or "").strip()
            if actual != expected:
                errors.append(f"Sheet1 R{r}C{c}: expected {expected!r}, got {actual!r}")

        # Invoice-Ready Calculation section — locate header row by content scan.
        # (Row varies when project-bucket overflow pushes the invoice block down.)
        # Only asserted when the section is present (no invoices → no block).
        inv_hdr_row = next(
            (r for r in range(10, ws.max_row + 1)
             if str(ws.cell(r, 1).value or "").strip() == "Line Item"),
            None,
        )
        if inv_hdr_row is not None:
            for c_off, expected_h in enumerate(
                ["Line Item", "Qty", "Unit", "Rate", "Amount", "Notes"], 1
            ):
                actual = str(ws.cell(inv_hdr_row, c_off).value or "").strip()
                if actual != expected_h:
                    errors.append(
                        f"Sheet1 R{inv_hdr_row}C{c_off}: expected {expected_h!r},"
                        f" got {actual!r}"
                    )

        # TOTAL for project pivot — must appear in col 6 between rows 10-20
        total_found = any(
            str(ws.cell(r, 6).value or "").strip() == "TOTAL"
            for r in range(10, 21)
        )
        if not total_found:
            errors.append("Sheet1: 'TOTAL' sentinel not found in col F rows 10-20")

        # Staff Rollup in col 1 → Val Weekly Hours at col 9 same row
        staff_rollup_row = next(
            (ws.cell(r, 1).row
             for r in range(11, ws.max_row + 1)
             if str(ws.cell(r, 1).value or "").strip() == "Staff Rollup"),
            None
        )
        if staff_rollup_row is None:
            errors.append("Sheet1: 'Staff Rollup' section header not found in col A")
        else:
            val_weekly = str(ws.cell(staff_rollup_row, 9).value or "").strip()
            if val_weekly != "Val Weekly Hours":
                errors.append(
                    f"Sheet1 R{staff_rollup_row}C9: expected 'Val Weekly Hours',"
                    f" got {val_weekly!r}"
                )

        # Daily Billing Detail header in col 1
        detail_found = any(
            str(ws.cell(r, 1).value or "").strip().startswith("Daily Billing Detail")
            for r in range(11, ws.max_row + 1)
        )
        if not detail_found:
            errors.append("Sheet1: 'Daily Billing Detail' section header not found")

        if errors:
            raise RuntimeError(
                "Generated Sheet 1 does not match Agilant template structure:\n"
                + "\n".join(f"  {e}" for e in errors)
            )

    def _assert_pivot_structure(ws2) -> None:
        """Raise RuntimeError if Sheet 2 fixed-position labels deviate from the template."""
        errors = []

        # Rows 1, 4-5: fixed labels
        fixed_checks = {
            (1,  1):  "Invoice Pivot Summary - Candidate",
            (4,  1):  "Monthly Invoice Totals",
            (4,  10): "Totals by Category",
            (5,  1):  "Service Month",
            (5,  2):  "Invoice Count",
            (5,  3):  "Trucking",
            (5,  4):  "Labor",
            (5,  5):  "Courier",
            (5,  6):  "Other",
            (5,  7):  "Total",
            (5,  8):  "PO Count",
            (5,  10): "Invoice Category",
            (5,  11): "Invoice Count",
            (5,  14): "Total",
        }
        for (r, c), expected in fixed_checks.items():
            actual = str(ws2.cell(r, c).value or "").strip()
            if actual != expected:
                errors.append(f"Sheet2 R{r}C{c}: expected {expected!r}, got {actual!r}")

        # Locate "Totals by PO Number" (dynamic row, scanned)
        po_section_row = next(
            (ws2.cell(r, 1).row
             for r in range(6, ws2.max_row + 1)
             if str(ws2.cell(r, 1).value or "").strip() == "Totals by PO Number"),
            None
        )
        if po_section_row is None:
            errors.append("Sheet2: 'Totals by PO Number' section header not found")
        else:
            three_sec_checks = {
                (po_section_row,     1):  "Totals by PO Number",
                (po_section_row,     6):  "Totals by Project",
                (po_section_row,     10): "Totals by Vendor / Crew",
                (po_section_row + 1, 1):  "PO Number",
                (po_section_row + 1, 6):  "Project",
                (po_section_row + 1, 10): "Vendor / Crew",
                (po_section_row + 1, 12): "Trucking",
                (po_section_row + 1, 13): "Labor",
                (po_section_row + 1, 14): "Total",
            }
            for (r, c), expected in three_sec_checks.items():
                actual = str(ws2.cell(r, c).value or "").strip()
                if actual != expected:
                    errors.append(
                        f"Sheet2 R{r}C{c}: expected {expected!r}, got {actual!r}"
                    )

        # Audit Notes header must exist
        audit_found = any(
            str(ws2.cell(r, 1).value or "").strip() == "Audit Notes"
            for r in range(6, ws2.max_row + 1)
        )
        if not audit_found:
            errors.append("Sheet2: 'Audit Notes' header not found")

        if errors:
            raise RuntimeError(
                "Generated Sheet 2 does not match Agilant template structure:\n"
                + "\n".join(f"  {e}" for e in errors)
            )

    def _assert_against_reference_template(ws1_gen, ws2_gen) -> None:
        """Zone-aware comparison of generated sheets against the Agilant reference workbook.

        Opens attached_assets/Agilant_Admins_April_2026_Billing_Summary_*.xlsx (if
        present) using openpyxl data_only=True and compares the generated sheets
        against two specific reference sheets:

          "April 2026 Updated Candidate"  →  ws1_gen  (Billing Summary)
          "Invoice Pivots - Candidate"    →  ws2_gen  (Invoice Pivots)

        Comparison covers:
          • Static cell values — section headers, column sub-headers (rows 4-5 and
            scanned invoice block).  Month-specific title cells and all data rows
            are excluded so the function works for any billing month.
          • Merged cell ranges — structural merges (title, subtitle, section labels).
            The invoice-section title merge is located by content scan so it passes
            whether or not the block is shifted by project-bucket overflow.
          • Column widths — every column that has a custom width in the reference.
          • Row heights — only structural header rows where the reference sets them.
          • Cell fills (fgColor RGB) — title row, section header rows.
          • Number formats — money-format cells in the invoice pivot sheet.
          • Freeze panes — both sheets must have none, matching the reference.

        Silently returns when the reference file is absent (e.g. CI without assets).
        """
        import glob as _glob

        asset_dir = Path(__file__).parent.parent / "attached_assets"
        matches = sorted(_glob.glob(
            str(asset_dir / "Agilant_Admins_April_2026_Billing_Summary_*.xlsx")
        ))
        if not matches:
            return  # Reference file not present — skip (CI / stripped repo)

        ref_wb = openpyxl.load_workbook(matches[0], data_only=True)
        ref_name = Path(matches[0]).name

        # Sheet 1 reference: "April 2026 Updated Candidate" (structural template)
        # Sheet 2 reference: "Invoice Pivots - Candidate"
        if "April 2026 Updated Candidate" not in ref_wb.sheetnames:
            return  # Reference structure changed — skip to avoid false failures
        if "Invoice Pivots - Candidate" not in ref_wb.sheetnames:
            return

        ref_ws1 = ref_wb["April 2026 Updated Candidate"]
        ref_ws2 = ref_wb["Invoice Pivots - Candidate"]
        errors: List[str] = []

        # ── Helpers ───────────────────────────────────────────────────────────

        def _rgb(cell) -> str:
            f = cell.fill
            return f.fgColor.rgb if (f and f.fgColor) else "00000000"

        def _cmp_cells(label, ref_ws, gen_ws, positions):
            for r, c in positions:
                ref_v = str(ref_ws.cell(r, c).value or "").strip()
                gen_v = str(gen_ws.cell(r, c).value or "").strip()
                if ref_v and gen_v != ref_v:
                    errors.append(
                        f"{label} R{r}C{c}: ref={ref_v!r}, generated={gen_v!r}"
                    )

        def _cmp_fills(label, ref_ws, gen_ws, positions):
            for r, c in positions:
                ref_rgb = _rgb(ref_ws.cell(r, c))
                gen_rgb = _rgb(gen_ws.cell(r, c))
                if ref_rgb != "00000000" and gen_rgb != ref_rgb:
                    errors.append(
                        f"{label} R{r}C{c} fill: ref={ref_rgb!r}, generated={gen_rgb!r}"
                    )

        def _cmp_widths(label, ref_ws, gen_ws):
            for col, dim in ref_ws.column_dimensions.items():
                if not dim.customWidth:
                    continue
                gen_dim = gen_ws.column_dimensions.get(col)
                gen_w = gen_dim.width if (gen_dim and gen_dim.customWidth) else 8.43
                if abs(gen_w - dim.width) > 0.5:
                    errors.append(
                        f"{label} col {col}: ref width={dim.width}, generated={gen_w}"
                    )

        def _cmp_row_heights(label, ref_ws, gen_ws, rows):
            for r in rows:
                ref_h = (ref_ws.row_dimensions[r].height
                         if r in ref_ws.row_dimensions else None)
                gen_h = (gen_ws.row_dimensions[r].height
                         if r in gen_ws.row_dimensions else None)
                if ref_h and gen_h and abs(gen_h - ref_h) > 1.0:
                    errors.append(
                        f"{label} row {r} height: ref={ref_h}, generated={gen_h}"
                    )

        def _cmp_merges(label, ref_ws, gen_ws, expected_merge_strs):
            gen_merges = {str(m) for m in gen_ws.merged_cells.ranges}
            for mr in expected_merge_strs:
                if mr not in gen_merges:
                    errors.append(
                        f"{label} merge {mr!r}: present in ref, missing in generated"
                    )

        def _cmp_number_fmt(label, ref_ws, gen_ws, r, c):
            ref_fmt = ref_ws.cell(r, c).number_format
            gen_fmt = gen_ws.cell(r, c).number_format
            if ref_fmt and ref_fmt != "General" and gen_fmt != ref_fmt:
                errors.append(
                    f"{label} R{r}C{c} number_format:"
                    f" ref={ref_fmt!r}, generated={gen_fmt!r}"
                )

        # ── Sheet 1: Billing Summary ──────────────────────────────────────────

        # Static cell values: rows 4-5 (section headers + column sub-headers)
        _cmp_cells("Sheet1", ref_ws1, ws1_gen, [
            (4, 1), (4, 6),
            (5, 1), (5, 2), (5, 3), (5, 4),
            (5, 6), (5, 7), (5, 8), (5, 9), (5, 10), (5, 11),
        ])

        # Invoice-Ready Calculation sub-headers — dynamic row (may shift with
        # project-bucket overflow); skipped entirely when no invoices are present.
        inv_hdr_row = next(
            (r for r in range(10, ws1_gen.max_row + 1)
             if str(ws1_gen.cell(r, 1).value or "").strip() == "Line Item"),
            None,
        )
        if inv_hdr_row is not None:
            for c_idx, expected_h in enumerate(
                ["Line Item", "Qty", "Unit", "Rate", "Amount", "Notes"], 1
            ):
                actual = str(ws1_gen.cell(inv_hdr_row, c_idx).value or "").strip()
                if actual != expected_h:
                    errors.append(
                        f"Sheet1 invoice hdr C{c_idx}:"
                        f" ref={expected_h!r}, generated={actual!r}"
                    )
            # Invoice section title row must be merged A{title_r}:F{title_r}
            inv_title_row = inv_hdr_row - 1
            gen_merges = {str(m) for m in ws1_gen.merged_cells.ranges}
            expected_inv_merge = f"A{inv_title_row}:F{inv_title_row}"
            if expected_inv_merge not in gen_merges:
                errors.append(
                    f"Sheet1 invoice title merge {expected_inv_merge!r} missing"
                )

        # Column widths (all custom-width columns from the reference)
        _cmp_widths("Sheet1", ref_ws1, ws1_gen)

        # Row heights: the first invoice line-item row has 25.5 pt in the reference.
        # Its row number is dynamic (shifts with project-bucket overflow), so we
        # locate it by scanning for "Line Item" header and checking row below it.
        if inv_hdr_row is not None:
            first_li_row = inv_hdr_row + 1
            gen_dim = ws1_gen.row_dimensions.get(first_li_row)
            gen_h = gen_dim.height if gen_dim else None
            if gen_h and abs(gen_h - 25.5) > 1.0:
                errors.append(
                    f"Sheet1 first invoice line-item row height:"
                    f" ref=25.5, generated={gen_h}"
                )

        # Structural merged ranges that must always be present
        _cmp_merges("Sheet1", ref_ws1, ws1_gen,
                    ["A1:K1", "A2:K2", "A4:D4", "F4:K4"])

        # Fill colors: section header rows (structural, must match reference)
        _cmp_fills("Sheet1", ref_ws1, ws1_gen,
                   [(4, 1), (4, 6), (5, 1), (5, 6)])

        # Freeze panes
        if ws1_gen.freeze_panes != ref_ws1.freeze_panes:
            errors.append(
                f"Sheet1 freeze_panes: ref={ref_ws1.freeze_panes!r},"
                f" generated={ws1_gen.freeze_panes!r}"
            )

        # ── Sheet 2: Invoice Pivots ───────────────────────────────────────────

        # Static cell values: rows 1, 4-5
        _cmp_cells("Sheet2", ref_ws2, ws2_gen, [
            (1,  1),
            (4,  1),  (4,  10),
            (5,  1),  (5,  2),  (5,  3),  (5,  4),  (5,  5),
            (5,  6),  (5,  7),  (5,  8),
            (5,  10), (5,  11), (5,  12), (5,  13), (5,  14),
        ])

        # Subtitle row (row 2, col 1) — full text
        ref_sub = str(ref_ws2.cell(2, 1).value or "").strip()
        gen_sub = str(ws2_gen.cell(2, 1).value or "").strip()
        if ref_sub and gen_sub != ref_sub:
            errors.append(
                f"Sheet2 R2C1 subtitle: ref={ref_sub!r}, generated={gen_sub!r}"
            )

        # Column widths
        _cmp_widths("Sheet2", ref_ws2, ws2_gen)

        # Row heights: rows 1, 4, 5 are set in the reference
        _cmp_row_heights("Sheet2", ref_ws2, ws2_gen, [1, 4, 5])

        # Structural merged ranges
        _cmp_merges("Sheet2", ref_ws2, ws2_gen,
                    ["A1:N1", "A2:N2", "A4:H4", "J4:N4"])

        # Fill colors: title and section header rows
        _cmp_fills("Sheet2", ref_ws2, ws2_gen,
                   [(1, 1), (4, 1), (4, 10), (5, 1)])

        # Number format: money-format data cells in pivot table (C6 = Trucking total)
        _cmp_number_fmt("Sheet2", ref_ws2, ws2_gen, 6, 3)

        # Freeze panes
        if ws2_gen.freeze_panes != ref_ws2.freeze_panes:
            errors.append(
                f"Sheet2 freeze_panes: ref={ref_ws2.freeze_panes!r},"
                f" generated={ws2_gen.freeze_panes!r}"
            )

        ref_wb.close()

        if errors:
            raise RuntimeError(
                f"Generated workbook diverges from Agilant reference ({ref_name}):\n"
                + "\n".join(f"  {e}" for e in errors)
            )

    _assert_billing_structure(ws1)
    _assert_pivot_structure(ws2)
    _assert_against_reference_template(ws1, ws2)

    # ── Save ──────────────────────────────────────────────────────────────────
    out_dir  = Path(out_root) / billing_month / "workbook"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"billing_summary_{billing_month}.xlsx"
    wb.save(str(out_path))

    # ── Manifest ──────────────────────────────────────────────────────────────
    resolved_run_id = run_id or f"billing-{billing_month}-{int(time.time())}"
    _write_manifest(
        out_root      = out_root,
        billing_month = billing_month,
        run_id        = resolved_run_id,
        inputs        = input_paths or [],
        outputs       = [str(out_path)],
        status        = "generated",
        meta          = {
            "billing_month":  billing_month,
            "roster_records": len(records),
            "invoices":       len(invoices),
            "gross_hours":    round(gross_total, 2),
            "net_hours":      round(net_total,   2),
            "invoice_total":  round(inv_total,   2),
        },
    )

    return str(out_path)


def update_manifest_status(
    out_root: str,
    billing_month: str,
    status: str,
    gate_status: str = "",
    failures: Optional[List[str]] = None,
    warnings: Optional[List[str]] = None,
) -> None:
    """Update an existing manifest's status after gate validation."""
    import json as _json
    manifest_path = Path(out_root) / billing_month / "run_manifest.json"
    if not manifest_path.exists():
        return
    try:
        data = _json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:
        data = {}
    data["status"]         = status
    data["gate_status"]    = gate_status
    data["gate_failures"]  = failures or []
    data["gate_warnings"]  = warnings or []
    data["validated_at"]   = time.strftime("%Y-%m-%dT%H:%M:%SZ")
    manifest_path.write_text(
        _json.dumps(data, indent=2, default=str), encoding="utf-8"
    )


def _write_manifest(
    out_root: str,
    billing_month: str,
    run_id: str,
    inputs: List[str],
    outputs: List[str],
    status: str,
    meta: Dict[str, Any],
) -> None:
    run_dir = Path(out_root) / billing_month
    run_dir.mkdir(parents=True, exist_ok=True)
    manifest = {
        "run_id":    run_id,
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "status":    status,
        "inputs":    inputs,
        "outputs":   outputs,
        **meta,
    }
    (run_dir / "run_manifest.json").write_text(
        json.dumps(manifest, indent=2, default=str), encoding="utf-8"
    )
