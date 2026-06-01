"""Read All-Wave and sprint dashboard workbooks into normalized target rows."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from triage.cybernet_targets.config import load_scope, normalize_site
from triage.cybernet_targets.models import TargetRow, target_key


def _require_openpyxl():
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from e
    return load_workbook


def _norm_wave(val: Any) -> Optional[int]:
    if val is None or val == "":
        return None
    try:
        return int(float(str(val).strip()))
    except (TypeError, ValueError):
        return None


def _row_dict(headers: List[str], row: tuple, row_num: int) -> Dict[str, Any]:
    d: Dict[str, Any] = {"_row_num": row_num}
    for i, h in enumerate(headers):
        if h:
            d[str(h).strip()] = row[i] if i < len(row) else None
    return d


def read_sheet_table(
    path: str | Path,
    sheet: str,
    header_row: int,
    data_start_row: int,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    load_workbook = _require_openpyxl()
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet not found: {sheet} in {path}")
    ws = wb[sheet]
    hdr_row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = [str(c).strip() if c is not None else "" for c in hdr_row]
    rows: List[Dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        if not any(c is not None and str(c).strip() for c in row):
            continue
        rows.append(_row_dict(headers, row, i))
    wb.close()
    return headers, rows


def read_all_wave_workbook(path: str | Path, scope: Dict[str, Any] | None = None) -> Dict[str, Any]:
    scope = scope or load_scope()
    anchors = scope["reader_anchors"]
    nc = anchors["neuron_cybernet"]
    ane = anchors["ane_ambulatory"]
    nc_headers, nc_rows = read_sheet_table(path, nc["sheet"], nc["header_row"], nc["data_start_row"])
    ane_headers, ane_rows = read_sheet_table(path, ane["sheet"], ane["header_row"], ane["data_start_row"])
    return {
        "path": str(Path(path).resolve()),
        "neuron_cybernet": {"headers": nc_headers, "rows": nc_rows},
        "ane_ambulatory": {"headers": ane_headers, "rows": ane_rows},
    }


def extract_wave3_targets(
    all_wave_data: Dict[str, Any],
    scope: Dict[str, Any] | None = None,
) -> List[TargetRow]:
    scope = scope or load_scope()
    aliases = scope.get("site_aliases", {})
    active_waves = set(scope.get("active_waves", [3]))
    active_scope = set(scope.get("active_scope", []))
    excluded_sites = {"Out of Scop HH", "Out of Scope HH"}
    source_path = all_wave_data["path"]
    rows = all_wave_data["neuron_cybernet"]["rows"]
    out: List[TargetRow] = []
    for rd in rows:
        wave = _norm_wave(rd.get("Wave"))
        if wave not in active_waves:
            continue
        source_site = str(rd.get("Site") or "").strip()
        if source_site in excluded_sites:
            continue
        site = normalize_site(source_site, aliases)
        if site not in active_scope:
            continue
        location = str(rd.get("Location") or "").strip()
        hostname = str(rd.get("PC Name") or rd.get("Hostname") or "").strip()
        device = str(rd.get("Device Type") or "").strip()
        row_num = str(rd.get("_row_num", ""))
        tr = TargetRow(
            target_id=target_key(site, location, "Neuron Cybernet", row_num),
            sprint_scope=site,
            wave=str(wave),
            site=site,
            source_site=source_site,
            location=location,
            source_workbook=Path(source_path).name,
            source_sheet="Neuron Cybernet",
            source_row=row_num,
            target_type=device or "Cybernet",
            cybernet_count=1 if device.lower() == "cybernet" else 0,
            neuron_count=1,
            hostname=hostname,
            kit_required="yes" if device.lower() == "cybernet" else "",
        )
        tr.apply_readiness()
        out.append(tr)
    return out


def read_sprint_dashboard(
    path: str | Path,
    scope: Dict[str, Any] | None = None,
) -> Dict[str, List[Dict[str, Any]]]:
    scope = scope or load_scope()
    tabs = scope["reader_anchors"].get("sprint_site_tabs", scope.get("active_scope", []))
    load_workbook = _require_openpyxl()
    wb = load_workbook(path, read_only=True, data_only=True)
    result: Dict[str, List[Dict[str, Any]]] = {}
    for tab in tabs:
        if tab not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sprint tab missing: {tab} in {path}")
        ws = wb[tab]
        hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(c).strip() if c is not None else "" for c in hdr]
        rows: List[Dict[str, Any]] = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            d = _row_dict(headers, row, i)
            d["_tab"] = tab
            rows.append(d)
        result[tab] = rows
    wb.close()
    return result
