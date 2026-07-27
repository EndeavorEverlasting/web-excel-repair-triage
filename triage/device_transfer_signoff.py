"""Generate shipping-ready Device Transfer / Stock Sign-Off workbooks.

The site config is the sole authority for shipment rows. The generator never
adds accessories, part numbers, quantities, or site metadata that are not
explicitly supplied. Serialized device rows are reconciled against a source
configuration workbook before output is written.
"""
from __future__ import annotations

import datetime as dt
import hashlib
import json
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Mapping, Sequence

from triage.nw_prj_neuron_track_hours.exporter import _repair_inlinestr

TITLE = "Device Transfer / Stock Sign-Off"
SHEET_NAME = "Sign-Off"
ARTIFACT_FAMILY = "device_transfer_signoff"
FORBIDDEN_OUTPUT_ROOTS = {"candidates", "active"}

_DARK = "111827"
_BLUE = "25608A"
_LIGHT = "F8FAFC"
_LINE = "CBD5E1"
_WHITE = "FFFFFF"
_MUTED = "D1D5DB"


class SignOffContractError(ValueError):
    """Raised when source data or site configuration violates the sign-off contract."""


@dataclass(frozen=True)
class SerializedItem:
    item: str
    qty: int
    serial_source: str
    serial_header: str
    serials: tuple[str, ...]


@dataclass(frozen=True)
class SignOffResult:
    workbook: Path
    manifest: Path
    preflight: Path
    report: Dict[str, Any]


def _require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc
    return Workbook, load_workbook, Alignment, Border, Font, PatternFill, Side, get_column_letter


def _read_json(path: Path) -> Dict[str, Any]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise SignOffContractError(f"site config not found: {path}") from exc
    except json.JSONDecodeError as exc:
        raise SignOffContractError(f"site config is not valid JSON: {path}: {exc}") from exc
    if not isinstance(data, dict):
        raise SignOffContractError("site config root must be an object")
    return data


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _positive_int(value: Any, label: str) -> int:
    try:
        ivalue = int(value)
    except (TypeError, ValueError) as exc:
        raise SignOffContractError(f"{label} must be an integer") from exc
    if ivalue <= 0:
        raise SignOffContractError(f"{label} must be > 0")
    return ivalue


def validate_config(config: Mapping[str, Any]) -> None:
    if config.get("artifact_family") != ARTIFACT_FAMILY:
        raise SignOffContractError(f"artifact_family must be {ARTIFACT_FAMILY!r}")
    site = config.get("site")
    if not isinstance(site, Mapping):
        raise SignOffContractError("site must be an object")
    required_site = (
        "name",
        "code",
        "address",
        "poc",
        "delivery_date",
        "delivery_time",
        "origin",
        "prepared_by",
        "signoff_id",
    )
    for key in required_site:
        if not _text(site.get(key)):
            raise SignOffContractError(f"site.{key} is required")
    try:
        dt.date.fromisoformat(_text(site["delivery_date"]))
    except ValueError as exc:
        raise SignOffContractError("site.delivery_date must be YYYY-MM-DD") from exc

    source = config.get("source")
    if not isinstance(source, Mapping):
        raise SignOffContractError("source must be an object")
    if not _text(source.get("device_type_header")):
        raise SignOffContractError("source.device_type_header is required")

    shipment = config.get("shipment")
    if not isinstance(shipment, list) or not shipment:
        raise SignOffContractError("shipment must be a non-empty array")

    seen: set[str] = set()
    serialized = 0
    for index, raw in enumerate(shipment, start=1):
        if not isinstance(raw, Mapping):
            raise SignOffContractError(f"shipment[{index}] must be an object")
        item = _text(raw.get("item"))
        if not item:
            raise SignOffContractError(f"shipment[{index}].item is required")
        key = item.casefold()
        if key in seen:
            raise SignOffContractError(f"duplicate shipment item: {item}")
        seen.add(key)
        _positive_int(raw.get("qty"), f"shipment[{index}].qty")
        serial_source = _text(raw.get("serial_source"))
        serial_header = _text(raw.get("serial_header"))
        if bool(serial_source) != bool(serial_header):
            raise SignOffContractError(
                f"shipment[{index}] must set both serial_source and serial_header, or neither"
            )
        if serial_source:
            serialized += 1
    if serialized > 2:
        raise SignOffContractError(
            "one-page sign-off supports at most two serialized shipment classes"
        )


def _select_source_sheet(wb, config: Mapping[str, Any], required_headers: Sequence[str]):
    source = config["source"]
    explicit = _text(source.get("sheet"))
    if explicit:
        if explicit not in wb.sheetnames:
            raise SignOffContractError(f"source sheet not found: {explicit}")
        ws = wb[explicit]
        headers = {_text(c.value) for c in ws[1] if _text(c.value)}
        missing = [h for h in required_headers if h not in headers]
        if missing:
            raise SignOffContractError(
                f"source sheet {explicit!r} is missing headers: {missing}"
            )
        return ws

    for ws in wb.worksheets:
        headers = {_text(c.value) for c in ws[1] if _text(c.value)}
        if all(h in headers for h in required_headers):
            return ws
    raise SignOffContractError(
        f"no worksheet contains the required headers: {list(required_headers)}"
    )


def _extract_serialized_items(
    source_workbook: Path,
    config: Mapping[str, Any],
) -> tuple[str, List[SerializedItem]]:
    _, load_workbook, *_ = _require_openpyxl()
    if not source_workbook.is_file():
        raise SignOffContractError(f"source workbook not found: {source_workbook}")

    serialized_specs = [
        row for row in config["shipment"] if _text(row.get("serial_source"))
    ]
    if not serialized_specs:
        return "", []

    device_type_header = _text(config["source"]["device_type_header"])
    required_headers = [device_type_header] + [
        _text(row["serial_header"]) for row in serialized_specs
    ]

    wb = load_workbook(source_workbook, data_only=True, read_only=True)
    try:
        ws = _select_source_sheet(wb, config, required_headers)
        header_map = {
            _text(cell.value): idx
            for idx, cell in enumerate(ws[1], start=1)
            if _text(cell.value)
        }

        results: List[SerializedItem] = []
        for raw in serialized_specs:
            item = _text(raw["item"])
            qty = _positive_int(raw["qty"], f"shipment.{item}.qty")
            source_value = _text(raw["serial_source"])
            serial_header = _text(raw["serial_header"])
            serials: List[str] = []
            seen: set[str] = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                dtype = _text(row[header_map[device_type_header] - 1])
                if dtype.casefold() != source_value.casefold():
                    continue
                serial = _text(row[header_map[serial_header] - 1])
                if not serial:
                    continue
                if serial in seen:
                    raise SignOffContractError(
                        f"duplicate {item} serial in source: {serial}"
                    )
                seen.add(serial)
                serials.append(serial)

            if len(serials) != qty:
                raise SignOffContractError(
                    f"{item} quantity/serial mismatch: shipment qty={qty}, "
                    f"source serials={len(serials)}"
                )
            results.append(
                SerializedItem(
                    item=item,
                    qty=qty,
                    serial_source=source_value,
                    serial_header=serial_header,
                    serials=tuple(serials),
                )
            )
        return ws.title, results
    finally:
        wb.close()


def _assert_safe_output(output: Path, source_workbook: Path, config_path: Path) -> None:
    resolved = output.resolve()
    if resolved in {source_workbook.resolve(), config_path.resolve()}:
        raise SignOffContractError("output must not overwrite an input")
    lower_parts = {part.casefold() for part in resolved.parts}
    if lower_parts & FORBIDDEN_OUTPUT_ROOTS:
        raise SignOffContractError(
            "generated sign-off output may not be written under Candidates/ or Active/"
        )


def _serial_layout(shipment_count: int, serialized: Sequence[SerializedItem]) -> Dict[str, int]:
    shipment_header = 8
    shipment_start = shipment_header + 1
    shipment_end = shipment_header + shipment_count
    serial_header = shipment_end + 3
    serial_start = serial_header + 1
    serial_rows = max(
        (max(1, (item.qty + 1) // 2) for item in serialized),
        default=1,
    )
    verification_header = serial_start + serial_rows + 2
    signature_header = verification_header + 6
    return {
        "shipment_header": shipment_header,
        "shipment_start": shipment_start,
        "shipment_end": shipment_end,
        "serial_header": serial_header,
        "serial_start": serial_start,
        "serial_rows": serial_rows,
        "verification_header": verification_header,
        "signature_header": signature_header,
        "last_row": signature_header + 4,
    }


def _write_serial_panel(
    ws,
    panel_index: int,
    item: SerializedItem,
    layout: Mapping[str, int],
    *,
    dark_fill,
    blue_font,
    body_border,
    center,
) -> None:
    base = 1 if panel_index == 0 else 7
    end = base + 5
    header_row = layout["serial_header"]
    ws.merge_cells(start_row=header_row, start_column=base, end_row=header_row, end_column=end)
    header = ws.cell(header_row, base, f"{item.item} Serial Numbers ({item.qty})")
    header.fill = dark_fill
    header.font = blue_font
    header.alignment = center

    left_count = (item.qty + 1) // 2
    for idx, serial in enumerate(item.serials):
        right = idx >= left_count
        offset = idx - left_count if right else idx
        row = layout["serial_start"] + offset
        number_col = base + (3 if right else 0)
        serial_start_col = number_col + 1
        serial_end_col = number_col + 2
        ws.cell(row, number_col, idx + 1).alignment = center
        ws.merge_cells(
            start_row=row,
            start_column=serial_start_col,
            end_row=row,
            end_column=serial_end_col,
        )
        ws.cell(row, serial_start_col, serial)
        for col in range(number_col, serial_end_col + 1):
            ws.cell(row, col).border = body_border


def build_workbook(
    config: Mapping[str, Any],
    serialized: Sequence[SerializedItem],
    output_path: Path,
) -> Dict[str, int]:
    (
        Workbook,
        _,
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
        get_column_letter,
    ) = _require_openpyxl()

    site = config["site"]
    shipment = config["shipment"]
    layout = _serial_layout(len(shipment), serialized)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.sheet_view.showGridLines = False

    dark_fill = PatternFill("solid", fgColor=_DARK)
    blue_fill = PatternFill("solid", fgColor=_BLUE)
    light_fill = PatternFill("solid", fgColor=_LIGHT)
    white_font = Font(color=_WHITE, bold=True)
    muted_font = Font(color=_MUTED)
    title_font = Font(color=_WHITE, bold=True, size=20)
    blue_font = Font(color="60A5FA", bold=True)
    bold = Font(bold=True)
    thin = Side(style="thin", color=_LINE)
    body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="center", wrap_text=True)

    ws.merge_cells("A1:L2")
    ws["A1"] = TITLE
    for row in ws["A1:L2"]:
        for cell in row:
            cell.fill = dark_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")

    ws.merge_cells("A3:L3")
    ws["A3"] = (
        "Epic Device Integration Logistics"
        f" | Delivery {site['delivery_date']} at {site['delivery_time']}"
        f" | Sign-Off ID: {site['signoff_id']}"
    )
    for cell in ws[3][:12]:
        cell.fill = dark_fill
    ws["A3"].font = muted_font

    details = [
        ("From / Origin", site["origin"]),
        ("To / Destination", site["name"]),
        ("Site POC", site["poc"]),
        ("Delivery Date", site["delivery_date"]),
        ("Delivery Address", site["address"]),
        ("Delivery Time", site["delivery_time"]),
        ("Prepared By", site["prepared_by"]),
        ("Site Code", site["code"]),
    ]
    for idx, (label, value) in enumerate(details):
        row = 5 + idx // 4
        slot = idx % 4
        col = 1 + slot * 3
        ws.cell(row, col, label).fill = blue_fill
        ws.cell(row, col).font = white_font
        ws.cell(row, col).alignment = wrap
        ws.merge_cells(start_row=row, start_column=col + 1, end_row=row, end_column=col + 2)
        ws.cell(row, col + 1, value).fill = light_fill
        ws.cell(row, col + 1).alignment = wrap

    header_row = layout["shipment_header"]
    for col, label in enumerate(["#", "Item Sent", "Qty", "Verification / Notes"], start=1):
        cell = ws.cell(header_row, col, label)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = body_border

    for idx, raw in enumerate(shipment, start=1):
        row = header_row + idx
        values = [idx, _text(raw["item"]), _positive_int(raw["qty"], f"shipment[{idx}].qty"), _text(raw.get("note"))]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = body_border
            cell.alignment = wrap if col in (2, 4) else center

    if serialized:
        for panel_index, item in enumerate(serialized):
            _write_serial_panel(
                ws,
                panel_index,
                item,
                layout,
                dark_fill=dark_fill,
                blue_font=blue_font,
                body_border=body_border,
                center=center,
            )
    else:
        ws.merge_cells(
            start_row=layout["serial_header"],
            start_column=1,
            end_row=layout["serial_header"],
            end_column=12,
        )
        ws.cell(layout["serial_header"], 1, "Serialized Device Detail: Not Required")
        ws.cell(layout["serial_header"], 1).fill = dark_fill
        ws.cell(layout["serial_header"], 1).font = blue_font

    vrow = layout["verification_header"]
    ws.merge_cells(start_row=vrow, start_column=1, end_row=vrow, end_column=12)
    ws.cell(vrow, 1, "Verification / Exceptions").fill = dark_fill
    ws.cell(vrow, 1).font = blue_font
    checks = [
        "☐ Quantities counted against shipment list",
        "☐ Serialized devices matched to source list",
        "☐ Items physically inspected at handoff",
        "☐ Exceptions documented below",
    ]
    for idx, label in enumerate(checks):
        row = vrow + 1 + idx // 2
        col = 1 if idx % 2 == 0 else 7
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 5)
        ws.cell(row, col, label).alignment = wrap
    ws.cell(vrow + 3, 1, "Exception Notes:").font = bold
    ws.merge_cells(start_row=vrow + 3, start_column=2, end_row=vrow + 3, end_column=6)
    ws.cell(vrow + 3, 2, "________________________________________________")
    ws.cell(vrow + 3, 7, "Additional Notes:").font = bold
    ws.merge_cells(start_row=vrow + 3, start_column=8, end_row=vrow + 3, end_column=12)
    ws.cell(vrow + 3, 8, "____________________________________________")

    srow = layout["signature_header"]
    ws.merge_cells(start_row=srow, start_column=1, end_row=srow, end_column=12)
    ws.cell(srow, 1, "Signatures").fill = dark_fill
    ws.cell(srow, 1).font = blue_font
    for idx, label in enumerate(["Role", "Printed Name", "Signature", "Date"]):
        col = 1 + idx * 3
        ws.merge_cells(start_row=srow + 1, start_column=col, end_row=srow + 1, end_column=col + 2)
        cell = ws.cell(srow + 1, col, label)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center

    roles = [("Released By", ""), ("Received By", site["poc"]), ("Verified By", "")]
    for ridx, (role, printed) in enumerate(roles, start=2):
        row = srow + ridx
        values = [role, printed, "", site["delivery_date"]]
        for idx, value in enumerate(values):
            col = 1 + idx * 3
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
            cell = ws.cell(row, col, value)
            cell.fill = light_fill
            cell.border = body_border
            cell.alignment = wrap

    for idx, width in enumerate([7, 22, 9, 20, 12, 12, 7, 22, 9, 7, 22, 9], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 12
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A8"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.print_area = f"A1:L{layout['last_row']}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    _repair_inlinestr(str(output_path))
    return layout


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _panel_serials(ws, panel_index: int, item: SerializedItem, layout: Mapping[str, int]) -> List[str]:
    base = 1 if panel_index == 0 else 7
    left_count = (item.qty + 1) // 2
    out: List[str] = []
    for idx in range(item.qty):
        right = idx >= left_count
        offset = idx - left_count if right else idx
        row = layout["serial_start"] + offset
        number_col = base + (3 if right else 0)
        serial = _text(ws.cell(row, number_col + 1).value)
        if serial:
            out.append(serial)
    return out


def preflight_workbook(
    workbook_path: Path,
    config: Mapping[str, Any],
    serialized: Sequence[SerializedItem],
    layout: Mapping[str, int],
) -> Dict[str, Any]:
    _, load_workbook, *_ = _require_openpyxl()
    result: Dict[str, Any] = {
        "artifact_family": ARTIFACT_FAMILY,
        "artifact": workbook_path.name,
        "exists": workbook_path.is_file(),
        "zip_valid": False,
        "sheet_names": [],
        "formula_cells": [],
        "inline_str": False,
        "has_calc_chain": False,
        "has_external_links": False,
        "shipment_exact": False,
        "serials_exact": False,
        "metadata_complete": False,
        "preflight_pass": False,
    }
    if not workbook_path.is_file():
        return result

    try:
        with zipfile.ZipFile(workbook_path, "r") as z:
            result["zip_valid"] = z.testzip() is None
            names = z.namelist()
            result["has_calc_chain"] = "xl/calcChain.xml" in names
            result["has_external_links"] = any("externalLink" in name for name in names)
            for name in names:
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                    if b"inlineStr" in z.read(name):
                        result["inline_str"] = True
                        break
    except zipfile.BadZipFile:
        return result

    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        result["sheet_names"] = list(wb.sheetnames)
        if wb.sheetnames != [SHEET_NAME]:
            return result
        ws = wb[SHEET_NAME]
        result["formula_cells"] = [
            cell.coordinate
            for row in ws.iter_rows()
            for cell in row
            if cell.data_type == "f"
        ]

        expected_ship = [
            (_text(row["item"]), _positive_int(row["qty"], f"shipment.{row.get('item')}.qty"))
            for row in config["shipment"]
        ]
        actual_ship: List[tuple[str, int]] = []
        for row in range(layout["shipment_start"], layout["shipment_end"] + 1):
            item = _text(ws.cell(row, 2).value)
            qty = ws.cell(row, 3).value
            if item:
                actual_ship.append((item, int(qty)))
        result["shipment_exact"] = actual_ship == expected_ship

        result["serials_exact"] = all(
            _panel_serials(ws, panel_index, item, layout) == list(item.serials)
            for panel_index, item in enumerate(serialized)
        )
        metadata_values = [
            _text(ws.cell(5, 2).value),
            _text(ws.cell(5, 5).value),
            _text(ws.cell(5, 8).value),
            _text(ws.cell(5, 11).value),
            _text(ws.cell(6, 2).value),
            _text(ws.cell(6, 5).value),
            _text(ws.cell(6, 8).value),
            _text(ws.cell(6, 11).value),
        ]
        result["metadata_complete"] = all(metadata_values)
        result["title_ok"] = _text(ws["A1"].value) == TITLE
        result["signoff_id_present"] = _text(config["site"]["signoff_id"]) in _text(ws["A3"].value)
    finally:
        wb.close()

    result["preflight_pass"] = bool(
        result["zip_valid"]
        and result["sheet_names"] == [SHEET_NAME]
        and not result["formula_cells"]
        and not result["inline_str"]
        and not result["has_calc_chain"]
        and not result["has_external_links"]
        and result["shipment_exact"]
        and result["serials_exact"]
        and result["metadata_complete"]
        and result.get("title_ok")
        and result.get("signoff_id_present")
    )
    return result


def run(
    source_workbook: str,
    site_config: str,
    *,
    output: str | None = None,
    out_dir: str = "Outputs/device_transfer_signoff",
) -> SignOffResult:
    source_path = Path(source_workbook)
    config_path = Path(site_config)
    config = _read_json(config_path)
    validate_config(config)

    site = config["site"]
    if output:
        workbook_path = Path(output)
    else:
        date_token = _text(site["delivery_date"]).replace("-", "")
        safe_code = re.sub(r"[^A-Za-z0-9_-]+", "_", _text(site["code"])).strip("_") or "SITE"
        workbook_path = Path(out_dir) / f"{safe_code}_Device_Transfer_SignOff_{date_token}.xlsx"

    _assert_safe_output(workbook_path, source_path, config_path)
    source_sheet, serialized = _extract_serialized_items(source_path, config)
    layout = build_workbook(config, serialized, workbook_path)
    preflight = preflight_workbook(workbook_path, config, serialized, layout)

    preflight_path = workbook_path.with_name(workbook_path.stem + "_preflight.json")
    manifest_path = workbook_path.with_name(workbook_path.stem + "_manifest.json")
    preflight_path.write_text(json.dumps(preflight, indent=2), encoding="utf-8")

    manifest = {
        "artifact_family": ARTIFACT_FAMILY,
        "schema_version": "device-transfer-signoff-manifest/v1",
        "generated_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
        "source_workbook": str(source_path.resolve()),
        "site_config": str(config_path.resolve()),
        "source_sheet": source_sheet,
        "site_code": site["code"],
        "site_name": site["name"],
        "signoff_id": site["signoff_id"],
        "shipment": [
            {"item": _text(row["item"]), "qty": _positive_int(row["qty"], "shipment.qty")}
            for row in config["shipment"]
        ],
        "serialized_counts": {item.item: item.qty for item in serialized},
        "outputs": {
            "workbook": str(workbook_path.resolve()),
            "preflight": str(preflight_path.resolve()),
            "manifest": str(manifest_path.resolve()),
        },
        "sha256": _sha256(workbook_path),
        "preflight_pass": bool(preflight["preflight_pass"]),
        "proof_ceiling": (
            "Repository generator, source reconciliation, package, and structural proof only; "
            "Excel for Web and physical handoff remain operator/runtime gates."
        ),
    }
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    if not preflight["preflight_pass"]:
        raise SignOffContractError(
            f"generated workbook failed preflight; inspect {preflight_path}"
        )

    return SignOffResult(
        workbook=workbook_path,
        manifest=manifest_path,
        preflight=preflight_path,
        report=manifest,
    )
