"""Conditional-formatting forensics for the May roster Live sheet.

Two responsibilities:

1. ``diff_cf`` - structural diff of CF blocks/rules between a candidate
   (current/bad) workbook and a visually-correct reference workbook.
2. ``sunday_bleed_report`` - a month-aware checker that flags CF rules which
   can cause a Sunday cell to highlight when the adjacent Monday cell is
   populated. It never hardcodes a single column pair: it derives the
   Sunday/Monday boundaries for the month from the live-sheet date headers.

All CF extraction reuses :mod:`triage.cf_engine` (byte/string-level, no
openpyxl reserialization). Header/date mapping uses openpyxl read-only.
"""
from __future__ import annotations

import calendar
import re
from dataclasses import dataclass, field
from datetime import date
from typing import Dict, List, Optional, Tuple

from triage.cf_engine import CFBlock, CFDictionary, extract_cf_dictionary
from triage.xlsx_utils import col_to_num, num_to_col

# Always-true rule formulas (paint regardless of cell content).
_ALWAYS_TRUE_FORMULAS = {"1", "true", "true()"}

_DATE_HEADER = re.compile(
    r"^([A-Za-z]+)\s+(\d{1,2})\s*[-\u2013]\s*(Clock\s*In|Clock\s*Out)\s*$",
    re.IGNORECASE,
)
_MONTH_ABBREVS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
# A column reference inside a CF formula, e.g. AI3, $AJ$3, AK10.
_FORMULA_COL_REF = re.compile(r"\$?([A-Z]{1,3})\$?\d+")


# ───────────────────────── column / range helpers ─────────────────────────


def expand_sqref_columns(sqref: str) -> List[str]:
    """Return the sorted set of column letters covered by a ``sqref``.

    Handles space-separated multi-range sqrefs and single cells.
    """
    cols: set[str] = set()
    for token in (sqref or "").split():
        m = re.match(r"^\$?([A-Z]+)\$?\d+(?::\$?([A-Z]+)\$?\d+)?$", token)
        if not m:
            continue
        c1 = m.group(1)
        c2 = m.group(2) or c1
        n1, n2 = col_to_num(c1), col_to_num(c2)
        for n in range(min(n1, n2), max(n1, n2) + 1):
            cols.add(num_to_col(n))
    return sorted(cols, key=col_to_num)


def formula_column_refs(formula: str) -> set[str]:
    """Column letters referenced by a CF formula (entities tolerated)."""
    return {m.group(1) for m in _FORMULA_COL_REF.finditer(formula or "")}


# ───────────────────────── date / column mapping ─────────────────────────


@dataclass
class DateColumns:
    day: date
    in_col: Optional[str] = None
    out_col: Optional[str] = None

    @property
    def columns(self) -> List[str]:
        return [c for c in (self.in_col, self.out_col) if c]


def map_live_date_columns(path: str, sheet_label: str, year: int) -> Dict[date, DateColumns]:
    """Map each date in a Live sheet to its clock-in/out column letters.

    Reads row-2 headers like ``"May 17 - Clock In"`` via openpyxl read-only.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = None
        for name in wb.sheetnames:
            if name.strip().lower() == sheet_label.strip().lower():
                ws = wb[name]
                break
        if ws is None:
            return {}
        result: Dict[date, DateColumns] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(2, c).value
            if not isinstance(v, str):
                continue
            mm = _DATE_HEADER.match(v.strip())
            if not mm:
                continue
            mon = _MONTH_ABBREVS.get(mm.group(1)[:3].lower())
            if mon is None:
                continue
            try:
                d = date(year, mon, int(mm.group(2)))
            except ValueError:
                continue
            letter = get_column_letter(c)
            dc = result.setdefault(d, DateColumns(day=d))
            if "in" in mm.group(3).lower():
                dc.in_col = letter
            else:
                dc.out_col = letter
        return result
    finally:
        wb.close()


def sunday_monday_boundaries(year: int, month: int) -> List[Tuple[date, Optional[date]]]:
    """All (Sunday, following-Monday-or-None) pairs for a month.

    The Monday is ``None`` when it falls outside the month (e.g. a trailing
    Sunday with no in-sheet Monday).
    """
    pairs: List[Tuple[date, Optional[date]]] = []
    days_in_month = calendar.monthrange(year, month)[1]
    for d in range(1, days_in_month + 1):
        cur = date(year, month, d)
        if cur.weekday() == 6:  # Sunday
            nxt = date(year, month, d + 1) if d + 1 <= days_in_month else None
            if nxt is not None and nxt.weekday() != 0:
                nxt = None
            pairs.append((cur, nxt))
    return pairs


# ───────────────────────── Sunday-bleed report ─────────────────────────


@dataclass
class BleedFinding:
    kind: str
    sheet: str
    sqref: str
    sunday: str
    monday: str
    detail: str
    rule_formula: str = ""
    priority: int = 0


@dataclass
class SundayBleedReport:
    path: str
    sheet: str
    year: int
    month: int
    boundaries: List[dict] = field(default_factory=list)
    findings: List[BleedFinding] = field(default_factory=list)

    @property
    def clean(self) -> bool:
        return not self.findings

    def to_dict(self) -> dict:
        return {
            "path": self.path,
            "sheet": self.sheet,
            "year": self.year,
            "month": self.month,
            "clean": self.clean,
            "boundaries": self.boundaries,
            "finding_count": len(self.findings),
            "findings": [f.__dict__ for f in self.findings],
        }


def sunday_bleed_report(
    path: str,
    sheet_label: str,
    year: int,
    month: int,
) -> SundayBleedReport:
    """Flag CF rules that can bleed Sunday highlighting from Monday data."""
    rpt = SundayBleedReport(path=path, sheet=sheet_label, year=year, month=month)

    date_cols = map_live_date_columns(path, sheet_label, year)
    cfd = extract_cf_dictionary(path)
    blocks = [b for b in cfd.blocks if (b.sheet_name or "").strip().lower() == sheet_label.strip().lower()]

    for sunday, monday in sunday_monday_boundaries(year, month):
        sun_cols = set(date_cols.get(sunday, DateColumns(sunday)).columns)
        mon_cols = set(date_cols.get(monday, DateColumns(monday)).columns) if monday else set()
        rpt.boundaries.append({
            "sunday": sunday.isoformat(),
            "monday": monday.isoformat() if monday else None,
            "sunday_columns": sorted(sun_cols, key=col_to_num),
            "monday_columns": sorted(mon_cols, key=col_to_num),
        })
        if not sun_cols:
            continue

        for b in blocks:
            block_cols = set(expand_sqref_columns(b.sqref))
            touches_sunday = bool(block_cols & sun_cols)
            if not touches_sunday:
                continue

            # 1. A single block range that spans Sunday AND Monday columns.
            if mon_cols and (block_cols & mon_cols) and _is_single_range(b.sqref):
                rpt.findings.append(BleedFinding(
                    kind="merged_range_crosses_sunday_monday",
                    sheet=sheet_label, sqref=b.sqref,
                    sunday=sunday.isoformat(), monday=monday.isoformat() if monday else "",
                    detail=f"sqref spans Sunday {sorted(sun_cols, key=col_to_num)} "
                           f"and Monday {sorted(mon_cols, key=col_to_num)} in one range",
                ))

            for r in b.rules:
                refs = formula_column_refs(r.formula) | formula_column_refs(r.formula2)
                # 2. A Sunday-covering rule that references a Monday column.
                if mon_cols and (refs & mon_cols):
                    rpt.findings.append(BleedFinding(
                        kind="sunday_rule_references_monday",
                        sheet=sheet_label, sqref=b.sqref,
                        sunday=sunday.isoformat(), monday=monday.isoformat() if monday else "",
                        detail=f"rule on Sunday columns references Monday column(s) "
                               f"{sorted(refs & mon_cols, key=col_to_num)}",
                        rule_formula=r.formula, priority=r.priority,
                    ))
                # 3. Always-true blanket rule over Sunday columns.
                if (r.formula or "").strip().lower() in _ALWAYS_TRUE_FORMULAS:
                    rpt.findings.append(BleedFinding(
                        kind="always_true_blanket_over_sunday",
                        sheet=sheet_label, sqref=b.sqref,
                        sunday=sunday.isoformat(), monday=monday.isoformat() if monday else "",
                        detail="always-true rule paints Sunday columns regardless of content",
                        rule_formula=r.formula, priority=r.priority,
                    ))
    return rpt


def _is_single_range(sqref: str) -> bool:
    return len((sqref or "").split()) == 1 and ":" in (sqref or "")


# ───────────────────────── CF diff ─────────────────────────


def _rule_signature(rule) -> tuple:
    return (rule.rule_type, rule.operator, rule.priority, rule.dxf_id,
            (rule.formula or "").strip(), (rule.formula2 or "").strip())


def _blocks_for_sheet(cfd: CFDictionary, sheet_label: str) -> Dict[str, CFBlock]:
    out: Dict[str, CFBlock] = {}
    for b in cfd.blocks:
        if (b.sheet_name or "").strip().lower() == sheet_label.strip().lower():
            out[b.sqref] = b
    return out


@dataclass
class CFDiffReport:
    sheet: str
    candidate_path: str
    reference_path: str
    only_in_candidate: List[dict] = field(default_factory=list)
    only_in_reference: List[dict] = field(default_factory=list)
    changed: List[dict] = field(default_factory=list)

    @property
    def identical(self) -> bool:
        return not (self.only_in_candidate or self.only_in_reference or self.changed)

    def to_dict(self) -> dict:
        return {
            "sheet": self.sheet,
            "candidate_path": self.candidate_path,
            "reference_path": self.reference_path,
            "identical": self.identical,
            "only_in_candidate": self.only_in_candidate,
            "only_in_reference": self.only_in_reference,
            "changed": self.changed,
        }


def diff_cf(candidate_path: str, reference_path: str, sheet_label: str) -> CFDiffReport:
    """Diff CF blocks of ``sheet_label`` between candidate and reference."""
    cand = _blocks_for_sheet(extract_cf_dictionary(candidate_path), sheet_label)
    ref = _blocks_for_sheet(extract_cf_dictionary(reference_path), sheet_label)
    rpt = CFDiffReport(sheet=sheet_label, candidate_path=candidate_path, reference_path=reference_path)

    for sqref, b in cand.items():
        if sqref not in ref:
            rpt.only_in_candidate.append({"sqref": sqref, "rules": len(b.rules)})
    for sqref, b in ref.items():
        if sqref not in cand:
            rpt.only_in_reference.append({"sqref": sqref, "rules": len(b.rules)})
    for sqref in set(cand) & set(ref):
        cand_sigs = [_rule_signature(r) for r in cand[sqref].rules]
        ref_sigs = [_rule_signature(r) for r in ref[sqref].rules]
        if cand_sigs != ref_sigs:
            rpt.changed.append({
                "sqref": sqref,
                "candidate_rules": [list(s) for s in cand_sigs],
                "reference_rules": [list(s) for s in ref_sigs],
            })
    return rpt
