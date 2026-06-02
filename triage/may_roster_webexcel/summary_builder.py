"""Roster reading + values-only share-safe summary generation.

``read_live_records`` extracts one record per (tech, date) punch pair from a
Live roster sheet. ``build_sharesafe_summary`` writes a values-only workbook
(no formulas, no external links, no hidden internal tabs) suitable for
leadership sharing. Purity is asserted afterwards via
:func:`triage.may_roster_webexcel.package_checks.run_package_preflight`.
"""
from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from triage.may_roster_webexcel.roster_rules import (
    OVERNIGHT_LABEL,
    STATUS_MALFORMED,
    STATUS_OVERNIGHT,
    UNASSIGNED_LABEL,
    build_unassigned_rows,
    classify_punch,
    is_unassigned,
)
from triage.xlsx_utils import fix_inlinestr

_DATE_HEADER = re.compile(
    r"^([A-Za-z]+)\s+(\d{1,2})\s*[-\u2013]\s*(Clock\s*In|Clock\s*Out)\s*$",
    re.IGNORECASE,
)
_MONTH_ABBREVS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

# Share-safe workbook sheet order (values-only).
SHARESAFE_SHEETS = [
    "Executive Summary",
    "Project Summary",
    "Tech Project Summary",
    "Daily Summary",
    "Exceptions Summary",
    "Unassigned Hours Summary",
]


@dataclass
class LiveRecord:
    tech: str
    date: Optional[date]
    project: str
    clock_in_raw: Any
    clock_out_raw: Any
    paid_hours: float
    day_type: str
    status: str

    def as_dict(self) -> dict:
        return {
            "tech": self.tech,
            "date": self.date,
            "project": self.project,
            "paid_hours": self.paid_hours,
            "day_type": self.day_type,
            "status": self.status,
        }


def _worked_project_lookup(ws, year: int) -> Dict[tuple, str]:
    lut: Dict[tuple, str] = {}
    if ws is None:
        return lut
    date_cols: Dict[int, date] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if isinstance(v, datetime):
            date_cols[c] = v.date()
        elif isinstance(v, date):
            date_cols[c] = v
    for r in range(3, ws.max_row + 1):
        staff = ws.cell(r, 1).value
        if not staff or not str(staff).strip():
            continue
        for c, d in date_cols.items():
            val = ws.cell(r, c).value
            if val is not None and str(val).strip():
                lut[(str(staff).strip(), d)] = str(val).strip()
    return lut


def read_live_records(path: str, sheet_label: str, year: int) -> List[LiveRecord]:
    """Read per-(tech, date) records from a Live roster sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        ws = None
        for name in wb.sheetnames:
            if name.strip().lower() == sheet_label.strip().lower():
                ws = wb[name]
                break
        if ws is None:
            return []

        worked_ws = None
        wp_label = sheet_label.replace("Live", "Worked Projects", 1)
        for name in wb.sheetnames:
            if name.strip().lower() == wp_label.strip().lower():
                worked_ws = wb[name]
                break
        worked = _worked_project_lookup(worked_ws, year)

        date_to_cols: Dict[date, Dict[str, int]] = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(2, c).value
            if not isinstance(h, str):
                continue
            mm = _DATE_HEADER.match(h.strip())
            if not mm:
                continue
            mon = _MONTH_ABBREVS.get(mm.group(1)[:3].lower())
            if mon is None:
                continue
            try:
                d = date(year, mon, int(mm.group(2)))
            except ValueError:
                continue
            direction = "in" if "in" in mm.group(3).lower() else "out"
            date_to_cols.setdefault(d, {})[direction] = c

        records: List[LiveRecord] = []
        for r in range(3, ws.max_row + 1):
            staff_val = ws.cell(r, 1).value
            if not staff_val or str(staff_val).strip() in ("", "None"):
                continue
            if isinstance(staff_val, (int, float)):
                continue
            tech = str(staff_val).strip()
            default_proj = str(ws.cell(r, 2).value or "").strip()

            for d, dirs in sorted(date_to_cols.items()):
                in_raw = ws.cell(r, dirs["in"]).value if "in" in dirs else None
                out_raw = ws.cell(r, dirs["out"]).value if "out" in dirs else None
                cls = classify_punch(in_raw, out_raw, is_weekend=d.weekday() >= 5)
                project = worked.get((tech, d), "") or default_proj
                # NOTE: do not exempt calendar weekends here. Weekend *no-work*
                # is handled by paid_hours <= 0; paid weekend work with no
                # project is genuinely unassigned. Only PTO/unpaid markers
                # carried in the project text exempt a paid row.
                proj_l = str(project).strip().lower()
                day_type = "pto" if ("pto" in proj_l or "unpaid" in proj_l) else ""
                paid = cls.gross_hours if cls.status in ("ok", STATUS_OVERNIGHT) else 0.0
                records.append(LiveRecord(
                    tech=tech, date=d, project=project,
                    clock_in_raw=in_raw, clock_out_raw=out_raw,
                    paid_hours=paid, day_type=day_type, status=cls.status,
                ))
        return records
    finally:
        wb.close()


# ───────────────────────── share-safe summary ─────────────────────────


def _safe(value: Any) -> Any:
    """Neutralize any formula-like leading characters for values-only output."""
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def _write_sheet(ws, header: List[str], rows: List[List[Any]]) -> None:
    ws.append([_safe(h) for h in header])
    for row in rows:
        ws.append([_safe(v) for v in row])


def build_sharesafe_summary(records: List[LiveRecord], out_path: str) -> str:
    """Write a values-only share-safe summary workbook."""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    by_project: Dict[str, float] = defaultdict(float)
    by_tech_project: Dict[tuple, float] = defaultdict(float)
    by_day: Dict[Optional[date], float] = defaultdict(float)
    total_hours = 0.0
    exceptions: List[List[Any]] = []
    for rec in records:
        proj = rec.project or "(blank)"
        by_project[proj] += rec.paid_hours
        by_tech_project[(rec.tech, proj)] += rec.paid_hours
        by_day[rec.date] += rec.paid_hours
        total_hours += rec.paid_hours
        if rec.status == STATUS_MALFORMED:
            exceptions.append([rec.tech, rec.date.isoformat() if rec.date else "", "Malformed punch"])
        elif rec.status == STATUS_OVERNIGHT:
            exceptions.append([rec.tech, rec.date.isoformat() if rec.date else "", OVERNIGHT_LABEL])

    unassigned = build_unassigned_rows([r.as_dict() for r in records])

    ws = wb.create_sheet("Executive Summary")
    _write_sheet(ws, ["Metric", "Value"], [
        ["Total Paid Hours", round(total_hours, 2)],
        ["Distinct Techs", len({r.tech for r in records})],
        ["Distinct Projects", len(by_project)],
        ["Unassigned Rows", len(unassigned)],
        ["Exception Rows", len(exceptions)],
    ])

    ws = wb.create_sheet("Project Summary")
    _write_sheet(ws, ["Project", "Paid Hours"],
                 [[p, round(h, 2)] for p, h in sorted(by_project.items())])

    ws = wb.create_sheet("Tech Project Summary")
    _write_sheet(ws, ["Tech", "Project", "Paid Hours"],
                 [[t, p, round(h, 2)] for (t, p), h in sorted(by_tech_project.items())])

    ws = wb.create_sheet("Daily Summary")
    _write_sheet(ws, ["Date", "Paid Hours"],
                 [[d.isoformat() if d else "", round(h, 2)] for d, h in sorted(by_day.items(), key=lambda kv: (kv[0] or date.min))])

    ws = wb.create_sheet("Exceptions Summary")
    _write_sheet(ws, ["Tech", "Date", "Exception"], exceptions)

    ws = wb.create_sheet("Unassigned Hours Summary")
    _write_sheet(ws, ["Tech", "Date", "Actual Paid Hours", "Current Project / Assignment", "Status"],
                 [[u.tech, u.date.isoformat() if u.date else "", u.paid_hours, u.project, u.status] for u in unassigned])

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    fix_inlinestr(str(out))
    return str(out)
