"""Fail-closed readiness checks for month-specific Neuron Track Hours inputs.

This module inspects workbook structure only. It does not infer hours from
presence matrices, assignments, device counts, or other non-attendance data.
"""
from __future__ import annotations

import re
from calendar import month_name
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional

from triage.month_validation import validate_month_key

_DATE_HEADER = re.compile(
    r"^([A-Za-z]+)\s+(\d{1,2})\s*[-\u2013]\s*(Clock\s*In|Clock\s*Out)\s*$",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class MonthReadiness:
    schema: str
    month_key: str
    month_label: str
    source_path: str
    status: str
    attendance_authority: str
    live_sheet: str
    worked_projects_sheet: str
    presence_only_sheets: List[str]
    paired_clock_dates: int
    warnings: List[str]
    blockers: List[str]

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


def _find_live_sheet(wb, label: str):
    target = f"live - {label}".lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return wb[name]
    month_word, year = label.split()
    for name in wb.sheetnames:
        low = name.strip().lower()
        if low.startswith("live") and month_word.lower() in low and year in name:
            return wb[name]
    return None


def _find_worked_projects_sheet(wb, label: str) -> str:
    target = f"worked projects - {label}".lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return name
    month_word, year = label.split()
    for name in wb.sheetnames:
        low = name.strip().lower()
        if low.startswith("worked projects") and month_word.lower() in low and year in name:
            return name
    return ""


def _presence_only_sheets(wb, label: str) -> List[str]:
    month_word, year = label.split()
    matches: List[str] = []
    for name in wb.sheetnames:
        low = name.strip().lower()
        if month_word.lower() not in low or year not in name:
            continue
        if "attendance" in low and not low.startswith("live"):
            matches.append(name)
    return matches


def _paired_clock_dates(ws) -> int:
    if ws is None:
        return 0
    directions: Dict[str, set[str]] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(2, col).value
        if not isinstance(value, str):
            continue
        match = _DATE_HEADER.match(value.strip())
        if not match:
            continue
        key = f"{match.group(1)[:3].lower()}-{int(match.group(2)):02d}"
        direction = "in" if "in" in match.group(3).lower() else "out"
        directions.setdefault(key, set()).add(direction)
    return sum(1 for found in directions.values() if found == {"in", "out"})


def inspect_month_readiness(roster_path: str | Path, month_key: str) -> MonthReadiness:
    try:
        year, month = validate_month_key(month_key)
    except ValueError as exc:
        raise ValueError(str(exc)) from exc

    path = Path(roster_path)
    if not path.exists():
        raise FileNotFoundError(f"roster-log not found: {path}")

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc

    label = f"{month_name[month]} {year}"
    wb = openpyxl.load_workbook(str(path), read_only=False, data_only=True)
    try:
        live_ws = _find_live_sheet(wb, label)
        worked_sheet = _find_worked_projects_sheet(wb, label)
        presence = _presence_only_sheets(wb, label)
        paired = _paired_clock_dates(live_ws)

        warnings: List[str] = []
        blockers: List[str] = []

        if live_ws is None:
            blockers.append("paid_hours_source_missing:no Live month sheet with clock-in/clock-out columns")
        elif paired == 0:
            blockers.append("paid_hours_source_invalid:Live month sheet has no paired Clock In/Clock Out date columns")

        if presence:
            warnings.append(
                "presence_only_attendance_detected: attendance matrices may corroborate worked/not-worked state but do not create paid hours"
            )
        if not worked_sheet:
            warnings.append(
                "worked_projects_missing: project classification will fall back to each Live row's default project"
            )

        ready = not blockers
        return MonthReadiness(
            schema="triage-nth-month-readiness/v1",
            month_key=month_key,
            month_label=label,
            source_path=str(path),
            status="READY" if ready else "NO_GO",
            attendance_authority="clock_in_out_live_sheet" if ready else "not_proven",
            live_sheet=live_ws.title if live_ws is not None else "",
            worked_projects_sheet=worked_sheet,
            presence_only_sheets=presence,
            paired_clock_dates=paired,
            warnings=warnings,
            blockers=blockers,
        )
    finally:
        wb.close()
