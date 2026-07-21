"""Layout-only compare against accepted reference workbook (validation only)."""
from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

from triage.nw_prj_admin_log.grid import (
    find_data_header_row,
    parse_date_columns,
    scan_forbidden_visible_text,
)
from triage.nw_prj_admin_log.visual_donor import PROJECT_TEAM_SHEET

FORBIDDEN_VISIBLE_PHRASES = (
    "roster log",
    "generated from",
    "source workbook",
    "automation",
    "internal",
    "holdover",
    "build notes",
    "sidecar",
    "manifest",
)


def _sheet_layout(ws) -> Dict[str, Any]:
    hdr_row = find_data_header_row(ws)
    dates = sorted(parse_date_columns(ws, hdr_row).values()) if hdr_row else []
    merged = [str(r) for r in ws.merged_cells.ranges][:20]
    row_heights = {
        str(r): ws.row_dimensions[r].height
        for r in range(1, 6)
        if ws.row_dimensions[r].height is not None
    }
    col_widths = {
        c: ws.column_dimensions[c].width
        for c in ("A", "B", "C", "D")
        if ws.column_dimensions[c].width is not None
    }
    return {
        "header_row": hdr_row,
        "first_weekly_block_dates": dates[:7],
        "merged_near_top": merged,
        "row_heights_1_5": row_heights,
        "column_widths_a_d": col_widths,
        "freeze_panes": ws.freeze_panes,
        "column_a_hidden": bool(ws.column_dimensions["A"].hidden),
        "forbidden_visible_phrases": scan_forbidden_visible_text(ws),
    }


def _package_counts(path: Path) -> Dict[str, int]:
    with zipfile.ZipFile(path, "r") as z:
        names = z.namelist()
        return {
            "drawing_part_count": len([n for n in names if n.startswith("xl/drawings/")]),
            "media_part_count": len([n for n in names if n.startswith("xl/media/")]),
        }


def run_visual_compare(
    candidate_path: str | Path,
    reference_path: Optional[str | Path] = None,
) -> Dict[str, Any]:
    cand = Path(candidate_path)
    report: Dict[str, Any] = {
        "candidate": str(cand.resolve()),
        "reference": str(Path(reference_path).resolve()) if reference_path else None,
        "visual_compare_pass": None,
        "failures": [],
        "candidate_layout": {},
        "reference_layout": {},
        "candidate_package": _package_counts(cand) if cand.is_file() else {},
    }

    wb = openpyxl.load_workbook(cand)
    visible = [s.title for s in wb.worksheets if s.sheet_state == "visible"]
    report["visible_sheet_count"] = len(visible)
    report["visible_sheet_names"] = visible
    if visible != [PROJECT_TEAM_SHEET]:
        report["failures"].append(f"visible_sheets:{visible}")
    if PROJECT_TEAM_SHEET in wb.sheetnames:
        report["candidate_layout"] = _sheet_layout(wb[PROJECT_TEAM_SHEET])
    wb.close()

    if report["candidate_layout"].get("forbidden_visible_phrases"):
        report["failures"].append("forbidden_visible_text_in_candidate")
    if report["candidate_layout"].get("freeze_panes") != "C1":
        report["failures"].append(
            f"freeze_not_C1:{report['candidate_layout'].get('freeze_panes')}"
        )
    if not report["candidate_layout"].get("column_a_hidden"):
        report["failures"].append("column_a_not_hidden")

    if reference_path and Path(reference_path).is_file():
        ref = Path(reference_path)
        report["reference_package"] = _package_counts(ref)
        rwb = openpyxl.load_workbook(ref)
        if PROJECT_TEAM_SHEET in rwb.sheetnames:
            report["reference_layout"] = _sheet_layout(rwb[PROJECT_TEAM_SHEET])
        rwb.close()
        if report["reference_package"].get("media_part_count", 0) > 0:
            if report["candidate_package"].get("media_part_count", 0) < 1:
                report["failures"].append("missing_media_vs_reference")
        ref_dates = report["reference_layout"].get("first_weekly_block_dates") or []
        cand_dates = report["candidate_layout"].get("first_weekly_block_dates") or []
        if ref_dates and cand_dates and cand_dates[0] > ref_dates[0]:
            report["failures"].append("first_block_starts_later_than_reference")

    report["visual_compare_pass"] = len(report["failures"]) == 0
    return report
