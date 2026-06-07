"""Visual donor workbook preparation — style only, not data authority."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

PROJECT_TEAM_SHEET = "Project Team"


def prepare_donor_workbook(
    donor_path: str | Path,
    output_path: str | Path,
) -> None:
    """Copy donor to output path; caller applies grid patches and layout fixes."""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(donor_path, output_path)


def apply_layout_rules(workbook_path: str | Path) -> None:
    """Hide column A, freeze at C1, keep only Project Team visible."""
    wb = openpyxl.load_workbook(workbook_path)
    if PROJECT_TEAM_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"missing sheet {PROJECT_TEAM_SHEET!r}")

    for name in wb.sheetnames:
        ws = wb[name]
        if name == PROJECT_TEAM_SHEET:
            ws.sheet_state = "visible"
        else:
            ws.sheet_state = "hidden"

    ws = wb[PROJECT_TEAM_SHEET]
    ws.column_dimensions["A"].hidden = True
    ws.freeze_panes = "C1"
    wb.save(workbook_path)
    wb.close()
