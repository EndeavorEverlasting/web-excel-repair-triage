"""Build the combined April+May billing summary workbook and delivery package.

Admin sheets stay clean (no raw punch notes). Notes, partials, exclusions, and
mismatches live only in the internal review queue CSV sidecar.
"""
from __future__ import annotations

import csv
import json
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

from triage.xlsx_utils import fix_inlinestr

from . import preflight as pf
from . import summary as summ
from .classifier import classify_rows
from .models import BillingReport, BillingRow, ReviewFlag
from .reader import read_billing_rows

DEFAULT_ARTIFACT = "NW_PRJ_Billing_Summary_April_May_2026"


@dataclass
class BillingResult:
    report: BillingReport
    outputs: Dict[str, str] = field(default_factory=dict)


def _bold_header(ws, ncols: int) -> None:
    from openpyxl.styles import Font

    for c in range(1, ncols + 1):
        ws.cell(1, c).font = Font(bold=True)
    ws.freeze_panes = "A2"
    if ncols >= 1 and ws.max_row >= 1:
        from openpyxl.utils import get_column_letter

        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(ws.max_row, 1)}"


def _write_table(ws, headers: List[str], rows: List[List[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _bold_header(ws, len(headers))


def build_workbook(report: BillingReport, output_path: str, invoices: List[Dict[str, Any]]) -> None:
    import openpyxl

    wb = openpyxl.Workbook()

    # --- Dashboard ---
    dash = wb.active
    dash.title = "Dashboard"
    dash.append(["NW PRJ Billing Summary", ""])
    dash.append(["Months", ", ".join(report.months)])
    dash.append(["Combined Gross Hours", report.combined_gross])
    dash.append(["Combined Lunch Deduction", report.combined_lunch])
    dash.append(["Combined Net Hours", report.combined_net])
    dash.append(["Project Team Rows", len(report.rows)])
    dash.append(["Invoices", report.invoice_count])
    dash.append(["Review Queue Items", len(report.review_flags)])
    dash.append([])
    dash.append(["Month", "Gross", "Lunch", "Net", "Staff", "Rows"])
    for ms in report.month_summaries:
        dash.append([ms.month_label, ms.gross_hours, ms.lunch_deduction,
                     ms.net_hours, ms.staff_count, ms.daily_rows])
    from openpyxl.styles import Font

    dash.cell(1, 1).font = Font(bold=True, size=14)
    dash.freeze_panes = "A2"

    # --- Per-month detail sheets (clean admin: no note column) ---
    detail_headers = ["Staff", "Project", "Date", "Day", "Clock In", "Clock Out",
                      "Gross", "Lunch", "Net", "Friday Batch"]
    for ms in report.month_summaries:
        ws = wb.create_sheet(title=ms.month_label[:31])
        rows = [
            [r.staff, r.project, r.date.isoformat(), r.date.strftime("%a"),
             r.clock_in, r.clock_out, r.gross_hours, r.lunch_deduction,
             r.net_hours, r.friday_batch.isoformat()]
            for r in report.rows if r.month_key == ms.month_key
        ]
        _write_table(ws, detail_headers, rows)

    # --- Friday Batches ---
    fb = wb.create_sheet(title="Friday Batches")
    fb_rows: List[List[Any]] = []
    for ms in report.month_summaries:
        for batch, net in ms.by_friday_batch.items():
            fb_rows.append([ms.month_label, batch, net])
    _write_table(fb, ["Month", "Friday Batch", "Net Hours"], fb_rows)

    # --- By Project ---
    bp = wb.create_sheet(title="By Project")
    bp_rows: List[List[Any]] = []
    for ms in report.month_summaries:
        for proj, net in ms.by_project.items():
            bp_rows.append([ms.month_label, proj, net])
    _write_table(bp, ["Month", "Project", "Net Hours"], bp_rows)

    # --- Invoice Pivot ---
    inv = wb.create_sheet(title="Invoice Pivot")
    pivot = summ.invoice_pivot(invoices)
    if pivot:
        _write_table(inv, ["Cost Category", "Vendor", "Total"],
                     [[p["cost_category"], p["vendor"], p["total"]] for p in pivot])
    else:
        _write_table(inv, ["Cost Category", "Vendor", "Total"], [])
        inv.append(["No invoices provided", "", ""])

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def run_export(
    roster_path: str,
    months: List[str],
    out_dir: str,
    *,
    websafe: bool = True,
    make_zip: bool = False,
    invoices: Optional[List[Dict[str, Any]]] = None,
    artifact: str = DEFAULT_ARTIFACT,
) -> BillingResult:
    invoices = invoices or []
    raw_rows, read_flags, warnings = read_billing_rows(roster_path, months)
    rows, class_flags = classify_rows(raw_rows)
    flags: List[ReviewFlag] = read_flags + class_flags

    month_summaries = summ.summarize_months(rows)
    gross, lunch, net = summ.combined_totals(rows)
    excluded = sorted({f.staff for f in flags if f.category == "excluded_name"})

    report = BillingReport(
        roster_path=str(Path(roster_path).resolve()),
        months=months,
        rows=rows,
        month_summaries=month_summaries,
        review_flags=flags,
        invoice_count=len(invoices),
        excluded_names=excluded,
        combined_gross=gross,
        combined_lunch=lunch,
        combined_net=net,
        warnings=warnings,
    )

    out_dir_p = Path(out_dir)
    out_dir_p.mkdir(parents=True, exist_ok=True)
    suffix = "_WEBSAFE" if websafe else ""
    wb_path = out_dir_p / f"{artifact}{suffix}.xlsx"
    report.output_workbook = str(wb_path.resolve())

    build_workbook(report, str(wb_path), invoices)

    if websafe:
        fix_inlinestr(str(wb_path))

    pre = pf.run_preflight(str(wb_path))
    report.webexcel_preflight_pass = pre.preflight_pass

    pre_path = out_dir_p / f"{artifact}_preflight.json"
    manifest_path = out_dir_p / f"{artifact}_manifest.json"
    review_path = out_dir_p / f"{artifact}_review_queue.csv"
    pre_path.write_text(json.dumps(pre.to_dict(), indent=2), encoding="utf-8")
    manifest_path.write_text(json.dumps(report.to_manifest(), indent=2), encoding="utf-8")
    _write_review_queue(review_path, report)

    outputs = {
        "workbook": str(wb_path.resolve()),
        "preflight": str(pre_path.resolve()),
        "manifest": str(manifest_path.resolve()),
        "review_queue": str(review_path.resolve()),
    }

    if make_zip:
        zip_path = out_dir_p / f"{artifact}_DELIVERY.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
            z.write(wb_path, wb_path.name)
            z.write(pre_path, pre_path.name)
            z.write(manifest_path, manifest_path.name)
            z.write(review_path, review_path.name)
        outputs["delivery_zip"] = str(zip_path.resolve())

    return BillingResult(report=report, outputs=outputs)


def _write_review_queue(path: Path, report: BillingReport) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["category", "staff", "date", "detail"])
        for f in report.review_flags:
            w.writerow([f.category, f.staff, f.date_iso, f.detail])
        for warn in report.warnings:
            w.writerow(["warning", "", "", warn])
