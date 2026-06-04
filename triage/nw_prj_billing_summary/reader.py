"""Read the Active Roster Log into resolved billing rows for April + May.

Reuses note-aware punch parsing from the Neuron engine and the override /
lunch-policy logic from the roster parser. Project resolution order:
  Worked-Projects cell  >  Assignments override  >  Live default project.
"""
from __future__ import annotations

import re
from calendar import month_name
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from triage.nw_prj_neuron_track_hours.reader import (
    _compute_gross,
    _format_clock,
    _worked_project_lookup,
    split_note_bearing_punch,
)
from triage.roster_parser import (
    _find_assignments_sheet,
    _load_assignments,
    _lunch_deduction,
)

from .classifier import friday_batch
from .models import BillingRow, ReviewFlag

_DATE_HEADER = re.compile(
    r"^([A-Za-z]+)\s+(\d{1,2})\s*[-\u2013]\s*(Clock\s*In|Clock\s*Out)\s*$",
    re.IGNORECASE,
)
_MONTH_ABBREVS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


class BillingReadError(Exception):
    pass


def _month_label(month_key: str) -> Tuple[str, int, int]:
    m = re.match(r"^(\d{4})-(\d{1,2})$", month_key.strip())
    if not m:
        raise BillingReadError(f"Invalid month key (expected YYYY-MM): {month_key}")
    return f"{month_name[int(m.group(2))]} {int(m.group(1))}", int(m.group(1)), int(m.group(2))


def _find_live_sheet(wb, label: str):
    target = f"live - {label}".lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return wb[name]
    month_word = label.split()[0].lower()
    year = label.split()[-1]
    for name in wb.sheetnames:
        low = name.strip().lower()
        if low.startswith("live") and month_word in low and year in name:
            return wb[name]
    return None


def _find_worked_sheet(wb, label: str):
    month_word = label.split()[0].lower()
    year = label.split()[-1]
    for name in wb.sheetnames:
        low = name.strip().lower()
        if low.startswith("worked projects") and month_word in low and year in name:
            return wb[name]
    return None


def read_month(wb, month_key: str) -> Tuple[List[BillingRow], List[ReviewFlag], List[str]]:
    label, year, mon = _month_label(month_key)
    warnings: List[str] = []
    flags: List[ReviewFlag] = []

    live_ws = _find_live_sheet(wb, label)
    if live_ws is None:
        warnings.append(f"missing_live_sheet:{label}")
        flags.append(ReviewFlag(category="missing_roster", staff="", detail=f"No Live sheet for {label}"))
        return [], flags, warnings

    worked = _worked_project_lookup(_find_worked_sheet(wb, label))
    assignments = _load_assignments(_find_assignments_sheet(wb, label))

    header_row = 2
    headers = [live_ws.cell(header_row, c).value for c in range(1, live_ws.max_column + 1)]
    date_to_cols: Dict[date, Dict[str, int]] = {}
    for i, h in enumerate(headers):
        if not isinstance(h, str):
            continue
        mm = _DATE_HEADER.match(h.strip())
        if not mm:
            continue
        mon_num = _MONTH_ABBREVS.get(mm.group(1)[:3].lower())
        if mon_num is None:
            continue
        try:
            d = date(year, mon_num, int(mm.group(2)))
        except ValueError:
            continue
        direction = "in" if "in" in mm.group(3).lower() else "out"
        date_to_cols.setdefault(d, {})[direction] = i

    if not date_to_cols:
        warnings.append(f"no_date_columns:{label}")
        return [], flags, warnings

    rows: List[BillingRow] = []
    for r in range(header_row + 1, live_ws.max_row + 1):
        staff_val = live_ws.cell(r, 1).value
        if not staff_val or str(staff_val).strip() in ("", "None", "0"):
            continue
        if isinstance(staff_val, (int, float)):
            continue
        staff = str(staff_val).strip()
        default_proj = str(live_ws.cell(r, 2).value or "").strip()
        if default_proj == "0":
            default_proj = ""

        for d, dirs in sorted(date_to_cols.items()):
            in_val = live_ws.cell(r, dirs["in"] + 1).value if "in" in dirs else None
            out_val = live_ws.cell(r, dirs["out"] + 1).value if "out" in dirs else None
            ci, note_in = split_note_bearing_punch(in_val)
            co, note_out = split_note_bearing_punch(out_val)
            if ci is None and co is None:
                continue

            # Project resolution: worked-project > assignment override > live default.
            resolved = (
                worked.get((staff, d))
                or assignments.get((d, staff))
                or default_proj
                or "Unassigned / Review"
            )
            project_source = (
                "worked" if worked.get((staff, d))
                else "override" if assignments.get((d, staff))
                else "live"
            )

            note = " ".join(n for n in (note_in, note_out) if n).strip()
            partial = (ci is None) != (co is None)  # exactly one punch present
            gross = _compute_gross(ci, co) if not partial else 0.0
            lunch = _lunch_deduction(gross)
            net = round(max(0.0, gross - lunch), 4)

            rows.append(
                BillingRow(
                    staff=staff,
                    project=resolved,
                    date=d,
                    month_key=month_key,
                    clock_in=_format_clock(ci),
                    clock_out=_format_clock(co),
                    gross_hours=round(gross, 2),
                    lunch_deduction=lunch,
                    net_hours=round(net, 2),
                    friday_batch=friday_batch(d),
                    weekend=d.weekday() >= 5,
                    project_source=project_source,
                    note=note,
                    partial=partial,
                )
            )

    return rows, flags, warnings


def read_billing_rows(
    roster_path: str | Path, months: List[str]
) -> Tuple[List[BillingRow], List[ReviewFlag], List[str]]:
    try:
        import openpyxl
    except ImportError as e:
        raise BillingReadError("openpyxl is required: pip install openpyxl") from e
    p = Path(roster_path)
    if not p.exists():
        raise BillingReadError(f"Roster file not found: {roster_path}")
    wb = openpyxl.load_workbook(str(p), data_only=True)
    all_rows: List[BillingRow] = []
    all_flags: List[ReviewFlag] = []
    all_warn: List[str] = []
    for mk in months:
        rows, flags, warn = read_month(wb, mk)
        all_rows.extend(rows)
        all_flags.extend(flags)
        all_warn.extend(warn)
    wb.close()
    return all_rows, all_flags, all_warn
