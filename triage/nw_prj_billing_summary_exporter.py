"""
NW PRJ Billing Summary Exporter — produce April/May billing summaries.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from triage.nw_prj_classifier import ClassificationResult
from triage.xlsx_utils import fix_inlinestr

def _require_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from e
    return Workbook, Font, PatternFill, Alignment, get_column_letter

def export_billing_summary(
    results: List[ClassificationResult], 
    month_label: str, 
    out_path: str
) -> str:
    Workbook, Font, PatternFill, Alignment, get_column_letter = _require_openpyxl()
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Billing Summary - {month_label}"

    # Title rows — no merge_cells; merged cells with styled strings produce
    # inlineStr in openpyxl which is a Web Excel stop-ship token.
    title_fill = PatternFill("solid", fgColor="1F365C")
    sub_fill   = PatternFill("solid", fgColor="EAF1F8")
    for col in range(1, 8):
        c1 = ws.cell(row=1, column=col)
        c1.fill = title_fill
        c2 = ws.cell(row=2, column=col)
        c2.fill = sub_fill
    ws.cell(row=1, column=1).value = f"{month_label} Billing Summary - Candidate - WEBSAFE"
    ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=1).value = "NW PRJ Billing Summary with Roster Reconciliation"
    
    # Header row (Row 4)
    headers = ["Tech", "Date", "Hours", "Status", "Reason", "Action Needed", "Notes"]
    header_fill = PatternFill("solid", fgColor="1F365C")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Data rows
    for r_idx, res in enumerate(results, 5):
        ws.cell(row=r_idx, column=1, value=res.tech)
        ws.cell(row=r_idx, column=2, value=res.date)
        
        h_cell = ws.cell(row=r_idx, column=3, value=res.resolved_hours)
        h_cell.number_format = "0.00"
        
        ws.cell(row=r_idx, column=4, value=res.status)
        ws.cell(row=r_idx, column=5, value=res.reason_code)
        ws.cell(row=r_idx, column=6, value=res.action_needed)
        ws.cell(row=r_idx, column=7, value=res.notes)

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 40
    
    # Filters and Panes
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, len(results)+4)}"

    # Add CF Dictionary tab
    ws_cf = wb.create_sheet("CF Dictionary")
    ws_cf["A1"] = "Conditional Formatting Dictionary"
    
    # Save and repair inlineStr (openpyxl 3.1.x Web Excel stop-ship fix)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    fix_inlinestr(out_path)

    return out_path
