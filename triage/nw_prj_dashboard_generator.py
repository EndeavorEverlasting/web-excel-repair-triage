"""
Generate NW PRJ Tech Roster Dashboard v6_x WEBSAFE workbooks.

Current status: carry-forward / skeleton generator. Active and archive rows
come almost entirely from prior dashboard carry-forward via
``nw_prj_artifact_compare``. Fresh row generation from admin scratch and
roster evidence lands with the readers and classifier in
``triage.nw_prj_admin_scratch_reader``, ``triage.nw_prj_roster_reader``, and
``triage.nw_prj_target_classifier``.
"""
from __future__ import annotations

import json
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

from triage.nw_prj_artifact_compare import CompareInputs, compare_artifacts
from triage.nw_prj_config import cf_palette, dashboard_schema, is_repair_filename, status_values
from triage.nw_prj_dashboard_validator import validate_nw_prj_dashboard


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from e
    return openpyxl, Workbook, load_workbook, Font, PatternFill, get_column_letter


def _write_sheet_table(ws, headers: List[str], data_rows: List[Dict[str, Any]]) -> None:
    _, _, _, Font, PatternFill, get_column_letter = _require_openpyxl()
    header_fill = PatternFill("solid", fgColor="FF4472C4")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = header_fill
    for r_idx, row in enumerate(data_rows, 2):
        for c, h in enumerate(headers, 1):
            ws.cell(row=r_idx, column=c, value=row.get(h, ""))
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16


def _apply_tab_colors(wb) -> None:
    palette = cf_palette().get("tab_colors", {})
    for name in wb.sheetnames:
        if name in palette:
            wb[name].sheet_properties.tabColor = palette[name]


def _cf_dictionary_rows() -> List[Dict[str, Any]]:
    return [
        {
            "Rule ID": "CF_A_DONE",
            "Priority": 1,
            "Applies To": "Active queues; Column A",
            "Condition (plain English)": "Review Status is Done, Confirmed Valid, Addressed, or Resolved",
            "Formula (audit)": '=$A2="Done"',
            "Fill / Font": "resolved_green",
            "Overrides": "Wins over all queue colors",
        },
        {
            "Rule ID": "CF_A_GRAY",
            "Priority": 2,
            "Applies To": "Active queues; Column A",
            "Condition (plain English)": "Review Status is Skipped/Gray or Gray/Skip",
            "Formula (audit)": '=$A2="Skipped/Gray"',
            "Fill / Font": "skipped_gray",
            "Overrides": "Demote to Resolved_Archive",
        },
    ]


@dataclass
class GenerateInputs:
    dashboard_path: Optional[str] = None
    roster_path: Optional[str] = None
    admin_scratch_path: str = ""
    official_admin_path: Optional[str] = None
    version_minor: str = "7"
    descriptor: str = "GENERATED"
    out_dir: str = "Outputs"


@dataclass
class GenerateResult:
    output_path: str = ""
    compare_report: Optional[Dict[str, Any]] = None
    validation: Optional[Dict[str, Any]] = None
    web_excel_safe: bool = False


def generate_dashboard(inputs: GenerateInputs) -> GenerateResult:
    if not inputs.admin_scratch_path:
        raise ValueError("admin_scratch_path is required")
    if is_repair_filename(Path(inputs.admin_scratch_path).name):
        raise ValueError(f"repaired admin scratch rejected: {inputs.admin_scratch_path}")

    openpyxl, Workbook, *_ = _require_openpyxl()
    schema = dashboard_schema()
    headers = schema["required_active_columns"]

    cmp = compare_artifacts(
        CompareInputs(
            dashboard_path=inputs.dashboard_path,
            roster_path=inputs.roster_path,
            admin_scratch_path=inputs.admin_scratch_path,
            official_admin_path=inputs.official_admin_path,
        )
    )

    out_name = (
        f"NW_PRJ_Tech_Roster_Dashboard_v6_{inputs.version_minor}_"
        f"{inputs.descriptor}_WEBSAFE.xlsx"
    )
    out_path = Path(inputs.out_dir) / out_name
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheet_plan = [
        ("Start Here", [["NW PRJ Dashboard v6 — edit In/Out cells only; do not overwrite Total formulas."]]),
        (f"Dashboard_Tool_v6_{inputs.version_minor}", [["Generated", time.strftime("%Y-%m-%d %H:%M")]]),
        ("Active_Admin_Targets", []),
        ("Partial_Hours_Active", []),
        ("Review_Guardrails", []),
        ("Quiet_Queues", []),
        ("Resolved_Archive", []),
        ("Tech_Summary", []),
        ("CF_Dictionary", []),
        ("Definitions_Current", []),
        (f"Visual_System_v6_{inputs.version_minor}", []),
        ("Dropdown_Values", []),
        ("Repo_Automation_Notes", []),
    ]

    for title, _ in sheet_plan:
        wb.create_sheet(title)

    wb["Start Here"]["A1"] = (
        "NW PRJ Dashboard v6 — edit In/Out cells only; do not overwrite Total formulas. "
        "Check Expected Total before submission."
    )

    _write_sheet_table(wb["Active_Admin_Targets"], headers, cmp.active_rows)
    _write_sheet_table(wb["Resolved_Archive"], headers, cmp.archive_rows)
    _write_sheet_table(wb["Review_Guardrails"], headers, cmp.rich_guard_rows)

    partials = [
        r for r in cmp.active_rows
        if str(r.get("Reason Code", "")).startswith("PARTIAL")
        or str(r.get("Work Queue Status", "")) == "AMBER"
    ]
    _write_sheet_table(wb["Partial_Hours_Active"], headers, partials)

    cf_headers = list(_cf_dictionary_rows()[0].keys())
    _write_sheet_table(wb["CF_Dictionary"], cf_headers, _cf_dictionary_rows())

    dv = status_values()
    _write_sheet_table(
        wb["Dropdown_Values"],
        ["List", "Values"],
        [
            {"List": "Review Status", "Values": ", ".join(dv["review_status"]["resolved_green"])},
            {"List": "Team Scope", "Values": "Cybernet/Neuron Active, Tracked Only, Out of Scope"},
        ],
    )

    notes = [
        ["Input", "Path"],
        ["admin_scratch", inputs.admin_scratch_path],
        ["dashboard", inputs.dashboard_path or ""],
        ["roster", inputs.roster_path or ""],
        ["official_admin", inputs.official_admin_path or ""],
        ["admin_authority", cmp.admin_authority],
    ]
    for r_idx, row in enumerate(notes, 1):
        for c_idx, val in enumerate(row, 1):
            wb["Repo_Automation_Notes"].cell(row=r_idx, column=c_idx, value=val)

    _apply_tab_colors(wb)
    wb.save(out_path)

    val = validate_nw_prj_dashboard(str(out_path))
    return GenerateResult(
        output_path=str(out_path),
        compare_report=cmp.to_dict(),
        validation=val.to_dict(),
        web_excel_safe=val.web_excel_safe,
    )


def main() -> None:
    import argparse

    ap = argparse.ArgumentParser(description="Generate NW PRJ dashboard v6 WEBSAFE workbook")
    ap.add_argument("--admin-scratch", required=True)
    ap.add_argument("--dashboard")
    ap.add_argument("--roster")
    ap.add_argument("--official-admin")
    ap.add_argument("--minor", default="7")
    ap.add_argument("--descriptor", default="GENERATED")
    ap.add_argument("--out-dir", default="Outputs")
    args = ap.parse_args()
    res = generate_dashboard(
        GenerateInputs(
            admin_scratch_path=args.admin_scratch,
            dashboard_path=args.dashboard,
            roster_path=args.roster,
            official_admin_path=args.official_admin,
            version_minor=args.minor,
            descriptor=args.descriptor,
            out_dir=args.out_dir,
        )
    )
    print(json.dumps({"output": res.output_path, "web_excel_safe": res.web_excel_safe}, indent=2))


if __name__ == "__main__":
    main()
