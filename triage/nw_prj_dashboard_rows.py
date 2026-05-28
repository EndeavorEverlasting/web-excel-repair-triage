"""Shared row model and dashboard sheet readers for NW PRJ v6."""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List

from triage.nw_prj_config import dashboard_schema


def _require_openpyxl():
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from e
    return load_workbook


@dataclass
class DashboardRow:
    review_status: str = ""
    work_queue_status: str = ""
    target_type: str = ""
    action_needed: str = ""
    target_workbook: str = ""
    edit_sheet: str = ""
    edit_row: str = ""
    edit_in_cell: str = ""
    proposed_in: str = ""
    edit_out_cell: str = ""
    proposed_out: str = ""
    total_cell: str = ""
    expected_total: str = ""
    tech: str = ""
    date: str = ""
    team_scope: str = ""
    current_admin_value: str = ""
    roster_latest_in: str = ""
    roster_latest_out: str = ""
    roster_latest_hours: str = ""
    roster_check: str = ""
    roster_check_notes: str = ""
    reason_code: str = ""
    confidence: str = ""
    submission_blocker: str = ""
    manual_note: str = ""

    @staticmethod
    def from_dict(d: Dict[str, Any]) -> "DashboardRow":
        return DashboardRow(
            review_status=str(d.get("Review Status", d.get("review_status", "")) or ""),
            work_queue_status=str(d.get("Work Queue Status", "") or ""),
            target_type=str(d.get("Target Type", "") or ""),
            action_needed=str(d.get("Action Needed", "") or ""),
            target_workbook=str(d.get("Target Workbook", "") or ""),
            edit_sheet=str(d.get("Edit Sheet", "") or ""),
            edit_row=str(d.get("Edit Row", "") or ""),
            edit_in_cell=str(d.get("Edit In Cell", "") or ""),
            proposed_in=str(d.get("Proposed In", "") or ""),
            edit_out_cell=str(d.get("Edit Out Cell", "") or ""),
            proposed_out=str(d.get("Proposed Out", "") or ""),
            total_cell=str(d.get("Total Cell", "") or ""),
            expected_total=str(d.get("Expected Total", "") or ""),
            tech=str(d.get("Tech", "") or ""),
            date=str(d.get("Date", "") or ""),
            team_scope=str(d.get("Team Scope", "") or ""),
            current_admin_value=str(d.get("Current Admin Value", "") or ""),
            roster_latest_in=str(d.get("Roster Latest In", "") or ""),
            roster_latest_out=str(d.get("Roster Latest Out", "") or ""),
            roster_latest_hours=str(d.get("Roster Latest Hours", "") or ""),
            roster_check=str(d.get("Roster Check", "") or ""),
            roster_check_notes=str(d.get("Roster Check Notes", "") or ""),
            reason_code=str(d.get("Reason Code", "") or ""),
            confidence=str(d.get("Confidence", "") or ""),
            submission_blocker=str(d.get("Submission Blocker", "") or ""),
            manual_note=str(d.get("Manual Note / Resolution Note", d.get("manual_note", "")) or ""),
        )

    def to_dict(self) -> Dict[str, Any]:
        return {
            "Review Status": self.review_status,
            "Work Queue Status": self.work_queue_status,
            "Target Type": self.target_type,
            "Action Needed": self.action_needed,
            "Target Workbook": self.target_workbook,
            "Edit Sheet": self.edit_sheet,
            "Edit Row": self.edit_row,
            "Edit In Cell": self.edit_in_cell,
            "Proposed In": self.proposed_in,
            "Edit Out Cell": self.edit_out_cell,
            "Proposed Out": self.proposed_out,
            "Total Cell": self.total_cell,
            "Expected Total": self.expected_total,
            "Tech": self.tech,
            "Date": self.date,
            "Team Scope": self.team_scope,
            "Current Admin Value": self.current_admin_value,
            "Roster Latest In": self.roster_latest_in,
            "Roster Latest Out": self.roster_latest_out,
            "Roster Latest Hours": self.roster_latest_hours,
            "Roster Check": self.roster_check,
            "Roster Check Notes": self.roster_check_notes,
            "Reason Code": self.reason_code,
            "Confidence": self.confidence,
            "Submission Blocker": self.submission_blocker,
            "Manual Note / Resolution Note": self.manual_note,
        }


def row_key(row: DashboardRow) -> str:
    return f"{row.tech}|{row.date}|{row.edit_sheet}|{row.edit_row}"


def _find_sheet(wb, base_name: str):
    if base_name in wb.sheetnames:
        return wb[base_name]
    for n in wb.sheetnames:
        if n.startswith(base_name.rstrip("_x")) or base_name.rstrip("_x") in n:
            return wb[n]
    return None


def load_dashboard_rows(path: str) -> List[DashboardRow]:
    load_workbook = _require_openpyxl()
    wb = load_workbook(path, data_only=True, read_only=True)
    schema = dashboard_schema()
    rows: List[DashboardRow] = []
    sheets = list(schema.get("active_queue_sheets", []))
    sheets.append(schema.get("archive_sheet", "Resolved_Archive"))
    for sheet_name in sheets:
        ws = _find_sheet(wb, sheet_name)
        if ws is None:
            continue
        headers: Dict[str, int] = {}
        for row_cells in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for i, h in enumerate(row_cells):
                if h:
                    headers[str(h).strip()] = i
        if "Review Status" not in headers:
            continue
        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            if not row_cells or not any(row_cells):
                continue
            d: Dict[str, Any] = {}
            for h, idx in headers.items():
                if idx < len(row_cells):
                    d[h] = row_cells[idx]
            rows.append(DashboardRow.from_dict(d))
    wb.close()
    return rows
