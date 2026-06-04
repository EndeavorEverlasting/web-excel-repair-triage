"""Build the clean, submission-grade Bonita-format Neuron Track Hours workbook.

Three tabs (``Apr 26`` / ``May 26`` / ``Rules & Legend``), two-line headers,
one row per included shift, values only. Layout follows the repair-free golden
profile (``configs/artifact_profiles/neuron_track_hours_repairfree_golden.json``).
"""
from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from triage.nw_prj_neuron_track_hours.bonita_resolver import BonitaResolution, BonitaShift
from triage.nw_prj_neuron_track_hours.exporter import _repair_inlinestr
from triage.nw_prj_neuron_track_hours.repairfree_profile_gate import (
    default_profile_path,
    load_repairfree_profile,
)

WEBSAFE_ALIAS_NAME = "Neuron_Track_Hours_April_May_2026_WEBSAFE.xlsx"

# Backward-compatible header constants (profile JSON is authoritative).
HEADER_TOP = ["", "TECH", "START", "END", "TOTAL", "PROJECT", "ASSIGNMENT"]
HEADER_BOTTOM = ["DATE", "NAME", "TIME", "TIME", "HOURS", "NAME", "TYPE"]

_MONTH_ABBR = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}
_TIME_FMT = "h:mm AM/PM"
_DATE_FMT = "mm-dd-yy"


def tab_name_for_month_key(month_key: str) -> str:
    year, mon = month_key.split("-")[0], int(month_key.split("-")[1])
    return f"{_MONTH_ABBR[mon]} {year[-2:]}"


def _require_openpyxl():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    return Workbook, Alignment, Font, PatternFill, get_column_letter


def _group_locators(shifts: List[BonitaShift]) -> List[object]:
    locators: List[object] = []
    n = len(shifts)
    i = 0
    while i < n:
        j = i
        while j < n and shifts[j].date == shifts[i].date:
            j += 1
        size = j - i
        if size == 1:
            locators.append(shifts[i].date)
        else:
            locators.append(shifts[i].date.strftime("%A").upper())
            locators.append(shifts[i].date)
            locators.extend([""] * (size - 2))
        i = j
    return locators


def _write_month_tab(wb, tab: str, shifts: List[BonitaShift], profile: Dict[str, Any]) -> None:
    Workbook, Alignment, Font, PatternFill, get_column_letter = _require_openpyxl()
    ws = wb.create_sheet(tab)
    fill1 = PatternFill("solid", fgColor=profile.get("header_fill_row1", "1F4E78"))
    fill2 = PatternFill("solid", fgColor=profile.get("header_fill_row2", "5B9BD5"))
    h1 = profile.get("header_row1") or HEADER_TOP
    h2 = profile.get("header_row2") or HEADER_BOTTOM
    asn_fills = profile.get("assignment_fills") or {}

    for c, (top, bottom) in enumerate(zip(h1, h2), 1):
        t = ws.cell(row=1, column=c, value=top)
        b = ws.cell(row=2, column=c, value=bottom)
        t.font = Font(bold=True, color="FFFFFF")
        t.fill = fill1
        t.alignment = Alignment(horizontal="center")
        b.font = Font(bold=True, color="FFFFFF")
        b.fill = fill2
        b.alignment = Alignment(horizontal="center")

    bold = Font(bold=True)
    locators = _group_locators(shifts)
    for r_idx, (shift, locator) in enumerate(zip(shifts, locators), 3):
        a = ws.cell(row=r_idx, column=1, value=locator)
        a.font = bold
        if isinstance(locator, date):
            a.number_format = _DATE_FMT

        tech = ws.cell(row=r_idx, column=2, value=shift.tech)
        if shift.start_time is not None:
            si = ws.cell(row=r_idx, column=3, value=shift.start_time)
            si.number_format = _TIME_FMT
        else:
            si = ws.cell(row=r_idx, column=3, value=shift.clock_in)
        if shift.end_time is not None:
            so = ws.cell(row=r_idx, column=4, value=shift.end_time)
            so.number_format = _TIME_FMT
        else:
            so = ws.cell(row=r_idx, column=4, value=shift.clock_out)

        tot = ws.cell(row=r_idx, column=5, value=round(shift.total_hours, 2))
        tot.number_format = "0.00"
        proj = ws.cell(row=r_idx, column=6, value=shift.project_name)
        asn = ws.cell(row=r_idx, column=7, value=shift.assignment_type)
        fill_hex = asn_fills.get(shift.assignment_type)
        if fill_hex:
            asn.fill = PatternFill("solid", fgColor=fill_hex)
        for cell in (tech, si, so, tot, proj, asn):
            cell.font = bold

    widths = profile.get("column_widths") or [12, 22, 12, 12, 10, 22, 32]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = profile.get("freeze_panes") or "A3"


def _write_rules_legend_tab(wb, profile: Dict[str, Any]) -> None:
    Workbook, Alignment, Font, PatternFill, get_column_letter = _require_openpyxl()
    name = profile.get("legend_sheet") or "Rules & Legend"
    ws = wb.create_sheet(name)
    fill1 = PatternFill("solid", fgColor=profile.get("header_fill_row1", "1F4E78"))
    rows = profile.get("legend_rows") or []
    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row[:2], 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill1
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 22
    ws.freeze_panes = profile.get("freeze_panes") or "A3"


def build_bonita_workbook(
    resolution: BonitaResolution,
    months: List[str],
    out_path: str,
    profile_path: Optional[str] = None,
    emit_websafe_alias: bool = True,
) -> Tuple[str, List[Tuple[str, str]]]:
    profile = load_repairfree_profile(profile_path or default_profile_path())
    Workbook, *_ = _require_openpyxl()
    wb = Workbook()
    wb.remove(wb.active)

    from triage.nw_prj_neuron_track_hours.reader import _month_label
    from calendar import month_name

    tabs: List[Tuple[str, str]] = []
    for mk in months:
        tab = tab_name_for_month_key(mk)
        _, _, mon = _month_label(mk)
        short = month_name[mon]
        shifts = resolution.shifts_for_month(short)
        _write_month_tab(wb, tab, shifts, profile)
        tabs.append((mk, tab))

    _write_rules_legend_tab(wb, profile)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    _repair_inlinestr(out_path)

    if emit_websafe_alias:
        shutil.copy2(out_path, Path(out_path).parent / WEBSAFE_ALIAS_NAME)

    return out_path, tabs
