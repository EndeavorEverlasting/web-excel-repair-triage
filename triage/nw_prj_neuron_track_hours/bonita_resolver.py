"""Per-date Neuron-project resolver for the Bonita submission workbook.

This is the inclusion engine for the clean two-tab Bonita output. It reuses the
proven note-aware punch parsing and clock helpers from ``reader`` and the
Assignments loader from ``roster_parser``; it adds project resolution,
classification (include / off-project / non-work-marker / excluded-name) and a
review trail so nothing is silently dropped.

Final submitted-artifact source hierarchy:
  1. ``Live - {Month}`` provides clock truth only.
  2. Monthly Assignments override table is the highest-priority project fix.
  3. ``Worked Projects - {Month}`` provides per-tech/per-date project scope.
  4. Assignments main grid fills behind Worked Projects.
  5. Live default project is a last-resort fallback only.
  6. Punch-note tags such as ``/ Neha``, ``/ Bonita``, and ``/ Josh`` are
     off-Neuron conflict evidence unless an explicit Assignments override
     confirms Neuron work.

A tech/date shift enters the Bonita workbook only when that resolved source
hierarchy says the work belongs to Neurons. This prevents stale/default Neuron
labels from pulling in off-project days such as Patricia / Neha work.
"""
from __future__ import annotations

import re
from calendar import month_name
from dataclasses import dataclass, field
from datetime import date, time
from pathlib import Path
from typing import Dict, List, Optional

from triage.neuron_work_context_rules import (
    CONFIGURATIONS,
    LOGISTICS,
    classify_neuron_work_context,
)
from triage.nw_prj_neuron_track_hours.reader import (
    RosterReadError,
    _compute_gross,
    _DATE_HEADER,
    _decimal_to_time,
    _find_sheet,
    _format_clock,
    _is_neuron,
    _month_label,
    _MONTH_ABBREVS,
    _worked_project_lookup,
    split_note_bearing_punch,
)

NEURON_INTERNAL_PROJECT = "Neuron Deployments"
NEURON_DISPLAY_NAME = "Northwell - Neurons"
DEFAULT_ASSIGNMENT_TYPE = CONFIGURATIONS
DELIVERY_ASSIGNMENT_TYPE = LOGISTICS
LONG_SHIFT_HOURS = 12.0

_OFF_PROJECT_NOTE = re.compile(r"\b(bonita|neha|josh)\b", re.IGNORECASE)
EXCLUDED_NAMES = {"yostinn minaya", "steven marques", "inventory"}
_NON_WORK_MARKER = re.compile(
    r"^\s*(pto|non[\s-]*pto|n/?a|out\s*sick|sick|vacation|off\b.*|holiday|"
    r"unpaid|leave|absent)\s*$",
    re.IGNORECASE,
)


@dataclass
class BonitaShift:
    month_key: str
    month_name: str
    date: date
    day: str
    tech: str
    clock_in: str
    clock_out: str
    total_hours: float
    project_name: str = NEURON_DISPLAY_NAME
    assignment_type: str = DEFAULT_ASSIGNMENT_TYPE
    note: str = ""
    long_shift: bool = False
    start_time: Optional[time] = None
    end_time: Optional[time] = None
    assignment_rule: str = ""
    assignment_confidence: str = ""


@dataclass
class BonitaReviewItem:
    category: str
    month_name: str
    date: Optional[date]
    day: str
    tech: str
    clock_in: str = ""
    clock_out: str = ""
    total_hours: float = 0.0
    project: str = ""
    note: str = ""
    source_cell: str = ""
    detail: str = ""

    def to_dict(self) -> Dict[str, object]:
        return {
            "Category": self.category,
            "Month": self.month_name,
            "Date": self.date.isoformat() if self.date else "",
            "Day": self.day,
            "Tech": self.tech,
            "Start Time": self.clock_in,
            "End Time": self.clock_out,
            "Total Hours": round(self.total_hours, 2),
            "Project": self.project,
            "Note": self.note,
            "Source Cell": self.source_cell,
            "Detail": self.detail,
        }


@dataclass
class BonitaResolution:
    shifts: List[BonitaShift] = field(default_factory=list)
    review: List[BonitaReviewItem] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    def shifts_for_month(self, month_name_: str) -> List[BonitaShift]:
        return [s for s in self.shifts if s.month_name == month_name_]

    def month_total(self, month_name_: str) -> float:
        return round(sum(round(s.total_hours, 2) for s in self.shifts_for_month(month_name_)), 2)

    def grand_total(self) -> float:
        return round(sum(round(s.total_hours, 2) for s in self.shifts), 2)


def _excluded_name(name: str) -> bool:
    low = name.strip().lower()
    if low in EXCLUDED_NAMES:
        return True
    return any(tok in low for tok in EXCLUDED_NAMES)


def _is_non_work_marker(text: str) -> bool:
    return bool(text) and bool(_NON_WORK_MARKER.match(text.strip()))


def _col_letter(idx0: int) -> str:
    idx = idx0 + 1
    out = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        out = chr(65 + rem) + out
    return out


def _add_review(
    resolution: BonitaResolution,
    *,
    category: str,
    month_name: str,
    work_date: Optional[date],
    day: str,
    tech: str,
    clock_in: str = "",
    clock_out: str = "",
    total_hours: float = 0.0,
    project: str = "",
    note: str = "",
    source_cell: str = "",
    detail: str = "",
) -> None:
    resolution.review.append(BonitaReviewItem(
        category=category,
        month_name=month_name,
        date=work_date,
        day=day,
        tech=tech,
        clock_in=clock_in,
        clock_out=clock_out,
        total_hours=total_hours,
        project=project,
        note=note,
        source_cell=source_cell,
        detail=detail,
    ))


def _resolve_month(wb, month_key: str, resolution: BonitaResolution) -> None:
    label, year, mon = _month_label(month_key)
    short = month_name[mon]

    live_ws = _find_sheet(wb, "Live", label)
    if live_ws is None:
        raise RosterReadError(f"Live sheet not found for {label}")

    worked_ws = _find_sheet(wb, "Worked Projects", label)
    if worked_ws is None:
        resolution.warnings.append(f"worked_projects_missing:{label}:falling_back_to_default_and_assignments")
    worked = _worked_project_lookup(worked_ws)

    assignments: Dict[tuple[date, str], str] = {}
    overrides: Dict[tuple[date, str], str] = {}
    try:
        from triage.roster_parser import _find_assignments_sheet, _load_assignments, _load_overrides_only
        assignments_sheet = _find_assignments_sheet(wb, label)
        assignments = _load_assignments(assignments_sheet)
        overrides = _load_overrides_only(assignments_sheet)
    except Exception as exc:  # pragma: no cover - assignments are optional
        resolution.warnings.append(f"assignments_unavailable:{label}:{exc}")

    header_row = 2
    headers = [live_ws.cell(header_row, c).value for c in range(1, live_ws.max_column + 1)]

    date_to_cols: Dict[date, Dict[str, int]] = {}
    for i, h in enumerate(headers):
        if not isinstance(h, str):
            continue
        mm = _DATE_HEADER.match(h.strip())
        if not mm:
            continue
        mon_num = _MONTH_ABBREVS.get(mm.group(1)[:3].lower())
        if mon_num is None:
            continue
        try:
            d = date(year, mon_num, int(mm.group(2)))
        except ValueError:
            continue
        direction = "in" if "in" in mm.group(3).lower() else "out"
        date_to_cols.setdefault(d, {})[direction] = i

    if not date_to_cols:
        raise RosterReadError(f"No date columns parsed in Live sheet for {label}")

    for r in range(header_row + 1, live_ws.max_row + 1):
        staff_val = live_ws.cell(r, 1).value
        if not staff_val or str(staff_val).strip() in ("", "None", "0"):
            continue
        if isinstance(staff_val, (int, float)):
            continue
        staff = str(staff_val).strip()
        default_proj = str(live_ws.cell(r, 2).value or "").strip()
        if default_proj == "0":
            default_proj = ""

        for d, dirs in sorted(date_to_cols.items()):
            in_idx = dirs.get("in")
            out_idx = dirs.get("out")
            in_val = live_ws.cell(r, in_idx + 1).value if in_idx is not None else None
            out_val = live_ws.cell(r, out_idx + 1).value if out_idx is not None else None
            ci, note_in = split_note_bearing_punch(in_val)
            co, note_out = split_note_bearing_punch(out_val)
            if ci is None and co is None:
                marker = (note_in or note_out or "").strip()
                if _is_non_work_marker(marker):
                    _add_review(
                        resolution,
                        category="non_work_marker",
                        month_name=short,
                        work_date=d,
                        day=d.strftime("%a"),
                        tech=staff,
                        project=default_proj,
                        note=marker,
                        source_cell=f"{_col_letter(in_idx or out_idx or 0)}{r}",
                        detail="non-work marker, not counted",
                    )
                continue

            note = " ".join(n for n in (note_in, note_out) if n).strip()
            day = d.strftime("%a")
            cell_ref = f"{_col_letter(in_idx if in_idx is not None else 0)}{r}"

            worked_label = worked.get((staff, d), "")
            assign_label = assignments.get((d, staff), "")
            override_label = overrides.get((d, staff), "")
            resolved = override_label or worked_label or assign_label or default_proj
            explicit_neuron_override = bool(override_label and _is_neuron(override_label))

            if _excluded_name(staff):
                if _is_neuron(resolved):
                    _add_review(
                        resolution,
                        category="excluded_name",
                        month_name=short,
                        work_date=d,
                        day=day,
                        tech=staff,
                        clock_in=_format_clock(ci),
                        clock_out=_format_clock(co),
                        total_hours=_compute_gross(ci, co),
                        project=resolved,
                        note=note,
                        source_cell=cell_ref,
                        detail="excluded name, not counted",
                    )
                continue

            if _OFF_PROJECT_NOTE.search(note) and not explicit_neuron_override:
                _add_review(
                    resolution,
                    category="off_project",
                    month_name=short,
                    work_date=d,
                    day=day,
                    tech=staff,
                    clock_in=_format_clock(ci),
                    clock_out=_format_clock(co),
                    total_hours=_compute_gross(ci, co),
                    project=resolved,
                    note=note,
                    source_cell=cell_ref,
                    detail="off-project punch note, excluded from Neuron totals",
                )
                continue

            if not _is_neuron(resolved):
                continue

            gross = _compute_gross(ci, co)
            decision = classify_neuron_work_context(
                work_date=d,
                start_hour=ci,
                end_hour=co,
                notes=note,
                worked_label=worked_label or assign_label,
                resolved_project=resolved,
                tech_name=staff,
            )
            long_shift = gross >= LONG_SHIFT_HOURS

            shift = BonitaShift(
                month_key=month_key,
                month_name=short,
                date=d,
                day=day,
                tech=staff,
                clock_in=_format_clock(ci),
                clock_out=_format_clock(co),
                total_hours=gross,
                project_name=NEURON_DISPLAY_NAME,
                assignment_type=decision.assignment_type,
                note=note,
                long_shift=long_shift,
                start_time=_decimal_to_time(ci),
                end_time=_decimal_to_time(co),
                assignment_rule=decision.rule,
                assignment_confidence=decision.confidence,
            )
            resolution.shifts.append(shift)

            if long_shift:
                _add_review(
                    resolution,
                    category="long_shift",
                    month_name=short,
                    work_date=d,
                    day=day,
                    tech=staff,
                    clock_in=shift.clock_in,
                    clock_out=shift.clock_out,
                    total_hours=gross,
                    project=NEURON_DISPLAY_NAME,
                    note=note,
                    source_cell=cell_ref,
                    detail=f"long shift {gross:g}h included; verify",
                )
            if decision.confidence == "low":
                _add_review(
                    resolution,
                    category="assignment_heuristic_low_confidence",
                    month_name=short,
                    work_date=d,
                    day=day,
                    tech=staff,
                    clock_in=shift.clock_in,
                    clock_out=shift.clock_out,
                    total_hours=gross,
                    project=NEURON_DISPLAY_NAME,
                    note="",
                    source_cell=cell_ref,
                    detail=f"assignment resolved by heuristic rule: {decision.rule}",
                )


def resolve_bonita_shifts(roster_path: str | Path, months: List[str]) -> BonitaResolution:
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover
        raise RosterReadError("openpyxl is required: pip install openpyxl") from e
    p = Path(roster_path)
    if not p.exists():
        raise RosterReadError(f"Roster file not found: {roster_path}")
    wb = openpyxl.load_workbook(str(p), data_only=True, read_only=False)
    resolution = BonitaResolution()
    try:
        for mk in months:
            _resolve_month(wb, mk, resolution)
    finally:
        wb.close()
    resolution.shifts.sort(key=lambda s: (s.month_key, s.date, s.tech))
    resolution.review.sort(key=lambda x: (x.category, x.date or date.min, x.tech))
    return resolution
