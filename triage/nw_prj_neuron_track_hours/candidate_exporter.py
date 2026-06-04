"""Candidate Neuron Track Hours workbook exporter.

Produces the polished Candidate artifact:

- `Apr 26` and `May 26` clean time sheets
- color-coded assignment rows
- Rezaul Roman mixed Inventory Management / Configurations rows emphasized
- unauthorized Client Coordination rows already removed by candidate_rules
- `Rules & Legend` sheet documenting applied rules
"""
from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from triage.neuron_work_context_rules import (
    CLIENT_COORDINATION,
    CONFIGURATIONS,
    DEPLOYMENTS,
    DOCUMENTATION,
    INVENTORY_MANAGEMENT,
    LOGISTICS,
    TICKET_FORWARDING,
    TROUBLESHOOTING,
)
from triage.nw_prj_neuron_track_hours.bonita_exporter import tab_name_for_month_key
from triage.nw_prj_neuron_track_hours.bonita_resolver import BonitaResolution, BonitaShift
from triage.nw_prj_neuron_track_hours.candidate_rules import is_rezaul_april_shift
from triage.nw_prj_neuron_track_hours.exporter import _repair_inlinestr
from triage.nw_prj_neuron_track_hours.reader import _month_label

HEADER_TOP = ["", "TECH", "START", "END", "TOTAL", "PROJECT", "ASSIGNMENT"]
HEADER_BOTTOM = ["DATE", "NAME", "TIME", "TIME", "HOURS", "NAME", "TYPE"]

HEADER_FILL = "1F365C"
BORDER_COLOR = "6E5494"
REZAUL_FONT = "7030A0"

ASSIGNMENT_FILLS: Dict[str, str] = {
    INVENTORY_MANAGEMENT: "D9EAD3",
    CONFIGURATIONS: "D9EAF7",
    TICKET_FORWARDING: "FFF2CC",
    CLIENT_COORDINATION: "EADCF8",
    LOGISTICS: "FCE4D6",
    DEPLOYMENTS: "DDEBF7",
    DOCUMENTATION: "E2F0D9",
    TROUBLESHOOTING: "F4CCCC",
}


def _require_openpyxl():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    return Workbook, Alignment, Border, Font, PatternFill, Side, get_column_letter


def _group_locators(shifts: List[BonitaShift]) -> List[object]:
    locators: List[object] = []
    i = 0
    while i < len(shifts):
        j = i
        while j < len(shifts) and shifts[j].date == shifts[i].date:
            j += 1
        size = j - i
        if size == 1:
            locators.append(shifts[i].date)
        else:
            locators.append(shifts[i].date)
            locators.extend([""] * (size - 1))
        i = j
    return locators


def _write_month_tab(ws, shifts: List[BonitaShift]) -> None:
    _, Alignment, Border, Font, PatternFill, Side, get_column_letter = _require_openpyxl()
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    white_bold = Font(bold=True, color="FFFFFF")

    for c, (top, bottom) in enumerate(zip(HEADER_TOP, HEADER_BOTTOM), 1):
        for r, value in ((1, top), (2, bottom)):
            cell = ws.cell(row=r, column=c, value=value)
            cell.fill = header_fill
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center", vertical="center")

    side = Side(style="thin", color="B7B7B7")
    standard_border = Border(left=side, right=side, top=side, bottom=side)
    rezaul_side = Side(style="medium", color=BORDER_COLOR)
    rezaul_border = Border(left=rezaul_side, right=rezaul_side, top=side, bottom=side)

    locators = _group_locators(shifts)
    for r_idx, (shift, locator) in enumerate(zip(shifts, locators), 3):
        fill = PatternFill("solid", fgColor=ASSIGNMENT_FILLS.get(shift.assignment_type, "FFFFFF"))
        rezaul = is_rezaul_april_shift(shift)
        font = Font(bold=True, color=REZAUL_FONT if rezaul else "000000")
        border = rezaul_border if rezaul else standard_border

        values = [
            locator,
            shift.tech,
            shift.start_time if shift.start_time is not None else shift.clock_in,
            shift.end_time if shift.end_time is not None else shift.clock_out,
            round(float(shift.total_hours or 0.0), 2),
            shift.project_name,
            shift.assignment_type,
        ]
        for c, value in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c, value=value)
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if c in (3, 4) and hasattr(value, "hour"):
                cell.number_format = "h:mm AM/PM"
            if c == 5:
                cell.number_format = "0.00"
            if c == 1 and isinstance(value, date):
                cell.number_format = "mm-dd-yy"

    widths = [14, 27, 12, 12, 11, 22, 28]
    for c, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:G{max(2, len(shifts) + 2)}"


def _write_rules_sheet(ws, stats: dict, output_name: str) -> None:
    _, Alignment, Border, Font, PatternFill, Side, get_column_letter = _require_openpyxl()
    ws.title = "Rules & Legend"
    ws.merge_cells("A1:G1")
    ws["A1"] = "Candidate Neuron Track Hours - Rules & Legend"
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws["A1"].alignment = Alignment(horizontal="center")

    rows = [
        ("Rule / Metric", "Details"),
        ("Candidate artifact name", output_name),
        ("Client Coordination allowed group", "Richard/Rich Perez, Khadejah Harrison, Alejandro Perales, and Geoff Gerber."),
        ("Coordination handling", "Client Coordination rows outside the approved group are removed from Apr 26 / May 26 clean time sheets."),
        ("Rows removed from clean sheets", stats.get("removed_client_coordination_rows", 0)),
        ("Rezaul Roman rows retained/added", stats.get("rezaul_rows", 0)),
        ("Rezaul Roman total hours", stats.get("rezaul_total_hours", 0)),
        ("Rezaul classification", "Mixed Inventory Management and Configurations, color-coded by assignment type. Rezaul rows use bold purple text with purple side borders."),
        ("", ""),
        ("Assignment Type", "Color / Meaning"),
    ]
    rows.extend((name, color) for name, color in ASSIGNMENT_FILLS.items())

    for r, (k, v) in enumerate(rows, 2):
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)
        ws.cell(r, 1).font = Font(bold=True)
        if k in ASSIGNMENT_FILLS:
            ws.cell(r, 1).fill = PatternFill("solid", fgColor=ASSIGNMENT_FILLS[k])
        if r == 2 or k == "Assignment Type":
            for c in range(1, 3):
                ws.cell(r, c).fill = PatternFill("solid", fgColor="D9E1F2")
                ws.cell(r, c).font = Font(bold=True)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 95
    ws.freeze_panes = "A2"


def build_candidate_workbook(
    resolution: BonitaResolution,
    months: List[str],
    out_path: str,
    stats: dict | None = None,
) -> Tuple[str, List[Tuple[str, str]]]:
    Workbook, *_ = _require_openpyxl()
    wb = Workbook()
    wb.remove(wb.active)

    from calendar import month_name

    tabs: List[Tuple[str, str]] = []
    for month_key in months:
        _, _, mon = _month_label(month_key)
        month = month_name[mon]
        tab = tab_name_for_month_key(month_key)
        ws = wb.create_sheet(tab)
        _write_month_tab(ws, resolution.shifts_for_month(month))
        tabs.append((month_key, tab))

    rules = wb.create_sheet("Rules & Legend")
    _write_rules_sheet(rules, stats or {}, Path(out_path).name)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    _repair_inlinestr(out_path)
    return out_path, tabs
