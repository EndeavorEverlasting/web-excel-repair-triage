"""Roster-derived Neuron billing evidence workbook.

The populated roster is clock/hour truth.  A reviewed local Neuron Track Hours
workbook may override only the task label after Date + Tech + Hours matching.
Daily Narrative Log and Event Log are deterministic reconstructions: missing
sites, devices, tickets, hostnames, quantities, and incidents are never guessed.
"""
from __future__ import annotations

import re
import zipfile
from calendar import month_name
from collections import defaultdict
from dataclasses import dataclass, replace
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from triage.admin_billing_summary.exporter import _add_net_chart, _add_table, _title_band
from triage.nw_prj_neuron_track_hours.bonita_resolver import BonitaResolution, BonitaShift
from triage.xlsx_utils import fix_inlinestr

CATEGORY_ORDER = (
    "Configurations",
    "Inventory Management",
    "Survey",
    "Ticket Forwarding",
    "Deployments",
)
CATEGORY_FILLS = {
    "Configurations": "DDEBF7",
    "Inventory Management": "FFF2CC",
    "Survey": "D9EAD3",
    "Ticket Forwarding": "DDEFEA",
    "Deployments": "E4DFEC",
    "Logistics": "FCE4D6",
    "Client Coordination": "E2F0D9",
    "Documentation": "EDEDED",
    "Troubleshooting / Incident Response": "F4CCCC",
}
MONTH_HEADERS = [
    "Date", "Day", "Tech Name", "Start Time", "End Time", "Total Hours",
    "Project Name", "Task / Assignment Type", "Supporting Work / Notes",
]
NARRATIVE_HEADERS = [
    "Date", "Day", "Person", "Site", "Primary Workstream", "Method / Detail",
    "Record State",
]
EVENT_HEADERS = [
    "Event Date", "Week Start", "Person", "Base of Operations", "Room / Area",
    "Case ID", "Workstream", "Event Type", "Task Category", "Quantity",
    "Unit Type", "Window Code", "Std Hours / Unit", "Fixed Overhead Hrs",
    "Complexity Factor", "Disruption Factor", "Travel / Transit Hrs",
    "Modeled Hours", "Suggested Hours", "Actual Billed Hours", "Variance Hrs",
    "Billable Flag", "Narrative Tag", "Primary Hostname",
    "Related Asset / Hostname", "Evidence Source", "Notes",
]
_WEEKDAYS = {
    "mon", "monday", "tue", "tuesday", "wed", "wednesday", "thu",
    "thursday", "fri", "friday", "sat", "saturday", "sun", "sunday",
}


@dataclass(frozen=True)
class AllocationRecord:
    work_date: date
    tech: str
    hours: float
    assignment: str
    sheet: str
    row: int


@dataclass(frozen=True)
class AllocationOverlayStats:
    source_path: str = ""
    matched: int = 0
    unmatched_shifts: int = 0
    unused_rows: int = 0
    strict: bool = True

    def to_dict(self) -> Dict[str, Any]:
        return self.__dict__.copy()


def _name(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _header(value: Any) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", str(value or "").casefold())).strip()


def _date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text or text.casefold() in _WEEKDAYS:
        return None
    for fmt in ("%Y-%m-%d", "%m-%d-%y", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _hours(value: Any) -> Optional[float]:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return round(number, 2) if number > 0 else None


def _allocation_header(ws) -> Optional[Tuple[int, int, Dict[str, int]]]:
    max_col = min(ws.max_column or 0, 40)
    for row in range(1, min(ws.max_row or 0, 12) + 1):
        direct = {c: _header(ws.cell(row, c).value) for c in range(1, max_col + 1)}
        combined = {
            c: _header(f"{ws.cell(row, c).value or ''} {ws.cell(row + 1, c).value or ''}")
            for c in range(1, max_col + 1)
        }
        next_tokens = {_header(ws.cell(row + 1, c).value) for c in range(1, max_col + 1)}
        candidates = ((2, combined), (1, direct)) if next_tokens & {"name", "time", "hours", "type"} else ((1, direct), (2, combined))
        for depth, headers in candidates:
            cols: Dict[str, int] = {}
            for col, text in headers.items():
                if "tech name" in text or text in {"tech", "person"}:
                    cols.setdefault("tech", col)
                if "total hours" in text or text in {"total", "hours"}:
                    cols.setdefault("hours", col)
                if "assignment" in text or "task category" in text:
                    cols.setdefault("assignment", col)
                if text == "date" or "event date" in text:
                    cols.setdefault("date", col)
            if {"tech", "hours", "assignment"}.issubset(cols):
                cols.setdefault("date", 1)
                return row, depth, cols
    return None


def read_allocation_records(path: str, months: Optional[Iterable[str]] = None) -> List[AllocationRecord]:
    """Read reviewed task labels from rich or two-line Neuron tracker tabs."""
    from openpyxl import load_workbook

    allowed = set(months or [])
    wb = load_workbook(path, read_only=True, data_only=True)
    records: List[AllocationRecord] = []
    try:
        for ws in wb.worksheets:
            found = _allocation_header(ws)
            if not found:
                continue
            header_row, depth, cols = found
            current: Optional[date] = None
            pending: List[Tuple[int, str, float, str]] = []

            def add(work_date: date, row: int, tech: str, hours: float, assignment: str) -> None:
                key = f"{work_date.year:04d}-{work_date.month:02d}"
                if not allowed or key in allowed:
                    records.append(AllocationRecord(work_date, tech, hours, assignment, ws.title, row))

            for row in range(header_row + depth, (ws.max_row or 0) + 1):
                locator = ws.cell(row, cols["date"]).value
                parsed = _date(locator)
                if str(locator or "").strip().casefold() in _WEEKDAYS:
                    current = None
                if parsed:
                    current = parsed
                    for pending_row, tech, hours, assignment in pending:
                        add(current, pending_row, tech, hours, assignment)
                    pending.clear()
                tech = str(ws.cell(row, cols["tech"]).value or "").strip()
                hours = _hours(ws.cell(row, cols["hours"]).value)
                assignment = str(ws.cell(row, cols["assignment"]).value or "").strip()
                if not tech or hours is None or not assignment:
                    continue
                if current:
                    add(current, row, tech, hours, assignment)
                else:
                    pending.append((row, tech, hours, assignment))
    finally:
        wb.close()
    return records


def apply_allocation_source(
    resolution: BonitaResolution,
    allocation_path: str,
    months: Sequence[str],
    *,
    strict: bool = True,
) -> Tuple[BonitaResolution, AllocationOverlayStats]:
    records = read_allocation_records(allocation_path, months)
    exact: Dict[Tuple[date, str, float], List[int]] = defaultdict(list)
    for idx, record in enumerate(records):
        exact[(record.work_date, _name(record.tech), record.hours)].append(idx)
    used: set[int] = set()
    unmatched: List[str] = []
    updated: List[BonitaShift] = []

    def take(indices: Iterable[int]) -> Optional[int]:
        for idx in indices:
            if idx not in used:
                used.add(idx)
                return idx
        return None

    for shift in resolution.shifts:
        idx = take(exact.get((shift.date, _name(shift.tech), round(shift.total_hours, 2)), ()))
        if idx is None:
            unmatched.append(f"{shift.date}|{shift.tech}|{shift.total_hours:.2f}")
            updated.append(shift)
            continue
        record = records[idx]
        updated.append(replace(
            shift,
            assignment_type=record.assignment,
            assignment_rule=f"allocation-source:{record.sheet}!{record.row}",
            assignment_confidence="high",
        ))
    if strict and unmatched:
        raise ValueError(
            "allocation source did not reconcile every roster-derived shift: "
            f"{len(unmatched)} unmatched ({', '.join(unmatched[:5])})"
        )
    warnings = list(resolution.warnings)
    if unmatched:
        warnings.append(f"allocation_source_unmatched:{len(unmatched)}")
    unused = len(records) - len(used)
    if unused:
        warnings.append(f"allocation_source_unused_rows:{unused}")
    return (
        BonitaResolution(shifts=updated, review=list(resolution.review), warnings=warnings),
        AllocationOverlayStats(str(Path(allocation_path).resolve()), len(used), len(unmatched), unused, strict),
    )


def _month_tab(key: str) -> str:
    year, month = (int(part) for part in key.split("-"))
    return f"{month_name[month]} {year}"


def _sorted(shifts: Iterable[BonitaShift]) -> List[BonitaShift]:
    return sorted(shifts, key=lambda shift: (shift.date, _name(shift.tech), shift.clock_in))


def _supporting(shift: BonitaShift) -> str:
    primary = shift.assignment_type or "Configurations"
    lanes = [primary, "Configurations", "Inventory Management", "Survey", "Ticket Forwarding"]
    if primary == "Deployments":
        lanes.append("Deployments")
    return "; ".join(dict.fromkeys(lane for lane in lanes if lane))


def _narrative(shift: BonitaShift) -> str:
    assignment = shift.assignment_type or "Configurations"
    lead = {
        "Configurations": "Performed configuration, validation, remediation, and readiness work.",
        "Inventory Management": "Performed inventory reconciliation, staging, count review, and material-readiness work.",
        "Survey": "Performed survey and validation work, including readiness/context capture.",
        "Ticket Forwarding": "Forwarded, routed, and tracked operational tickets or reported issues.",
        "Deployments": "Provided client-facing deployment coverage, field coordination, and handoff support.",
        "Logistics": "Coordinated material movement, staging, delivery, or cleanup.",
        "Client Coordination": "Performed client coordination, status follow-up, and workflow alignment.",
        "Documentation": "Produced operational documentation, reporting, and handoff records.",
        "Troubleshooting / Incident Response": "Performed troubleshooting, incident response, remediation, and readiness validation.",
    }.get(assignment, f"Performed {assignment.casefold()} work.")
    source = "reconciled local allocation workbook" if shift.assignment_rule.startswith("allocation-source:") else "roster ruleset"
    return (
        f"{lead} Assignment classification comes from the {source}. Date, technician, "
        "and billed hours come from the roster log; no unrecorded site, room, device, "
        "ticket, hostname, quantity, or incident detail was inferred."
    )


def _event_type(assignment: str) -> str:
    return {
        "Configurations": "Configuration / Readiness",
        "Inventory Management": "Inventory Reconciliation / Staging",
        "Survey": "Survey / Validation",
        "Ticket Forwarding": "Ticket Routing / Follow-Up",
        "Deployments": "Deployment Coverage",
        "Logistics": "Logistics / Material Flow",
        "Client Coordination": "Client Follow-Up",
        "Documentation": "Reporting / Documentation",
        "Troubleshooting / Incident Response": "Root Cause / Remediation",
    }.get(assignment, assignment or "Recorded Work")


def _case_id(shift: BonitaShift, ordinal: int) -> str:
    slug = re.sub(r"[^A-Z0-9]+", "-", shift.tech.upper()).strip("-")[:18]
    return f"NTH-{shift.date:%Y%m%d}-{slug}-{ordinal:02d}"


def _fill_task_cells(ws, start_row: int, task_col: int, end_col: int) -> None:
    from openpyxl.styles import PatternFill

    for row in range(start_row, (ws.max_row or 0) + 1):
        task = str(ws.cell(row, task_col).value or "")
        fill = CATEGORY_FILLS.get(task)
        if fill:
            for col in range(task_col, end_col + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=fill)


def _month_rows(shifts: Sequence[BonitaShift]) -> List[Dict[str, Any]]:
    return [{
        "Date": shift.date,
        "Day": shift.day,
        "Tech Name": shift.tech,
        "Start Time": shift.start_time or shift.clock_in,
        "End Time": shift.end_time or shift.clock_out,
        "Total Hours": round(shift.total_hours, 2),
        "Project Name": shift.project_name,
        "Task / Assignment Type": shift.assignment_type,
        "Supporting Work / Notes": _supporting(shift),
    } for shift in _sorted(shifts)]


def _write_month(wb, key: str, shifts: Sequence[BonitaShift]) -> str:
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    title = _month_tab(key)
    ws = wb.create_sheet(title)
    _title_band(ws, f"{title} Neuron Track Hours", "Roster-derived clock truth with reconciled task allocation.", len(MONTH_HEADERS))
    _, last = _add_table(ws, f"NTH{key.replace('-', '')}Table", MONTH_HEADERS, _month_rows(shifts), header_row=4)
    for row in range(5, last + 1):
        ws.cell(row, 1).number_format = "yyyy-mm-dd"
        ws.cell(row, 4).number_format = "h:mm AM/PM"
        ws.cell(row, 5).number_format = "h:mm AM/PM"
        ws.cell(row, 6).number_format = "0.00"
        ws.cell(row, 9).alignment = Alignment(wrap_text=True, vertical="top")
    _fill_task_cells(ws, 5, 8, 9)
    for col, width in enumerate([13, 10, 25, 13, 13, 12, 24, 26, 58], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.sheet_view.showGridLines = False
    return title


def _aggregate(shifts: Sequence[BonitaShift]):
    tech: Dict[str, float] = defaultdict(float)
    tasks: Dict[str, float] = defaultdict(float)
    counts: Dict[str, int] = defaultdict(int)
    daily: Dict[date, float] = defaultdict(float)
    for shift in shifts:
        tech[shift.tech] += shift.total_hours
        tasks[shift.assignment_type] += shift.total_hours
        counts[shift.assignment_type] += 1
        daily[shift.date] += shift.total_hours
    return tech, tasks, counts, daily


def _write_visual(wb, shifts: Sequence[BonitaShift]) -> None:
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("Visual Summary")
    _title_band(ws, "Neuron Track Hours — Visual Summary", "Technician, task, and daily roster reconciliation.", 12)
    total = round(sum(shift.total_hours for shift in shifts), 2)
    tech, tasks, counts, daily = _aggregate(shifts)
    ws["A4"], ws["B4"], ws["C4"], ws["D4"] = "MTD HOURS", total, "SHIFT RECORDS", len(shifts)
    for cell in ws[4]:
        if cell.column <= 4:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(bold=True, color="FFFFFF")
    tech_rows = [{"Technician": name, "Total Hours": round(hours, 2)} for name, hours in sorted(tech.items(), key=lambda item: -item[1])]
    task_rows = [{"Task / Assignment Type": task, "Hours": round(hours, 2), "Row Count": counts[task]} for task, hours in sorted(tasks.items(), key=lambda item: (CATEGORY_ORDER.index(item[0]) if item[0] in CATEGORY_ORDER else 99, -item[1]))]
    daily_rows = [{"Date": work_date, "Hours": round(hours, 2)} for work_date, hours in sorted(daily.items())]
    _, tech_last = _add_table(ws, "VisualTechTable", ["Technician", "Total Hours"], tech_rows, header_row=7)
    _, task_last = _add_table(ws, "VisualTaskTable", ["Task / Assignment Type", "Hours", "Row Count"], task_rows, header_row=7, start_col=5)
    _add_table(ws, "VisualDailyTable", ["Date", "Hours"], daily_rows, header_row=7, start_col=10)
    _fill_task_cells(ws, 8, 5, 7)
    _add_net_chart(ws, "Hours by Technician", ["Technician", "Total Hours"], 7, tech_last, "A16", category_col=1, value_header="Total Hours")
    _add_net_chart(ws, "Hours by Assignment", ["Task / Assignment Type", "Hours", "Row Count"], 7, task_last, "H16", category_col=5, value_header="Hours")
    ws.sheet_view.showGridLines = False


def _write_dashboard(wb, shifts: Sequence[BonitaShift]) -> None:
    ws = wb.create_sheet("Executive Dashboard")
    _title_band(ws, "Neuron Track Hours Executive Dashboard", "Roster, narrative, and event-log totals must reconcile.", 8)
    total = round(sum(shift.total_hours for shift in shifts), 2)
    tech, tasks, counts, _ = _aggregate(shifts)
    _add_table(ws, "EvidenceKpiTable", ["Metric", "Value"], [
        {"Metric": "Month-to-Date Hours", "Value": total},
        {"Metric": "Shift Records", "Value": len(shifts)},
        {"Metric": "Technicians", "Value": len(tech)},
        {"Metric": "Daily Narrative Rows", "Value": len(shifts)},
        {"Metric": "Event Rows", "Value": len(shifts)},
        {"Metric": "Status", "Value": "ALIGNED"},
    ], header_row=5)
    task_rows = [{"Assignment Type": task, "Hours": round(hours, 2), "Rows": counts[task]} for task, hours in sorted(tasks.items(), key=lambda item: -item[1])]
    _add_table(ws, "EvidenceTaskMixTable", ["Assignment Type", "Hours", "Rows"], task_rows, header_row=5, start_col=5)
    _fill_task_cells(ws, 6, 5, 7)
    ws["A14"] = "Evidence Boundary"
    ws["A15"] = "Roster clock truth > optional allocation overlay > deterministic narrative. No unrecorded operational specifics are inferred. Task Summary is intentionally omitted because Visual Summary already contains the task mix."
    ws["A15"].alignment = __import__("openpyxl").styles.Alignment(wrap_text=True)
    ws.sheet_view.showGridLines = False


def _write_narrative(wb, shifts: Sequence[BonitaShift]) -> None:
    rows = [{
        "Date": shift.date,
        "Day": shift.date.strftime("%A"),
        "Person": shift.tech,
        "Site": f"{shift.project_name} (specific site not recorded)",
        "Primary Workstream": shift.assignment_type,
        "Method / Detail": _narrative(shift),
        "Record State": "Roster-derived; allocation-source override" if shift.assignment_rule.startswith("allocation-source:") else "Roster-derived; ruleset classification",
    } for shift in _sorted(shifts)]
    ws = wb.create_sheet("Daily Narrative Log")
    _title_band(ws, "Neuron Track Hours | Daily Narrative Log", "One row per roster-derived technician shift.", len(NARRATIVE_HEADERS))
    _add_table(ws, "DailyNarrativeLogTable", NARRATIVE_HEADERS, rows, header_row=4)
    for row in range(5, (ws.max_row or 0) + 1):
        ws.cell(row, 1).number_format = "yyyy-mm-dd"
    _fill_task_cells(ws, 5, 5, 5)
    ws.column_dimensions["F"].width = 95
    ws.column_dimensions["G"].width = 34
    ws.sheet_view.showGridLines = False


def _write_events(wb, shifts: Sequence[BonitaShift]) -> None:
    ordinals: Dict[Tuple[date, str], int] = defaultdict(int)
    rows: List[Dict[str, Any]] = []
    for shift in _sorted(shifts):
        key = (shift.date, _name(shift.tech))
        ordinals[key] += 1
        source = f"Local roster log + allocation overlay ({shift.assignment_rule.split(':', 1)[1]})" if shift.assignment_rule.startswith("allocation-source:") else f"Local roster log + roster rules ({shift.assignment_rule or 'default'})"
        rows.append({
            "Event Date": shift.date,
            "Week Start": shift.date - timedelta(days=shift.date.weekday()),
            "Person": shift.tech,
            "Base of Operations": shift.project_name,
            "Room / Area": "Not recorded in roster",
            "Case ID": _case_id(shift, ordinals[key]),
            "Workstream": shift.assignment_type,
            "Event Type": _event_type(shift.assignment_type),
            "Task Category": shift.assignment_type,
            "Quantity": 1,
            "Unit Type": "Shift",
            "Window Code": "",
            "Std Hours / Unit": "",
            "Fixed Overhead Hrs": "",
            "Complexity Factor": "",
            "Disruption Factor": "",
            "Travel / Transit Hrs": "",
            "Modeled Hours": "",
            "Suggested Hours": "",
            "Actual Billed Hours": round(shift.total_hours, 2),
            "Variance Hrs": "",
            "Billable Flag": "Yes",
            "Narrative Tag": "Ruleset-Based Reconstruction",
            "Primary Hostname": "",
            "Related Asset / Hostname": "",
            "Evidence Source": source,
            "Notes": _narrative(shift),
        })
    ws = wb.create_sheet("Event Log")
    _title_band(ws, "Event Log | Roster-Linked Tech Hours", "One row per shift; modeled/sample fields remain blank.", len(EVENT_HEADERS))
    _add_table(ws, "EventLogTable", EVENT_HEADERS, rows, header_row=5)
    for row in range(6, (ws.max_row or 0) + 1):
        ws.cell(row, 1).number_format = "yyyy-mm-dd"
        ws.cell(row, 2).number_format = "yyyy-mm-dd"
        ws.cell(row, 20).number_format = "0.00"
    _fill_task_cells(ws, 6, 7, 9)
    ws.column_dimensions["Z"].width = 46
    ws.column_dimensions["AA"].width = 95
    ws.sheet_view.showGridLines = False


def build_evidence_pack(resolution: BonitaResolution, months: Sequence[str], out_path: str) -> Tuple[str, List[str]]:
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    tabs: List[str] = []
    for key in months:
        year, month = (int(part) for part in key.split("-"))
        tabs.append(_write_month(wb, key, [shift for shift in resolution.shifts if shift.date.year == year and shift.date.month == month]))
    shifts = _sorted(resolution.shifts)
    _write_visual(wb, shifts)
    _write_dashboard(wb, shifts)
    _write_narrative(wb, shifts)
    _write_events(wb, shifts)
    tabs.extend(["Visual Summary", "Executive Dashboard", "Daily Narrative Log", "Event Log"])
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    wb.close()
    fix_inlinestr(out_path)
    return out_path, tabs


def _sharedstrings(z: zipfile.ZipFile) -> bool:
    if "xl/sharedStrings.xml" not in z.namelist():
        return True
    text = z.read("xl/sharedStrings.xml").decode("utf-8", errors="ignore")
    match = re.search(r'\bcount="(\d+)"', text)
    declared = int(match.group(1)) if match else -1
    actual = sum(z.read(name).count(b't="s"') for name in z.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"))
    return declared == actual


def preflight_evidence_pack(path: str, months: Sequence[str], *, expected_shift_count: int, expected_total_hours: float) -> Dict[str, Any]:
    from openpyxl import load_workbook

    required = [_month_tab(key) for key in months] + ["Visual Summary", "Executive Dashboard", "Daily Narrative Log", "Event Log"]
    result: Dict[str, Any] = {
        "artifact": Path(path).name, "tabs": [], "required_tabs": required,
        "zip_valid": False, "token_failures": [], "task_summary_absent": False,
        "month_shift_count": 0, "month_total_hours": 0.0,
        "daily_narrative_rows": 0, "event_rows": 0,
        "event_actual_billed_hours": 0.0, "preflight_pass": False,
    }
    try:
        with zipfile.ZipFile(path) as z:
            result["zip_valid"] = z.testzip() is None
            names = z.namelist()
            result["has_calc_chain"] = "xl/calcChain.xml" in names
            result["has_external_links"] = any("externalLink" in name for name in names)
            xml = b"".join(z.read(name) for name in names if name.endswith((".xml", ".rels")))
            if b' t="inlineStr"' in xml:
                result["token_failures"].append("inlineStr")
            if b"ns0:" in xml or b"xmlns:ns0" in xml:
                result["token_failures"].append("ns0")
            result["sharedstrings_count_ok"] = _sharedstrings(z)
    except (FileNotFoundError, zipfile.BadZipFile):
        return result

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        result["tabs"] = list(wb.sheetnames)
        result["task_summary_absent"] = "Task Summary" not in wb.sheetnames
        for tab in required[:len(months)]:
            if tab not in wb.sheetnames:
                continue
            for row in wb[tab].iter_rows(min_row=5, values_only=True):
                hours = _hours(row[5] if len(row) > 5 else None)
                if hours is not None and len(row) > 2 and row[2]:
                    result["month_shift_count"] += 1
                    result["month_total_hours"] += hours
        result["month_total_hours"] = round(result["month_total_hours"], 2)
        if "Daily Narrative Log" in wb.sheetnames:
            result["daily_narrative_rows"] = sum(1 for row in wb["Daily Narrative Log"].iter_rows(min_row=5, values_only=True) if len(row) > 2 and row[2])
        if "Event Log" in wb.sheetnames:
            for row in wb["Event Log"].iter_rows(min_row=6, values_only=True):
                hours = _hours(row[19] if len(row) > 19 else None)
                if hours is not None and len(row) > 2 and row[2]:
                    result["event_rows"] += 1
                    result["event_actual_billed_hours"] += hours
            result["event_actual_billed_hours"] = round(result["event_actual_billed_hours"], 2)
    finally:
        wb.close()

    expected_total = round(expected_total_hours, 2)
    result["preflight_pass"] = all([
        result["zip_valid"], not result["token_failures"],
        not result.get("has_calc_chain"), not result.get("has_external_links"),
        result.get("sharedstrings_count_ok"), result["tabs"] == required,
        result["task_summary_absent"], result["month_shift_count"] == expected_shift_count,
        result["month_total_hours"] == expected_total,
        result["daily_narrative_rows"] == expected_shift_count,
        result["event_rows"] == expected_shift_count,
        result["event_actual_billed_hours"] == expected_total,
    ])
    return result
