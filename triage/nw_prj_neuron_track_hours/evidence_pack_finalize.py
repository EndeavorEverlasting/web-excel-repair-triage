"""Final repair-free chart and allocation guards for the evidence-pack CLI."""
from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Sequence, Tuple

from triage.nw_prj_neuron_track_hours.bonita_resolver import BonitaResolution
from triage.nw_prj_neuron_track_hours.evidence_pack import read_allocation_records
from triage.xlsx_utils import fix_inlinestr


def _key(work_date, tech: str, hours: float) -> Tuple[object, str, float]:
    return work_date, " ".join(str(tech or "").split()).casefold(), round(float(hours), 2)


def validate_allocation_source_exact(
    resolution: BonitaResolution,
    allocation_path: str,
    months: Sequence[str],
) -> None:
    """Require a one-to-one Date + Tech + Hours allocation reconciliation."""
    expected = Counter(_key(s.date, s.tech, s.total_hours) for s in resolution.shifts)
    actual = Counter(
        _key(r.work_date, r.tech, r.hours)
        for r in read_allocation_records(allocation_path, months)
    )
    missing = expected - actual
    extra = actual - expected
    if missing or extra:
        missing_text = ", ".join(
            f"{day}|{tech}|{hours:.2f} x{count}"
            for (day, tech, hours), count in list(missing.items())[:5]
        )
        extra_text = ", ".join(
            f"{day}|{tech}|{hours:.2f} x{count}"
            for (day, tech, hours), count in list(extra.items())[:5]
        )
        raise ValueError(
            "allocation source must reconcile exactly by Date + Tech + Hours; "
            f"missing=[{missing_text}] extra=[{extra_text}]"
        )


def _last_data_row(ws, column: int, start: int) -> int:
    row = start
    while row <= (ws.max_row or start) and ws.cell(row, column).value not in (None, ""):
        row += 1
    return row - 1


def repair_visual_summary_charts(path: str) -> None:
    """Rebuild Visual Summary chart references against their actual tables."""
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference

    workbook_path = Path(path)
    wb = load_workbook(workbook_path)
    try:
        if "Visual Summary" not in wb.sheetnames:
            return
        ws = wb["Visual Summary"]
        tech_last = _last_data_row(ws, 1, 8)
        task_last = _last_data_row(ws, 5, 8)
        ws._charts = []

        if tech_last >= 8:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Hours by Technician"
            chart.y_axis.title = "Total Hours"
            chart.height = 9
            chart.width = 20
            chart.add_data(
                Reference(ws, min_col=2, min_row=7, max_row=tech_last),
                titles_from_data=True,
            )
            chart.set_categories(Reference(ws, min_col=1, min_row=8, max_row=tech_last))
            chart.legend = None
            ws.add_chart(chart, "A16")

        if task_last >= 8:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Hours by Assignment"
            chart.y_axis.title = "Hours"
            chart.height = 9
            chart.width = 20
            chart.add_data(
                Reference(ws, min_col=6, min_row=7, max_row=task_last),
                titles_from_data=True,
            )
            chart.set_categories(Reference(ws, min_col=5, min_row=8, max_row=task_last))
            chart.legend = None
            ws.add_chart(chart, "H16")

        wb.save(workbook_path)
    finally:
        wb.close()
    fix_inlinestr(str(workbook_path))
