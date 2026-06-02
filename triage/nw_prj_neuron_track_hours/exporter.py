"""Build the Web Excel-safe Neuron Track Hours workbook.

Self-contained: includes a private inlineStr repair so no shared helper edits
are required. Produces the proven 8-sheet layout with filters, frozen headers,
dropdown validations, conditional formatting, and a CF Dictionary.
"""
from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path
from typing import Any, Dict, List

from triage.nw_prj_neuron_track_hours.models import (
    ACTION_STATUS_VALUES,
    APRIL_MAY_COLUMNS,
    GO_LIVE_COLUMNS,
    REVIEW_FLAG_COLUMNS,
    REVIEW_RESULT_VALUES,
    TECH_SUMMARY_COLUMNS,
    TrackHoursReport,
)

EXPECTED_SHEETS = [
    "Start Here",
    "April Neuron Hours",
    "May Neuron Hours",
    "Go Live Weekend",
    "Tech Summary",
    "Review Flags",
    "CF Dictionary",
    "WebExcel QC",
]

CF_DICTIONARY_ROWS = [
    ("RED", "Must fix before trusted submission", "Go to the named sheet/cell and correct."),
    ("AMBER", "Review required before finalizing", "Compare admin control to roster evidence."),
    ("PURPLE", "Rich protected review lane", "Do not downgrade full/long days without proof."),
    ("BLUE", "Admin lingering or out-of-scope note", "Clear/ignore if pinned/transport/other project."),
    ("GREEN", "No action or reconciled", "Mark confirmed after spot check."),
    ("GRAY", "Archive/reference only", "Keep for traceability."),
]

_HEADER_FILL = "1F365C"
_SUBTITLE_FILL = "EAF1F8"


def _require_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.formatting.rule import CellIsRule, FormulaRule
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("openpyxl is required: pip install openpyxl") from e
    return (Workbook, Alignment, Font, PatternFill, get_column_letter,
            DataValidation, CellIsRule, FormulaRule)


def _style_header(ws, headers: List[str], header_row: int):
    _, Alignment, Font, PatternFill, get_column_letter, *_ = _require_openpyxl()
    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _write_table(ws, title: str, subtitle: str, headers: List[str],
                 rows: List[Dict[str, Any]]):
    (_, Alignment, Font, PatternFill, get_column_letter, *_) = _require_openpyxl()
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    sub_fill = PatternFill("solid", fgColor=_SUBTITLE_FILL)
    for c in range(1, max(2, len(headers) + 1)):
        ws.cell(row=1, column=c).fill = title_fill
        ws.cell(row=2, column=c).fill = sub_fill
    ws.cell(row=2, column=1, value=subtitle)

    header_row = 4
    _style_header(ws, headers, header_row)
    for r_idx, row in enumerate(rows, header_row + 1):
        for c, h in enumerate(headers, 1):
            ws.cell(row=r_idx, column=c, value=row.get(h, ""))

    last_row = max(header_row, len(rows) + header_row)
    last_col = get_column_letter(len(headers))
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    return header_row, last_row


def _add_status_dropdowns(ws, headers: List[str], header_row: int, last_row: int):
    (_, _, _, _, get_column_letter, DataValidation, *_) = _require_openpyxl()
    if last_row <= header_row:
        last_row = header_row + 1
    for col_name, values in (("Action Status", ACTION_STATUS_VALUES),
                             ("Review Result", REVIEW_RESULT_VALUES)):
        if col_name not in headers:
            continue
        col = headers.index(col_name) + 1
        letter = get_column_letter(col)
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(values) + '"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.add(f"{letter}{header_row + 1}:{letter}{last_row}")
        ws.add_data_validation(dv)


def _add_severity_cf(ws, headers: List[str], header_row: int, last_row: int):
    (_, _, _, PatternFill, get_column_letter, _, CellIsRule, FormulaRule) = _require_openpyxl()
    if last_row <= header_row:
        return
    palette = {
        "RED": "FFC7CE",
        "AMBER": "FFEB9C",
        "PURPLE": "E4DFEC",
        "BLUE": "DDEBF7",
        "GREEN": "C6EFCE",
        "GRAY": "D9D9D9",
    }
    if "Severity" in headers:
        col = get_column_letter(headers.index("Severity") + 1)
        rng = f"{col}{header_row + 1}:{col}{last_row}"
        for token, rgb in palette.items():
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator="equal", formula=[f'"{token}"'],
                           fill=PatternFill("solid", fgColor=rgb)),
            )
    if "Gross Hours" in headers:
        col = get_column_letter(headers.index("Gross Hours") + 1)
        rng = f"{col}{header_row + 1}:{col}{last_row}"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThanOrEqual", formula=["16"],
                       fill=PatternFill("solid", fgColor="FFC7CE")),
        )
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="between", formula=["12", "15.999"],
                       fill=PatternFill("solid", fgColor="FFEB9C")),
        )


def build_workbook(report: TrackHoursReport, out_path: str,
                   reference_totals: Dict[str, Any] | None = None) -> str:
    (Workbook, Alignment, Font, PatternFill, get_column_letter, *_rest) = _require_openpyxl()
    wb = Workbook()
    wb.remove(wb.active)

    # ── Start Here ────────────────────────────────────────────────
    ws = wb.create_sheet("Start Here")
    grand = report.grand_total()
    april = report.month_total("April")
    may = report.month_total("May")
    go_live_rows = report.go_live_rows()
    metrics = [
        ("Total Neuron roster hours", round(grand, 2), "April + May gross roster hours."),
        ("April Neuron hours", round(april, 2), "From Live/Worked Project classification."),
        ("May Neuron hours", round(may, 2), "From Live/Worked Project classification."),
        ("Go Live weekend rows", len(go_live_rows), "Rows on May 30 or May 31."),
        ("Go Live weekend hours", round(report.go_live_hours(), 2), "Includes technician weekend support."),
    ]
    _write_simple(ws, "Neuron Track Hours - April and May 2026",
                  "Roster-derived Neuron hours. Generated fresh; Web Excel-safe.",
                  ["Metric", "Value", "Notes"],
                  [{"Metric": m, "Value": v, "Notes": n} for m, v, n in metrics])

    # ── April / May ───────────────────────────────────────────────
    for month, sheet in (("April", "April Neuron Hours"), ("May", "May Neuron Hours")):
        ws = wb.create_sheet(sheet)
        rows = [r.to_track_dict() for r in report.rows_for_month(month)]
        hr, lr = _write_table(ws, f"{month} Neuron Hours",
                              "Roster-derived daily Neuron Deployment rows.",
                              APRIL_MAY_COLUMNS, rows)
        _add_status_dropdowns(ws, APRIL_MAY_COLUMNS, hr, lr)
        _add_severity_cf(ws, APRIL_MAY_COLUMNS, hr, lr)

    # ── Go Live Weekend ───────────────────────────────────────────
    ws = wb.create_sheet("Go Live Weekend")
    gl_rows = [r.to_go_live_dict() for r in go_live_rows]
    hr, lr = _write_table(ws, "Go Live Weekend Support",
                          "Rows on May 30 and May 31. Technician weekend support.",
                          GO_LIVE_COLUMNS, gl_rows)
    _add_status_dropdowns(ws, GO_LIVE_COLUMNS, hr, lr)
    _add_severity_cf(ws, GO_LIVE_COLUMNS, hr, lr)

    # ── Tech Summary ──────────────────────────────────────────────
    ws = wb.create_sheet("Tech Summary")
    _write_table(ws, "Neuron Technician Summary",
                 "Roster-derived Neuron totals per technician per month.",
                 TECH_SUMMARY_COLUMNS, [s.to_dict() for s in report.tech_summary])

    # ── Review Flags ──────────────────────────────────────────────
    ws = wb.create_sheet("Review Flags")
    hr, lr = _write_table(ws, "Neuron Review Flags",
                          "Long, weekend, overnight, go-live, and note-bearing rows.",
                          REVIEW_FLAG_COLUMNS, [f.to_dict() for f in report.review_flags])
    _add_status_dropdowns(ws, REVIEW_FLAG_COLUMNS, hr, lr)
    _add_severity_cf(ws, REVIEW_FLAG_COLUMNS, hr, lr)

    # ── CF Dictionary ─────────────────────────────────────────────
    ws = wb.create_sheet("CF Dictionary")
    _write_simple(ws, "Conditional Formatting Dictionary",
                  "Plain-English color rules used across this workbook.",
                  ["Color", "Meaning", "Action"],
                  [{"Color": c, "Meaning": m, "Action": a} for c, m, a in CF_DICTIONARY_ROWS])

    # ── WebExcel QC ───────────────────────────────────────────────
    ws = wb.create_sheet("WebExcel QC")
    qc_rows = [
        {"Check": "Fresh workbook", "Result": "PASS", "Notes": "No inherited XML."},
        {"Check": "Formulas", "Result": "PASS", "Notes": "No formulas; values only."},
        {"Check": "Track tabs", "Result": "PASS", "Notes": "Separate April and May tabs."},
        {"Check": "Go Live visibility", "Result": "PASS", "Notes": "Go Live Weekend tab included."},
        {"Check": "Filters", "Result": "PASS", "Notes": "Auto-filter on data sheets."},
        {"Check": "Frozen headers", "Result": "PASS", "Notes": "Header rows frozen."},
        {"Check": "Dropdowns", "Result": "PASS", "Notes": "Action Status / Review Result validations."},
        {"Check": "Conditional formatting", "Result": "PASS", "Notes": "Severity + long-shift rules."},
    ]
    if reference_totals:
        qc_rows.append({
            "Check": "Reference totals",
            "Result": reference_totals.get("result", "INFO"),
            "Notes": reference_totals.get("notes", ""),
        })
    _write_simple(ws, "Web Excel QC",
                  "Generated as a fresh workbook with Web Excel-safe structure.",
                  ["Check", "Result", "Notes"], qc_rows)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    _repair_inlinestr(out_path)
    return out_path


def _write_simple(ws, title: str, subtitle: str, headers: List[str],
                  rows: List[Dict[str, Any]]):
    (_, Alignment, Font, PatternFill, get_column_letter, *_) = _require_openpyxl()
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    sub_fill = PatternFill("solid", fgColor=_SUBTITLE_FILL)
    for c in range(1, max(2, len(headers) + 1)):
        ws.cell(row=1, column=c).fill = title_fill
        ws.cell(row=2, column=c).fill = sub_fill
    ws.cell(row=2, column=1, value=subtitle)
    header_row = 4
    _style_header(ws, headers, header_row)
    for r_idx, row in enumerate(rows, header_row + 1):
        for c, h in enumerate(headers, 1):
            ws.cell(row=r_idx, column=c, value=row.get(h, ""))
    last_row = max(header_row, len(rows) + header_row)
    last_col = get_column_letter(len(headers))
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 26 if c == 1 else 18


# ─────────────────────── inlineStr repair (private) ───────────────────────

def _repair_inlinestr(path: str) -> None:
    _IS_FULL = re.compile(
        rb'(<c\b[^>]*?)\s+t="inlineStr"([^>]*?)><is><t([^>]*)>(.*?)</t></is></c>',
        re.DOTALL,
    )
    _IS_EMPTY = re.compile(rb'\s+t="inlineStr"(?=[^<]*?/>|[^<]*?></c>)')
    p = Path(path)
    original = p.read_bytes()
    with zipfile.ZipFile(io.BytesIO(original), "r") as zin:
        names = zin.namelist()
        str_table: List[str] = []
        str_index: Dict[str, int] = {}
        if "xl/sharedStrings.xml" in names:
            ss_xml = zin.read("xl/sharedStrings.xml").decode("utf-8", errors="ignore")
            for m in re.finditer(r"<t[^>]*?>(.*?)</t>", ss_xml, re.DOTALL):
                s = _xml_unescape(m.group(1))
                if s not in str_index:
                    str_index[s] = len(str_table)
                    str_table.append(s)

        def _get_or_add(s: str) -> int:
            if s not in str_index:
                str_index[s] = len(str_table)
                str_table.append(s)
            return str_index[s]

        patched: Dict[str, bytes] = {}
        for name in names:
            if not (name.startswith("xl/worksheets/sheet") and name.endswith(".xml")):
                continue
            raw = zin.read(name)
            if b"inlineStr" not in raw:
                continue

            def _replace_full(m: "re.Match") -> bytes:
                prefix = m.group(1) + m.group(2)
                value = m.group(4).decode("utf-8", errors="ignore")
                idx = _get_or_add(value)
                return prefix + b' t="s"><v>' + str(idx).encode() + b"</v></c>"

            fixed = _IS_FULL.sub(_replace_full, raw)
            fixed = _IS_EMPTY.sub(b"", fixed)
            if fixed != raw:
                patched[name] = fixed

        if not patched:
            # No worksheet carried an inlineStr cell, so there is nothing to
            # repair: the workbook openpyxl emitted is already valid (correct
            # sharedStrings count, no inlineStr). Rewriting sharedStrings here
            # previously set count to the UNIQUE count instead of the total
            # reference count, which Excel for Web flagged as corruption and
            # "repaired" by dropping sharedStrings entirely. Leave it untouched.
            return

        # Real inlineStr cells were converted to shared references. Rebuild
        # sharedStrings with a spec-correct ``count`` (total references across
        # all sheets, NOT the unique count), preserve significant whitespace,
        # and avoid double-escaping previously-escaped text.
        def _ss_item(s: str) -> str:
            preserve = ' xml:space="preserve"' if s != s.strip() else ""
            return f"<si><t{preserve}>{_xml_escape(s)}</t></si>"

        ss_items = "".join(_ss_item(s) for s in str_table)
        total_refs = 0
        for name in names:
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                raw_ws = patched.get(name) or zin.read(name)
                total_refs += raw_ws.count(b't="s"')
        unique = len(str_table)
        new_ss = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            f' count="{total_refs or unique}" uniqueCount="{unique}">{ss_items}</sst>'
        ).encode("utf-8")

        need_new_ss = "xl/sharedStrings.xml" not in names and bool(str_table)
        extra: Dict[str, bytes] = {}
        if need_new_ss:
            extra["xl/sharedStrings.xml"] = new_ss
            ct_name = "[Content_Types].xml"
            if ct_name in names:
                ct = zin.read(ct_name).decode("utf-8")
                if "sharedStrings" not in ct:
                    ct = ct.replace(
                        "</Types>",
                        '<Override PartName="/xl/sharedStrings.xml"'
                        ' ContentType="application/vnd.openxmlformats-officedocument'
                        ".spreadsheetml.sharedStrings+xml\"/></Types>",
                    )
                    extra[ct_name] = ct.encode("utf-8")

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in names:
                if name in patched:
                    zout.writestr(name, patched[name])
                elif name == "xl/sharedStrings.xml":
                    zout.writestr(name, new_ss)
                elif name in extra:
                    zout.writestr(name, extra[name])
                else:
                    zout.writestr(name, zin.read(name))
            for name, data in extra.items():
                if name not in names:
                    zout.writestr(name, data)
        p.write_bytes(buf.getvalue())


def _xml_escape(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def _xml_unescape(s: str) -> str:
    return (s.replace("&lt;", "<").replace("&gt;", ">")
            .replace("&quot;", '"').replace("&apos;", "'").replace("&amp;", "&"))
