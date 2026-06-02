"""Read the Active Roster Log and derive daily Neuron Deployment hours.

Source layout (proven against the real roster workbook):
  - "Live - {Month YYYY}"  : wide form, row 2 headers
        Staff Name | Project | "Apr 01 - Clock In" | "Apr 01 - Clock Out" | ...
  - "Worked Projects - {Month YYYY}" : per-date project classification
        Staff Name | Default Project | <date columns with project text per day>

Neuron classification rule (matches TRUE_NEURON_RECON_POLICY):
  Resolve each staff/date project as Worked-Projects cell when present, else the
  Live default project. A staff/date enters scope only when the resolved project
  is documented as a Neuron Deployment.
"""
from __future__ import annotations

import re
from calendar import month_name
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from triage.nw_prj_neuron_track_hours.models import (
    GO_LIVE_WEEKEND_DATES,
    NeuronHoursRow,
)


class RosterReadError(Exception):
    pass


_DATE_HEADER = re.compile(
    r"^([A-Za-z]+)\s+(\d{1,2})\s*[-\u2013]\s*(Clock\s*In|Clock\s*Out)\s*$",
    re.IGNORECASE,
)
_MONTH_ABBREVS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
_PUNCH_RE = re.compile(r"(\d{1,2}):(\d{2})(?::(\d{2}))?\s*(AM|PM)?", re.IGNORECASE)


def split_note_bearing_punch(value: Any) -> Tuple[Optional[float], str]:
    """Parse a punch cell that may carry a trailing note.

    Examples:
        "9:28:00 AM/ Bonita" -> (9.4667, "Bonita")
        "5:00 PM - covered"  -> (17.0, "covered")
        "9:00 AM"            -> (9.0, "")
    Returns (decimal_hours_or_None, note_text).
    """
    if value is None:
        return None, ""
    if isinstance(value, datetime):
        return value.hour + value.minute / 60.0 + value.second / 3600.0, ""
    if isinstance(value, time):
        return value.hour + value.minute / 60.0 + value.second / 3600.0, ""
    if isinstance(value, timedelta):
        return value.total_seconds() / 3600.0, ""
    if isinstance(value, (int, float)):
        f = float(value)
        if 0.0 <= f < 2.0:
            return f * 24.0, ""
        return None, ""
    if isinstance(value, str):
        s = value.strip()
        m = _PUNCH_RE.match(s)
        if not m:
            return None, s
        hour = int(m.group(1))
        minute = int(m.group(2))
        second = int(m.group(3)) if m.group(3) else 0
        ampm = (m.group(4) or "").upper()
        if ampm == "PM" and hour != 12:
            hour += 12
        elif ampm == "AM" and hour == 12:
            hour = 0
        decimal = hour + minute / 60.0 + second / 3600.0
        note = s[m.end():].strip()
        note = re.sub(r"^[\s/\-\u2013:]+", "", note).strip()
        return decimal, note
    return None, ""


def _format_clock(decimal_hours: Optional[float]) -> str:
    if decimal_hours is None:
        return ""
    total_min = int(round(decimal_hours * 60))
    total_min %= 24 * 60
    hour24 = total_min // 60
    minute = total_min % 60
    ampm = "AM" if hour24 < 12 else "PM"
    hour12 = hour24 % 12
    if hour12 == 0:
        hour12 = 12
    return f"{hour12}:{minute:02d} {ampm}"


def _decimal_to_time(decimal_hours: Optional[float]) -> Optional[time]:
    """Decimal hours (e.g. 9.5) -> a real ``datetime.time`` for true time cells."""
    if decimal_hours is None:
        return None
    total_min = int(round(decimal_hours * 60)) % (24 * 60)
    return time(total_min // 60, total_min % 60)


def _compute_gross(clock_in: Optional[float], clock_out: Optional[float]) -> float:
    if clock_in is None or clock_out is None:
        return 0.0
    diff = clock_out - clock_in
    if diff < 0:
        diff += 24.0
    return round(diff, 4)


def _month_label(month_key: str) -> Tuple[str, int, int]:
    """'2026-04' -> ('April 2026', 2026, 4)."""
    m = re.match(r"^(\d{4})-(\d{1,2})$", month_key.strip())
    if not m:
        raise RosterReadError(f"Invalid month key (expected YYYY-MM): {month_key}")
    year = int(m.group(1))
    mon = int(m.group(2))
    return f"{month_name[mon]} {year}", year, mon


def _find_sheet(wb, prefix: str, label: str):
    target = f"{prefix} - {label}".lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return wb[name]
    # tolerant match: prefix + month word
    month_word = label.split()[0].lower()
    for name in wb.sheetnames:
        low = name.strip().lower()
        if low.startswith(prefix.lower()) and month_word in low and label.split()[-1] in name:
            return wb[name]
    return None


def _worked_project_lookup(ws) -> Dict[Tuple[str, date], str]:
    lut: Dict[Tuple[str, date], str] = {}
    if ws is None:
        return lut
    header_row = 2
    date_cols: Dict[int, date] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, datetime):
            date_cols[c] = v.date()
        elif isinstance(v, date):
            date_cols[c] = v
    if not date_cols:
        return lut
    for r in range(header_row + 1, ws.max_row + 1):
        staff = ws.cell(r, 1).value
        if not staff or not str(staff).strip():
            continue
        staff_name = str(staff).strip()
        for c, d in date_cols.items():
            val = ws.cell(r, c).value
            if val is not None and str(val).strip():
                lut[(staff_name, d)] = str(val).strip()
    return lut


def _is_neuron(project: str) -> bool:
    return "neuron" in (project or "").lower()


def read_month(
    wb,
    month_key: str,
    pinned_techs: Optional[List[str]] = None,
) -> Tuple[List[NeuronHoursRow], List[str]]:
    label, year, mon = _month_label(month_key)
    warnings: List[str] = []
    live_ws = _find_sheet(wb, "Live", label)
    if live_ws is None:
        raise RosterReadError(f"Live sheet not found for {label}")
    worked_ws = _find_sheet(wb, "Worked Projects", label)
    if worked_ws is None:
        warnings.append(f"worked_projects_missing:{label}:falling_back_to_default_project")
    worked = _worked_project_lookup(worked_ws)

    header_row = 2
    headers = [live_ws.cell(header_row, c).value for c in range(1, live_ws.max_column + 1)]

    idx_staff = 0
    idx_project = 1
    date_to_cols: Dict[date, Dict[str, int]] = {}
    for i, h in enumerate(headers):
        if not isinstance(h, str):
            continue
        mm = _DATE_HEADER.match(h.strip())
        if not mm:
            continue
        mon_num = _MONTH_ABBREVS.get(mm.group(1)[:3].lower())
        if mon_num is None:
            continue
        try:
            d = date(year, mon_num, int(mm.group(2)))
        except ValueError:
            continue
        direction = "in" if "in" in mm.group(3).lower() else "out"
        date_to_cols.setdefault(d, {})[direction] = i

    if not date_to_cols:
        raise RosterReadError(f"No date columns parsed in Live sheet for {label}")

    rows: List[NeuronHoursRow] = []
    month_short = month_name[mon]
    for r in range(header_row + 1, live_ws.max_row + 1):
        staff_val = live_ws.cell(r, idx_staff + 1).value
        if not staff_val or str(staff_val).strip() in ("", "None", "0"):
            continue
        if isinstance(staff_val, (int, float)):
            continue
        staff = str(staff_val).strip()
        default_proj = str(live_ws.cell(r, idx_project + 1).value or "").strip()
        if default_proj == "0":
            default_proj = ""

        for d, dirs in sorted(date_to_cols.items()):
            in_val = live_ws.cell(r, dirs["in"] + 1).value if "in" in dirs else None
            out_val = live_ws.cell(r, dirs["out"] + 1).value if "out" in dirs else None
            ci, note_in = split_note_bearing_punch(in_val)
            co, note_out = split_note_bearing_punch(out_val)
            if ci is None and co is None:
                continue

            resolved = worked.get((staff, d), "") or default_proj
            if not _is_neuron(resolved):
                continue

            note = " ".join(n for n in (note_in, note_out) if n).strip()
            gross = _compute_gross(ci, co)
            rows.append(NeuronHoursRow(
                month=month_short,
                date=d,
                day=d.strftime("%a"),
                tech=staff,
                project="Neuron Deployments",
                clock_in=_format_clock(ci),
                clock_out=_format_clock(co),
                gross_hours=gross,
                weekend=d.weekday() >= 5,
                go_live_weekend=d in GO_LIVE_WEEKEND_DATES,
                live_sheet=live_ws.title,
                roster_row=r,
                note=note,
            ))

    return rows, warnings


def read_track_hours(
    roster_path: str | Path,
    months: List[str],
    pinned_techs: Optional[List[str]] = None,
) -> Tuple[List[NeuronHoursRow], List[str]]:
    try:
        import openpyxl
    except ImportError as e:
        raise RosterReadError("openpyxl is required: pip install openpyxl") from e
    p = Path(roster_path)
    if not p.exists():
        raise RosterReadError(f"Roster file not found: {roster_path}")
    wb = openpyxl.load_workbook(str(p), data_only=True, read_only=False)
    all_rows: List[NeuronHoursRow] = []
    warnings: List[str] = []
    for mk in months:
        rows, warn = read_month(wb, mk, pinned_techs=pinned_techs)
        all_rows.extend(rows)
        warnings.extend(warn)
    wb.close()
    return all_rows, warnings
