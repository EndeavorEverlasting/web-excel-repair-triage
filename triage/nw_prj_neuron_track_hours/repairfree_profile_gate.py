"""Structural gate for repair-free Bonita-format Neuron Track Hours workbooks."""
from __future__ import annotations

import json
import zipfile
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, List, Optional

_FORMULA_ERRORS = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#NULL!", "#NUM!", "#N/A")
_DEFAULT_PROFILE = (
    Path(__file__).resolve().parent.parent.parent
    / "configs"
    / "artifact_profiles"
    / "neuron_track_hours_repairfree_golden.json"
)


def load_repairfree_profile(path: Optional[str | Path] = None) -> Dict[str, Any]:
    p = Path(path) if path else _DEFAULT_PROFILE
    return json.loads(p.read_text(encoding="utf-8"))


def default_profile_path() -> Path:
    return _DEFAULT_PROFILE


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_numeric_like(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float, datetime, date, time)):
        return True
    if isinstance(value, str):
        try:
            float(value)
            return True
        except ValueError:
            return False
    return False


def run_repairfree_profile_gate(
    workbook_path: str | Path,
    profile_path: Optional[str | Path] = None,
) -> Dict[str, Any]:
    profile = load_repairfree_profile(profile_path)
    path = Path(workbook_path)
    failures: List[str] = []

    if not path.is_file():
        return {
            "profile": profile.get("profile"),
            "profile_pass": False,
            "profile_failures": ["file_not_found"],
        }

    pkg = profile.get("package_stop_ship") or {}
    try:
        with zipfile.ZipFile(path, "r") as z:
            if z.testzip() is not None:
                failures.append("zip_testzip_failed")
            names = z.namelist()
            all_text = ""
            for name in names:
                if name.endswith((".xml", ".rels")):
                    all_text += z.read(name).decode("utf-8", errors="ignore")
            if pkg.get("no_calc_chain") and "xl/calcChain.xml" in names:
                failures.append("calc_chain_present")
            if pkg.get("no_external_links") and any("externalLink" in n for n in names):
                failures.append("external_links_present")
            if pkg.get("no_vba") and "xl/vbaProject.bin" in names:
                failures.append("vba_present")
            if pkg.get("no_formula_error_tokens"):
                for tok in _FORMULA_ERRORS:
                    if tok in all_text:
                        failures.append(f"formula_error_token:{tok}")
            if pkg.get("no_inline_str"):
                for name in names:
                    if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                        ws_text = z.read(name).decode("utf-8", errors="ignore")
                        if 't="inlineStr"' in ws_text:
                            failures.append("inlineStr_present")
                            break
    except zipfile.BadZipFile:
        return {
            "profile": profile.get("profile"),
            "profile_pass": False,
            "profile_failures": failures + ["bad_zip"],
        }

    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        for sheet in profile.get("required_sheets") or []:
            if sheet not in wb.sheetnames:
                failures.append(f"missing_sheet:{sheet}")

        month_tabs = profile.get("month_sheet_names") or []
        h1 = profile.get("header_row1") or []
        h2 = profile.get("header_row2") or []
        allowed_asn = set(profile.get("allowed_assignment_types") or [])
        freeze = profile.get("freeze_panes") or "A3"

        for tab in month_tabs:
            if tab not in wb.sheetnames:
                continue
            ws = wb[tab]
            if ws.freeze_panes != freeze:
                failures.append(f"{tab}:freeze_panes_not_{freeze}")

            for col_idx, (top, bottom) in enumerate(zip(h1, h2), 1):
                if _cell_str(ws.cell(row=1, column=col_idx).value) != _cell_str(top):
                    failures.append(f"{tab}:header_row1_col{col_idx}_mismatch")
                if _cell_str(ws.cell(row=2, column=col_idx).value) != _cell_str(bottom):
                    failures.append(f"{tab}:header_row2_col{col_idx}_mismatch")

            if ws.max_column > 7:
                failures.append(f"{tab}:columns_beyond_G")

            if pkg.get("no_merged_cells_in_month_sheets") and ws.merged_cells.ranges:
                failures.append(f"{tab}:merged_cells_present")

            if pkg.get("no_conditional_formatting_in_month_sheets") and ws.conditional_formatting:
                failures.append(f"{tab}:conditional_formatting_present")

            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
                vals = [c.value for c in row]
                if all(v is None or str(v).strip() == "" for v in vals):
                    continue
                if pkg.get("no_formulas_in_month_sheets"):
                    for c in row:
                        if getattr(c, "data_type", None) == "f":
                            failures.append(f"{tab}:formula_at_{c.coordinate}")
                            break
                tech, start, end, hours, project, asn = vals[1:7]
                if not _cell_str(tech):
                    failures.append(f"{tab}:blank_tech_row")
                if start is None or (isinstance(start, str) and not start.strip()):
                    failures.append(f"{tab}:blank_start")
                if end is None or (isinstance(end, str) and not end.strip()):
                    failures.append(f"{tab}:blank_end")
                if hours is None:
                    failures.append(f"{tab}:blank_hours")
                elif not _is_numeric_like(hours):
                    failures.append(f"{tab}:non_numeric_hours")
                if not _cell_str(project):
                    failures.append(f"{tab}:blank_project")
                asn_s = _cell_str(asn)
                if not asn_s:
                    failures.append(f"{tab}:blank_assignment")
                elif allowed_asn and asn_s not in allowed_asn:
                    failures.append(f"{tab}:assignment_not_allowed:{asn_s}")
    finally:
        wb.close()

    return {
        "profile": profile.get("profile"),
        "profile_pass": len(failures) == 0,
        "profile_failures": failures,
    }
