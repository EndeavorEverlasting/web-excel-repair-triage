"""
NW PRJ Neuron Track Hours Exporter — produce the monthly track hours workbook.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from triage.nw_prj_classifier import ClassificationResult
from triage.xlsx_utils import fix_inlinestr

def _require_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from e
    return Workbook, Font, PatternFill, Alignment, get_column_letter

def export_neuron_track_hours(
    april_results: List[ClassificationResult],
    may_results: List[ClassificationResult],
    out_path: str
) -> str:
    Workbook, Font, PatternFill, Alignment, get_column_letter = _require_openpyxl()
    
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)
    
    # 1. Summary Tab
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Neuron Track Hours - Summary"
    ws_sum["A1"].font = Font(bold=True, size=14)
    
    # 2. April 2026 Tab
    _write_month_tab(wb, "April 2026", april_results)
    
    # 3. May 2026 Tab
    _write_month_tab(wb, "May 2026", may_results)
    
    # 4. Go Live Weekend Support Tab
    ws_gl = wb.create_sheet("Go Live Weekend Support")
    _write_weekend_support(ws_gl, april_results + may_results)
    
    # 5. CF Dictionary Tab
    ws_cf = wb.create_sheet("CF Dictionary")
    ws_cf["A1"] = "Conditional Formatting Dictionary"
    
    # 6. WebExcel QC Tab
    ws_qc = wb.create_sheet("WebExcel QC")
    ws_qc["A1"] = "Web Excel Quality Control"

    # Save and repair inlineStr (openpyxl 3.1.x Web Excel stop-ship fix)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    fix_inlinestr(out_path)
    return out_path

def _write_month_tab(wb: Any, title: str, results: List[ClassificationResult]):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    ws = wb.create_sheet(title)
    
    # Header Row 4
    headers = ["Tech", "Date", "Hours", "Status", "Reason", "Action Needed", "Notes"]
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        
    # Data Rows
    for r_idx, res in enumerate(results, 5):
        ws.cell(row=r_idx, column=1, value=res.tech)
        ws.cell(row=r_idx, column=2, value=res.date)
        ws.cell(row=r_idx, column=3, value=res.resolved_hours)
        ws.cell(row=r_idx, column=4, value=res.status)
        ws.cell(row=r_idx, column=5, value=res.reason_code)
        ws.cell(row=r_idx, column=6, value=res.action_needed)
        ws.cell(row=r_idx, column=7, value=res.notes)
        
    # Filters and Panes
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, len(results)+4)}"

def _write_weekend_support(ws: Any, all_results: List[ClassificationResult]):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    ws["A1"] = "Go Live Weekend Support"
    ws["A1"].font = Font(bold=True, size=12)
    
    headers = ["Tech", "Date", "Hours", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h).font = Font(bold=True)
    
    # Filter for weekends or specific Go Live dates (e.g. May 2-3 2026?)
    # For now, placeholder based on common weekend check
    import datetime
    weekend_rows = []
    for res in all_results:
        try:
            # Assuming date is YYYY-MM-DD
            dt = datetime.datetime.strptime(res.date, "%Y-%m-%d")
            if dt.weekday() >= 5: # Saturday or Sunday
                weekend_rows.append(res)
        except:
            pass
            
    for r_idx, res in enumerate(weekend_rows, 5):
        ws.cell(row=r_idx, column=1, value=res.tech)
        ws.cell(row=r_idx, column=2, value=res.date)
        ws.cell(row=r_idx, column=3, value=res.resolved_hours)
        ws.cell(row=r_idx, column=4, value=res.notes)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, len(weekend_rows)+4)}"
