"""Clean-render workbook generator for One Marcus inventory recon."""
from __future__ import annotations

import re
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font

from triage.xlsx_utils import fix_inlinestr

from . import formula_relink as fr
from .config import (
    PART_NUMBERS_SHEET,
    PN_DATA_END,
    PIVOT_SHEET,
    ROLLUP_DATA_START,
    ROLLUP_HEADER_ROW,
    visual_font_color,
)
from .reader import PartNumbersSnapshot, read_integrated_workbook
from .visual_field import write_rollup_table


def _copy_part_numbers_sheet(source_path: str, source_tab: str, target_ws, output_tab: str) -> int:
    """Copy Part Numbers sheet content; rewrite tab refs in formulas. Returns last data row."""
    wb = openpyxl.load_workbook(source_path, data_only=False)
    try:
        ws = wb[source_tab]
        max_col = max(ws.max_column or 1, 29)
        max_row = min(ws.max_row or 1, PN_DATA_END)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                src = ws.cell(r, c)
                dst = target_ws.cell(r, c, src.value)
                if src.number_format and src.number_format != "General":
                    dst.number_format = src.number_format
                if src.data_type == "f" and isinstance(dst.value, str):
                    new_val, _patched, _loc, _ext = fr._rewrite_refs_in_text(
                        dst.value, output_tab, source_tab if source_tab != output_tab else None
                    )
                    dst.value = new_val
        return max_row
    finally:
        wb.close()


def _write_pivot_shell(ws, snapshot: PartNumbersSnapshot, pn_last_row: int) -> None:
    m, d, y = snapshot.inferred_date.split("-")
    ws["A1"] = f"1M Recon — {int(m)}/{int(d)}/{y}"
    ws["A2"] = "Executive inventory review. Update source rows on Part Numbers."
    ws["A3"] = "Visual bars show relative stock quantity in the executive rollup."
    ws["A10"] = "Leadership view: quantity rollup and operational posture."
    write_rollup_table(
        ws,
        snapshot.rollup_keys,
        pn_tab=PART_NUMBERS_SHEET,
        pn_last_row=pn_last_row,
    )
    visual_color = visual_font_color()
    last_row = ROLLUP_DATA_START + max(len(snapshot.rollup_keys), 1) - 1
    if snapshot.rollup_keys:
        last_row = ROLLUP_DATA_START + len(snapshot.rollup_keys) - 1
        for row in range(ROLLUP_DATA_START, last_row + 1):
            ws.cell(row, 3).font = Font(color=visual_color[2:] if visual_color.startswith("FF") else visual_color)


def build_workbook(
    snapshot: PartNumbersSnapshot,
    out_path: str,
    *,
    source_path: Optional[str] = None,
) -> str:
    """Generate a two-sheet recon workbook from a Part Numbers snapshot."""
    src = source_path or snapshot.source_path
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)
    pivot_ws = out_wb.create_sheet(PIVOT_SHEET, 0)
    pn_ws = out_wb.create_sheet(PART_NUMBERS_SHEET, 1)
    pn_last_row = _copy_part_numbers_sheet(src, snapshot.source_tab, pn_ws, PART_NUMBERS_SHEET)
    _write_pivot_shell(pivot_ws, snapshot, pn_last_row)
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(str(out))
    out_wb.close()
    fix_inlinestr(str(out))
    return str(out.resolve())


def load_snapshot(
    input_path: str,
    *,
    cli_date: str = "auto",
    part_number_tab: Optional[str] = None,
    strict: bool = False,
) -> PartNumbersSnapshot:
    return read_integrated_workbook(
        input_path,
        cli_date=cli_date,
        part_number_tab=part_number_tab,
        strict=strict,
    )
