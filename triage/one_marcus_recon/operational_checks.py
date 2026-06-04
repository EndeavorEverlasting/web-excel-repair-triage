"""Operational surface checks for generated One Marcus recon workbooks."""
from __future__ import annotations

import dataclasses
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List

import openpyxl

from .config import (
    EXPECTED_SHEETS,
    FORBIDDEN_WORKBOOK_TEXT,
    PART_NUMBERS_SHEET,
    PIVOT_SHEET,
    ROLLUP_DATA_START,
    ROLLUP_HEADER_ROW,
    load_inventory_visual_config,
)


@dataclass
class OperationalCheckResult:
    path: str
    operational_pass: bool = False
    failures: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict:
        return dataclasses.asdict(self)


def _scan_forbidden_text(path: str) -> List[str]:
    hits: List[str] = []
    with zipfile.ZipFile(path, "r") as z:
        for name in z.namelist():
            if not name.endswith(".xml"):
                continue
            text = z.read(name).decode("utf-8", errors="ignore").lower()
            for token in FORBIDDEN_WORKBOOK_TEXT:
                if token in text:
                    hits.append(f"{token} in {name}")
    return hits


def run_operational_checks(path: str, *, min_visual_rows: int = 1) -> OperationalCheckResult:
    """Validate executive operational surface per success checkpoint + visual config."""
    res = OperationalCheckResult(path=str(Path(path).resolve()))
    cfg = load_inventory_visual_config()
    visual_cfg = cfg.get("executive_visual_field", {})
    expected_visual_header = str(visual_cfg.get("header", "Visual"))

    if not Path(path).is_file():
        res.failures.append("workbook_missing")
        return res

    forbidden = _scan_forbidden_text(path)
    res.failures.extend(forbidden)

    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        for sheet in EXPECTED_SHEETS:
            if sheet not in wb.sheetnames:
                res.failures.append(f"missing_sheet:{sheet}")

        if PIVOT_SHEET not in wb.sheetnames or PART_NUMBERS_SHEET not in wb.sheetnames:
            res.operational_pass = False
            return res

        pivot = wb[PIVOT_SHEET]
        headers = [pivot.cell(ROLLUP_HEADER_ROW, c).value for c in range(1, 8)]
        header_text = [str(h or "").strip() for h in headers]
        if "Total Qty" not in header_text:
            res.failures.append("missing_header:Total Qty")
        if expected_visual_header not in header_text:
            res.failures.append(f"missing_header:{expected_visual_header}")
            visual_idx = -1
        else:
            visual_idx = header_text.index(expected_visual_header)
            qty_idx = header_text.index("Total Qty") if "Total Qty" in header_text else -1
            if qty_idx >= 0 and visual_idx != qty_idx + 1:
                res.failures.append("visual_not_after_total_qty")

        if visual_idx >= 0:
            visual_rows = 0
            for row in range(ROLLUP_DATA_START, ROLLUP_DATA_START + 200):
                key = pivot.cell(row, 1).value
                if key is None or str(key).strip() == "":
                    break
                visual_cell = pivot.cell(row, visual_idx + 1)
                formula = visual_cell.value
                if isinstance(formula, str) and "REPT" in formula.upper():
                    visual_rows += 1
                elif formula not in (None, ""):
                    visual_rows += 1
            if visual_rows < min_visual_rows:
                res.failures.append(f"visual_rows_below_minimum:{visual_rows}<{min_visual_rows}")
    finally:
        wb.close()

    res.operational_pass = not res.failures
    return res
