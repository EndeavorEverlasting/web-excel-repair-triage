"""Read Part Numbers data from an integrated recon workbook."""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

from . import date_inference as di
from . import formula_relink as fr
from .config import (
    COL_INCLUDE,
    COL_PIVOT_KEY,
    COL_QTY_NUM,
    PART_NUMBERS_SHEET,
    PN_DATA_END,
    PN_DATA_START,
    PN_HEADER_ROW,
)


@dataclass
class PartNumbersSnapshot:
    source_path: str
    source_tab: str
    output_tab: str = PART_NUMBERS_SHEET
    inferred_date: str = ""
    date_source: str = ""
    headers: List[Any] = field(default_factory=list)
    rows: List[Tuple[Any, ...]] = field(default_factory=list)
    rollup_keys: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


def _cell_value(cell) -> Any:
    val = cell.value
    if isinstance(val, str):
        return val.strip()
    return val


def _derive_key(row_values: dict[int, Any]) -> str:
    for col in (COL_PIVOT_KEY, 4, 6, 2):  # S, D part, F desc, B type
        val = row_values.get(col)
        if val is not None and str(val).strip():
            return str(val).strip()
    return ""


def _is_included(row_values: dict[int, Any]) -> bool:
    flag = row_values.get(COL_INCLUDE)
    if flag is None:
        return True
    return str(flag).strip().lower() == "include"


def read_integrated_workbook(
    path: str,
    *,
    cli_date: str = "auto",
    part_number_tab: Optional[str] = None,
    strict: bool = False,
) -> PartNumbersSnapshot:
    """Extract Part Numbers rows and rollup keys from an integrated workbook."""
    p = Path(path)
    wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        chosen, _candidates, warnings = di.infer_update_date(
            str(p), cli_date, sheet_names, strict=strict
        )
        tab = fr.choose_source_tab(
            sheet_names,
            explicit_tab=part_number_tab,
            chosen_date_iso=chosen.date_iso,
            target_label=chosen.tab_label,
        )
        if not tab:
            raise ValueError("no Part Numbers candidate tab found in workbook")
        ws = wb[tab]
        max_col = max(ws.max_column or 1, 29)
        headers = [
            _cell_value(ws.cell(PN_HEADER_ROW, c))
            for c in range(1, max_col + 1)
        ]
        rows: List[Tuple[Any, ...]] = []
        rollup_keys: List[str] = []
        seen_keys: set[str] = set()
        for r in range(PN_DATA_START, min((ws.max_row or PN_DATA_START) + 1, PN_DATA_END + 1)):
            row_map = {
                c: _cell_value(ws.cell(r, c))
                for c in range(1, max_col + 1)
            }
            if not any(v not in (None, "") for v in row_map.values()):
                continue
            row_tuple = tuple(row_map.get(c) for c in range(1, max_col + 1))
            rows.append(row_tuple)
            if _is_included(row_map):
                key = _derive_key(row_map)
                if key and key not in seen_keys:
                    seen_keys.add(key)
                    rollup_keys.append(key)
        return PartNumbersSnapshot(
            source_path=str(p.resolve()),
            source_tab=tab,
            inferred_date=chosen.date_iso,
            date_source=chosen.source,
            headers=headers,
            rows=rows,
            rollup_keys=rollup_keys,
            warnings=list(warnings),
        )
    finally:
        wb.close()


def copy_row_formulas_from_source(
    source_path: str,
    source_tab: str,
    target_ws,
    *,
    target_tab_name: str,
) -> None:
    """Copy Part Numbers cell values/formulas from source into target worksheet."""
    wb = openpyxl.load_workbook(source_path, data_only=False, read_only=False)
    try:
        ws = wb[source_tab]
        max_col = max(ws.max_column or 1, 29)
        max_row = min(ws.max_row or 1, PN_DATA_END)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                src = ws.cell(r, c)
                dst = target_ws.cell(r, c)
                dst.value = src.value
                if src.has_style:
                    dst._style = src._style
        # Rewrite dated tab references in formulas to stable output tab name.
        if source_tab != target_tab_name:
            for row in target_ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    if cell.data_type == "f" and isinstance(cell.value, str):
                        cell.value = cell.value.replace(f"'{source_tab}'!", f"'{target_tab_name}'!")
    finally:
        wb.close()
