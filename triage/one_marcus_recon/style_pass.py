"""Formula-safe styling pass for generated One Marcus workbooks."""
from __future__ import annotations

import hashlib
import json
import zipfile
from pathlib import Path
from typing import Dict, Tuple

import openpyxl

from triage.spreadsheet_style import apply_openpyxl_tab_colors, openpyxl_style_primitives

from .config import PART_NUMBERS_SHEET, PIVOT_SHEET, ROLLUP_HEADER_ROW


def _formula_fingerprint(path: str) -> str:
    """Hash of all formula strings in workbook (stable compare aid)."""
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    formulas = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f" and cell.value is not None:
                        formulas.append(f"{name}!{cell.coordinate}={cell.value}")
    finally:
        wb.close()
    payload = json.dumps(sorted(formulas), separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def apply_style_pass(path: str) -> Tuple[str, str]:
    """Apply tab colors and executive header styling. Returns before/after formula hashes."""
    before = _formula_fingerprint(path)
    wb = openpyxl.load_workbook(path)
    apply_openpyxl_tab_colors(wb)
    styles = openpyxl_style_primitives()
    if PIVOT_SHEET in wb.sheetnames:
        pivot = wb[PIVOT_SHEET]
        for col in range(1, 9):
            cell = pivot.cell(ROLLUP_HEADER_ROW, col)
            cell.fill = styles["table_header_fill"]
            cell.font = styles["header_font"]
        pivot["A1"].font = styles["title_font"]
    if PART_NUMBERS_SHEET in wb.sheetnames:
        pn = wb[PART_NUMBERS_SHEET]
        for col in range(1, pn.max_column + 1):
            cell = pn.cell(1, col)
            if cell.value:
                cell.fill = styles["table_header_fill"]
                cell.font = styles["header_font"]
    wb.save(path)
    wb.close()
    after = _formula_fingerprint(path)
    if before != after:
        raise RuntimeError("style pass altered workbook formulas")
    return before, after
