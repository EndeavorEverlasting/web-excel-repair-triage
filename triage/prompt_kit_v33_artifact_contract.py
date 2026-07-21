"""Validate the generated AI Harness Prompt Kit V33 workbook UX contract.

This module is read-only and independent from the finalizer implementation. The
shared JSON file is the product contract: accepted tab order/colors/protection,
required prompt tabs, and the repaired P02 plus P45-P49 declarative payloads.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import range_boundaries

DEFAULT_SPEC_PATH = (
    Path(__file__).resolve().parents[1]
    / "configs"
    / "prompt_kit"
    / "v33_gnhf_harness_prompts.json"
)
LIBRARY_SHEET = "Prompt_Library"
LIBRARY_PROMPT_ID_COLUMN = 3
LIBRARY_SHEET_NAME_COLUMN = 15
LIBRARY_LAST_COLUMN = 16
PROMPT_ID_RE = re.compile(r"^P\d{2,}$")
PROMPT_SHEET_RE = re.compile(r"^(P\d{2,})_COPY_SAFE$")
PROMPT_RANGE_RE = re.compile(r"^A1:A([1-9]\d*)$")
HYPERLINK_FORMULA_RE = re.compile(r'^=HYPERLINK\("((?:[^"]|"")*)"\s*,', re.IGNORECASE)


@dataclass(frozen=True)
class ArtifactContractResult:
    workbook: str
    sha256: str
    prompt_count: int
    prompt_ids: tuple[str, ...]
    findings: tuple[str, ...]

    @property
    def passed(self) -> bool:
        return not self.findings

    def to_dict(self) -> dict:
        return {
            "workbook": self.workbook,
            "sha256": self.sha256,
            "status": "PASS" if self.passed else "FAIL",
            "prompt_count": self.prompt_count,
            "prompt_ids": list(self.prompt_ids),
            "findings": list(self.findings),
            "proof_ceiling": (
                "OOXML workbook structure, exact range links, accepted sheet order, "
                "theme/RGB tab colors, prompt presence, and worksheet protection. Excel "
                "Desktop/Web click-selection behavior, clipboard behavior, and operator "
                "acceptance remain runtime gates."
            ),
        }


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _load_spec(path: Path) -> dict:
    data = json.loads(path.read_text(encoding="utf-8"))
    if data.get("schema_version") != 1:
        raise ValueError("unsupported prompt extension schema_version")
    return data


def _target(cell: Cell) -> str:
    if cell.hyperlink:
        return cell.hyperlink.target
    value = cell.value
    if isinstance(value, str):
        match = HYPERLINK_FORMULA_RE.match(value)
        if match:
            return match.group(1).replace('""', '"')
    return ""


def _find_footer_row(library) -> int:
    for row in range(2, library.max_row + 1):
        text = " | ".join(
            str(library.cell(row=row, column=column).value)
            for column in range(1, LIBRARY_LAST_COLUMN + 1)
            if library.cell(row=row, column=column).value is not None
        )
        if "End of Prompt Library" in text or "↑ Top" in text:
            return row
    return library.max_row + 1


def _prompt_rows(library, footer_row: int) -> dict[str, tuple[int, str]]:
    prompts: dict[str, tuple[int, str]] = {}
    for row in range(2, footer_row):
        prompt_value = library.cell(row=row, column=LIBRARY_PROMPT_ID_COLUMN).value
        prompt_id = prompt_value if isinstance(prompt_value, str) and PROMPT_ID_RE.fullmatch(prompt_value) else ""
        if not prompt_id:
            sequence = library.cell(row=row, column=2).value
            if isinstance(sequence, int) or (isinstance(sequence, str) and sequence.isdigit()):
                prompt_id = f"P{int(sequence):02d}"
        sheet_name = library.cell(row=row, column=LIBRARY_SHEET_NAME_COLUMN).value
        if not prompt_id:
            continue
        if prompt_id in prompts:
            raise ValueError(f"duplicate Prompt Library ID: {prompt_id}")
        prompts[prompt_id] = (row, str(sheet_name or ""))
    return prompts


def _expected_prompt_ids(sheet_order: Sequence[str]) -> tuple[str, ...]:
    result = []
    for sheet_name in sheet_order:
        match = PROMPT_SHEET_RE.fullmatch(sheet_name)
        if match:
            result.append(match.group(1))
    return tuple(result)


def _color_matches(color, expected: Mapping[str, object]) -> bool:
    if color is None:
        return False
    if "rgb" in expected:
        return color.type == "rgb" and str(color.rgb).upper()[-6:] == str(expected["rgb"]).upper()[-6:]
    if color.type != "theme" or int(color.theme) != int(expected["theme"]):
        return False
    return abs(float(color.tint or 0.0) - float(expected.get("tint", 0.0))) < 1e-12


def _validate_protection(
    workbook,
    editable_ranges: Mapping[str, str],
    findings: list[str],
) -> None:
    if workbook.security is None or not workbook.security.lockStructure:
        findings.append("workbook structure is not locked")
    for ws in workbook.worksheets:
        if not ws.protection.sheet:
            findings.append(f"worksheet is not protected: {ws.title}")
            continue
        editable = editable_ranges.get(ws.title)
        bounds = range_boundaries(editable) if editable else None
        sheet_failed = False
        for row in ws.iter_rows():
            for cell in row:
                allowed = bool(
                    bounds
                    and bounds[0] <= cell.column <= bounds[2]
                    and bounds[1] <= cell.row <= bounds[3]
                )
                if allowed and cell.protection.locked:
                    findings.append(f"editable cell remained locked: {ws.title}!{cell.coordinate}")
                    sheet_failed = True
                    break
                if not allowed and not cell.protection.locked:
                    findings.append(f"unexpected editable cell: {ws.title}!{cell.coordinate}")
                    sheet_failed = True
                    break
            if sheet_failed:
                break


def validate_artifact(path: Path, spec_path: Path = DEFAULT_SPEC_PATH) -> ArtifactContractResult:
    if not path.exists():
        raise FileNotFoundError(path)
    spec = _load_spec(spec_path)
    workbook = load_workbook(path, keep_links=True, data_only=False)
    findings: list[str] = []
    expected_order = tuple(str(name) for name in spec["sheet_order"])
    if tuple(workbook.sheetnames) != expected_order:
        findings.append("worksheet order does not match the accepted V33 layout contract")
    if LIBRARY_SHEET not in workbook.sheetnames:
        return ArtifactContractResult(
            workbook=str(path),
            sha256=_sha256(path),
            prompt_count=0,
            prompt_ids=(),
            findings=tuple(findings + [f"missing required sheet: {LIBRARY_SHEET}"]),
        )

    library = workbook[LIBRARY_SHEET]
    footer_row = _find_footer_row(library)
    try:
        prompts = _prompt_rows(library, footer_row)
    except ValueError as exc:
        findings.append(str(exc))
        prompts = {}

    expected_corners = {
        "A1": f"#'{LIBRARY_SHEET}'!A{footer_row}",
        "P1": f"#'{LIBRARY_SHEET}'!P{footer_row}",
        f"A{footer_row}": f"#'{LIBRARY_SHEET}'!A1",
        f"P{footer_row}": f"#'{LIBRARY_SHEET}'!P1",
    }
    for coordinate, expected in expected_corners.items():
        actual = _target(library[coordinate])
        if actual != expected:
            findings.append(f"{LIBRARY_SHEET}!{coordinate} target {actual!r} != {expected!r}")

    required_prompt_ids = _expected_prompt_ids(expected_order)
    for required in required_prompt_ids:
        if required not in prompts:
            findings.append(f"missing required prompt: {required}")

    generated_prompt_ids = {str(prompt["prompt_id"]) for prompt in spec.get("prompts", [])}
    for prompt_id, (row, sheet_name) in sorted(prompts.items()):
        if not sheet_name:
            findings.append(f"{prompt_id} has no copy-safe sheet name")
            continue
        if sheet_name not in workbook.sheetnames:
            findings.append(f"{prompt_id} references missing sheet: {sheet_name}")
            continue
        forward = _target(library.cell(row=row, column=LIBRARY_PROMPT_ID_COLUMN))
        prefix = f"#'{sheet_name}'!"
        if not forward.startswith(prefix):
            findings.append(f"{prompt_id} forward link {forward!r} does not target {sheet_name}")
            continue
        prompt_range = forward[len(prefix):]
        range_match = PROMPT_RANGE_RE.fullmatch(prompt_range)
        if not range_match:
            findings.append(f"{prompt_id} forward link must select full A1:A<n> range: {forward!r}")
            continue
        last_row = int(range_match.group(1))
        ws = workbook[sheet_name]
        if not any(ws.cell(row=index, column=1).value not in (None, "") for index in range(1, last_row + 1)):
            findings.append(f"{sheet_name} prompt payload range is blank: {prompt_range}")
        expected_back = f"#'{LIBRARY_SHEET}'!A{row}:P{row}"
        for coordinate in ("B1", "E1", f"B{last_row}", f"E{last_row}"):
            actual = _target(ws[coordinate])
            if actual != expected_back:
                findings.append(f"{sheet_name}!{coordinate} target {actual!r} != {expected_back!r}")
        expected_self = f"#'{sheet_name}'!{prompt_range}"
        for coordinate in ("C1", f"C{last_row}"):
            actual = _target(ws[coordinate])
            if actual != expected_self:
                findings.append(f"{sheet_name}!{coordinate} range-recovery target {actual!r} != {expected_self!r}")
            if ws[coordinate].value != f"Copy {prompt_range} only":
                findings.append(f"{sheet_name}!{coordinate} range label is not canonical")
        if prompt_id in generated_prompt_ids:
            row_fills = {
                (
                    library.cell(row=row, column=column).fill.fill_type,
                    str(library.cell(row=row, column=column).fill.fgColor.rgb or "").upper()[-6:],
                )
                for column in range(1, 17)
            }
            if len(row_fills) != 1 or next(iter(row_fills))[0] != "solid":
                findings.append(f"{prompt_id} Prompt Library row does not use one coordinated solid fill")
            prompt_cell = library.cell(row=row, column=LIBRARY_PROMPT_ID_COLUMN)
            if not prompt_cell.font.bold or prompt_cell.font.underline != "single":
                findings.append(f"{prompt_id} Prompt Library link is not bold and underlined")
        if (ws.column_dimensions["A"].width or 0) < 60:
            findings.append(f"{sheet_name} copy-safe column A is too narrow")

    expected_colors = {str(name): dict(value) for name, value in spec["tab_colors"].items()}
    for ws in workbook.worksheets:
        expected = expected_colors.get(ws.title)
        actual = ws.sheet_properties.tabColor
        if expected is None and actual is not None:
            findings.append(f"unexpected tab color: {ws.title}")
        elif expected is not None and not _color_matches(actual, expected):
            findings.append(f"{ws.title} tab color does not match the accepted contract")

    _validate_protection(
        workbook,
        {str(name): str(cell_range) for name, cell_range in spec["editable_ranges"].items()},
        findings,
    )
    prompt_ids = tuple(sorted(prompts))
    return ArtifactContractResult(
        workbook=str(path),
        sha256=_sha256(path),
        prompt_count=len(prompt_ids),
        prompt_ids=prompt_ids,
        findings=tuple(findings),
    )


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Validate AI Harness Prompt Kit V33 range recovery, order, colors, prompts, and protection."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--spec", type=Path, default=DEFAULT_SPEC_PATH)
    parser.add_argument("--report", type=Path)
    args = parser.parse_args(argv)
    result = validate_artifact(args.workbook, args.spec)
    payload = result.to_dict()
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(payload, indent=2))
    return 0 if result.passed else 1


if __name__ == "__main__":
    raise SystemExit(main())
