"""Validate exact V33 copy ranges while allowing deliberate navigation rails."""
from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook

PROMPT_ID_RE = re.compile(r"^P\d{2}$")
RANGE_RE = re.compile(r"^#?'([^']+)'!A1:A([1-9]\d*)$")
HYPERLINK_FORMULA_RE = re.compile(r'^=HYPERLINK\("((?:[^"]|"")*)"\s*,', re.IGNORECASE)


@dataclass(frozen=True)
class V33CopySurfaceResult:
    workbook: str
    prompt_count: int
    findings: tuple[str, ...]

    @property
    def passed(self) -> bool:
        return not self.findings

    def to_dict(self) -> dict:
        return {
            "workbook": self.workbook,
            "status": "PASS" if self.passed else "FAIL",
            "prompt_count": self.prompt_count,
            "findings": list(self.findings),
            "navigation_policy": "B/C/E navigation cells are outside the linked column-A copy range",
        }


def _target(cell) -> str:
    if cell.hyperlink:
        return cell.hyperlink.target
    if isinstance(cell.value, str):
        match = HYPERLINK_FORMULA_RE.match(cell.value)
        if match:
            return match.group(1).replace('""', '"')
    return ""


def _prompt_id(library, row: int) -> str:
    value = library.cell(row=row, column=3).value
    if isinstance(value, str) and PROMPT_ID_RE.fullmatch(value):
        return value
    sequence = library.cell(row=row, column=2).value
    if isinstance(sequence, int) or (isinstance(sequence, str) and sequence.isdigit()):
        return f"P{int(sequence):02d}"
    return ""


def validate_v33_copy_surfaces(path: Path) -> V33CopySurfaceResult:
    workbook = load_workbook(path, data_only=False, keep_links=True)
    findings: list[str] = []
    if "Prompt_Library" not in workbook.sheetnames:
        return V33CopySurfaceResult(str(path), 0, ("missing Prompt_Library",))
    library = workbook["Prompt_Library"]
    prompt_count = 0
    for row in range(2, library.max_row + 1):
        prompt_id = _prompt_id(library, row)
        if not prompt_id:
            continue
        prompt_count += 1
        sheet_name = str(library.cell(row=row, column=15).value or "")
        target = _target(library.cell(row=row, column=3))
        match = RANGE_RE.fullmatch(target)
        if not match or match.group(1) != sheet_name:
            findings.append(f"{prompt_id} Prompt Library link is not an exact column-A range: {target!r}")
            continue
        last_row = int(match.group(2))
        if sheet_name not in workbook.sheetnames:
            findings.append(f"{prompt_id} copy sheet is missing: {sheet_name}")
            continue
        sheet = workbook[sheet_name]
        if sheet["A1"].value in (None, "") or sheet.cell(row=last_row, column=1).value in (None, ""):
            findings.append(f"{prompt_id} copy range endpoints must be non-empty: A1:A{last_row}")
        trailing = [
            cell.row
            for cell in sheet["A"]
            if cell.row > last_row and cell.value not in (None, "")
        ]
        if trailing:
            findings.append(f"{prompt_id} has column-A payload after A{last_row}: {trailing}")
        self_target = f"#'{sheet_name}'!A1:A{last_row}"
        for coordinate in ("C1", f"C{last_row}"):
            if _target(sheet[coordinate]) != self_target:
                findings.append(f"{prompt_id} {coordinate} recovery link is not {self_target}")
        library_target = f"#'Prompt_Library'!A{row}:P{row}"
        for coordinate in ("B1", "E1", f"B{last_row}", f"E{last_row}"):
            if _target(sheet[coordinate]) != library_target:
                findings.append(f"{prompt_id} {coordinate} backlink is not {library_target}")
    expected_ids = {f"P{number:02d}" for number in range(50)}
    actual_ids = {
        _prompt_id(library, row)
        for row in range(2, library.max_row + 1)
        if _prompt_id(library, row)
    }
    if actual_ids != expected_ids:
        findings.append("Prompt Library copy surfaces must cover P00 through P49 exactly")
    return V33CopySurfaceResult(str(path), prompt_count, tuple(findings))


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args(argv)
    result = validate_v33_copy_surfaces(args.workbook)
    print(json.dumps(result.to_dict(), indent=2))
    return 0 if result.passed else 1


if __name__ == "__main__":
    raise SystemExit(main())
