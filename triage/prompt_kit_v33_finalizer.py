"""Finalize a prompt-kit workbook with V33+ navigation, prompts, styling, and protection.

The finalizer intentionally operates on a source workbook instead of inventing the
prompt kit from a blank file. It adds or updates declarative prompt definitions,
repairs internal range hyperlinks, coordinates Prompt Library row styles, assigns
selected cream tab colors, and protects every worksheet while leaving only the
configured operator-entry ranges unlocked.
"""
from __future__ import annotations

import argparse
import json
import shutil
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Mapping, MutableMapping, Sequence

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import range_boundaries
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.worksheet import Worksheet

from triage.prompt_kit_v33_prompt_contract import validate_prompt_contract_data

DEFAULT_SPEC_PATH = (
    Path(__file__).resolve().parents[1]
    / "configs"
    / "prompt_kit"
    / "v33_gnhf_harness_prompts.json"
)

LIBRARY_SHEET = "Prompt_Library"
LIBRARY_COLUMNS: Mapping[str, int] = {
    "seq": 2,
    "prompt_id": 3,
    "prompt_type": 4,
    "prompt_class": 5,
    "sprint_path_role": 6,
    "use_for_progress": 7,
    "prompt_name": 8,
    "use_this_when": 9,
    "inspect_first": 10,
    "expected_output": 11,
    "next_step": 12,
    "proof_gate": 13,
    "color": 14,
    "sheet_name": 15,
}

PALETTE: Mapping[str, tuple[str, str]] = {
    "Slate": ("F1F5F9", "334155"),
    "Gray": ("E5E7EB", "374151"),
    "Sky": ("E0F2FE", "075985"),
    "Amber": ("FEF3C7", "92400E"),
    "Blue": ("DBEAFE", "1D4ED8"),
    "Green": ("DCFCE7", "166534"),
    "Rose": ("FFE4E6", "9F1239"),
    "Purple": ("F3E8FF", "6B21A8"),
    "Peach": ("FFEDD5", "9A3412"),
    "Teal": ("CCFBF1", "0F766E"),
    "Lavender": ("EDE9FE", "5B21B6"),
    "Cyan": ("CFFAFE", "0E7490"),
    "Indigo": ("E0E7FF", "3730A3"),
    "Blue-Green": ("CCFBF1", "0F766E"),
    "Gold": ("FEF3C7", "92400E"),
    "Sand": ("FDE68A", "854D0E"),
    "Cream": ("FFF2CC", "7C5C00"),
    "Orange": ("FED7AA", "9A3412"),
    "Emerald": ("D1FAE5", "047857"),
    "Coral": ("FFE4E6", "BE123C"),
    "Ocean": ("DBEAFE", "1D4ED8"),
    "Mint": ("D1FAE5", "047857"),
    "Night": ("E0E7FF", "3730A3"),
    "Violet": ("F3E8FF", "6B21A8"),
}


@dataclass(frozen=True)
class FinalizeResult:
    output: str
    prompt_ids: tuple[str, ...]
    prompt_ranges: Mapping[str, str]
    library_rows: Mapping[str, int]
    cream_tabs: tuple[str, ...]
    editable_ranges: Mapping[str, str]
    validation_passed: bool = True

    def to_dict(self) -> dict:
        return {
            "output": self.output,
            "prompt_ids": list(self.prompt_ids),
            "prompt_ranges": dict(self.prompt_ranges),
            "library_rows": dict(self.library_rows),
            "cream_tabs": list(self.cream_tabs),
            "editable_ranges": dict(self.editable_ranges),
            "validation_passed": self.validation_passed,
            "proof_ceiling": (
                "generated workbook structure, internal hyperlinks, styles, tab colors, "
                "prompt payloads, and protection state; Excel Desktop and Excel for Web "
                "open, clipboard, and operator acceptance remain runtime gates"
            ),
        }


def load_spec(path: Path = DEFAULT_SPEC_PATH) -> dict:
    data = json.loads(path.read_text(encoding="utf-8"))
    if data.get("schema_version") != 1:
        raise ValueError("unsupported prompt extension schema_version")
    contract = validate_prompt_contract_data(data, spec=str(path))
    if not contract.passed:
        raise ValueError("prompt source contract failed: " + "; ".join(contract.findings))
    prompts = data.get("prompts")
    if not isinstance(prompts, list) or not prompts:
        raise ValueError("prompt extension spec must contain prompts")
    seen: set[str] = set()
    for prompt in prompts:
        prompt_id = prompt.get("prompt_id", "")
        if not isinstance(prompt_id, str) or not prompt_id.startswith("P"):
            raise ValueError(f"invalid prompt_id: {prompt_id!r}")
        if prompt_id in seen:
            raise ValueError(f"duplicate prompt_id: {prompt_id}")
        seen.add(prompt_id)
        lines = prompt.get("lines")
        if not isinstance(lines, list) or not lines or not all(isinstance(line, str) for line in lines):
            raise ValueError(f"{prompt_id} lines must be a non-empty list of strings")
        if not prompt.get("sheet_name"):
            raise ValueError(f"{prompt_id} is missing sheet_name")
    return data


def _find_footer_row(ws: Worksheet) -> int:
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row=row, column=column).value for column in range(1, 17)]
        text = " | ".join(str(value) for value in values if value is not None)
        if "End of Prompt Library" in text or "↑ Top" in text:
            return row
    return ws.max_row + 1


def _prompt_row_map(ws: Worksheet) -> Dict[str, int]:
    result: Dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=LIBRARY_COLUMNS["prompt_id"]).value
        if isinstance(value, str) and value.startswith("P"):
            result[value] = row
    return result


def _copy_row_height_and_dimensions(ws: Worksheet, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height or 82
    for column in range(1, 17):
        source = ws.cell(row=source_row, column=column)
        target = ws.cell(row=target_row, column=column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def _write_library_row(ws: Worksheet, row: int, prompt: Mapping[str, object]) -> None:
    color_name = str(prompt["color"])
    fill_color, font_color = PALETTE.get(color_name, PALETTE["Slate"])
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 82)
    for key, column in LIBRARY_COLUMNS.items():
        ws.cell(row=row, column=column).value = prompt.get(key, "")
    for column in range(1, 17):
        cell = ws.cell(row=row, column=column)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        size = 10
        bold = column in {2, 3, 4, 5, 7, 8, 14, 15}
        if column == 3:
            size = 28
        elif column == 8:
            size = 12
        cell.font = Font(name="Aptos", size=size, bold=bold, color=font_color)
        cell.protection = Protection(locked=True)
    ws.cell(row=row, column=3).font = Font(
        name="Aptos", size=28, bold=True, underline="single", color=font_color
    )


def _set_internal_link(cell: Cell, target: str, label: str) -> None:
    cell.value = label
    cell.hyperlink = target
    cell.font = Font(name="Aptos", size=10, bold=True, underline="single", color="0563C1")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.protection = Protection(locked=True)


def _ensure_prompt_sheet(workbook, prompt: Mapping[str, object], library_row: int) -> str:
    prompt_id = str(prompt["prompt_id"])
    sheet_name = str(prompt["sheet_name"])
    lines = [str(line) for line in prompt["lines"]]
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
                cell.hyperlink = None
    else:
        ws = workbook.create_sheet(sheet_name)
    last_row = len(lines)
    for row, line in enumerate(lines, start=1):
        cell = ws.cell(row=row, column=1)
        cell.value = line
        cell.font = Font(name="Aptos", size=10, color="111827")
        cell.alignment = Alignment(vertical="top", wrap_text=False)
        cell.protection = Protection(locked=True)
    ws.column_dimensions["A"].width = 110
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 23
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 25
    ws.freeze_panes = "A2"

    library_target = f"#'{LIBRARY_SHEET}'!A{library_row}:P{library_row}"
    prompt_range = f"A1:A{last_row}"
    for coordinate, label in {
        "B1": f"← Prompt Library · {prompt_id}",
        "E1": f"{prompt_id} · Prompt Library →",
        f"B{last_row}": f"← Prompt Library · {prompt_id}",
        f"E{last_row}": f"{prompt_id} · Prompt Library →",
    }.items():
        _set_internal_link(ws[coordinate], library_target, label)
    for coordinate in ("C1", f"C{last_row}"):
        ws[coordinate] = f"Copy {prompt_range} only"
        ws[coordinate].font = Font(name="Aptos", size=10, bold=True, color="334155")
        ws[coordinate].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[coordinate].protection = Protection(locked=True)
    ws.protection.sheet = True
    ws.protection.enable()
    return prompt_range


def _apply_library_navigation(ws: Worksheet, footer_row: int) -> None:
    for coordinate, target, label in (
        ("A1", f"#'{LIBRARY_SHEET}'!A{footer_row}", "↓ Bottom"),
        ("P1", f"#'{LIBRARY_SHEET}'!P{footer_row}", "↓ Bottom"),
        (f"A{footer_row}", f"#'{LIBRARY_SHEET}'!A1", "↑ Top"),
        (f"P{footer_row}", f"#'{LIBRARY_SHEET}'!P1", "↑ Top"),
    ):
        _set_internal_link(ws[coordinate], target, label)
        ws[coordinate].fill = PatternFill("solid", fgColor="0F766E")
        ws[coordinate].font = Font(
            name="Aptos", size=10, bold=True, underline="single", color="FFFFFF"
        )


def _unlock_range(ws: Worksheet, cell_range: str) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.protection = Protection(locked=False)


def _apply_protection(workbook, editable_ranges: Mapping[str, str]) -> None:
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True)
        if ws.title in editable_ranges:
            _unlock_range(ws, editable_ranges[ws.title])
        ws.protection.sheet = True
        ws.protection.enable()
    if workbook.security is None:
        workbook.security = WorkbookProtection()
    workbook.security.lockStructure = True


def validate_finalized_workbook(path: Path, spec_path: Path = DEFAULT_SPEC_PATH) -> List[str]:
    """Return finalization contract violations; an empty list means PASS."""
    spec = load_spec(spec_path)
    workbook = load_workbook(path, keep_links=True)
    findings: List[str] = []
    if workbook.security is None or not workbook.security.lockStructure:
        findings.append("workbook structure is not locked")
    if LIBRARY_SHEET not in workbook.sheetnames:
        return [f"missing required sheet: {LIBRARY_SHEET}"]
    library = workbook[LIBRARY_SHEET]
    row_map = _prompt_row_map(library)
    footer_row = _find_footer_row(library)

    expected_corners = {
        "A1": f"#'{LIBRARY_SHEET}'!A{footer_row}",
        "P1": f"#'{LIBRARY_SHEET}'!P{footer_row}",
        f"A{footer_row}": f"#'{LIBRARY_SHEET}'!A1",
        f"P{footer_row}": f"#'{LIBRARY_SHEET}'!P1",
    }
    for coordinate, expected in expected_corners.items():
        actual = library[coordinate].hyperlink.target if library[coordinate].hyperlink else ""
        if actual != expected:
            findings.append(f"{LIBRARY_SHEET}!{coordinate} target {actual!r} != {expected!r}")

    for prompt in spec["prompts"]:
        prompt_id = str(prompt["prompt_id"])
        sheet_name = str(prompt["sheet_name"])
        lines = [str(line) for line in prompt["lines"]]
        if prompt_id not in row_map:
            findings.append(f"missing Prompt Library row: {prompt_id}")
            continue
        if sheet_name not in workbook.sheetnames:
            findings.append(f"missing prompt sheet: {sheet_name}")
            continue
        row = row_map[prompt_id]
        prompt_range = f"A1:A{len(lines)}"
        expected_forward = f"#'{sheet_name}'!{prompt_range}"
        prompt_cell = library.cell(row=row, column=3)
        actual_forward = prompt_cell.hyperlink.target if prompt_cell.hyperlink else ""
        if actual_forward != expected_forward:
            findings.append(f"{prompt_id} forward link {actual_forward!r} != {expected_forward!r}")
        ws = workbook[sheet_name]
        actual_lines = [ws.cell(row=index, column=1).value or "" for index in range(1, len(lines) + 1)]
        if actual_lines != lines:
            findings.append(f"{sheet_name} payload does not match declarative prompt spec")
        expected_back = f"#'{LIBRARY_SHEET}'!A{row}:P{row}"
        for coordinate in ("B1", "E1", f"B{len(lines)}", f"E{len(lines)}"):
            actual_back = ws[coordinate].hyperlink.target if ws[coordinate].hyperlink else ""
            if actual_back != expected_back:
                findings.append(f"{sheet_name}!{coordinate} target {actual_back!r} != {expected_back!r}")

    cream = str(spec.get("cream_tab_color", "FFF2CC")).upper()[-6:]
    for sheet_name in spec.get("cream_tabs", []):
        if sheet_name not in workbook.sheetnames:
            findings.append(f"missing cream-tab sheet: {sheet_name}")
            continue
        tab_color = workbook[sheet_name].sheet_properties.tabColor
        actual = (tab_color.rgb or "").upper()[-6:] if tab_color is not None else ""
        if actual != cream:
            findings.append(f"{sheet_name} tab color {actual!r} != {cream!r}")

    editable = dict(spec.get("editable_ranges", {}))
    for ws in workbook.worksheets:
        if not ws.protection.sheet:
            findings.append(f"worksheet is not protected: {ws.title}")
        allowed = editable.get(ws.title)
        bounds = range_boundaries(str(allowed)) if allowed else None
        for row in ws.iter_rows():
            for cell in row:
                is_editable = bool(
                    bounds
                    and bounds[0] <= cell.column <= bounds[2]
                    and bounds[1] <= cell.row <= bounds[3]
                )
                if is_editable and cell.protection.locked:
                    findings.append(f"editable cell remained locked: {ws.title}!{cell.coordinate}")
                    break
                if not is_editable and not cell.protection.locked:
                    findings.append(f"unexpected editable cell: {ws.title}!{cell.coordinate}")
                    break
            if findings and findings[-1].startswith(("editable cell", "unexpected editable")):
                break
    for sheet_name, cell_range in editable.items():
        if sheet_name not in workbook.sheetnames:
            findings.append(f"missing editable-range sheet: {sheet_name}")
            continue
        ws = workbook[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.protection.locked:
                    findings.append(f"editable cell remained locked: {sheet_name}!{cell.coordinate}")
                    break
            if findings and findings[-1].startswith("editable cell remained locked"):
                break
    return findings


def finalize_workbook(source: Path, output: Path, spec_path: Path = DEFAULT_SPEC_PATH) -> FinalizeResult:
    spec = load_spec(spec_path)
    if not source.exists():
        raise FileNotFoundError(source)
    output.parent.mkdir(parents=True, exist_ok=True)
    if source.resolve() == output.resolve():
        raise ValueError("output must not overwrite source workbook")
    shutil.copy2(source, output)
    workbook = load_workbook(output, keep_links=True)
    if LIBRARY_SHEET not in workbook.sheetnames:
        raise ValueError(f"missing required sheet: {LIBRARY_SHEET}")
    library = workbook[LIBRARY_SHEET]
    footer_row = _find_footer_row(library)
    row_map = _prompt_row_map(library)
    prompts: Sequence[Mapping[str, object]] = spec["prompts"]
    missing = [prompt for prompt in prompts if prompt["prompt_id"] not in row_map]
    if missing:
        library.insert_rows(footer_row, amount=len(missing))
        source_style_row = max(2, footer_row - 1)
        for offset, prompt in enumerate(missing):
            target_row = footer_row + offset
            _copy_row_height_and_dimensions(library, source_style_row, target_row)
            row_map[str(prompt["prompt_id"])] = target_row
        footer_row += len(missing)
    prompt_ranges: MutableMapping[str, str] = {}
    for prompt in prompts:
        prompt_id = str(prompt["prompt_id"])
        library_row = row_map[prompt_id]
        _write_library_row(library, library_row, prompt)
        prompt_range = _ensure_prompt_sheet(workbook, prompt, library_row)
        prompt_ranges[prompt_id] = prompt_range
        target_sheet = str(prompt["sheet_name"])
        target = f"#'{target_sheet}'!{prompt_range}"
        color_name = str(prompt["color"])
        _, font_color = PALETTE.get(color_name, PALETTE["Slate"])
        prompt_cell = library.cell(row=library_row, column=3)
        prompt_cell.value = prompt_id
        prompt_cell.hyperlink = target
        prompt_cell.font = Font(
            name="Aptos", size=28, bold=True, underline="single", color=font_color
        )
        sheet_cell = library.cell(row=library_row, column=15)
        sheet_cell.value = target_sheet
        sheet_cell.hyperlink = target
        sheet_cell.font = Font(
            name="Aptos", size=10, bold=True, underline="single", color=font_color
        )
    _apply_library_navigation(library, footer_row)
    cream_color = str(spec.get("cream_tab_color", "FFF2CC"))
    applied_cream_tabs: List[str] = []
    for sheet_name in spec.get("cream_tabs", []):
        if sheet_name in workbook.sheetnames:
            workbook[sheet_name].sheet_properties.tabColor = cream_color
            applied_cream_tabs.append(sheet_name)
    editable_ranges = {
        str(name): str(cell_range)
        for name, cell_range in dict(spec.get("editable_ranges", {})).items()
        if name in workbook.sheetnames
    }
    _apply_protection(workbook, editable_ranges)
    library.freeze_panes = "C2"
    workbook.save(output)
    findings = validate_finalized_workbook(output, spec_path)
    if findings:
        raise RuntimeError("finalized workbook contract failed: " + "; ".join(findings))
    return FinalizeResult(
        output=str(output),
        prompt_ids=tuple(str(prompt["prompt_id"]) for prompt in prompts),
        prompt_ranges=dict(prompt_ranges),
        library_rows={prompt_id: row_map[prompt_id] for prompt_id in prompt_ranges},
        cream_tabs=tuple(applied_cream_tabs),
        editable_ranges=editable_ranges,
    )


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Finalize a prompt-kit workbook with GNHF harness prompts and workbook UX contracts."
    )
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--spec", type=Path, default=DEFAULT_SPEC_PATH)
    parser.add_argument("--report", type=Path)
    args = parser.parse_args(argv)
    result = finalize_workbook(args.source, args.output, args.spec)
    payload = result.to_dict()
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(payload, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
