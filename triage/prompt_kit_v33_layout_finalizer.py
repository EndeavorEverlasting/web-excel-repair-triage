"""Canonicalize the accepted V33 prompt-kit workbook layout in place.

This is the layout stage after ``prompt_kit_v33_finalizer`` has repaired P02 and
added P45-P49.
It does not generate prompt content. It makes both range labels clickable,
enforces the accepted tab order and theme colors, and protects every sheet except
for the exact operator-edit range declared in the source contract.
"""
from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Color, Font, Protection
from openpyxl.utils import range_boundaries
from openpyxl.workbook.protection import WorkbookProtection

from triage.prompt_kit_v33_artifact_contract import validate_artifact

DEFAULT_SPEC_PATH = (
    Path(__file__).resolve().parents[1]
    / "configs"
    / "prompt_kit"
    / "v33_gnhf_harness_prompts.json"
)
LIBRARY_SHEET = "Prompt_Library"
PROMPT_ID_COLUMN = 3
SHEET_NAME_COLUMN = 15
PROMPT_RANGE_RE = re.compile(r"^A1:A([1-9]\d*)$")
PROMPT_ID_RE = re.compile(r"^P\d{2,}$")
HYPERLINK_FORMULA_RE = re.compile(r'^=HYPERLINK\("((?:[^"]|"")*)"\s*,', re.IGNORECASE)


@dataclass(frozen=True)
class LayoutResult:
    workbook: str
    prompt_ranges: Mapping[str, str]
    sheet_order: tuple[str, ...]
    colored_tabs: tuple[str, ...]
    editable_ranges: Mapping[str, str]

    def to_dict(self) -> dict:
        return {
            "workbook": self.workbook,
            "prompt_ranges": dict(self.prompt_ranges),
            "sheet_order": list(self.sheet_order),
            "colored_tabs": list(self.colored_tabs),
            "editable_ranges": dict(self.editable_ranges),
            "unprotected_sheets": [],
            "validation_passed": True,
            "proof_ceiling": (
                "exact internal range links, accepted sheet order, tab colors, and "
                "worksheet protection state; Excel Desktop/Web click-selection and "
                "operator acceptance remain runtime gates"
            ),
        }


def _load_spec(path: Path) -> dict:
    data = json.loads(path.read_text(encoding="utf-8"))
    if data.get("schema_version") != 1:
        raise ValueError("unsupported prompt extension schema_version")
    order = data.get("sheet_order")
    if not isinstance(order, list) or not order or len(order) != len(set(order)):
        raise ValueError("sheet_order must be a non-empty unique list")
    colors = data.get("tab_colors")
    if not isinstance(colors, dict):
        raise ValueError("tab_colors must be an object")
    editable = data.get("editable_ranges")
    if not isinstance(editable, dict) or editable != {"Opportunity_Discovery": "A1:R100"}:
        raise ValueError("editable_ranges must declare only Opportunity_Discovery!A1:R100")
    return data


def _target(cell: Cell) -> str:
    if cell.hyperlink:
        return cell.hyperlink.target
    value = cell.value
    if isinstance(value, str):
        match = HYPERLINK_FORMULA_RE.match(value)
        if match:
            return match.group(1).replace('""', '"')
    return ""


def _prompt_rows(library) -> Dict[str, int]:
    rows: Dict[str, int] = {}
    for row in range(2, library.max_row + 1):
        value = library.cell(row=row, column=PROMPT_ID_COLUMN).value
        prompt_id = value if isinstance(value, str) and PROMPT_ID_RE.fullmatch(value) else ""
        if not prompt_id:
            sequence = library.cell(row=row, column=2).value
            if isinstance(sequence, int) or (isinstance(sequence, str) and sequence.isdigit()):
                prompt_id = f"P{int(sequence):02d}"
        if prompt_id:
            rows[prompt_id] = row
    return rows


def _set_range_link(ws, coordinate: str, prompt_range: str) -> None:
    cell = ws[coordinate]
    cell.value = f"Copy {prompt_range} only"
    cell.hyperlink = f"#'{ws.title}'!{prompt_range}"
    cell.font = Font(name="Aptos", size=10, bold=True, underline="single", color="0563C1")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.protection = Protection(locked=True)


def _apply_range_links(workbook) -> Dict[str, str]:
    if LIBRARY_SHEET not in workbook.sheetnames:
        raise ValueError(f"missing required sheet: {LIBRARY_SHEET}")
    library = workbook[LIBRARY_SHEET]
    prompt_ranges: Dict[str, str] = {}
    for prompt_id, row in sorted(_prompt_rows(library).items(), key=lambda item: item[1]):
        sheet_name = str(library.cell(row=row, column=SHEET_NAME_COLUMN).value or "")
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"{prompt_id} references missing sheet: {sheet_name!r}")
        forward = _target(library.cell(row=row, column=PROMPT_ID_COLUMN))
        prefix = f"#'{sheet_name}'!"
        if not forward.startswith(prefix):
            raise ValueError(f"{prompt_id} forward link does not target {sheet_name}: {forward!r}")
        prompt_range = forward[len(prefix):]
        match = PROMPT_RANGE_RE.fullmatch(prompt_range)
        if not match:
            raise ValueError(f"{prompt_id} forward link must select A1:A<n>: {forward!r}")
        last_row = int(match.group(1))
        ws = workbook[sheet_name]
        _set_range_link(ws, "C1", prompt_range)
        _set_range_link(ws, f"C{last_row}", prompt_range)
        prompt_ranges[prompt_id] = prompt_range
    return prompt_ranges


def _color_from_spec(spec: Mapping[str, object]) -> Color:
    if "rgb" in spec:
        return Color(rgb=str(spec["rgb"]))
    return Color(theme=int(spec["theme"]), tint=float(spec.get("tint", 0.0)))


def _apply_order(workbook, order: Sequence[str]) -> None:
    expected = list(order)
    missing = [name for name in expected if name not in workbook.sheetnames]
    unexpected = [name for name in workbook.sheetnames if name not in expected]
    if missing or unexpected:
        raise ValueError(f"sheet set does not match layout contract; missing={missing}, unexpected={unexpected}")
    workbook._sheets = [workbook[name] for name in expected]


def _apply_colors(workbook, colors: Mapping[str, Mapping[str, object]]) -> tuple[str, ...]:
    for ws in workbook.worksheets:
        ws.sheet_properties.tabColor = None
    for name, color_spec in colors.items():
        if name not in workbook.sheetnames:
            raise ValueError(f"missing tab-color sheet: {name}")
        workbook[name].sheet_properties.tabColor = _color_from_spec(color_spec)
    return tuple(colors)


def _apply_protection(workbook, editable_ranges: Mapping[str, str]) -> dict[str, str]:
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True)
        editable_range = editable_ranges.get(ws.title)
        if editable_range:
            min_col, min_row, max_col, max_row = range_boundaries(editable_range)
            for row in ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                for cell in row:
                    cell.protection = Protection(locked=False)
        ws.protection.sheet = True
        ws.protection.enable()
    if workbook.security is None:
        workbook.security = WorkbookProtection()
    workbook.security.lockStructure = True
    return {
        name: cell_range
        for name, cell_range in editable_ranges.items()
        if name in workbook.sheetnames
    }


def canonicalize_layout(workbook_path: Path, spec_path: Path = DEFAULT_SPEC_PATH) -> LayoutResult:
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    spec = _load_spec(spec_path)
    workbook = load_workbook(workbook_path, keep_links=True, data_only=False)
    prompt_ranges = _apply_range_links(workbook)
    _apply_order(workbook, [str(name) for name in spec["sheet_order"]])
    colored_tabs = _apply_colors(workbook, spec["tab_colors"])
    editable_ranges = _apply_protection(
        workbook,
        {str(name): str(cell_range) for name, cell_range in spec["editable_ranges"].items()},
    )
    workbook.save(workbook_path)

    result = validate_artifact(workbook_path, spec_path)
    if not result.passed:
        raise RuntimeError("layout contract failed: " + "; ".join(result.findings))
    return LayoutResult(
        workbook=str(workbook_path),
        prompt_ranges=prompt_ranges,
        sheet_order=tuple(spec["sheet_order"]),
        colored_tabs=colored_tabs,
        editable_ranges=editable_ranges,
    )


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Canonicalize V33 prompt-range links, tab order/colors, and worksheet protection."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--spec", type=Path, default=DEFAULT_SPEC_PATH)
    parser.add_argument("--report", type=Path)
    args = parser.parse_args(argv)
    result = canonicalize_layout(args.workbook, args.spec)
    payload = result.to_dict()
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(payload, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
