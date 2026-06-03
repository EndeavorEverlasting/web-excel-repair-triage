"""Header row style comparison (section D)."""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

from triage.roster_log_compare.load import load_workbook

_HEADER_SHEET_RE = re.compile(
    r"^(Live|Worked Projects|Assignments|Expected Hours)\s*[-–]",
    re.IGNORECASE,
)


def _style_fp(cell) -> Dict[str, Any]:
    fill = cell.fill
    font = cell.font
    align = cell.alignment
    fg = None
    if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
        fg = str(fill.fgColor.rgb)
    return {
        "value": cell.value,
        "fill": fg,
        "font_bold": bool(font.bold) if font else False,
        "font_size": font.size if font else None,
        "font_color": str(font.color.rgb) if font and font.color and font.color.rgb else None,
        "align_h": align.horizontal if align else None,
        "align_v": align.vertical if align else None,
    }


def _header_sheets(names: List[str]) -> List[str]:
    return [n for n in names if _HEADER_SHEET_RE.match(n.strip())]


def compare_header_styles(left_path: Path, right_path: Path, *, max_col: int = 60) -> Dict[str, Any]:
    wl = load_workbook(left_path, data_only=False)
    wr = load_workbook(right_path, data_only=False)
    results: List[Dict[str, Any]] = []
    try:
        sheets = sorted(set(_header_sheets(wl.sheetnames)) | set(_header_sheets(wr.sheetnames)))
        for name in sheets:
            if name not in wl.sheetnames or name not in wr.sheetnames:
                results.append({
                    "sheet": name,
                    "header_identical": False,
                    "present": "partial",
                    "changes": [{"aspect": "sheet_presence"}],
                })
                continue
            ws_l, ws_r = wl[name], wr[name]
            changes: List[Dict[str, Any]] = []
            cols = min(max_col, max(ws_l.max_column, ws_r.max_column, 1))
            for r in (1, 2):
                h_l = ws_l.row_dimensions[r].height if r in ws_l.row_dimensions else None
                h_r = ws_r.row_dimensions[r].height if r in ws_r.row_dimensions else None
                if h_l != h_r and (h_l or h_r):
                    changes.append({
                        "cell": f"row{r}",
                        "aspect": "row_height",
                        "left": h_l,
                        "right": h_r,
                    })
                for c in range(1, cols + 1):
                    fl = _style_fp(ws_l.cell(r, c))
                    fr = _style_fp(ws_r.cell(r, c))
                    if fl != fr:
                        from openpyxl.utils import get_column_letter
                        changes.append({
                            "cell": f"{get_column_letter(c)}{r}",
                            "aspect": "style_or_value",
                            "left": fl,
                            "right": fr,
                        })
            merges_l: Set[str] = {str(m) for m in ws_l.merged_cells.ranges if m.min_row <= 2}
            merges_r: Set[str] = {str(m) for m in ws_r.merged_cells.ranges if m.min_row <= 2}
            if merges_l != merges_r:
                changes.append({
                    "cell": "rows1-2",
                    "aspect": "merged_cells",
                    "left": sorted(merges_l),
                    "right": sorted(merges_r),
                })
            results.append({
                "sheet": name,
                "header_identical": len(changes) == 0,
                "changes": changes,
            })
        return {"sheets": results}
    finally:
        wl.close()
        wr.close()
