"""Live sheet punch/value comparison (section C)."""
from __future__ import annotations

import re
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from triage.roster_log_compare.load import load_workbook
from triage.roster_parser import (
    _extract_year_from_sheet,
    _find_header_row,
    _parse_date_header,
    _time_to_hours,
)

_LIVE_RE = re.compile(r"^Live\s*[-–]\s*(.+)$", re.IGNORECASE)


def _live_sheets(wb) -> Dict[str, str]:
    out = {}
    for name in wb.sheetnames:
        m = _LIVE_RE.match(name.strip())
        if m:
            out[name] = m.group(1).strip()
    return out


def _normalize_live_value(val: Any) -> Any:
    if val is None or val == "":
        return None
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        th = _time_to_hours(s)
        if th is not None:
            return round(th, 6)
        return s
    if isinstance(val, (int, float)):
        th = _time_to_hours(val)
        if th is not None:
            return round(th, 6)
        return val
    try:
        import datetime as dt
        if isinstance(val, dt.time):
            return round(_time_to_hours(val) or 0, 6)
        if isinstance(val, dt.datetime):
            return round(_time_to_hours(val) or 0, 6)
    except Exception:
        pass
    return val


def _display_value(val: Any) -> Any:
    if val is None:
        return ""
    return val


def compare_live(
    left_path: Path,
    right_path: Path,
    *,
    include_formatting: bool = False,
) -> Dict[str, Any]:
    wl = load_workbook(left_path, data_only=True)
    wr = load_workbook(right_path, data_only=True)
    diffs: List[Dict[str, Any]] = []
    stats = {"left_nonempty_punches": 0, "right_nonempty_punches": 0}
    try:
        left_live = _live_sheets(wl)
        right_live = _live_sheets(wr)
        paired = sorted(set(left_live) & set(right_live))
        only_left = sorted(set(left_live) - set(right_live))
        only_right = sorted(set(right_live) - set(left_live))

        for sheet_name in paired:
            ws_l, ws_r = wl[sheet_name], wr[sheet_name]
            year = _extract_year_from_sheet(sheet_name)
            hdr_l = _find_header_row(ws_l)
            hdr_r = _find_header_row(ws_r)
            if hdr_l is None or hdr_r is None:
                continue
            headers_l = [ws_l.cell(hdr_l, c).value for c in range(1, ws_l.max_column + 1)]
            headers_r = [ws_r.cell(hdr_r, c).value for c in range(1, ws_r.max_column + 1)]
            date_cols: Dict[date, Dict[str, Tuple[int, int, str]]] = {}
            for i, h in enumerate(headers_l):
                if h is None:
                    continue
                res = _parse_date_header(str(h), year)
                if not res:
                    continue
                d, direction = res
                field = "Clock In" if direction == "in" else "Clock Out"
                date_cols.setdefault(d, {})[direction] = (i + 1, i + 1, field)

            staff_col = 1
            max_row = max(ws_l.max_row, ws_r.max_row)
            for r in range(hdr_l + 1, max_row + 1):
                staff_l = ws_l.cell(r, staff_col).value if r <= ws_l.max_row else None
                staff_r = ws_r.cell(r, staff_col).value if r <= ws_r.max_row else None
                staff = staff_r or staff_l
                if not staff or str(staff).strip() in ("", "None"):
                    continue
                staff_s = str(staff).strip()
                for d, dirs in sorted(date_cols.items()):
                    for direction, (col_l, _, field) in dirs.items():
                        col_r = col_l
                        if col_r > ws_r.max_column:
                            continue
                        vl = ws_l.cell(r, col_l).value if r <= ws_l.max_row else None
                        vr = ws_r.cell(r, col_r).value if r <= ws_r.max_row else None
                        nl, nr = _normalize_live_value(vl), _normalize_live_value(vr)
                        if nl is not None:
                            stats["left_nonempty_punches"] += 1
                        if nr is not None:
                            stats["right_nonempty_punches"] += 1
                        if nl == nr and not include_formatting:
                            continue
                        if nl != nr:
                            from openpyxl.utils import get_column_letter
                            addr = f"{get_column_letter(col_l)}{r}"
                            diffs.append({
                                "sheet": sheet_name,
                                "date": d.isoformat(),
                                "staff": staff_s,
                                "field": field,
                                "cell": addr,
                                "left_value": _display_value(vl),
                                "right_value": _display_value(vr),
                            })

        return {
            "diffs": diffs,
            "stats": stats,
            "paired_sheets": paired,
            "only_left": only_left,
            "only_right": only_right,
        }
    finally:
        wl.close()
        wr.close()
