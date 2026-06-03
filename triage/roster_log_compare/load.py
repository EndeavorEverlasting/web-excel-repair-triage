"""Read-only workbook loading helpers."""
from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Optional, Tuple


def resolve_path(path: str, repo_root: Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else (repo_root / p).resolve()


def file_mtime_iso(path: Path) -> Optional[str]:
    try:
        ts = path.stat().st_mtime
        return datetime.fromtimestamp(ts, tz=timezone.utc).isoformat()
    except OSError:
        return None


def filename_date_token(name: str) -> Optional[str]:
    m = re.search(r"(20\d{2})[-_](\d{2})[-_](\d{2})", name)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def load_workbook(path: Path, *, data_only: bool):
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from e
    return openpyxl.load_workbook(str(path), read_only=False, data_only=data_only)


def workbook_side_meta(path: Path, wb) -> Dict[str, Any]:
    props = wb.properties
    created = props.created
    modified = props.modified
    return {
        "path": str(path),
        "filename": path.name,
        "file_mtime_utc": file_mtime_iso(path),
        "filename_date_token": filename_date_token(path.name),
        "workbook_created": created.isoformat() if created else None,
        "workbook_modified": modified.isoformat() if modified else None,
        "creator": props.creator,
        "sheetnames": list(wb.sheetnames),
        "sheet_count": len(wb.sheetnames),
    }
