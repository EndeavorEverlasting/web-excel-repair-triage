"""Web Excel-safe comparison workbook report."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

_TITLE_FILL = "1F365C"
_RISK_STOP = "FFC7CE"
_RISK_WARN = "FFEB9C"


def _xl():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    return Workbook, Font, PatternFill, Alignment, get_column_letter, Table, TableStyleInfo


def _write_table(ws, title: str, headers: List[str], rows: List[Dict[str, Any]],
                 table_name: str, key_order: List[str] | None = None) -> None:
    Workbook, Font, PatternFill, Alignment, get_column_letter, Table, TableStyleInfo = _xl()
    fill = PatternFill("solid", fgColor=_TITLE_FILL)
    ws.cell(1, 1, title).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(1, 1).fill = fill
    hdr_row = 3
    keys = key_order or headers
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hdr_row, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, hdr_row + 1):
        for ci, k in enumerate(keys, 1):
            ws.cell(ri, ci, row.get(k, ""))
    last = max(hdr_row, hdr_row + len(rows))
    if rows:
        c0, c1 = get_column_letter(1), get_column_letter(len(headers))
        tab = Table(displayName=table_name[:255], ref=f"{c0}{hdr_row}:{c1}{last}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1).coordinate
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22


def write_comparison_workbook(result: Dict[str, Any], out_path: Path) -> None:
    Workbook, *_ = _xl()
    from openpyxl.styles import PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    v = result.get("verdict") or {}
    ws_v = wb.create_sheet("Verdict")
    ws_v["A1"] = "Roster Log Comparison Verdict"
    ws_v["A3"] = "Recommendation"
    ws_v["B3"] = v.get("recommendation")
    ws_v["A4"] = "Confidence"
    ws_v["B4"] = v.get("confidence")
    ws_v["A5"] = "Left score"
    ws_v["B5"] = v.get("left_score")
    ws_v["A6"] = "Right score"
    ws_v["B6"] = v.get("right_score")
    ws_v["A8"] = "Reasons"
    for i, r in enumerate(v.get("reasons") or [], 9):
        ws_v.cell(i, 1, r)

    meta = result.get("sections", {}).get("metadata") or {}
    meta_rows = []
    for side in ("left", "right"):
        m = meta.get(side) or {}
        meta_rows.append({
            "Side": side,
            "Filename": m.get("filename"),
            "File mtime UTC": m.get("file_mtime_utc"),
            "Workbook created": m.get("workbook_created"),
            "Workbook modified": m.get("workbook_modified"),
            "Sheet count": m.get("sheet_count"),
        })
    ws_m = wb.create_sheet("Workbook Metadata")
    _write_table(ws_m, "Workbook Metadata",
                 ["Side", "Filename", "File mtime UTC", "Workbook created",
                  "Workbook modified", "Sheet count"],
                 meta_rows, "MetaTbl",
                 ["Side", "Filename", "File mtime UTC", "Workbook created",
                  "Workbook modified", "Sheet count"])

    struct_rows = []
    for row in (result.get("sections", {}).get("structure") or {}).get("rows") or []:
        struct_rows.append({
            "Sheet": row.get("sheet"),
            "Status": row.get("status"),
            "Row delta": row.get("row_delta"),
            "Col delta": row.get("col_delta"),
        })
    ws_s = wb.create_sheet("Sheet Structure Diff")
    _write_table(ws_s, "Sheet Structure Diff",
                 ["Sheet", "Status", "Row delta", "Col delta"],
                 struct_rows, "StructTbl",
                 ["Sheet", "Status", "Row delta", "Col delta"])

    live_rows = result.get("sections", {}).get("live", {}).get("diffs") or []
    ws_l = wb.create_sheet("Live Date Diffs")
    _write_table(ws_l, "Live Date Diffs",
                 ["Date", "Staff", "Field", "Cell", "Left value", "Right value", "Sheet"],
                 live_rows, "LiveTbl",
                 ["date", "staff", "field", "cell", "left_value", "right_value", "sheet"])

    cf_rows = result.get("sections", {}).get("conditional_formatting", {}).get("per_sheet") or []
    flat_cf = []
    for r in cf_rows:
        flat_cf.append({
            "Sheet": r.get("sheet"),
            "Left rules": r.get("left_rules"),
            "Right rules": r.get("right_rules"),
            "Ranges added on right": ", ".join(r.get("ranges_added_on_right") or []),
        })
    ws_cf = wb.create_sheet("CF Summary")
    _write_table(ws_cf, "CF Summary",
                 ["Sheet", "Left rules", "Right rules", "Ranges added on right"],
                 flat_cf, "CFTbl",
                 ["Sheet", "Left rules", "Right rules", "Ranges added on right"])

    ov_rows = []
    for row in (result.get("sections", {}).get("override_table") or {}).get("rows") or []:
        for side in ("left", "right"):
            s = row.get(side) or {}
            ov_rows.append({
                "Sheet": row.get("sheet"),
                "Side": side,
                "Override present": s.get("override_table_present"),
                "Refs override range": s.get("formulas_reference_override_range"),
                "Structural OK": s.get("structurally_functional"),
            })
    ws_o = wb.create_sheet("Override Table Check")
    _write_table(ws_o, "Override Table Check",
                 ["Sheet", "Side", "Override present", "Refs override range", "Structural OK"],
                 ov_rows, "OvTbl",
                 ["Sheet", "Side", "Override present", "Refs override range", "Structural OK"])

    eh_rows = []
    for row in (result.get("sections", {}).get("expected_hours") or {}).get("rows") or []:
        for side in ("left", "right"):
            s = row.get(side) or {}
            if not s.get("present"):
                continue
            eh_rows.append({
                "Sheet": row.get("sheet"),
                "Side": side,
                "Mode": s.get("mode"),
                "Max date": s.get("max_date"),
                "Stale warning": s.get("stale_snapshot_warning"),
            })
    ws_e = wb.create_sheet("Expected Hours Check")
    _write_table(ws_e, "Expected Hours Check",
                 ["Sheet", "Side", "Mode", "Max date", "Stale warning"],
                 eh_rows, "EHTbl",
                 ["Sheet", "Side", "Mode", "Max date", "Stale warning"])

    risk_rows = result.get("risk_flags") or []
    ws_r = wb.create_sheet("Risk Flags")
    stop_fill = PatternFill("solid", fgColor=_RISK_STOP)
    warn_fill = PatternFill("solid", fgColor=_RISK_WARN)
    _write_table(ws_r, "Risk Flags",
                 ["Severity", "Code", "Detail"],
                 [{"Severity": r.get("severity"), "Code": r.get("code"),
                   "Detail": str(r.get("detail"))} for r in risk_rows],
                 "RiskTbl",
                 ["Severity", "Code", "Detail"])
    for ri, r in enumerate(risk_rows, 4):
        fill = stop_fill if r.get("severity") == "stop" else warn_fill
        for c in range(1, 4):
            ws_r.cell(ri, c).fill = fill

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    try:
        from triage.xlsx_utils import fix_inlinestr
        fix_inlinestr(str(out_path))
    except Exception:
        pass
