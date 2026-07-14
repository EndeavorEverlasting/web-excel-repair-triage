"""Build a new roster review workbook shell for ``--mode blank``."""
from __future__ import annotations

import json
from calendar import month_abbr, month_name, monthrange
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from triage.month_validation import validate_month_key

_REVIEW_RULES_SEED = (
    Path(__file__).resolve().parents[2]
    / "configs"
    / "roster_log_review_queue"
    / "review_rules_seed.json"
)

REVIEW_QUEUE_HEADERS = [
    "Review ID",
    "Month",
    "Date",
    "Staff",
    "Source Sheet",
    "Source Cells",
    "Rule Code",
    "Severity",
    "Current Status",
    "Detected Value",
    "Expected Value",
    "Suggested Resolution",
    "Resolution Value",
    "Resolution Source",
    "Owner",
    "Last Reviewed",
    "Notes",
]

CF_DICTIONARY: List[Tuple[str, str, str]] = [
    (
        "Red",
        "Missing project or invalid punch pair",
        "MISSING_PROJECT_CORRECTION, INVALID_PUNCH_PAIR",
    ),
    (
        "Green",
        "Documented non-work marker",
        "PTO_ACCEPTED, OFF_DAY_ACCEPTED, DAY_OFF_EVIDENCE_ACCEPTED, "
        "CASR_DAY_OFF_ACCEPTED",
    ),
    ("Blue", "Shift is 12 hours or longer", "LONG_SHIFT_REVIEW"),
    ("Purple", "Potential overtime or extended shift", "OT_REVIEW"),
    ("Amber", "Partial shift requires review", "PARTIAL_HOURS_REVIEW"),
    (
        "Light Blue",
        "Punch contains reviewable note text",
        "NOTE_BEARING_PUNCH",
    ),
]


def _load_review_rules() -> List[Tuple[str, str, str, str]]:
    """Load the canonical review-rule contract without inventing rule codes."""
    payload = json.loads(_REVIEW_RULES_SEED.read_text(encoding="utf-8"))
    codes = [str(code).strip() for code in payload.get("rule_codes", []) if str(code).strip()]
    if not codes:
        raise ValueError(f"review rule seed has no rule_codes: {_REVIEW_RULES_SEED}")

    rows_by_code: Dict[str, dict] = {}
    for row in payload.get("rows", []):
        if not isinstance(row, dict):
            continue
        code = str(row.get("rule_code") or row.get("Rule Code") or "").strip()
        if code:
            rows_by_code[code] = row

    rules: List[Tuple[str, str, str, str]] = []
    for code in codes:
        row = rows_by_code.get(code, {})
        rules.append(
            (
                code,
                str(row.get("severity") or row.get("Severity") or ""),
                str(row.get("trigger") or row.get("Trigger") or ""),
                str(
                    row.get("suggested_resolution")
                    or row.get("Suggested Resolution")
                    or ""
                ),
            )
        )
    return rules


def _normalize_months(months: Iterable[str]) -> List[Tuple[str, int, int, str]]:
    values = list(months or [])
    if not values:
        raise ValueError("--months is required for blank mode")

    result: List[Tuple[str, int, int, str]] = []
    seen: set[str] = set()
    for key in values:
        year, month = validate_month_key(key)
        normalized = f"{year:04d}-{month:02d}"
        if normalized in seen:
            continue
        seen.add(normalized)
        result.append((normalized, year, month, f"{month_name[month]} {year}"))
    return result


def _style_header(ws, row: int, *, fill, font, alignment) -> None:
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def build_blank_roster(path: str | Path, months: Iterable[str]) -> Dict[str, object]:
    """Create a review-first blank roster shell and return build metadata.

    The new workbook is intentionally generated with openpyxl. Existing workbook
    modes remain package/XML-only and never pass through an openpyxl save.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    month_specs = _normalize_months(months)
    review_rules = _load_review_rules()
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    dashboard = wb.create_sheet("Review Dashboard")
    dashboard["A1"] = "Roster Log Review Dashboard"
    dashboard["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    dashboard["A1"].fill = header_fill
    dashboard.merge_cells("A1:D1")
    dashboard.append([])
    dashboard.append(["Workbook State", "Blank operator shell"])
    dashboard.append(["Months", ", ".join(spec[0] for spec in month_specs)])
    dashboard.append(["Review Queue", "No source rows loaded"])
    dashboard.append(
        [
            "Instructions",
            "Enter staff, project, and punches on the Live tabs. Run full or "
            "review-only mode against a populated roster to build review evidence.",
        ]
    )
    dashboard.column_dimensions["A"].width = 22
    dashboard.column_dimensions["B"].width = 92
    dashboard.freeze_panes = "A3"

    queue = wb.create_sheet("Review Queue")
    queue.append(REVIEW_QUEUE_HEADERS)
    _style_header(queue, 1, fill=header_fill, font=header_font, alignment=header_alignment)
    queue.freeze_panes = "A2"
    queue.auto_filter.ref = f"A1:{get_column_letter(len(REVIEW_QUEUE_HEADERS))}1"
    widths = [16, 14, 14, 24, 24, 18, 24, 12, 16, 24, 24, 34, 24, 24, 20, 18, 40]
    for idx, width in enumerate(widths, 1):
        queue.column_dimensions[get_column_letter(idx)].width = width

    rules = wb.create_sheet("Review Rules")
    rules.append(["Rule Code", "Severity", "Trigger", "Suggested Resolution"])
    for row in review_rules:
        rules.append(row)
    _style_header(rules, 1, fill=header_fill, font=header_font, alignment=header_alignment)
    rules.freeze_panes = "A2"
    for col, width in zip("ABCD", [34, 14, 52, 52]):
        rules.column_dimensions[col].width = width

    dictionary = wb.create_sheet("CF Dictionary")
    dictionary.append(["Color", "Meaning", "Rule Codes"])
    for row in CF_DICTIONARY:
        dictionary.append(row)
    _style_header(
        dictionary, 1, fill=header_fill, font=header_font, alignment=header_alignment
    )
    dictionary.freeze_panes = "A2"
    for col, width in zip("ABC", [18, 44, 72]):
        dictionary.column_dimensions[col].width = width

    live_sheets: List[str] = []
    for _, year, month, label in month_specs:
        title = f"Live - {label}"
        ws = wb.create_sheet(title)
        live_sheets.append(title)
        ws.append([f"{label} - Attendance"])
        headers = ["Staff Name", "Project"]
        for day in range(1, monthrange(year, month)[1] + 1):
            prefix = f"{month_abbr[month]} {day:02d}"
            headers.extend([f"{prefix} - Clock In", f"{prefix} - Clock Out"])
        ws.append(headers)
        _style_header(ws, 2, fill=header_fill, font=header_font, alignment=header_alignment)
        ws.freeze_panes = "C3"
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 28
        for col in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 17
        ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}202"

    wb.save(out)
    wb.close()
    return {
        "months": [spec[0] for spec in month_specs],
        "live_sheets": live_sheets,
        "review_rules_rows": len(review_rules),
        "cf_dictionary_rows": len(CF_DICTIONARY),
    }
