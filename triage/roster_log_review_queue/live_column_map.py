"""Detect clock-in/out column pairs on Live roster sheets."""
from __future__ import annotations

import re
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .column_utils import col_index_to_letter
from .models import ClockPair

_DATE_HEADER = re.compile(
    r"^([A-Za-z]+)\s+(\d{1,2})\s*[-–]\s*(Clock\s+In|Clock\s+Out)\s*$",
    re.IGNORECASE,
)

_MONTH_ABBREVS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _extract_year(sheet_name: str) -> int:
    m = re.search(r"(20\d{2})", sheet_name)
    return int(m.group(1)) if m else 2026


def _find_header_row(ws) -> Optional[int]:
    for r in range(1, 11):
        for c in range(1, min(ws.max_column or 1, 20) + 1):
            val = ws.cell(r, c).value
            if isinstance(val, str) and "staff" in val.lower() and "name" in val.lower():
                return r
    return None


def detect_clock_pairs_from_workbook(path: str, sheet_name: str) -> Tuple[List[ClockPair], int, str]:
    """Return clock pairs, data row start, and project column letter for a Live sheet."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        hdr_row = _find_header_row(ws)
        if hdr_row is None:
            raise ValueError(f"{sheet_name}: header row not found")

        year = _extract_year(sheet_name)
        headers = [ws.cell(hdr_row, c).value for c in range(1, (ws.max_column or 0) + 1)]

        project_col = "B"
        date_cols: dict = {}
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_str = str(h).strip()
            if i == 0 or ("staff" in h_str.lower() and "name" in h_str.lower()):
                continue
            if "project" in h_str.lower():
                project_col = col_index_to_letter(i + 1)
                continue
            m = _DATE_HEADER.match(h_str)
            if not m:
                continue
            mon = _MONTH_ABBREVS.get(m.group(1)[:3].lower())
            if mon is None:
                continue
            direction = "in" if "in" in m.group(3).lower() else "out"
            date_cols.setdefault((mon, int(m.group(2))), {})[direction] = i + 1

        pairs: List[ClockPair] = []
        for key in sorted(date_cols.keys()):
            cols = date_cols[key]
            if "in" not in cols or "out" not in cols:
                continue
            in_idx, out_idx = cols["in"], cols["out"]
            pairs.append(
                ClockPair(
                    in_col=col_index_to_letter(in_idx),
                    out_col=col_index_to_letter(out_idx),
                    in_index=in_idx,
                    out_index=out_idx,
                )
            )

        data_row_start = hdr_row + 1
        return pairs, data_row_start, project_col
    finally:
        wb.close()


def detect_clock_pairs(path: str, sheet_name: str) -> List[ClockPair]:
    pairs, _, _ = detect_clock_pairs_from_workbook(path, sheet_name)
    return pairs


def last_clock_column(pairs: List[ClockPair]) -> str:
    return pairs[-1].out_col if pairs else ""
