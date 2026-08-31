"""Generate the Roster Log V2 workbook from normalized JSON state."""
from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any, Dict, List

from triage.xlsx_utils import fix_inlinestr

from .preflight import preflight_roster_v2
from .schema import normalize_state, reconcile_state


def _iso_date(value: str):
    return date.fromisoformat(value)


def build_roster_workbook(
    payload: Dict[str, Any],
    output_path: str | Path,
    *,
    require_reconciled: bool = False,
) -> Dict[str, Any]:
    """Build a new editable roster without mutating any predecessor workbook.

    One project is the default. Explicit allocation rows make a day multi-project;
    that is normal. Only attendance/allocation variance is reviewable.
    """
    from openpyxl import Workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    state = normalize_state(payload)
    reconciliation = reconcile_state(state)
    if require_reconciled:
        bad = [row for row in reconciliation if not row.reconciled]
        if bad:
            raise ValueError("require_reconciled requested with unresolved allocation variance")

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="163A5F")
    accent_fill = PatternFill("solid", fgColor="DDEBF7")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    bad_fill = PatternFill("solid", fgColor="FCE4D6")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(vertical="top", wrap_text=True)

    dashboard = wb.create_sheet("Dashboard")
    dashboard.append(["Roster Log V2", "Normalized attendance + project allocation ledger"])
    dashboard.append(["State", "CURRENT CANDIDATE — promote only after operator acceptance"])
    dashboard.append(["Attendance days", len(state["attendance"])])
    dashboard.append(["Allocation rows", len(state["allocations"])])
    dashboard.append(["Reconciled days", sum(1 for row in reconciliation if row.reconciled)])
    dashboard.append(["Days requiring allocation repair", sum(1 for row in reconciliation if not row.reconciled)])
    dashboard.append(["Rule", "One project is the default. Add project rows only when the day should be split."])
    dashboard.append(["Rule", "Multi-project days are supported and are not review flags by themselves."])
    dashboard.append(["Rule", "Allocated project hours must reconcile to paid attendance before closeout."])
    dashboard.column_dimensions["A"].width = 32
    dashboard.column_dimensions["B"].width = 96
    dashboard.freeze_panes = "A2"
    for cell in dashboard[1]:
        cell.fill = header_fill
        cell.font = header_font

    attendance = wb.create_sheet("Attendance")
    attendance_headers = [
        "Date",
        "Staff",
        "Clock In",
        "Clock Out",
        "Paid Hours",
        "Default Project",
        "Project Mode",
        "Allocated Hours",
        "Variance",
        "Reconciled?",
        "Notes",
    ]
    attendance.append(attendance_headers)
    for raw in state["attendance"]:
        attendance.append(
            [
                _iso_date(raw["date"]),
                raw["staff"],
                raw.get("clock_in", ""),
                raw.get("clock_out", ""),
                raw["paid_hours"],
                raw["default_project"],
                None,
                None,
                None,
                None,
                raw.get("notes", ""),
            ]
        )
    for row in range(2, max(attendance.max_row, 201) + 1):
        attendance.cell(row, 7).value = (
            f'=IF(COUNTIFS(\'Project Allocations\'!$B:$B,A{row},\'Project Allocations\'!$C:$C,B{row})>1,"MULTI","SINGLE")'
        )
        attendance.cell(row, 8).value = (
            f'=SUMIFS(\'Project Allocations\'!$F:$F,\'Project Allocations\'!$B:$B,A{row},\'Project Allocations\'!$C:$C,B{row})'
        )
        attendance.cell(row, 9).value = f'=IF(OR(A{row}="",B{row}=""),"",E{row}-H{row})'
        attendance.cell(row, 10).value = f'=IF(I{row}="","",IF(ABS(I{row})<=0.01,"YES","NO"))'
    attendance.freeze_panes = "A2"
    attendance.auto_filter.ref = f"A1:K{max(attendance.max_row, 2)}"
    attendance.column_dimensions["A"].width = 13
    attendance.column_dimensions["B"].width = 24
    attendance.column_dimensions["C"].width = 11
    attendance.column_dimensions["D"].width = 11
    attendance.column_dimensions["E"].width = 12
    attendance.column_dimensions["F"].width = 38
    attendance.column_dimensions["G"].width = 14
    attendance.column_dimensions["H"].width = 15
    attendance.column_dimensions["I"].width = 12
    attendance.column_dimensions["J"].width = 13
    attendance.column_dimensions["K"].width = 42
    attendance["A2"].number_format = "yyyy-mm-dd"
    for row in attendance.iter_rows(min_row=2, max_row=attendance.max_row, min_col=1, max_col=1):
        row[0].number_format = "yyyy-mm-dd"

    allocations = wb.create_sheet("Project Allocations")
    allocation_headers = [
        "Allocation ID",
        "Date",
        "Staff",
        "Project / Billing Scope",
        "Workstream",
        "Allocated Hours",
        "Status",
        "Notes",
    ]
    allocations.append(allocation_headers)
    for raw in state["allocations"]:
        allocations.append(
            [
                raw["allocation_id"],
                _iso_date(raw["date"]),
                raw["staff"],
                raw["project"],
                raw.get("workstream", ""),
                raw["hours"],
                raw.get("status", "RECONCILED"),
                raw.get("notes", ""),
            ]
        )
    allocations.freeze_panes = "A2"
    allocations.auto_filter.ref = f"A1:H{max(allocations.max_row, 2)}"
    widths = [24, 13, 24, 40, 36, 15, 18, 48]
    for idx, width in enumerate(widths, 1):
        allocations.column_dimensions[chr(64 + idx)].width = width
    for row in allocations.iter_rows(min_row=2, max_row=allocations.max_row, min_col=2, max_col=2):
        row[0].number_format = "yyyy-mm-dd"

    dictionaries = wb.create_sheet("Dictionaries")
    dictionaries.append(["Projects", "Workstreams", "Allocation Status"])
    projects = list(dict.fromkeys([str(x) for x in state.get("projects", []) if str(x).strip()] + [r["default_project"] for r in state["attendance"] if r["default_project"]] + [r["project"] for r in state["allocations"]]))
    workstreams = list(dict.fromkeys([str(x) for x in state.get("workstreams", []) if str(x).strip()] + [str(r.get("workstream") or "") for r in state["allocations"] if str(r.get("workstream") or "").strip()]))
    statuses = ["RECONCILED", "DRAFT", "PENDING CLOSE"]
    for idx in range(max(len(projects), len(workstreams), len(statuses), 1)):
        dictionaries.append([
            projects[idx] if idx < len(projects) else None,
            workstreams[idx] if idx < len(workstreams) else None,
            statuses[idx] if idx < len(statuses) else None,
        ])
    dictionaries.freeze_panes = "A2"
    dictionaries.column_dimensions["A"].width = 44
    dictionaries.column_dimensions["B"].width = 40
    dictionaries.column_dimensions["C"].width = 20

    # Range-backed validation: mutable dictionaries stay in the workbook rather
    # than being duplicated into stale inline validation strings.
    if projects:
        default_dv = DataValidation(type="list", formula1=f"=Dictionaries!$A$2:$A${len(projects)+1}", allow_blank=True)
        alloc_project_dv = DataValidation(type="list", formula1=f"=Dictionaries!$A$2:$A${len(projects)+1}", allow_blank=False)
        attendance.add_data_validation(default_dv)
        allocations.add_data_validation(alloc_project_dv)
        default_dv.add("F2:F1000")
        alloc_project_dv.add("D2:D2000")
    if workstreams:
        workstream_dv = DataValidation(type="list", formula1=f"=Dictionaries!$B$2:$B${len(workstreams)+1}", allow_blank=True)
        allocations.add_data_validation(workstream_dv)
        workstream_dv.add("E2:E2000")
    status_dv = DataValidation(type="list", formula1="=Dictionaries!$C$2:$C$4", allow_blank=False)
    allocations.add_data_validation(status_dv)
    status_dv.add("G2:G2000")

    review = wb.create_sheet("Review Queue")
    review.append(["Date", "Staff", "Rule", "Paid Hours", "Allocated Hours", "Variance", "Action"])
    for rec in reconciliation:
        if rec.reconciled:
            continue
        review.append([
            _iso_date(rec.work_date),
            rec.staff,
            "ALLOCATION_VARIANCE",
            rec.paid_hours,
            rec.allocated_hours,
            rec.variance,
            "Adjust project allocation rows until total allocated hours equals paid attendance.",
        ])
    review.freeze_panes = "A2"
    review.column_dimensions["A"].width = 13
    review.column_dimensions["B"].width = 24
    review.column_dimensions["C"].width = 24
    review.column_dimensions["D"].width = 14
    review.column_dimensions["E"].width = 16
    review.column_dimensions["F"].width = 12
    review.column_dimensions["G"].width = 76

    readme = wb.create_sheet("Read Me")
    readme_rows = [
        ["Roster Log V2 operating contract"],
        ["One project is the default. A paid attendance day automatically receives one full-day allocation when no explicit allocation is supplied."],
        ["Multi-project days are supported. Add one Project Allocations row per project/workstream that should receive part of the attendance day."],
        ["A multi-project day is not an error. The review condition is arithmetic: allocated hours must reconcile to paid attendance."],
        ["Operator project decisions and explicit overrides are authoritative inputs. This workbook does not manufacture an 80/20 split or second-guess a deliberate full-day project decision."],
        ["Attendance owns paid hours. Project Allocations explain where those hours belong; allocation rows cannot create additional paid hours."],
        ["Use the Dictionaries sheet for reusable projects/workstreams. Validation points to worksheet ranges so the vocabulary can evolve without stale inline lists."],
        ["The prior roster remains untouched. Promote this V2 workbook to CURRENT only after the operator confirms the new workflow is functional."],
    ]
    for row in readme_rows:
        readme.append(row)
    readme.column_dimensions["A"].width = 116
    readme.freeze_panes = "A2"

    for ws in (attendance, allocations, dictionaries, review):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = wrap
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = wrap

    attendance.conditional_formatting.add("J2:J1000", FormulaRule(formula=['$J2="YES"'], fill=ok_fill))
    attendance.conditional_formatting.add("J2:J1000", FormulaRule(formula=['$J2="NO"'], fill=bad_fill))
    attendance.conditional_formatting.add("G2:G1000", FormulaRule(formula=['$G2="MULTI"'], fill=accent_fill))
    review.conditional_formatting.add("A2:G1000", FormulaRule(formula=['$C2="ALLOCATION_VARIANCE"'], fill=warn_fill))

    wb.save(out)
    wb.close()

    # Triage-owned package normalization and Web Excel package checks.
    fix_inlinestr(str(out))
    preflight = preflight_roster_v2(out)
    if not preflight["preflight_pass"]:
        raise ValueError(f"Roster Log V2 preflight failed: {preflight['errors']}")

    return {
        "path": str(out),
        "schema_version": state["schema_version"],
        "attendance_days": len(state["attendance"]),
        "allocation_rows": len(state["allocations"]),
        "reconciled_days": sum(1 for row in reconciliation if row.reconciled),
        "unreconciled_days": sum(1 for row in reconciliation if not row.reconciled),
        "preflight": preflight,
    }
