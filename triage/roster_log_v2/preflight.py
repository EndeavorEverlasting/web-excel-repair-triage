"""Fail-closed Roster Log V2 workbook preflight."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from triage.web_excel_compatibility_rules import inspect_web_excel_package

REQUIRED_SHEETS = [
    "Dashboard",
    "Attendance",
    "Project Allocations",
    "Dictionaries",
    "Review Queue",
    "Read Me",
]


def preflight_roster_v2(path: str | Path) -> Dict[str, Any]:
    from openpyxl import load_workbook

    p = Path(path)
    errors: List[str] = []
    if not p.exists():
        return {"preflight_pass": False, "errors": ["file_not_found"]}

    package_issues = inspect_web_excel_package(p)
    errors.extend(f"webexcel:{issue.code}:{issue.part}" for issue in package_issues)

    try:
        wb = load_workbook(p, read_only=False, data_only=False)
    except Exception as exc:  # pragma: no cover - package gate normally catches this
        return {"preflight_pass": False, "errors": errors + [f"open:{exc}"]}

    try:
        for name in REQUIRED_SHEETS:
            if name not in wb.sheetnames:
                errors.append(f"missing_sheet:{name}")
        if all(name in wb.sheetnames for name in ("Attendance", "Project Allocations")):
            attendance_headers = [cell.value for cell in wb["Attendance"][1]]
            allocation_headers = [cell.value for cell in wb["Project Allocations"][1]]
            for header in ("Default Project", "Allocated Hours", "Variance", "Reconciled?"):
                if header not in attendance_headers:
                    errors.append(f"attendance_header:{header}")
            for header in ("Allocation ID", "Project / Billing Scope", "Allocated Hours"):
                if header not in allocation_headers:
                    errors.append(f"allocation_header:{header}")
        if "Read Me" in wb.sheetnames:
            text = " ".join(
                str(wb["Read Me"].cell(row=r, column=1).value or "")
                for r in range(1, min(wb["Read Me"].max_row, 20) + 1)
            )
            for phrase in (
                "One project is the default",
                "Multi-project days are supported",
                "allocated hours must reconcile",
            ):
                if phrase not in text:
                    errors.append(f"missing_contract_text:{phrase}")
    finally:
        wb.close()

    return {
        "preflight_pass": not errors,
        "errors": errors,
        "web_excel_issue_count": len(package_issues),
        "required_sheets": REQUIRED_SHEETS,
    }
