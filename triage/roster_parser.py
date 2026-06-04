"""
triage/roster_parser.py
-----------------------
Parse the Active Roster Log (.xlsx) — wide-form layout.

Sheet pattern: "Live - {Month YYYY}"

Wide-form layout:
  Row 1: Title
  Row 2: Staff Name | Project | <Apr 01 - Clock In> | <Apr 01 - Clock Out> | ...
  Row 3+: One row per staff member; clock columns hold datetime.time values.

Returns a list of dicts per staff/date record:
    {staff, project, date, clock_in, clock_out, gross_hours,
     lunch_deduction, net_hours, long_shift}

Raises RosterParseError if expected structure is missing.
"""
from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple


class RosterParseError(Exception):
    pass


# ── Lunch deduction policy ────────────────────────────────────────────────────

def _lunch_deduction(gross_hours: float) -> float:
    if gross_hours >= 8.0:
        return 1.0
    if gross_hours >= 6.0:
        return 0.5
    return 0.0


# ── Time helpers ──────────────────────────────────────────────────────────────

def _time_to_hours(value: Any) -> Optional[float]:
    """Convert any clock-in/out cell value to decimal hours (0–24).

    Handles:
      - datetime.time / datetime.datetime / datetime.timedelta objects
      - Excel time fractions (float 0.0–1.0)
      - Strings like '9:28:00 AM', '17:00', '9:28:00 AM/ Bonita'
        (appended notes after the time are silently ignored)
    """
    import datetime as dt
    if value is None:
        return None
    if isinstance(value, dt.time):
        return value.hour + value.minute / 60.0 + value.second / 3600.0
    if isinstance(value, dt.datetime):
        return value.hour + value.minute / 60.0 + value.second / 3600.0
    if isinstance(value, dt.timedelta):
        return value.total_seconds() / 3600.0
    if isinstance(value, (int, float)):
        frac = float(value)
        if 0.0 <= frac < 2.0:          # Excel time fraction (0.0–1.0)
            return frac * 24.0
        return None                     # Probably an unrelated number
    if isinstance(value, str):
        # Extract leading HH:MM or HH:MM:SS, optionally followed by AM/PM.
        # Any trailing text (e.g. "/ Bonita", "- note") is ignored.
        m = re.match(
            r"(\d{1,2}):(\d{2})(?::(\d{2}))?\s*(AM|PM)?",
            value.strip(),
            re.IGNORECASE,
        )
        if not m:
            return None
        hour   = int(m.group(1))
        minute = int(m.group(2))
        second = int(m.group(3)) if m.group(3) else 0
        ampm   = (m.group(4) or "").upper()
        if ampm == "PM" and hour != 12:
            hour += 12
        elif ampm == "AM" and hour == 12:
            hour = 0
        return hour + minute / 60.0 + second / 3600.0
    return None


def _compute_gross(clock_in: Optional[float], clock_out: Optional[float]) -> float:
    """Compute gross hours from decimal-hour clock-in and clock-out values."""
    if clock_in is None or clock_out is None:
        return 0.0
    diff = clock_out - clock_in
    if diff < 0:
        diff += 24.0      # Overnight shift
    return round(diff, 4)


def _is_overnight(clock_in: Optional[float], clock_out: Optional[float]) -> bool:
    """Return True when clock_out < clock_in (shift crossed midnight)."""
    if clock_in is None or clock_out is None:
        return False
    return clock_out < clock_in


# ── Header parsing ────────────────────────────────────────────────────────────

# Matches headers like:
#   "Apr 01 - Clock In"   "April 1 - Clock Out"   "Apr 01 - Clock In"
_DATE_HEADER = re.compile(
    r"^([A-Za-z]+)\s+(\d{1,2})\s*[-–]\s*(Clock\s+In|Clock\s+Out)\s*$",
    re.IGNORECASE,
)

_MONTH_ABBREVS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _parse_date_header(header: str, year_hint: int) -> Optional[Tuple[date, str]]:
    """
    Parse a column header like 'Apr 01 - Clock In'.
    Returns (date_obj, 'in'|'out') or None if not a date header.
    """
    m = _DATE_HEADER.match((header or "").strip())
    if not m:
        return None
    month_str = m.group(1)[:3].lower()
    month_num  = _MONTH_ABBREVS.get(month_str)
    if month_num is None:
        return None
    day_num  = int(m.group(2))
    direction = "in" if "in" in m.group(3).lower() else "out"
    try:
        d = date(year_hint, month_num, day_num)
    except ValueError:
        return None
    return d, direction


def _find_header_row(ws) -> Optional[int]:
    """Return the 1-based row index of the header row (contains 'Staff Name')."""
    for r in range(1, min(ws.max_row + 1, 10)):
        val = ws.cell(r, 1).value
        if isinstance(val, str) and "staff" in val.lower():
            return r
    return None


def _extract_year_from_sheet(sheet_name: str) -> int:
    """Extract year from sheet name like 'Live - April 2026'."""
    m = re.search(r"\b(20\d{2})\b", sheet_name)
    return int(m.group(1)) if m else date.today().year


def _month_variants(month_str: str) -> list:
    """
    Return a list of strings to try when matching a month label.
    Handles full names ('April'), abbreviations ('Apr'), and combined
    'Month YYYY' → ['April 2026', 'Apr 2026', 'april', 'apr'] etc.
    """
    import calendar as _cal
    variants = [month_str.lower().strip()]
    parts = month_str.strip().split()
    if parts:
        month_word = parts[0]
        # Try to expand abbreviation → full name and vice-versa
        for i in range(1, 13):
            full = _cal.month_name[i].lower()
            abbr = _cal.month_abbr[i].lower()
            if month_word.lower() == full:
                # add abbreviated version
                alt = " ".join([_cal.month_abbr[i]] + parts[1:])
                variants.append(alt.lower())
                variants.append(abbr)
                break
            elif month_word.lower() == abbr:
                # add full-name version
                alt = " ".join([_cal.month_name[i]] + parts[1:])
                variants.append(alt.lower())
                variants.append(full)
                break
    return list(dict.fromkeys(variants))  # preserve order, deduplicate


def _find_assignments_sheet(wb, month_label: str):
    """
    Locate the 'Assignments - {Month YYYY}' worksheet that matches month_label.
    Returns the worksheet object, or None if not found (non-fatal).
    """
    pattern = re.compile(r"^Assignments\s*[-–]\s*(.+)$", re.IGNORECASE)
    candidates = [
        (name, m.group(1).strip())
        for name in wb.sheetnames
        if (m := pattern.match(name.strip()))
    ]
    if not candidates or not month_label:
        return None
    variants = _month_variants(month_label)
    if re.search(r"\d{4}", month_label):
        variants = [v for v in variants if re.search(r"\d{4}", v)]
    for variant in variants:
        for name, label in candidates:
            if variant in label.lower():
                return wb[name]
    return None


def _load_assignments(ws) -> Dict[Tuple[date, str], str]:
    """
    Parse an Assignments sheet and return a per-day project lookup:

        {(date, staff_name): project_label}

    Two sources are merged (overrides win over the main table):
      1. Main auto-fill table  — wide form: staff rows × date columns
      2. Overrides sub-table   — rows: [Override Staff Name, Override Date,
                                         Override Project, Notes, ...]
    """
    import datetime as dt

    lookup: Dict[Tuple[date, str], str] = {}
    if ws is None:
        return lookup

    max_row = ws.max_row
    max_col = ws.max_column

    # ── Locate the header row and the overrides section ───────────────────────
    hdr_row: Optional[int] = None
    overrides_start: Optional[int] = None

    for r in range(1, min(max_row + 1, 20)):
        val = ws.cell(r, 1).value
        if val is None:
            continue
        val_s = str(val).strip().lower()
        if hdr_row is None and ("staff name" in val_s or ("staff" in val_s and "name" in val_s)):
            hdr_row = r
        elif "override" in val_s and overrides_start is None:
            overrides_start = r
            break

    if hdr_row is None:
        return lookup

    # ── Build date-column index from the header row ───────────────────────────
    date_cols: Dict[int, date] = {}
    for c in range(1, max_col + 1):
        val = ws.cell(hdr_row, c).value
        if isinstance(val, dt.datetime):
            date_cols[c] = val.date()
        elif isinstance(val, dt.date):
            date_cols[c] = val

    if not date_cols:
        return lookup

    # ── Parse the main auto-fill table ───────────────────────────────────────
    main_end = (overrides_start - 1) if overrides_start else max_row
    for r in range(hdr_row + 1, main_end + 1):
        staff_val = ws.cell(r, 1).value
        if not staff_val or str(staff_val).strip() in ("", "None", "0"):
            continue
        if isinstance(staff_val, (int, float)):
            continue
        staff_name = str(staff_val).strip()
        for c, d in date_cols.items():
            proj_val = ws.cell(r, c).value
            if proj_val is None or str(proj_val).strip() in ("", "0"):
                continue
            lookup[(d, staff_name)] = str(proj_val).strip()

    # ── Parse the Overrides sub-table ─────────────────────────────────────────
    if overrides_start is None:
        return lookup

    # Find the overrides header row (directly below the section title)
    override_hdr: Optional[int] = None
    for r in range(overrides_start, min(overrides_start + 5, max_row + 1)):
        val = ws.cell(r, 1).value
        if val and "override staff" in str(val).strip().lower():
            override_hdr = r
            break
        if val and "staff" in str(val).strip().lower() and r != overrides_start:
            override_hdr = r
            break

    if override_hdr is None:
        return lookup

    for r in range(override_hdr + 1, max_row + 1):
        staff_val   = ws.cell(r, 1).value
        date_val    = ws.cell(r, 2).value
        project_val = ws.cell(r, 3).value
        if not staff_val or not date_val or not project_val:
            continue
        staff_name = str(staff_val).strip()
        project    = str(project_val).strip()
        if isinstance(date_val, dt.datetime):
            override_date = date_val.date()
        elif isinstance(date_val, dt.date):
            override_date = date_val
        else:
            continue
        lookup[(override_date, staff_name)] = project

    return lookup


def _load_overrides_only(ws) -> Dict[Tuple[date, str], str]:
    """Parse only the Overrides sub-table of an Assignments sheet.

    Returns ``{(date, staff_name): override_project}`` for reviewed overrides,
    ignoring the main auto-fill table. Used where a reviewed override must
    outrank a resolved worked-project value.
    """
    import datetime as dt

    lookup: Dict[Tuple[date, str], str] = {}
    if ws is None:
        return lookup

    max_row = ws.max_row

    overrides_start: Optional[int] = None
    for r in range(1, max_row + 1):
        val = ws.cell(r, 1).value
        if val is None:
            continue
        if "override" in str(val).strip().lower():
            overrides_start = r
            break

    if overrides_start is None:
        return lookup

    override_hdr: Optional[int] = None
    for r in range(overrides_start, min(overrides_start + 5, max_row + 1)):
        val = ws.cell(r, 1).value
        if val and "override staff" in str(val).strip().lower():
            override_hdr = r
            break
        if val and "staff" in str(val).strip().lower() and r != overrides_start:
            override_hdr = r
            break

    if override_hdr is None:
        return lookup

    for r in range(override_hdr + 1, max_row + 1):
        staff_val   = ws.cell(r, 1).value
        date_val    = ws.cell(r, 2).value
        project_val = ws.cell(r, 3).value
        if not staff_val or not date_val or not project_val:
            continue
        staff_name = str(staff_val).strip()
        project    = str(project_val).strip()
        if isinstance(date_val, dt.datetime):
            override_date = date_val.date()
        elif isinstance(date_val, dt.date):
            override_date = date_val
        else:
            continue
        lookup[(override_date, staff_name)] = project

    return lookup


def _find_live_sheet(wb, target_month: Optional[str] = None):
    """
    Locate the 'Live - {Month YYYY}' worksheet.
    If target_month is given (e.g. 'April 2026'), prefer matching sheet.
    Handles both full ('April') and abbreviated ('Apr') month names.
    Falls back to the most recently dated Live sheet when no target given.
    """
    pattern = re.compile(r"^Live\s*[-–]\s*(.+)$", re.IGNORECASE)
    candidates = [
        (name, m.group(1).strip())
        for name in wb.sheetnames
        if (m := pattern.match(name.strip()))
    ]

    if not candidates:
        raise RosterParseError(
            f"No 'Live - {{Month YYYY}}' sheet found. "
            f"Available sheets: {wb.sheetnames}"
        )

    if target_month:
        variants = _month_variants(target_month)
        # When the target includes a 4-digit year, restrict matching to
        # year-qualified variants only.  Without this, a bare abbreviation
        # like "may" would match "Live - May 2025" when the caller wants
        # "May 2026", silently returning wrong-year data instead of raising.
        if re.search(r'\d{4}', target_month):
            variants = [v for v in variants if re.search(r'\d{4}', v)]
        for variant in variants:
            for name, label in candidates:
                if variant in label.lower():
                    return wb[name]
        raise RosterParseError(
            f"No Live sheet matching '{target_month}'. "
            f"Found: {[c[0] for c in candidates]}"
        )

    # Return the last candidate (most recent)
    return wb[candidates[-1][0]]


# ── Public API ────────────────────────────────────────────────────────────────

def parse_roster(
    path: str,
    target_month: Optional[str] = None,
    target_week_start: Optional[date] = None,
    target_week_end: Optional[date] = None,
    malformed_out: Optional[List[str]] = None,
    overnight_out: Optional[List[Dict[str, Any]]] = None,
    long_shift_threshold_hours: float = 12.0,
) -> List[Dict[str, Any]]:
    """
    Parse the wide-form Roster Log and return a list of attendance records.

    Parameters
    ----------
    path              : path to the .xlsx file
    target_month      : e.g. 'April 2026' — selects the matching Live sheet
    target_week_start : optional date filter (inclusive)
    target_week_end   : optional date filter (inclusive)
    malformed_out     : if provided, malformed clock pairs are appended here
                        and parsing continues (collect-and-warn mode).
                        If None (default), a RosterParseError is raised on
                        the first malformed pair (strict mode).
    overnight_out     : if provided, dicts describing overnight-shift records
                        are appended here as they are encountered.  Each dict
                        has the same keys as a normal record plus a human-
                        readable 'note' string.  Overnight records are still
                        included in the normal returned list.
    long_shift_threshold_hours
                      : gross hours above this threshold are flagged with
                        long_shift=True for extra review.

    Returns
    -------
    List of dicts with keys:
        staff, project, date, clock_in, clock_out,
        gross_hours, lunch_deduction, net_hours, long_shift

    Overnight records are included with correct gross hours (24h added), and
    are also reported via overnight_out when that parameter is provided.
    """
    try:
        import openpyxl
    except ImportError:
        raise RosterParseError("openpyxl is required: pip install openpyxl")

    from pathlib import Path
    p = Path(path)
    if not p.exists():
        raise RosterParseError(f"Roster file not found: {path}")

    try:
        wb = openpyxl.load_workbook(str(p), data_only=True)
    except Exception as exc:
        raise RosterParseError(f"Cannot open workbook '{path}': {exc}")

    # When target_week spans two calendar months, we need to parse both Live
    # sheets and merge the results so cross-month weeks (e.g. Apr 27–May 1)
    # include all days from both months.
    sheets_to_parse = []
    if (
        target_week_start is not None
        and target_week_end is not None
        and target_week_start.month != target_week_end.month
    ):
        # Build month labels for both months and collect matching sheets.
        import calendar as _cal
        start_label = f"{_cal.month_name[target_week_start.month]} {target_week_start.year}"
        end_label   = f"{_cal.month_name[target_week_end.month]}   {target_week_end.year}"
        # Both month sheets are REQUIRED for a cross-month week.  Silently
        # skipping a missing sheet would produce a partial report that omits
        # whole calendar days — that is worse than a clear error.
        missing_months = []
        for label in (start_label, end_label):
            try:
                ws_ = _find_live_sheet(wb, label.strip())
                sheets_to_parse.append(ws_)
            except RosterParseError:
                missing_months.append(label.strip())
        if missing_months:
            raise RosterParseError(
                f"Cross-month week {target_week_start} – {target_week_end} "
                f"requires roster sheets for both months, but the following "
                f"Live sheet(s) are missing from the workbook: "
                + ", ".join(f"'Live - {m}'" for m in missing_months)
                + ". Add the missing sheet(s) or choose a single-month week."
            )
    else:
        sheets_to_parse.append(_find_live_sheet(wb, target_month))

    # ── Load per-day project assignments (non-fatal if sheet absent) ───────────
    # Build a combined lookup from all relevant Assignments sheets so that
    # cross-month weeks also get the right project labels for each half.
    assignments_lookup: Dict[Tuple[date, str], str] = {}
    for live_ws in sheets_to_parse:
        month_label_for_ws = live_ws.title.split("-", 1)[-1].strip()
        asn_ws = _find_assignments_sheet(wb, month_label_for_ws)
        if asn_ws is not None:
            assignments_lookup.update(_load_assignments(asn_ws))

    # ── Parse each sheet (usually one; two for cross-month weeks) ──────────────
    records: List[Dict[str, Any]] = []
    seen_keys: set = set()   # (staff, date) — deduplicate if sheets overlap

    def _parse_sheet(ws) -> None:
        year_hint = _extract_year_from_sheet(ws.title)

        hdr_row = _find_header_row(ws)
        if hdr_row is None:
            raise RosterParseError(
                f"Sheet '{ws.title}': cannot find header row with 'Staff Name'. "
                f"Scanned first 10 rows."
            )

        headers = [ws.cell(hdr_row, c).value for c in range(1, ws.max_column + 1)]

        idx_staff   = None
        idx_project = None
        for i, h in enumerate(headers):
            h_str = str(h).lower().strip() if h else ""
            if "staff" in h_str or "name" in h_str:
                idx_staff = i
            elif "project" in h_str or "team" in h_str or "bucket" in h_str:
                idx_project = i

        if idx_staff is None:
            raise RosterParseError(
                f"Sheet '{ws.title}': 'Staff Name' column not found. "
                f"Headers: {headers[:10]}"
            )

        date_cols: Dict[int, Tuple[date, str]] = {}
        for i, h in enumerate(headers):
            if h is None:
                continue
            result = _parse_date_header(str(h), year_hint)
            if result:
                date_cols[i] = result

        if not date_cols:
            raise RosterParseError(
                f"Sheet '{ws.title}': no date columns found. "
                "Expected headers like 'Apr 01 - Clock In'. "
                f"Found headers: {headers[:10]}"
            )

        date_to_cols: Dict[date, Dict[str, int]] = {}
        for col_idx, (d, direction) in date_cols.items():
            date_to_cols.setdefault(d, {})[direction] = col_idx

        for r in range(hdr_row + 1, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]

            staff_val = row[idx_staff] if idx_staff < len(row) else None
            if not staff_val or str(staff_val).strip() in ("", "None", "0"):
                continue
            if isinstance(staff_val, (int, float)):
                continue

            live_project = (
                str(row[idx_project]).strip()
                if idx_project is not None and idx_project < len(row)
                and row[idx_project] is not None
                else ""
            )
            if live_project in ("0", ""):
                live_project = ""

            staff_name = str(staff_val).strip()

            for record_date, dirs in sorted(date_to_cols.items()):
                if target_week_start and record_date < target_week_start:
                    continue
                if target_week_end and record_date > target_week_end:
                    continue

                # Resolve project: per-day Assignments sheet entry wins over
                # the Live sheet's default project column.
                project_val = assignments_lookup.get(
                    (record_date, staff_name), live_project
                )

                # Key includes project so that staff with multiple project
                # assignments on the same date each produce a separate record.
                # Dedup is only to prevent the exact same (staff, date, project)
                # row from appearing twice when two sheets share overlapping rows.
                key = (staff_name, record_date, project_val)
                if key in seen_keys:
                    continue  # exact duplicate row from overlapping sheets

                in_col  = dirs.get("in")
                out_col = dirs.get("out")

                in_val  = row[in_col]  if in_col  is not None and in_col  < len(row) else None
                out_val = row[out_col] if out_col is not None and out_col < len(row) else None

                clock_in  = _time_to_hours(in_val)
                clock_out = _time_to_hours(out_val)

                if clock_in is None and clock_out is None:
                    continue

                if clock_in is None or clock_out is None:
                    missing = "Clock In" if clock_in is None else "Clock Out"
                    msg = (
                        f"Malformed row for '{staff_name}' on "
                        f"{record_date.isoformat()}: {missing} is blank while "
                        f"the other time is present — row excluded."
                    )
                    if malformed_out is not None:
                        malformed_out.append(msg)
                        continue
                    raise RosterParseError(msg)

                gross        = _compute_gross(clock_in, clock_out)
                is_overnight = _is_overnight(clock_in, clock_out)
                is_long_shift = gross > long_shift_threshold_hours
                lunch = _lunch_deduction(gross)
                net   = round(max(0.0, gross - lunch), 4)

                seen_keys.add(key)
                rec = {
                    "staff":           staff_name,
                    "project":         project_val,
                    "date":            record_date,
                    "clock_in":        clock_in,
                    "clock_out":       clock_out,
                    "gross_hours":     round(gross, 4),
                    "lunch_deduction": lunch,
                    "net_hours":       net,
                    "long_shift":      is_long_shift,
                }
                records.append(rec)

                if is_overnight and overnight_out is not None:
                    def _fmt_h(h: float) -> str:
                        total_minutes = int(round(h * 60)) % (24 * 60)
                        hh, mm = divmod(total_minutes, 60)
                        return f"{hh:02d}:{mm:02d}"
                    overnight_out.append({
                        **rec,
                        "overnight": True,
                        "note": (
                            f"{staff_name} on {record_date.isoformat()}: "
                            f"clock-in {_fmt_h(clock_in)} → clock-out {_fmt_h(clock_out)} "
                            f"(overnight, gross {gross:.2f}h)"
                        ),
                    })

    for ws in sheets_to_parse:
        _parse_sheet(ws)

    return records


def week_bounds(target_date: Optional[date] = None) -> Tuple[date, date]:
    """
    Return (monday, friday) of the ISO week containing target_date.
    Defaults to the most recently completed Friday week.
    """
    if target_date is None:
        today = date.today()
        days_since_friday = (today.weekday() - 4) % 7
        target_date = today - timedelta(days=days_since_friday)
    monday = target_date - timedelta(days=target_date.weekday())
    friday = monday + timedelta(days=4)
    return monday, friday
