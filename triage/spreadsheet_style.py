"""Shared spreadsheet styling helpers for generated Web Excel-safe workbooks.

This module is intentionally small and boring. It centralizes the workbook
presentation palette without changing formulas, ranges, pivots, or calculation
behavior.
"""
from __future__ import annotations

import json
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, Optional

_REPO_ROOT = Path(__file__).resolve().parent.parent
_CONFIG_PATH = _REPO_ROOT / "configs" / "spreadsheet_style_v1.json"


@lru_cache(maxsize=1)
def style_config() -> Dict[str, Any]:
    """Load the global spreadsheet style contract."""
    with _CONFIG_PATH.open(encoding="utf-8") as f:
        return json.load(f)


def color(key: str) -> str:
    """Return an ARGB color from the global palette."""
    colors = style_config()["colors"]
    try:
        return colors[key]
    except KeyError as exc:
        raise KeyError(f"unknown spreadsheet style color: {key}") from exc


def tab_color_for_sheet(sheet_name: str, *, role: Optional[str] = None) -> str:
    """Return the preferred tab color for a sheet name or semantic role.

    Exact sheet-name matches win. Role fallback supports new generator sheets
    without hardcoding every future worksheet title.
    """
    cfg = style_config()
    exact = cfg.get("tab_colors", {}).get(sheet_name)
    if exact:
        return exact
    role_defaults = cfg.get("role_defaults", {})
    if role:
        role_key = f"{role}_sheet_tab"
        color_key = role_defaults.get(role_key)
        if color_key:
            return color(color_key)
    return color("table_header")


def apply_openpyxl_tab_colors(workbook) -> None:
    """Apply configured tab colors to an openpyxl workbook.

    This is a style-only helper. It must not modify formulas, values, tables,
    ranges, relationships, or calculation settings.
    """
    for name in workbook.sheetnames:
        workbook[name].sheet_properties.tabColor = tab_color_for_sheet(name)


def openpyxl_style_primitives():
    """Return reusable openpyxl style primitives.

    Importing openpyxl inside the function keeps the module cheap to import for
    package-level validation paths that do not need workbook rendering.
    """
    try:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as e:
        raise RuntimeError("openpyxl is required for workbook styling: pip install openpyxl") from e

    cfg = style_config()
    layout = cfg["layout"]
    fonts = cfg["fonts"]

    thin_grid = Side(style="thin", color=layout["border_color"])
    border = Border(left=thin_grid, right=thin_grid, top=thin_grid, bottom=thin_grid)

    return {
        "title_font": Font(name=fonts["title"], size=16, bold=True, color=color("white")),
        "section_font": Font(name=fonts["title"], size=12, bold=True, color=color("white")),
        "header_font": Font(name=fonts["body"], size=10, bold=True, color=color("white")),
        "body_font": Font(name=fonts["body"], size=10, color=color("ink")),
        "note_font": Font(name=fonts["body"], size=9, color=color("muted_ink")),
        "title_fill": PatternFill("solid", fgColor=color("header_primary")),
        "section_fill": PatternFill("solid", fgColor=color("header_secondary")),
        "table_header_fill": PatternFill("solid", fgColor=color("table_header")),
        "panel_fill": PatternFill("solid", fgColor=color("panel")),
        "body_fill": PatternFill("solid", fgColor=color("white")),
        "border": border,
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
        "wrap": Alignment(horizontal="left", vertical="top", wrap_text=True),
    }
