"""
triage/tech_hours_parser.py
---------------------------
Parse the NW_PRJ_Tech_hours workbook to extract per-tech, per-date hours.

Target sheet: "Project Team" (contains April 2026 data).

Sheet structure:
  Row ~10: [None, 'Techs ', date1, None, 'Monday', None, date2, None, 'Tuesday', ...]
  Row ~11: [None, None, 'In', 'Out', 'Total', None, 'In', 'Out', 'Total', ...]
  Row ~12+: [None, staff_name, clock_in, clock_out, total_h, None, ...]

Column layout per date group (4 columns wide):
  col+0 = date (in header row)  → In  (in data rows)
  col+1 = None/day_name        → Out
  col+2 = day_name             → Total (pre-computed; used as fallback)
  col+3 = None (spacer)

Returns a list of dicts per staff/date record (same schema as roster_parser):
  {staff, project, date, clock_in, clock_out, gross_hours,
   lunch_deduction, net_hours, source}

Raises TechHoursParseError on structural failures.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple


class TechHoursParseError(Exception):
    pass


# ── Time helpers (identical policy to roster_parser) ─────────────────────────

def _time_to_hours(value: Any) -> Optional[float]:
    import datetime as dt
    if value is None:
        return None
    if isinstance(value, dt.time):
        return value.hour + value.minute / 60.0 + value.second / 3600.0
    if isinstance(value, dt.datetime):
        return value.hour + value.minute / 60.0 + value.second / 3600.0
    if isinstance(value, dt.timedelta):
        return value.total_seconds() / 3600.0
    if isinstance(value, (int, float)):
        f = float(value)
        if 0.0 <= f < 2.0:          # openpyxl time fraction
            return f * 24.0
        if 0.0 <= f <= 24.0:        # already decimal hours
            return f
        return None
    return None


def _lunch_deduction(gross: float) -> float:
    if gross >= 8.0:
        return 1.0
    if gross >= 6.0:
        return 0.5
    return 0.0


def _find_data_header_row(ws) -> Optional[int]:
    """Find the row that contains 'Techs' (the column-group header row)."""
    for r in range(1, min(ws.max_row + 1, 30)):
        for c in range(1, min(ws.max_column + 1, 10)):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower().startswith("tech"):
                return r
    return None


def _parse_date_columns(ws, hdr_row: int) -> Dict[int, date]:
    """
    Return {col_index: date} for all date columns in the header row.
    Date columns contain datetime objects.
    """
    result: Dict[int, date] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hdr_row, c).value
        if isinstance(v, datetime):
            result[c] = v.date()
        elif isinstance(v, date):
            result[c] = v
    return result


def parse_tech_hours(
    path: str,
    sheet_name: str = "Project Team",
    project: str = "Projects Team",
    target_month: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Parse the Tech Hours workbook and return attendance records.

    Parameters
    ----------
    path         : path to the .xlsx file
    sheet_name   : sheet to parse (default 'Project Team')
    project      : project name to assign to all records
    target_month : optional 'YYYY-MM' filter; skips dates outside this month

    Returns
    -------
    List of record dicts with same schema as roster_parser.parse_roster().
    Raises TechHoursParseError on structural failures.
    """
    try:
        import openpyxl
    except ImportError:
        raise TechHoursParseError("openpyxl is required: pip install openpyxl")

    from pathlib import Path
    p = Path(path)
    if not p.exists():
        raise TechHoursParseError(f"Tech Hours file not found: {path}")

    try:
        wb = openpyxl.load_workbook(str(p), data_only=True)
    except Exception as exc:
        raise TechHoursParseError(f"Cannot open Tech Hours workbook '{path}': {exc}")

    if sheet_name not in wb.sheetnames:
        available = [s for s in wb.sheetnames
                     if "project" in s.lower() or "tech" in s.lower()]
        raise TechHoursParseError(
            f"Sheet '{sheet_name}' not found. "
            f"Project-related sheets: {available or wb.sheetnames[:5]}"
        )

    ws = wb[sheet_name]

    # Locate header row
    hdr_row = _find_data_header_row(ws)
    if hdr_row is None:
        raise TechHoursParseError(
            f"Sheet '{sheet_name}': cannot find 'Techs' header row."
        )

    # Parse date columns from header row
    date_cols = _parse_date_columns(ws, hdr_row)
    if not date_cols:
        raise TechHoursParseError(
            f"Sheet '{sheet_name}': no date columns found in row {hdr_row}."
        )

    # Find staff name column (contains "Techs" in the header row)
    staff_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hdr_row, c).value
        if isinstance(v, str) and v.strip().lower().startswith("tech"):
            staff_col = c
            break
    if staff_col is None:
        raise TechHoursParseError(
            f"Sheet '{sheet_name}': cannot find staff name column."
        )

    # The In/Out/Total columns follow each date column in groups of 3.
    # For a date at col D: In=D, Out=D+1, Total=D+2
    # (verified from actual structure: date col, then In, Out, Total in same cols)
    # But the date col in row 10 is different from what row 11 says:
    # row 10: col 3 = date, col 5 = day_name, col 6 = None
    # row 11: col 3 = 'In', col 4 = 'Out', col 5 = 'Total'
    # So In/Out/Total are AT the date column and date+1 and date+2.
    # Total is at date_col + 2 in row 11.

    # Build {date: (in_col, out_col, total_col)} mapping
    date_col_groups: Dict[date, Tuple[int, int, int]] = {}
    for dc, d in date_cols.items():
        date_col_groups[d] = (dc, dc + 1, dc + 2)

    # Month filter
    month_filter: Optional[Tuple[int, int]] = None
    if target_month:
        try:
            year, month = target_month.split("-")
            month_filter = (int(year), int(month))
        except (ValueError, AttributeError):
            pass

    # Parse data rows (below the In/Out/Total sub-header row)
    sub_hdr_row = hdr_row + 1  # row with 'In', 'Out', 'Total'
    records: List[Dict[str, Any]] = []

    for r in range(sub_hdr_row + 1, ws.max_row + 1):
        staff_val = ws.cell(r, staff_col).value
        if not staff_val or not isinstance(staff_val, str):
            continue
        staff_name = staff_val.strip()
        _sn_lower = staff_name.lower()
        if (not staff_name
                or _sn_lower in ("", "none", "total", "totals")
                or _sn_lower.startswith("tech total")):
            continue

        for record_date, (in_c, out_c, tot_c) in sorted(date_col_groups.items()):
            # Apply month filter
            if month_filter:
                if (record_date.year, record_date.month) != month_filter:
                    continue

            in_val  = ws.cell(r, in_c).value
            out_val = ws.cell(r, out_c).value
            tot_val = ws.cell(r, tot_c).value  # fallback

            clock_in  = _time_to_hours(in_val)
            clock_out = _time_to_hours(out_val)

            # Skip days with no data
            if clock_in is None and clock_out is None:
                # Try total as fallback
                tot = _time_to_hours(tot_val)
                if tot is None or tot <= 0:
                    continue
                # Only total available — record as synthetic
                gross = float(tot)
                lunch = _lunch_deduction(gross)
                net   = round(max(0.0, gross - lunch), 4)
                records.append({
                    "staff":           staff_name,
                    "project":         project,
                    "date":            record_date,
                    "clock_in":        None,
                    "clock_out":       None,
                    "gross_hours":     round(gross, 4),
                    "lunch_deduction": lunch,
                    "net_hours":       net,
                    "source":          "tech_hours_total_only",
                })
                continue

            # Compute from In/Out
            if clock_in is not None and clock_out is not None:
                diff = clock_out - clock_in
                if diff < 0:
                    diff += 24.0
                gross = round(diff, 4)
            elif _time_to_hours(tot_val) is not None and _time_to_hours(tot_val) > 0:
                gross = float(_time_to_hours(tot_val))
            else:
                continue

            if gross <= 0:
                continue

            lunch = _lunch_deduction(gross)
            net   = round(max(0.0, gross - lunch), 4)

            records.append({
                "staff":           staff_name,
                "project":         project,
                "date":            record_date,
                "clock_in":        clock_in,
                "clock_out":       clock_out,
                "gross_hours":     round(gross, 4),
                "lunch_deduction": lunch,
                "net_hours":       net,
                "source":          "tech_hours",
            })

    return records
