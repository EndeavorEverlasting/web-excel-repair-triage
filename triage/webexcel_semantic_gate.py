"""Shared semantic integrity gate for Web Excel workbook artifacts.

Detects sharedStrings collapse (Column1..ColumnN pattern), validates sentinel
cells across billing and Bonita profiles, scores semantic density, and checks
repair preservation. Called from:

    triage/admin_billing_summary/preflight.py
    triage/nw_prj_neuron_track_hours/bonita_cli.py  (preflight_bonita)
    triage/nw_prj_neuron_track_hours/preflight.py   (run_preflight)

Profiles
--------
"admin_billing"  — Admin Billing Summary (Internal / Client variants)
"bonita"         — Bonita two-tab Neuron Track Hours workbook
"""
from __future__ import annotations

import io
import re
import shutil
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional

_SS_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_GENERIC_RE = re.compile(r"^Column\d+$")

# ───────────────────────── sentinel profiles ──────────────────────────────────
#
# Each entry: (sheet_name_fragment, row, col_1based, check, expected_value)
#   check:  "nonblank" | "equals" | "contains"
#   Sheet is matched by case-insensitive substring search against tab names.

ADMIN_BILLING_SENTINELS: List = [
    ("Start Here",          1, 1, "nonblank",  None),
    ("Executive Dashboard", 1, 1, "nonblank",  None),
    ("Monthly Summary",     5, 1, "equals",    "Month"),
    ("Project Summary",     5, 1, "equals",    "Month"),
    ("Tech Summary",        5, 1, "equals",    "Month"),
]

# Bonita sentinels are checked dynamically (see _check_bonita_sentinels).
BONITA_SENTINELS: List = []

# Neuron Track Hours dashboard sentinels are checked dynamically
# (see _check_neuron_track_sentinels). The "* Neuron Hours" tabs carry their
# real headers on row 4 (APRIL_MAY_COLUMNS), so "Month"/"Tech" survival proves
# the sharedStrings did not collapse to ColumnN.
NEURON_TRACK_SENTINELS: List = [
    ("Start Here", 1, 1, "nonblank", None),
]

_META_TABS = frozenset({"CF Dictionary", "CF_Dictionary", "WebExcel QC", "Review Flags"})


# ───────────────────────── XML-aware sharedStrings extraction ─────────────────


def extract_shared_strings(xml_bytes: bytes) -> List[str]:
    """Extract all shared strings from sharedStrings.xml bytes.

    Handles both plain ``<si><t>text</t></si>`` and rich-text
    ``<si><r><t>text</t></r></si>`` nodes by gathering all ``<t>``
    descendants under each ``<si>``.  Falls back to regex on broken XML.
    """
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        text = xml_bytes.decode("utf-8", errors="ignore")
        return re.findall(r"<t[^>]*>(.*?)</t>", text, re.DOTALL)

    sis = root.findall(f"{{{_SS_NS}}}si")
    if not sis:
        sis = root.findall("si")

    results: List[str] = []
    for si in sis:
        ts = si.findall(f".//{{{_SS_NS}}}t")
        if not ts:
            ts = si.findall(".//t")
        text = "".join((t.text or "") for t in ts)
        results.append(text)
    return results


# ───────────────────────── semantic density scoring ───────────────────────────


def _classify_strings(strings: List[str]) -> Dict[str, Any]:
    total = len(strings)
    generic = [s for s in strings if _GENERIC_RE.match(s.strip())]
    meaningful = total - len(generic)
    ratio = meaningful / total if total > 0 else 1.0
    return {
        "shared_string_count": total,
        "generic_column_string_count": len(generic),
        "meaningful_shared_string_count": meaningful,
        "meaningful_shared_string_ratio": round(ratio, 4),
        "generic_column_strings_only": bool(total > 0 and meaningful == 0),
    }


# ───────────────────────── sentinel cell checks ───────────────────────────────


def _col_letter(col: int) -> str:
    letters = ""
    c = col
    while c:
        c, r = divmod(c - 1, 26)
        letters = chr(65 + r) + letters
    return letters


def _rc(row: int, col: int) -> str:
    return f"{_col_letter(col)}{row}"


def _check_admin_billing_sentinels(path: str, tabs: List[str]) -> List[str]:
    """Return list of sentinel failure descriptions (empty list = pass)."""
    try:
        import openpyxl
    except ImportError:
        return ["openpyxl_not_installed"]

    failures: List[str] = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return [f"workbook_load_error:{exc}"]

    try:
        for sheet_frag, row, col, check, expected in ADMIN_BILLING_SENTINELS:
            matched = next(
                (n for n in wb.sheetnames if sheet_frag.lower() in n.lower()), None
            )
            if matched is None:
                failures.append(f"missing_sheet:{sheet_frag}")
                continue
            val = wb[matched].cell(row=row, column=col).value
            addr = f"{matched}!{_rc(row, col)}"
            if check == "nonblank":
                if val is None or str(val).strip() == "":
                    failures.append(f"{addr} is blank")
            elif check == "equals":
                if str(val or "").strip() != expected:
                    failures.append(f"{addr} expected '{expected}', got '{val}'")
            elif check == "contains":
                if expected and expected.lower() not in str(val or "").lower():
                    failures.append(f"{addr} does not contain '{expected}'")

        # Dynamic: any tab containing "neuron hours" should have A5 == "Month"
        neuron_tab = next(
            (n for n in wb.sheetnames if "neuron hours" in n.lower()), None
        )
        if neuron_tab:
            val = wb[neuron_tab].cell(row=5, column=1).value
            if str(val or "").strip() != "Month":
                failures.append(
                    f"{neuron_tab}!A5 expected 'Month', got '{val}'"
                )
    finally:
        wb.close()

    return failures


def _check_bonita_sentinels(path: str, tabs: List[str]) -> List[str]:
    """Return list of sentinel failure descriptions for Bonita workbooks."""
    try:
        import openpyxl
    except ImportError:
        return ["openpyxl_not_installed"]

    data_tabs = [t for t in tabs if t not in _META_TABS]
    failures: List[str] = []
    if len(data_tabs) < 2:
        failures.append(f"expected_at_least_2_data_tabs:got_{len(data_tabs)}")

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return failures + [f"workbook_load_error:{exc}"]

    try:
        for tab in data_tabs:
            if tab not in wb.sheetnames:
                continue
            ws = wb[tab]
            # Bonita layout: col A = date (blank header); col B = TECH
            header = ws.cell(row=1, column=2).value
            if header is None or str(header).strip() == "":
                failures.append(f"{tab}!B1 header is blank")
            found_tech = False
            for ws_row in ws.iter_rows(
                min_row=3, max_row=20, min_col=2, max_col=2, values_only=True
            ):
                if ws_row[0] is not None and str(ws_row[0]).strip():
                    found_tech = True
                    break
            if not found_tech:
                failures.append(
                    f"{tab}: no nonblank tech name in rows 3-20 col B"
                )
    finally:
        wb.close()

    return failures


def _check_neuron_track_sentinels(path: str, tabs: List[str]) -> List[str]:
    """Return sentinel failures for the full NTH dashboard workbook.

    Structural proof that the dashboard kept real content after any Web Excel
    round-trip (rather than only scoring string density):

      * "Start Here" title cell (A1) is non-blank.
      * At least one "* Neuron Hours" month tab is present.
      * The Neuron Hours header row (row 4) still carries the real "Month" and
        "Tech" headers (not collapsed to ColumnN).
    """
    try:
        import openpyxl
    except ImportError:
        return ["openpyxl_not_installed"]

    failures: List[str] = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return [f"workbook_load_error:{exc}"]

    try:
        start = next((n for n in wb.sheetnames if "start here" in n.lower()), None)
        if start is None:
            failures.append("missing_sheet:Start Here")
        else:
            val = wb[start].cell(row=1, column=1).value
            if val is None or str(val).strip() == "":
                failures.append(f"{start}!A1 is blank")

        neuron_tabs = [n for n in wb.sheetnames if "neuron hours" in n.lower()]
        if not neuron_tabs:
            failures.append("missing_tab:* Neuron Hours")
        else:
            ws = wb[neuron_tabs[0]]
            header = [
                str(ws.cell(row=4, column=c).value or "").strip()
                for c in range(1, ws.max_column + 1)
            ]
            for required in ("Month", "Tech"):
                if required not in header:
                    failures.append(
                        f"{neuron_tabs[0]}!row4 header missing '{required}'"
                    )
    finally:
        wb.close()

    return failures


# ───────────────────────── repair preservation ────────────────────────────────


def _snapshot_sentinel_texts(path: str, profile: str) -> Dict[str, str]:
    """Return ``{cell_addr: text_value}`` for sentinel cells of the given profile."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return {}

    result: Dict[str, str] = {}
    try:
        if profile == "admin_billing":
            for sheet_frag, row, col, _, _ in ADMIN_BILLING_SENTINELS:
                matched = next(
                    (n for n in wb.sheetnames if sheet_frag.lower() in n.lower()), None
                )
                if matched:
                    val = wb[matched].cell(row=row, column=col).value
                    result[f"{matched}!{_rc(row, col)}"] = str(val or "")
        elif profile == "bonita":
            for tab in wb.sheetnames[:4]:
                # Bonita layout: col A = date (blank header); col B = TECH
                val = wb[tab].cell(row=1, column=2).value
                result[f"{tab}!B1"] = str(val or "")
    finally:
        wb.close()
    return result


def check_repair_preservation(path: str, profile: str) -> bool:
    """Return True if ``fix_inlinestr`` mutates sentinel cell text (text loss).

    Copies the workbook to a temp path, runs ``fix_inlinestr``, reloads, and
    compares sentinel cell values before and after.  Returns False on a clean
    no-op (expected for well-formed workbooks).
    """
    from triage.xlsx_utils import fix_inlinestr

    p = Path(path)
    tmp = None
    try:
        fd, tmp = tempfile.mkstemp(suffix=".xlsx")
        import os
        os.close(fd)
        shutil.copy2(str(p), tmp)
        before = _snapshot_sentinel_texts(path, profile)
        fix_inlinestr(tmp)
        after = _snapshot_sentinel_texts(tmp, profile)
        if before != after:
            return True
        for key, bval in before.items():
            aval = after.get(key, "")
            if bval.strip() and _GENERIC_RE.match(aval.strip()):
                return True
        return False
    except Exception:
        return False
    finally:
        if tmp:
            try:
                Path(tmp).unlink(missing_ok=True)
            except Exception:
                pass


# ───────────────────────── main gate entrypoint ───────────────────────────────


def run_semantic_gate(path: str, profile: str = "admin_billing") -> Dict[str, Any]:
    """Run the full semantic integrity gate on a workbook.

    Parameters
    ----------
    path:
        Absolute or relative path to the ``.xlsx`` file.
    profile:
        ``"admin_billing"`` (Admin Billing Summary workbooks) or
        ``"bonita"`` (Bonita Neuron Track Hours two-tab workbooks).

    Returns
    -------
    dict with keys:

    ==========================================  ==============================
    ``semantic_integrity``                      ``"PASS"`` or ``"FAIL"``
    ``sentinel_failures``                       list of failure descriptions
    ``shared_string_count``                     total ``<si>`` entries
    ``generic_column_string_count``             entries matching ``Column\\d+``
    ``meaningful_shared_string_count``          total minus generic
    ``meaningful_shared_string_ratio``          0.0–1.0
    ``generic_column_strings_only``             bool — all strings are generic
    ``post_repair_text_loss``                   bool — fix_inlinestr mutates
    ``excel_for_web_manual_check``              always ``"NOT_PROVEN"``
    ==========================================  ==============================
    """
    base: Dict[str, Any] = {
        "semantic_integrity": "FAIL",
        "sentinel_failures": [],
        "shared_string_count": 0,
        "generic_column_string_count": 0,
        "meaningful_shared_string_count": 0,
        "meaningful_shared_string_ratio": 1.0,
        "generic_column_strings_only": False,
        "post_repair_text_loss": False,
        "excel_for_web_manual_check": "NOT_PROVEN",
    }

    p = Path(path)
    if not p.exists():
        base["sentinel_failures"].append("file_not_found")
        return base

    density_failures: List[str] = []
    tabs: List[str] = []

    # 1. sharedStrings semantic analysis
    try:
        with zipfile.ZipFile(path, "r") as z:
            names = z.namelist()
            if "xl/workbook.xml" in names:
                wb_text = z.read("xl/workbook.xml").decode("utf-8", errors="ignore")
                tabs = re.findall(r'<sheet[^>]*name="([^"]+)"', wb_text)
            if "xl/sharedStrings.xml" in names:
                ss_bytes = z.read("xl/sharedStrings.xml")
                strings = extract_shared_strings(ss_bytes)
                density = _classify_strings(strings)
                base.update(density)
                if density["generic_column_strings_only"]:
                    density_failures.append(
                        "generic_column_strings_only:all_shared_strings_are_ColumnN"
                    )
                if (
                    density["shared_string_count"] > 0
                    and density["meaningful_shared_string_ratio"] < 0.50
                ):
                    density_failures.append(
                        f"meaningful_ratio_below_threshold:"
                        f"{density['meaningful_shared_string_ratio']:.2f}<0.50"
                    )
    except zipfile.BadZipFile:
        base["sentinel_failures"].append("bad_zip")
        return base
    except Exception as exc:
        base["sentinel_failures"].append(f"sharedstrings_read_error:{exc}")

    # 2. Sentinel cell checks
    if profile == "admin_billing":
        sentinel_failures = _check_admin_billing_sentinels(path, tabs)
    elif profile == "bonita":
        sentinel_failures = _check_bonita_sentinels(path, tabs)
    elif profile == "neuron_track":
        sentinel_failures = _check_neuron_track_sentinels(path, tabs)
    else:
        sentinel_failures = []

    # 3. Repair preservation
    post_repair_text_loss = False
    try:
        post_repair_text_loss = check_repair_preservation(path, profile)
    except Exception:
        pass
    base["post_repair_text_loss"] = post_repair_text_loss

    all_failures: List[str] = density_failures + sentinel_failures
    if post_repair_text_loss:
        all_failures.append(
            "post_repair_text_loss:sentinel_text_mutated_by_fix_inlinestr"
        )

    base["sentinel_failures"] = all_failures
    base["semantic_integrity"] = "PASS" if not all_failures else "FAIL"
    return base
